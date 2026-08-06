"""
MSS Excel Report Engine
Version : 1.0
Sprint : 36.0
Compatible : v0.31
"""

from pathlib import Path

from openpyxl import Workbook

from mss.domain.report import Report


class ExcelReportEngine:

    def build(
        self,
        report: Report,
        output_folder="reports",
    ) -> str:

        folder = Path(output_folder)

        folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        filename = folder / "Backtest_Report.xlsx"

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Summary"

        #
        # Summary
        #

        sheet["A1"] = report.title

        sheet["A3"] = "Processed Candles"
        sheet["B3"] = report.processed_candles

        sheet["A4"] = "Generated Signals"
        sheet["B4"] = report.generated_signals

        sheet["A5"] = "Executed Trades"
        sheet["B5"] = report.executed_trades

        sheet["A6"] = "Execution Time"
        sheet["B6"] = report.execution_time

        premium_discount = report.premium_discount

        sheet["D3"] = "Premium Zone"
        sheet["E3"] = self._format_zone(premium_discount.premium_zone)

        sheet["D4"] = "Discount Zone"
        sheet["E4"] = self._format_zone(premium_discount.discount_zone)

        sheet["D5"] = "Equilibrium"
        sheet["E5"] = premium_discount.equilibrium

        sheet["D6"] = "Current Zone"
        sheet["E6"] = premium_discount.current_zone

        sheet["D7"] = "Distance to Equilibrium"
        sheet["E7"] = premium_discount.distance_to_equilibrium

        kill_zone_status = report.kill_zone_status

        sheet["D9"] = "Current Session"
        sheet["E9"] = kill_zone_status.current_session

        sheet["D10"] = "Active Kill Zone"
        sheet["E10"] = kill_zone_status.active_kill_zone

        sheet["D11"] = "Remaining Time"
        sheet["E11"] = self._format_duration(kill_zone_status.remaining_time)

        sheet["D12"] = "Kill Zone Active"
        sheet["E12"] = kill_zone_status.active

        session_bias = report.session_bias

        sheet["D14"] = "Bias Session"
        sheet["E14"] = session_bias.current_session

        sheet["D15"] = "Session Bias"
        sheet["E15"] = session_bias.bias

        sheet["D16"] = "Bias Strength"
        sheet["E16"] = session_bias.strength

        sheet["D17"] = "Bias Confidence"
        sheet["E17"] = session_bias.confidence

        news_risk = report.news_risk_status

        sheet["D19"] = "Next Economic Event"
        sheet["E19"] = news_risk.next_event

        sheet["D20"] = "Event Impact"
        sheet["E20"] = news_risk.event_impact

        sheet["D21"] = "Minutes Remaining"
        sheet["E21"] = news_risk.minutes_remaining

        sheet["D22"] = "News Trading Status"
        sheet["E22"] = news_risk.trading_status

        portfolio = report.portfolio_exposure

        sheet["D24"] = "Portfolio Exposure"
        sheet["E24"] = portfolio.portfolio_exposure

        sheet["D25"] = "Currency Exposure"
        sheet["E25"] = self._format_exposure(portfolio.currency_exposure)

        sheet["D26"] = "Asset Exposure"
        sheet["E26"] = self._format_exposure(portfolio.asset_exposure)

        sheet["D27"] = "Correlation Level"
        sheet["E27"] = portfolio.correlation_level

        sheet["D28"] = "Portfolio Risk Score"
        sheet["E28"] = portfolio.portfolio_risk_score

        sheet["D29"] = "Portfolio Risk Level"
        sheet["E29"] = portfolio.risk_level

        risk = report.risk_profile

        sheet["D31"] = "Risk Approved"
        sheet["E31"] = risk.valid

        sheet["D32"] = "Risk Amount"
        sheet["E32"] = risk.risk_amount

        sheet["D33"] = "Lot Size"
        sheet["E33"] = risk.lot_size

        sheet["D34"] = "Risk Status"
        sheet["E34"] = risk.trading_status

        sheet["D35"] = "Risk Reason"
        sheet["E35"] = risk.reason

        optimization = report.optimization_result

        sheet["D37"] = "Optimizer Valid"
        sheet["E37"] = optimization.valid

        sheet["D38"] = "Optimizer Cases"
        sheet["E38"] = optimization.total_cases

        sheet["D39"] = "Best Parameters"
        sheet["E39"] = (
            str(optimization.best_case.parameters)
            if optimization.best_case is not None
            else ""
        )

        sheet["D41"] = "Final Decision"
        sheet["E41"] = report.final_decision

        sheet["D42"] = "Decision Reason"
        sheet["E42"] = report.decision_reason

        #
        # Performance
        #

        s = report.statistics

        row = 9

        values = [

            ("Total Trades", s.total_trades),

            ("Winning Trades", s.winning_trades),

            ("Losing Trades", s.losing_trades),

            ("Breakeven Trades", s.breakeven_trades),

            ("Gross Profit", s.gross_profit),

            ("Gross Loss", s.gross_loss),

            ("Net Profit", s.net_profit),

            ("Win Rate", s.win_rate),

            ("Profit Factor", s.profit_factor),

            ("Expectancy", s.expectancy),

            ("Max Drawdown", s.max_drawdown),

        ]

        for name, value in values:

            sheet.cell(row=row, column=1).value = name

            sheet.cell(row=row, column=2).value = value

            row += 1

        #
        # Equity Curve
        #

        row += 2

        sheet.cell(row=row, column=1).value = "Equity Curve"

        row += 1

        for value in s.equity_curve:

            sheet.cell(row=row, column=1).value = value

            row += 1

        paper_sheet = workbook.create_sheet("Paper Trades")
        headers = [
            "Ticket", "Symbol", "Direction", "Volume", "Entry",
            "Stop Loss", "Take Profit", "Status", "Profit",
        ]
        paper_sheet.append(headers)

        for position in report.paper_positions:
            paper_sheet.append([
                position.ticket,
                position.symbol,
                position.direction,
                position.volume,
                position.entry_price,
                position.stop_loss,
                position.take_profit,
                position.status,
                position.profit,
            ])

        workbook.save(filename)

        return str(filename)

    @staticmethod
    def _format_zone(zone) -> str:
        if zone is None:
            return ""

        return f"{zone[0]} - {zone[1]}"

    @staticmethod
    def _format_duration(duration) -> str:
        if duration is None:
            return ""

        total_seconds = max(0, int(duration.total_seconds()))
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    @staticmethod
    def _format_exposure(exposure) -> str:
        return ", ".join(
            f"{name}: {value}"
            for name, value in sorted(exposure.items())
        )
