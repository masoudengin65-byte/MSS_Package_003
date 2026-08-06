"""Sprint 88 diagnostic-only analysis of pre-registered context combinations."""

from __future__ import annotations

import copy
import json
import math
import random
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine
from mss.analysis.feature_discovery import FeatureDiscovery
from mss.analysis.mtf_evidence_engine import MTFEvidenceEngine
from mss.analysis.shadow_score_validation import ShadowScoreValidation
from mss.analysis.smart_money_evidence_engine import SmartMoneyEvidenceEngine
from mss.domain.context_combination_result import ContextCombinationResult


class ContextCombinationAnalysis(ShadowScoreValidation):
    """Join immutable evidence and test fixed combinations without production use."""

    VERSION = "SPRINT_88_CONTEXT_COMBINATION_V1"
    SEED = 880088
    BOOTSTRAP_ITERATIONS = 2000
    PERMUTATION_ITERATIONS = 2000
    MIN_GROUP = 15
    MIN_SLICE = 8
    ALPHA = 0.05
    NOT_AVAILABLE = ContextCaptureEngine.NOT_AVAILABLE
    REQUIRED_SHEETS = (
        "Summary", "Combination Ranking", "Winner Loser Statistics",
        "Symbol Analysis", "Temporal Stability", "Sample Quality",
        "Statistical Tests", "Feature Availability", "Configuration",
        "Diagnostics", "Audit",
    )
    SOURCE_KEYS = (
        "historical", "expanded", "mtf", "smart_money",
        "feature_discovery", "shadow_validation",
    )
    COMBINATIONS = (
        ("MTF_M15_H1", "Multi-timeframe alignment", ("m15_trend", "h1_trend")),
        ("MTF_M15_H4", "Multi-timeframe alignment", ("m15_trend", "h4_trend")),
        ("MTF_M15_D1", "Multi-timeframe alignment", ("m15_trend", "d1_trend")),
        ("MTF_ALIGNMENT_TRIPLET", "Multi-timeframe alignment", ("h1_alignment", "h4_alignment", "d1_alignment")),
        ("OB_LIQUIDITY", "Smart Money", ("ob_state", "liquidity_state")),
        ("OB_LIQUIDITY_ZONE", "Smart Money", ("ob_state", "liquidity_state", "market_zone")),
        ("LIQUIDITY_ZONE_BOS", "Smart Money", ("liquidity_state", "market_zone", "bos_direction")),
        ("OB_BOS", "Smart Money", ("ob_state", "bos_direction")),
        ("STRUCTURE_SESSION", "Market context", ("structure", "session")),
        ("STRUCTURE_KILL_ZONE", "Market context", ("structure", "kill_zone")),
        ("STRUCTURE_VOLATILITY", "Market context", ("structure", "volatility_regime")),
        ("SESSION_KILL_VOLATILITY", "Market context", ("session", "kill_zone", "volatility_regime")),
        ("STRUCTURE_H1_ZONE", "Composite context", ("structure", "h1_alignment", "market_zone")),
        ("STRUCTURE_MTF_LIQUIDITY", "Composite context", ("structure", "overall_alignment", "liquidity_state")),
        ("FULL_CONTEXT_SIGNATURE", "Composite context", ("structure", "overall_alignment", "market_zone", "liquidity_state", "ob_present")),
        ("BOS_H1_LIQUIDITY", "Composite context", ("bos_direction", "h1_alignment", "liquidity_state")),
    )
    FEATURE_SOURCES = {
        "m15_trend": "MSS_MTF_Context_v1.xlsx",
        "h1_trend": "MSS_MTF_Context_v1.xlsx",
        "h4_trend": "MSS_MTF_Context_v1.xlsx",
        "d1_trend": "MSS_MTF_Context_v1.xlsx",
        "h1_alignment": "MSS_MTF_Context_v1.xlsx",
        "h4_alignment": "MSS_MTF_Context_v1.xlsx",
        "d1_alignment": "MSS_MTF_Context_v1.xlsx",
        "overall_alignment": "MSS_MTF_Context_v1.xlsx",
        "ob_state": "MSS_SmartMoney_Evidence_v1.xlsx",
        "ob_present": "MSS_SmartMoney_Evidence_v1.xlsx",
        "liquidity_state": "MSS_SmartMoney_Evidence_v1.xlsx",
        "structure": "MSS_Historical_Backtest_Context_v1.xlsx",
        "bos_direction": "MSS_Historical_Backtest_Context_v1.xlsx",
        "market_zone": "MSS_Historical_Backtest_Context_v1.xlsx",
        "session": "MSS_Historical_Backtest_Context_v1.xlsx",
        "kill_zone": "MSS_Historical_Backtest_Context_v1.xlsx",
        "volatility_regime": "Sprint 84 Feature Discovery definition over immutable relative_volatility",
    }

    def run(self, source_paths, excel_path, json_path):
        paths = {key: Path(source_paths[key]) for key in self.SOURCE_KEYS}
        outputs = {Path(excel_path).resolve(), Path(json_path).resolve()}
        if outputs & {path.resolve() for path in paths.values()}:
            raise ValueError("Outputs must not overwrite immutable input artifacts")
        before = {key: self.file_sha256(path) for key, path in paths.items()}
        records, validation, audit, prior = self.load_sources(paths)
        records_before = copy.deepcopy(records)
        result = self.analyze(records, validation, audit, prior)
        if records != records_before:
            raise RuntimeError("Joined immutable evidence was mutated")
        after = {key: self.file_sha256(path) for key, path in paths.items()}
        if before != after:
            raise RuntimeError("At least one immutable input artifact changed")
        result["diagnostics"]["input_artifacts"] = [
            {
                "source": key, "path": str(paths[key]),
                "sha256_before": before[key], "sha256_after": after[key],
                "unchanged": before[key] == after[key],
            }
            for key in self.SOURCE_KEYS
        ]
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    def load_sources(self, paths):
        historical_rows, historical_validation = ContextExpansionEngine().load(paths["historical"])
        historical = {(row["symbol"].upper(), row["trade_id"]): row for row in historical_rows}
        expanded, expanded_audit = self._load_evidence_workbook(
            paths["expanded"],
            ("MTF Context", "Order Block Context", "FVG Context", "Liquidity Context", "Session Context", "Economic Context"),
            expected_schema_count=len(ContextExpansionEngine.FIELDS),
        )
        mtf, mtf_audit = self._load_evidence_workbook(
            paths["mtf"], ("M15 Context", "H1 Context", "H4 Context", "D1 Context"),
            expected_schema_count=len(MTFEvidenceEngine.FIELDS),
        )
        smart, smart_audit = self._load_evidence_workbook(
            paths["smart_money"],
            ("Order Block Evidence", "FVG Evidence", "Liquidity Evidence"),
            expected_schema_count=len(SmartMoneyEvidenceEngine.FIELDS),
        )
        feature_prior = json.loads(paths["feature_discovery"].read_text(encoding="utf-8"))
        shadow_prior = json.loads(paths["shadow_validation"].read_text(encoding="utf-8"))
        key_sets = {
            "historical": set(historical), "expanded": set(expanded),
            "mtf": set(mtf), "smart_money": set(smart),
        }
        if any(keys != key_sets["historical"] for keys in key_sets.values()):
            details = {name: len(keys ^ key_sets["historical"]) for name, keys in key_sets.items()}
            raise ValueError(f"Evidence join-key mismatch: {details}")
        joined = []
        future_violations = []
        cross_source_conflicts = []
        for key in sorted(historical, key=lambda item: (historical[item]["source_index"], item)):
            base, expanded_row, mtf_row, smart_row = historical[key], expanded[key], mtf[key], smart[key]
            decision = base["decision_time"]
            for label, row in (("expanded", expanded_row), ("mtf", mtf_row), ("smart_money", smart_row)):
                other_time = self._time_value(row["decision_time"])
                if other_time != decision:
                    cross_source_conflicts.append({"key": key, "source": label, "field": "decision_time"})
                other_direction = str(row.get("direction", base["direction"])).upper()
                if other_direction != base["direction"]:
                    cross_source_conflicts.append({"key": key, "source": label, "field": "direction"})
            self._validate_auxiliary_times(key, decision, mtf_row, smart_row, future_violations)
            context = base["frozen_context"]
            features = {
                "m15_trend": self._available(mtf_row.get("m15_trend_direction")),
                "h1_trend": self._available(mtf_row.get("h1_trend_direction")),
                "h4_trend": self._available(mtf_row.get("h4_trend_direction")),
                "d1_trend": self._available(mtf_row.get("d1_trend_direction")),
                "h1_alignment": self._available(mtf_row.get("h1_alignment_with_m15")),
                "h4_alignment": self._available(mtf_row.get("h4_alignment_with_m15")),
                "d1_alignment": self._available(mtf_row.get("d1_alignment_with_m15")),
                "ob_state": self._available(smart_row.get("ob_mitigation_state")),
                "ob_present": self._presence(smart_row.get("ob_detected")),
                "liquidity_state": self._available(smart_row.get("liquidity_lifecycle_state")),
                "structure": self._available(context.get("structure")),
                "bos_direction": self._available(context.get("bos_direction")),
                "market_zone": self._available(context.get("current_zone")),
                "session": self._available(context.get("session")),
                "kill_zone": self._available(context.get("kill_zone")),
                "volatility_regime": FeatureDiscovery._volatility_regime(self._number(context.get("relative_volatility"))),
            }
            features["overall_alignment"] = self._overall_alignment(
                features["h1_alignment"], features["h4_alignment"], features["d1_alignment"],
            )
            joined.append({
                "source_index": base["source_index"], "trade_id": base["trade_id"],
                "symbol": base["symbol"], "direction": base["direction"],
                "status": base["status"], "profit": base["profit"],
                "decision_time": decision, "features": features,
            })
        if cross_source_conflicts:
            raise ValueError(f"Cross-source identity conflicts: {cross_source_conflicts[:5]}")
        if future_violations:
            raise ValueError(f"Future-data contamination: {future_violations[:5]}")
        prior = self._validate_prior_results(feature_prior, shadow_prior)
        validation = {
            **historical_validation,
            "joined_trade_count": len(joined),
            "joined_unique_key_count": len({(row["symbol"], row["trade_id"]) for row in joined}),
            "source_key_counts": {name: len(keys) for name, keys in key_sets.items()},
            "join_key_mismatch_count": 0, "cross_source_conflict_count": 0,
            "future_data_violation_count": 0,
            "source_context_field_count": len(ContextCaptureEngine.FIELDS),
            "expanded_context_field_count": len(ContextExpansionEngine.FIELDS),
            "mtf_context_field_count": len(MTFEvidenceEngine.FIELDS),
            "smart_money_context_field_count": len(SmartMoneyEvidenceEngine.FIELDS),
        }
        audit = {
            "historical": historical_validation,
            "expanded": expanded_audit, "mtf": mtf_audit,
            "smart_money": smart_audit,
            "join_key": ["symbol", "trade_id"],
        }
        return joined, validation, audit, prior

    def _load_evidence_workbook(self, path, sheet_names, expected_schema_count):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Schema" not in workbook.sheetnames:
            raise ValueError(f"{path} is missing Schema")
        schema = workbook["Schema"]
        schema_count = schema.max_row - 1
        if schema_count != expected_schema_count:
            raise ValueError(f"{path} schema count {schema_count} != {expected_schema_count}")
        merged = {}
        sheet_audit = []
        expected_keys = None
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"{path} is missing {sheet_name}")
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in next(sheet.iter_rows())]
            required = ("trade_id", "symbol", "direction", "decision_time")
            self.validate_columns(headers, required, sheet_name)
            keys = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values))
                key = (str(row["symbol"]).upper(), int(row["trade_id"]))
                keys.append(key)
                target = merged.setdefault(key, {})
                for field, value in row.items():
                    if field in target and target[field] != value:
                        if field == "decision_time" and self._time_value(target[field]) == self._time_value(value):
                            continue
                        raise ValueError(f"{path} conflicting {field} for {key}")
                    target[field] = value
            if len(keys) != len(set(keys)):
                raise ValueError(f"{path} duplicate keys in {sheet_name}")
            if expected_keys is not None and set(keys) != expected_keys:
                raise ValueError(f"{path} record sheets do not share identical keys")
            expected_keys = set(keys)
            sheet_audit.append({
                "sheet": sheet_name, "row_count": len(keys),
                "column_count": len(headers), "duplicate_key_count": 0,
            })
        workbook.close()
        return merged, {
            "path": str(path), "schema_field_count": schema_count,
            "record_count": len(merged), "sheets": sheet_audit,
        }

    def _validate_auxiliary_times(self, key, decision, mtf, smart, violations):
        for timeframe in ("m15", "h1", "h4", "d1"):
            field = f"{timeframe}_latest_candle_close_time"
            value = self._available(mtf.get(field))
            if value is not None and self._time_value(value) > decision:
                violations.append({"key": key, "field": field, "value": value})
        for field in (
            "ob_formation_timestamp", "ob_detection_timestamp",
            "fvg_formation_timestamp", "fvg_detection_timestamp",
            "liquidity_formation_timestamp", "liquidity_sweep_timestamp",
        ):
            value = self._available(smart.get(field))
            if value is not None and self._time_value(value) > decision:
                violations.append({"key": key, "field": field, "value": value})

    def _validate_prior_results(self, feature, shadow):
        if feature.get("recommendations", {}).get("production_consideration_justified") is not False:
            raise ValueError("Sprint 84 production guardrail is missing")
        if shadow.get("verdicts", {}).get("shadow_score") != "NOT_RELIABLE":
            raise ValueError("Sprint 83 shadow-score guardrail is missing")
        if shadow.get("verdicts", {}).get("shadow_confidence") != "NOT_RELIABLE":
            raise ValueError("Sprint 83 shadow-confidence guardrail is missing")
        return {
            "feature_discovery_schema_version": feature.get("schema_version"),
            "feature_discovery_measurable_features": feature.get("recommendations", {}).get("measurable_features", []),
            "feature_discovery_production_consideration": False,
            "shadow_validation_schema_version": shadow.get("schema_version"),
            "shadow_score_classification": "NOT_RELIABLE",
            "shadow_confidence_classification": "NOT_RELIABLE",
            "shadow_per_trade_values_reconstructed": False,
        }

    def analyze(self, records, validation=None, audit=None, prior=None):
        before = copy.deepcopy(records)
        closed = [row for row in records if row["status"] == "CLOSED" and row["profit"] is not None]
        unresolved = [row for row in records if not (row["status"] == "CLOSED" and row["profit"] is not None)]
        assignments, availability = self._assign_patterns(records, closed)
        ranking = []
        statistical_tests = []
        symbol_analysis = []
        temporal_stability = []
        sample_quality = []
        combination_map = {name: (category, fields) for name, category, fields in self.COMBINATIONS}
        ordered_closed = sorted(closed, key=lambda row: (row["decision_time"], row["source_index"]))
        midpoint = len(ordered_closed) // 2
        periods = (("FIRST_HALF", ordered_closed[:midpoint]), ("SECOND_HALF", ordered_closed[midpoint:]))
        for name, category, fields in self.COMBINATIONS:
            usable = [(row, self._pattern(row, fields)) for row in closed]
            usable = [(row, pattern) for row, pattern in usable if pattern is not None]
            patterns = sorted({pattern for _, pattern in usable})
            for pattern in patterns:
                group = [row for row, value in usable if value == pattern]
                complement = [row for row, value in usable if value != pattern]
                stats = self._pattern_statistics(name, category, fields, pattern, group, complement, len(usable))
                sample_quality.append({
                    "combination": name, "pattern": pattern,
                    "available_closed_count": len(usable), "unique_pattern_count": len(patterns),
                    "group_count": len(group), "complement_count": len(complement),
                    "group_share_percent": 100.0 * len(group) / len(usable) if usable else 0.0,
                    "warning": stats["sample_warning"],
                })
                test = {
                    "combination": name, "pattern": pattern,
                    "group_count": len(group), "complement_count": len(complement),
                    "mann_whitney_u": stats["mann_whitney_u"],
                    "p_value": stats["mann_whitney_p_value"],
                    "permutation_win_rate_p_value": stats["permutation_win_rate_p_value"],
                }
                statistical_tests.append(test)
                symbol_rows, symbol_stable = self._symbol_slices(name, pattern, fields, closed, stats["win_rate_difference_pp"])
                temporal_rows, temporal_stable = self._temporal_slices(name, pattern, fields, periods, stats["win_rate_difference_pp"])
                symbol_analysis.extend(symbol_rows)
                temporal_stability.extend(temporal_rows)
                stats["symbol_stable"] = symbol_stable
                stats["temporal_stable"] = temporal_stable
                ranking.append(stats)
        self._apply_corrections(statistical_tests)
        corrected = {(row["combination"], row["pattern"]): row for row in statistical_tests}
        for row in ranking:
            tests = corrected[(row["combination"], row["pattern"])]
            row["mann_whitney_adjusted_p_value"] = tests["adjusted_p_value"]
            row["permutation_adjusted_p_value"] = tests["permutation_adjusted_p_value"]
            row["classification"] = self._classification(row)
            row["future_investigation_candidate"] = row["classification"] == "PROMISING_BUT_LIMITED"
        ranking.sort(key=lambda row: (-row["ranking_score"], -row["group_count"], row["combination"], row["pattern"]))
        ranking = [ContextCombinationResult.create(row).to_dict() for row in ranking]
        adequate = [row for row in ranking if row["sample_warning"] == "NONE"]
        best = max(adequate, key=lambda row: row["win_rate_difference_pp"], default=None)
        worst = min(adequate, key=lambda row: row["win_rate_difference_pp"], default=None)
        promising = [row for row in ranking if row["classification"] == "PROMISING_BUT_LIMITED"]
        if records != before:
            raise RuntimeError("Combination analysis mutated its inputs")
        data_validation = validation or {
            "trade_count": len(records), "closed_trade_count": len(closed),
            "unresolved_trade_count": len(unresolved),
        }
        result = {
            "schema_version": self.VERSION, "mode": "DIAGNOSTIC_ONLY",
            "summary": {
                "trade_count": len(records), "closed_trade_count": len(closed),
                "unresolved_trade_count": len(unresolved),
                "winner_count": sum(row["profit"] > 0 for row in closed),
                "loser_count": sum(row["profit"] <= 0 for row in closed),
                "combination_definition_count": len(self.COMBINATIONS),
                "observed_pattern_count": len(ranking),
                "adequate_sample_pattern_count": len(adequate),
                "promising_pattern_count": len(promising),
                "reliable_pattern_count": 0,
                "best_descriptive_pattern": self._summary_pattern(best),
                "worst_descriptive_pattern": self._summary_pattern(worst),
                "future_investigation_justified": bool(promising),
                "production_change_justified": False,
            },
            "combination_definitions": [
                {"combination": name, "category": category, "fields": list(fields)}
                for name, category, fields in self.COMBINATIONS
            ],
            "combination_ranking": ranking,
            "winner_loser_statistics": [self._winner_loser_row(row) for row in ranking],
            "symbol_analysis": symbol_analysis,
            "temporal_stability": temporal_stability,
            "sample_quality": sample_quality,
            "statistical_tests": statistical_tests,
            "feature_availability": availability,
            "configuration": {
                "seed": self.SEED, "bootstrap_iterations": self.BOOTSTRAP_ITERATIONS,
                "permutation_iterations": self.PERMUTATION_ITERATIONS,
                "minimum_group_sample": self.MIN_GROUP,
                "minimum_symbol_or_temporal_slice": self.MIN_SLICE,
                "multiple_comparison_method": "BENJAMINI_HOCHBERG_FDR",
                "alpha": self.ALPHA, "ranking_metric": "ABSOLUTE_WIN_RATE_DIFFERENCE_PP",
                "closed_only_for_outcome_statistics": True,
                "unresolved_preserved_for_audit": True,
                "out_of_sample_validation_performed": False,
                "maximum_possible_classification": "PROMISING_BUT_LIMITED",
            },
            "prior_results": prior or {}, "data_validation": data_validation,
            "diagnostics": {
                "input_objects_unchanged": True, "future_data_used": False,
                "outcome_fields_used_as_combination_components": False,
                "unresolved_excluded_from_statistics": len(closed) + len(unresolved) == len(records),
                "production_modules_import_analysis": False,
                "production_decisions_changed": False,
                "detector_logic_changed": False, "strategy_behavior_changed": False,
                "audit": audit or {},
            },
            "audit": {
                "unresolved_trades": [
                    {"trade_id": row["trade_id"], "symbol": row["symbol"], "status": row["status"]}
                    for row in unresolved
                ],
                "pre_registered_combination_count": len(self.COMBINATIONS),
                "post_outcome_combination_generation": False,
                "prior_guardrails": prior or {},
            },
            "conclusion": {
                "classification": "PROMISING_BUT_LIMITED" if promising else "NOT_RELIABLE",
                "strong_enough_for_future_investigation": bool(promising),
                "production_use_justified": False,
                "reason": (
                    "At least one in-sample combination passed corrected descriptive safeguards, but independent out-of-sample validation is still required."
                    if promising else
                    "No pre-registered combination passed corrected sample, bootstrap, and stability safeguards."
                ),
            },
        }
        return self._clean(result)

    def _assign_patterns(self, records, closed):
        assignments = []
        availability = []
        all_features = sorted(self.FEATURE_SOURCES)
        for feature in all_features:
            all_count = sum(self._available(row["features"].get(feature)) is not None for row in records)
            closed_count = sum(self._available(row["features"].get(feature)) is not None for row in closed)
            availability.append({
                "feature": feature, "source": self.FEATURE_SOURCES[feature],
                "trade_count": len(records), "available_count": all_count,
                "availability_percent": 100.0 * all_count / len(records) if records else 0.0,
                "closed_available_count": closed_count,
            })
        availability.extend((
            {
                "feature": "sprint_84_feature_results", "source": "MSS_Feature_Discovery.json",
                "trade_count": len(records), "available_count": 0,
                "availability_percent": 0.0, "closed_available_count": 0,
                "note": "Summary guardrail only; per-trade results were not reconstructed.",
            },
            {
                "feature": "sprint_83_shadow_results", "source": "MSS_Shadow_Score_Validation.json",
                "trade_count": len(records), "available_count": 0,
                "availability_percent": 0.0, "closed_available_count": 0,
                "note": "Summary guardrail only; per-trade shadow values were not reconstructed.",
            },
        ))
        for row in records:
            for name, _, fields in self.COMBINATIONS:
                assignments.append({
                    "trade_id": row["trade_id"], "symbol": row["symbol"],
                    "combination": name, "pattern": self._pattern(row, fields),
                })
        return assignments, availability

    def _pattern_statistics(self, name, category, fields, pattern, group, complement, available_count):
        group_profit = [float(row["profit"]) for row in group]
        complement_profit = [float(row["profit"]) for row in complement]
        group_win = [1.0 if value > 0 else 0.0 for value in group_profit]
        complement_win = [1.0 if value > 0 else 0.0 for value in complement_profit]
        group_metrics = self._metrics(group_profit)
        complement_metrics = self._metrics(complement_profit)
        warning = self._sample_warning(len(group), len(complement), available_count)
        key = f"{name}|{pattern}"
        group_win_ci = self._bootstrap_single(group_win, key + "|group_win")
        group_expectancy_ci = self._bootstrap_single(group_profit, key + "|group_pnl")
        win_diff_ci = self._bootstrap_difference_values(group_win, complement_win, key + "|win_diff", scale=100.0)
        expectancy_diff_ci = self._bootstrap_difference_values(group_profit, complement_profit, key + "|pnl_diff")
        win_diff = 100.0 * (statistics.mean(group_win) - statistics.mean(complement_win)) if group_win and complement_win else 0.0
        expectancy_diff = group_metrics["expectancy"] - complement_metrics["expectancy"] if group_profit and complement_profit else 0.0
        u, mw_p = self.mann_whitney(group_profit, complement_profit) if warning == "NONE" else (None, None)
        permutation_p = self._permutation_win_rate(group_win, complement_win, key) if warning == "NONE" else None
        return {
            "combination": name, "category": category, "fields": list(fields),
            "pattern": pattern, "available_closed_count": available_count,
            "group_count": len(group), "complement_count": len(complement),
            "group_wins": group_metrics["wins"], "group_losses": group_metrics["losses"],
            "group_win_rate_percent": group_metrics["win_rate_percent"],
            "group_win_rate_ci_low": 100.0 * group_win_ci[0] if group_win_ci[0] is not None else None,
            "group_win_rate_ci_high": 100.0 * group_win_ci[1] if group_win_ci[1] is not None else None,
            "group_expectancy": group_metrics["expectancy"],
            "group_expectancy_ci_low": group_expectancy_ci[0],
            "group_expectancy_ci_high": group_expectancy_ci[1],
            "group_profit_factor": group_metrics["profit_factor"],
            "group_cumulative_pnl": group_metrics["cumulative_pnl"],
            "complement_wins": complement_metrics["wins"],
            "complement_losses": complement_metrics["losses"],
            "complement_win_rate_percent": complement_metrics["win_rate_percent"],
            "complement_expectancy": complement_metrics["expectancy"],
            "complement_profit_factor": complement_metrics["profit_factor"],
            "complement_cumulative_pnl": complement_metrics["cumulative_pnl"],
            "win_rate_difference_pp": win_diff,
            "win_rate_difference_ci_low": win_diff_ci[0],
            "win_rate_difference_ci_high": win_diff_ci[1],
            "expectancy_difference": expectancy_diff,
            "expectancy_difference_ci_low": expectancy_diff_ci[0],
            "expectancy_difference_ci_high": expectancy_diff_ci[1],
            "mann_whitney_u": u, "mann_whitney_p_value": mw_p,
            "permutation_win_rate_p_value": permutation_p,
            "ranking_score": abs(win_diff), "sample_warning": warning,
        }

    def _symbol_slices(self, name, pattern, fields, closed, overall_difference):
        output, usable_signs = [], []
        for symbol in sorted({row["symbol"] for row in closed}):
            rows = [row for row in closed if row["symbol"] == symbol]
            group = [row for row in rows if self._pattern(row, fields) == pattern]
            complement = [row for row in rows if self._pattern(row, fields) not in (None, pattern)]
            difference = self._win_rate_difference(group, complement)
            adequate = len(group) >= self.MIN_SLICE and len(complement) >= self.MIN_SLICE
            if adequate:
                usable_signs.append(self._sign(difference))
            output.append({
                "combination": name, "pattern": pattern, "symbol": symbol,
                "group_count": len(group), "complement_count": len(complement),
                "group_win_rate_percent": self._row_win_rate(group),
                "complement_win_rate_percent": self._row_win_rate(complement),
                "win_rate_difference_pp": difference,
                "sample_adequate": adequate,
            })
        overall_sign = self._sign(overall_difference)
        stable = len(usable_signs) == 2 and overall_sign != 0 and all(sign == overall_sign for sign in usable_signs)
        return output, stable

    def _temporal_slices(self, name, pattern, fields, periods, overall_difference):
        output, usable_signs = [], []
        for period, rows in periods:
            group = [row for row in rows if self._pattern(row, fields) == pattern]
            complement = [row for row in rows if self._pattern(row, fields) not in (None, pattern)]
            difference = self._win_rate_difference(group, complement)
            adequate = len(group) >= self.MIN_SLICE and len(complement) >= self.MIN_SLICE
            if adequate:
                usable_signs.append(self._sign(difference))
            output.append({
                "combination": name, "pattern": pattern, "period": period,
                "group_count": len(group), "complement_count": len(complement),
                "group_win_rate_percent": self._row_win_rate(group),
                "complement_win_rate_percent": self._row_win_rate(complement),
                "win_rate_difference_pp": difference,
                "sample_adequate": adequate,
                "start": rows[0]["decision_time"].isoformat() if rows else None,
                "end": rows[-1]["decision_time"].isoformat() if rows else None,
            })
        overall_sign = self._sign(overall_difference)
        stable = len(usable_signs) == 2 and overall_sign != 0 and all(sign == overall_sign for sign in usable_signs)
        return output, stable

    def _apply_corrections(self, tests):
        mann = [{"p_value": row["p_value"]} for row in tests]
        self.apply_bh(mann)
        permutation = [{"p_value": row["permutation_win_rate_p_value"]} for row in tests]
        self.apply_bh(permutation)
        for row, mw, perm in zip(tests, mann, permutation):
            row["adjusted_p_value"] = mw["adjusted_p_value"]
            row["permutation_adjusted_p_value"] = perm["adjusted_p_value"]

    def _classification(self, row):
        if row["sample_warning"] != "NONE":
            return "INSUFFICIENT_DATA"
        win_ci = (row["win_rate_difference_ci_low"], row["win_rate_difference_ci_high"])
        expectancy_ci = (row["expectancy_difference_ci_low"], row["expectancy_difference_ci_high"])
        win_excludes_zero = None not in win_ci and (win_ci[0] > 0 or win_ci[1] < 0)
        expectancy_excludes_zero = None not in expectancy_ci and (expectancy_ci[0] > 0 or expectancy_ci[1] < 0)
        corrected = [row.get("mann_whitney_adjusted_p_value"), row.get("permutation_adjusted_p_value")]
        corrected_support = any(value is not None and value < self.ALPHA for value in corrected)
        stability_not_contradicted = row.get("symbol_stable") or row.get("temporal_stable")
        if win_excludes_zero and expectancy_excludes_zero and corrected_support and stability_not_contradicted:
            return "PROMISING_BUT_LIMITED"
        return "NOT_RELIABLE"

    def _bootstrap_single(self, values, key):
        if not values:
            return None, None
        rng = random.Random(self._seed("single|" + key))
        samples = [statistics.mean(values[rng.randrange(len(values))] for _ in values) for _ in range(self.BOOTSTRAP_ITERATIONS)]
        return self.percentile_ci(samples)

    def _bootstrap_difference_values(self, left, right, key, scale=1.0):
        if not left or not right:
            return None, None
        rng = random.Random(self._seed("difference|" + key))
        samples = []
        for _ in range(self.BOOTSTRAP_ITERATIONS):
            left_mean = statistics.mean(left[rng.randrange(len(left))] for _ in left)
            right_mean = statistics.mean(right[rng.randrange(len(right))] for _ in right)
            samples.append(scale * (left_mean - right_mean))
        return self.percentile_ci(samples)

    def _permutation_win_rate(self, left, right, key):
        if not left or not right:
            return None
        observed = abs(statistics.mean(left) - statistics.mean(right))
        combined = list(left) + list(right)
        size = len(left)
        rng = random.Random(self._seed("permutation|" + key))
        extreme = 0
        for _ in range(self.PERMUTATION_ITERATIONS):
            rng.shuffle(combined)
            difference = abs(statistics.mean(combined[:size]) - statistics.mean(combined[size:]))
            extreme += difference >= observed
        return (extreme + 1) / (self.PERMUTATION_ITERATIONS + 1)

    @staticmethod
    def _metrics(profits):
        if not profits:
            return {
                "wins": 0, "losses": 0, "win_rate_percent": None,
                "expectancy": None, "profit_factor": None, "cumulative_pnl": 0.0,
            }
        wins = sum(value > 0 for value in profits)
        gross_profit = sum(value for value in profits if value > 0)
        gross_loss = -sum(value for value in profits if value < 0)
        return {
            "wins": wins, "losses": len(profits) - wins,
            "win_rate_percent": 100.0 * wins / len(profits),
            "expectancy": statistics.mean(profits),
            "profit_factor": gross_profit / gross_loss if gross_loss else None,
            "cumulative_pnl": sum(profits),
        }

    def _sample_warning(self, group_count, complement_count, available_count):
        warnings = []
        if group_count < self.MIN_GROUP:
            warnings.append(f"GROUP_BELOW_{self.MIN_GROUP}")
        if complement_count < self.MIN_GROUP:
            warnings.append(f"COMPLEMENT_BELOW_{self.MIN_GROUP}")
        if available_count and group_count / available_count < 0.05:
            warnings.append("RARE_PATTERN_BELOW_5_PERCENT")
        return "|".join(warnings) if warnings else "NONE"

    def _pattern(self, row, fields):
        values = []
        for field in fields:
            value = self._available(row["features"].get(field))
            if value is None:
                return None
            values.append(f"{field}={self._canonical(value)}")
        return " | ".join(values)

    @staticmethod
    def _canonical(value):
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        return str(value).upper()

    @classmethod
    def _available(cls, value):
        return None if value in (None, "", cls.NOT_AVAILABLE) else value

    @classmethod
    def _number(cls, value):
        value = cls._available(value)
        if value is None or isinstance(value, bool):
            return None
        try:
            number = float(value)
            return number if math.isfinite(number) else None
        except (TypeError, ValueError):
            return None

    @classmethod
    def _overall_alignment(cls, h1, h4, d1):
        values = (h1, h4, d1)
        if any(value is None for value in values):
            return None
        if all(value == "ALIGNED" for value in values):
            return "ALIGNED"
        if all(value == "CONFLICTING" for value in values):
            return "CONFLICTING"
        return "NEUTRAL"

    @classmethod
    def _presence(cls, value):
        if value in (True, 1, "TRUE", "True", "true"):
            return "PRESENT"
        if value in (False, 0, "FALSE", "False", "false"):
            return "ABSENT"
        return None

    @staticmethod
    def _row_win_rate(rows):
        return 100.0 * sum(row["profit"] > 0 for row in rows) / len(rows) if rows else None

    @classmethod
    def _win_rate_difference(cls, group, complement):
        left, right = cls._row_win_rate(group), cls._row_win_rate(complement)
        return left - right if left is not None and right is not None else None

    @staticmethod
    def _summary_pattern(row):
        if row is None:
            return None
        return {
            "combination": row["combination"], "pattern": row["pattern"],
            "group_count": row["group_count"],
            "win_rate_difference_pp": row["win_rate_difference_pp"],
            "expectancy_difference": row["expectancy_difference"],
            "classification": row.get("classification"),
        }

    @staticmethod
    def _winner_loser_row(row):
        keys = (
            "combination", "pattern", "group_count", "group_wins", "group_losses",
            "group_win_rate_percent", "group_win_rate_ci_low", "group_win_rate_ci_high",
            "group_expectancy", "group_expectancy_ci_low", "group_expectancy_ci_high",
            "group_profit_factor", "group_cumulative_pnl", "complement_count",
            "complement_wins", "complement_losses", "complement_win_rate_percent",
            "complement_expectancy", "win_rate_difference_pp", "expectancy_difference",
            "classification",
        )
        return {key: row.get(key) for key in keys}

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        summary = workbook.create_sheet("Summary")
        self._key_values(summary, {
            "Mode": "DIAGNOSTIC_ONLY", "Schema Version": self.VERSION,
            "Trades": result["summary"]["trade_count"],
            "Closed": result["summary"]["closed_trade_count"],
            "Unresolved": result["summary"]["unresolved_trade_count"],
            "Combinations Defined": result["summary"]["combination_definition_count"],
            "Observed Patterns Tested": result["summary"]["observed_pattern_count"],
            "Adequate Samples": result["summary"]["adequate_sample_pattern_count"],
            "Promising Patterns": result["summary"]["promising_pattern_count"],
            "Best Descriptive Pattern": result["summary"]["best_descriptive_pattern"],
            "Worst Descriptive Pattern": result["summary"]["worst_descriptive_pattern"],
            "Future Investigation Justified": result["summary"]["future_investigation_justified"],
            "Production Change Justified": False,
        })
        mappings = {
            "Combination Ranking": result["combination_ranking"],
            "Winner Loser Statistics": result["winner_loser_statistics"],
            "Symbol Analysis": result["symbol_analysis"],
            "Temporal Stability": result["temporal_stability"],
            "Sample Quality": result["sample_quality"],
            "Statistical Tests": result["statistical_tests"],
            "Feature Availability": result["feature_availability"],
            "Configuration": {
                **result["configuration"], "combination_definitions": result["combination_definitions"],
                "prior_results": result["prior_results"],
            },
            "Diagnostics": {**result["data_validation"], **result["diagnostics"]},
            "Audit": {**result["audit"], "conclusion": result["conclusion"]},
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
        for sheet in workbook.worksheets:
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()
