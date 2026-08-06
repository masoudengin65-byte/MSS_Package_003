"""Sprint 78 diagnostic-only classification of historical backtest trades.

The engine consumes a Sprint 77 workbook.  It does not invoke or modify the
strategy pipeline, and it never drops source trade rows.
"""

from __future__ import annotations

import json
import math
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mss.analysis.kill_zone_engine import KillZoneEngine
from mss.analysis.session_engine import SessionEngine


class TradeClassificationEngine:
    REQUIRED_SHEETS = [
        "Summary", "Classified Trades", "Symbol Analysis", "Detector Analysis",
        "Session Analysis", "Time Analysis", "Loss Categories", "Cost Analysis",
        "Diagnostics",
    ]
    SOURCE_SHEET = "Trades"
    NOT_CAPTURED = "NOT_CAPTURED"
    GROUP_FIELDS = [
        "Symbol", "Direction", "Structure", "Session", "Kill Zone",
        "Premium / Discount", "Detector Combination", "Score Band",
        "Confidence Band", "Entry Hour", "Day Of Week",
    ]
    METRIC_HEADERS = [
        "Dimension", "Group", "Trade Count", "Win Rate %", "Profit Factor",
        "Expectancy", "Average R", "Net Profit", "Drawdown Contribution",
        "Reliability",
    ]

    def classify_workbook(self, source, output="reports/MSS_Trade_Classification.xlsx"):
        source = Path(source)
        workbook = load_workbook(source, read_only=True, data_only=True)
        if self.SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"Missing source sheet: {self.SOURCE_SHEET}")
        sheet = workbook[self.SOURCE_SHEET]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        source_rows = [dict(zip(headers, (cell.value for cell in row))) for row in sheet.iter_rows(min_row=2)]
        classified = [self.classify_trade(row, index) for index, row in enumerate(source_rows, 1)]
        self.build(classified, output, source=source)
        return classified, str(Path(output))

    def classify_trade(self, source, source_row=None):
        states = source.get("Detector and Context States", {}) or {}
        if isinstance(states, str):
            try:
                states = json.loads(states)
            except json.JSONDecodeError:
                states = {}
        entry = source.get("Entry Time")
        exit_time = source.get("Exit Time")
        profit = self._number(source.get("Profit/Loss"))
        score = self._number(source.get("Score"))
        confidence = self._number(source.get("Confidence"))
        spread = abs(self._number(source.get("Spread")))
        commission = abs(self._number(source.get("Commission")))
        slippage = abs(self._number(source.get("Slippage")))
        volume = abs(self._number(source.get("Volume")))
        stop_distance = abs(self._number(source.get("Entry Price")) - self._number(source.get("Stop Loss")))
        holding = ((exit_time - entry).total_seconds() / 60.0) if isinstance(entry, datetime) and isinstance(exit_time, datetime) else None
        session = SessionEngine().detect(entry) if isinstance(entry, datetime) else None
        kill_zone = KillZoneEngine().detect(entry) if isinstance(entry, datetime) else None
        detector_combo = self._detector_combination(states)

        row = {
            "Source Row": source_row,
            "Trade ID": source.get("Trade ID"),
            "Symbol": source.get("Symbol"),
            "Timeframe": source.get("Timeframe"),
            "Direction": source.get("Direction"),
            "Signal Time": source.get("Signal Time"),
            "Entry Time": entry,
            "Exit Time": exit_time,
            "Status": source.get("Status"),
            "Exit Reason": source.get("Exit Reason"),
            "Entry Price": source.get("Entry Price"),
            "Stop Loss": source.get("Stop Loss"),
            "Take Profit": source.get("Take Profit"),
            "Exit Price": source.get("Exit Price"),
            "Profit/Loss": profit,
            "R Multiple": self._number(source.get("R Multiple")),
            "Structure": states.get("structure", self.NOT_CAPTURED),
            "BOS": states.get("bos", self.NOT_CAPTURED),
            "CHOCH": states.get("choch", self.NOT_CAPTURED),
            "Liquidity": states.get("liquidity", self.NOT_CAPTURED),
            "Liquidity Side": states.get("liquidity_side", self.NOT_CAPTURED),
            "Liquidity Sweep": states.get("liquidity_sweep", self.NOT_CAPTURED),
            "Order Block": states.get("order_block", self.NOT_CAPTURED),
            "Fair Value Gap": states.get("fair_value_gap", self.NOT_CAPTURED),
            "Score": score,
            "Score Band": self._score_band(score),
            "Confidence": confidence,
            "Confidence Band": self._confidence_band(confidence),
            "Session": session.name if session and session.active else "OFF_SESSION",
            "Kill Zone": kill_zone.name if kill_zone and kill_zone.active else "NONE",
            "Premium / Discount": states.get("premium_discount", self.NOT_CAPTURED),
            "Multi Timeframe Alignment": states.get("multi_timeframe_alignment", self.NOT_CAPTURED),
            "Session Bias": states.get("session_bias", self.NOT_CAPTURED),
            "News Status": states.get("news_status", self.NOT_CAPTURED),
            "Portfolio Risk": states.get("portfolio_risk", self.NOT_CAPTURED),
            "Entry Hour": entry.hour if isinstance(entry, datetime) else self.NOT_CAPTURED,
            "Day Of Week": entry.strftime("%A") if isinstance(entry, datetime) else self.NOT_CAPTURED,
            "Holding Minutes": holding,
            "Holding Duration Band": self._holding_band(holding),
            "Spread": spread,
            "Spread Band": self._relative_band(spread, stop_distance, "spread"),
            "Commission": commission,
            "Slippage": slippage,
            "Volume": volume,
            "Stop Distance": stop_distance,
            "Stop Distance Band": self._stop_band(stop_distance, source.get("Symbol")),
            "Detector Combination": detector_combo,
        }
        row["Setup Combination"] = " | ".join(str(row[name]) for name in (
            "Symbol", "Direction", "Structure", "Session", "Kill Zone",
            "Detector Combination", "Score Band", "Confidence Band",
        ))
        category, evidence = self._failure_category(row, states)
        row["Failure Category"] = category
        row["Classification Evidence"] = evidence
        return row

    def build(self, rows, output, source=None):
        output = Path(output)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        closed = [r for r in rows if r["Status"] == "CLOSED"]
        self._write_summary(ws, rows, closed, source)
        self._write_rows(wb.create_sheet("Classified Trades"), rows)
        grouped = {field: self.group_statistics(closed, field) for field in self.GROUP_FIELDS}
        self._write_analysis(wb.create_sheet("Symbol Analysis"), {"Symbol": grouped["Symbol"], "Direction": grouped["Direction"], "Structure": grouped["Structure"]})
        self._write_analysis(wb.create_sheet("Detector Analysis"), {"Detector Combination": grouped["Detector Combination"], "Score Band": grouped["Score Band"], "Confidence Band": grouped["Confidence Band"]})
        self._write_analysis(wb.create_sheet("Session Analysis"), {"Session": grouped["Session"], "Kill Zone": grouped["Kill Zone"], "Premium / Discount": grouped["Premium / Discount"]})
        self._write_analysis(wb.create_sheet("Time Analysis"), {"Entry Hour": grouped["Entry Hour"], "Day Of Week": grouped["Day Of Week"]})
        self._write_loss_categories(wb.create_sheet("Loss Categories"), closed)
        self._write_cost_analysis(wb.create_sheet("Cost Analysis"), closed)
        self._write_diagnostics(wb.create_sheet("Diagnostics"), rows, closed, source)
        for sheet in wb.worksheets:
            self._format(sheet)
        wb.save(output)

    @classmethod
    def group_statistics(cls, rows, field):
        groups = {}
        for row in rows:
            groups.setdefault(str(row.get(field)), []).append(row)
        result = []
        for value, trades in sorted(groups.items()):
            count = len(trades)
            wins = [t for t in trades if t["Profit/Loss"] > 0]
            losses = [t for t in trades if t["Profit/Loss"] < 0]
            gross_profit = sum(t["Profit/Loss"] for t in wins)
            gross_loss = abs(sum(t["Profit/Loss"] for t in losses))
            net = sum(t["Profit/Loss"] for t in trades)
            result.append({
                "Dimension": field, "Group": value, "Trade Count": count,
                "Win Rate %": round(len(wins) / count * 100, 4) if count else 0.0,
                "Profit Factor": round(gross_profit / gross_loss, 4) if gross_loss else ("INF" if gross_profit else 0.0),
                "Expectancy": round(net / count, 4) if count else 0.0,
                "Average R": round(sum(t["R Multiple"] for t in trades) / count, 4) if count else 0.0,
                "Net Profit": round(net, 2),
                "Drawdown Contribution": round(gross_loss, 2),
                "Reliability": "RELIABLE" if count >= 20 else "LOW_SAMPLE",
            })
        return result

    @staticmethod
    def _number(value):
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _score_band(score):
        if score >= 95: return "95-110"
        if score >= 80: return "80-94"
        if score >= 65: return "65-79"
        if score >= 45: return "45-64"
        if score > 0: return "1-44"
        return "0"

    @staticmethod
    def _confidence_band(value):
        lower = int(value // 10 * 10)
        return f"{lower:02d}-{min(lower + 9, 100):02d}%"

    @staticmethod
    def _holding_band(minutes):
        if minutes is None: return "NOT_CAPTURED"
        if minutes <= 60: return "<=1H"
        if minutes <= 240: return "1-4H"
        if minutes <= 720: return "4-12H"
        if minutes <= 1440: return "12-24H"
        return ">24H"

    @staticmethod
    def _relative_band(value, stop, kind):
        if stop <= 0: return "NOT_AVAILABLE"
        ratio = value / stop * 100
        if ratio <= 1: return "<=1%_OF_STOP"
        if ratio <= 2: return "1-2%_OF_STOP"
        if ratio <= 5: return "2-5%_OF_STOP"
        return ">5%_OF_STOP"

    @staticmethod
    def _stop_band(distance, symbol):
        if distance <= 0: return "NOT_AVAILABLE"
        # Instrument-aware display units only; no strategy threshold is used.
        unit = distance * (10000 if symbol == "EURUSD" else 10)
        if unit <= 10: return "<=10_PIPS_OR_0.1"
        if unit <= 25: return "10-25_PIPS_OR_0.1"
        if unit <= 50: return "25-50_PIPS_OR_0.1"
        return ">50_PIPS_OR_0.1"

    @staticmethod
    def _detector_combination(states):
        names = (("BOS", "bos"), ("CHOCH", "choch"), ("LIQUIDITY", "liquidity"), ("ORDER_BLOCK", "order_block"), ("FVG", "fair_value_gap"))
        active = [label for label, key in names if states.get(key) is True]
        return "+".join(active) if active else "NONE"

    def _failure_category(self, row, states):
        if row["Status"] != "CLOSED":
            return "UNCLASSIFIED", "TRADE_NOT_CLOSED"
        if row["Profit/Loss"] > 0:
            return "CLEAN_WIN", "CLOSED_PROFIT_GT_0"
        direction = row["Direction"]
        bos_direction = states.get("bos_direction")
        structure = row["Structure"]
        if direction == "BUY" and bos_direction == "BEARISH" or direction == "SELL" and bos_direction == "BULLISH":
            return "WRONG_DIRECTION", f"DIRECTION={direction};BOS_DIRECTION={bos_direction}"
        if direction == "BUY" and structure == "DOWNTREND" or direction == "SELL" and structure == "UPTREND":
            return "COUNTER_TREND", f"DIRECTION={direction};STRUCTURE={structure}"
        if structure in {"RANGE", "RANGING"} and states.get("bos") is True:
            return "RANGE_FALSE_BREAK", f"STRUCTURE={structure};BOS=True"
        if states.get("context_valid") is False:
            return "INVALID_CONTEXT", "CAPTURED_CONTEXT_VALID=False"
        active_optional = sum(states.get(k) is True for k in ("choch", "liquidity", "order_block", "fair_value_gap"))
        if active_optional == 0:
            return "LOW_CONFLUENCE", "LOSS;BOS_ONLY;CHOCH=False;LIQUIDITY=False;ORDER_BLOCK=False;FVG=False"
        return "CLEAN_LOSS", f"LOSS;NO_HIGHER_PRECEDENCE_EVIDENCE;DETECTORS={row['Detector Combination']}"

    def _write_summary(self, ws, rows, closed, source):
        wins = sum(r["Profit/Loss"] > 0 for r in closed)
        gross_profit = sum(max(0, r["Profit/Loss"]) for r in closed)
        gross_loss = abs(sum(min(0, r["Profit/Loss"]) for r in closed))
        metrics = [
            ("Source Workbook", str(source) if source else "IN_MEMORY"),
            ("Source Trade Rows", len(rows)), ("Closed Trades Used In Statistics", len(closed)),
            ("Unresolved/Open Trades Preserved", len(rows) - len(closed)),
            ("Win Rate %", round(wins / len(closed) * 100, 4) if closed else 0),
            ("Profit Factor", round(gross_profit / gross_loss, 4) if gross_loss else "INF"),
            ("Net Profit", round(sum(r["Profit/Loss"] for r in closed), 2)),
            ("LOW_SAMPLE Rule", "Trade Count < 20"),
            ("Drawdown Contribution Definition", "Sum of absolute negative P/L within group"),
            ("Unavailable Context Policy", "NOT_CAPTURED; never inferred"),
        ]
        ws.append(["Metric", "Value"])
        for item in metrics: ws.append(list(item))
        ws.append([]); ws.append(["Executive Summary Question", "Evidence-Based Answer"])
        for q, answer in self._executive_summary(closed): ws.append([q, answer])
        ws.append([]); ws.append(["Ranking", "Group", "Trades", "Net Profit", "Reliability"])
        setups = self.group_statistics(closed, "Setup Combination")
        rankings = [
            ("Top Profitable Setup Combination", sorted(setups, key=lambda x: x["Net Profit"], reverse=True)[:5]),
            ("Top Losing Setup Combination", sorted(setups, key=lambda x: x["Net Profit"])[:5]),
        ]
        for dimension, best_label, worst_label in (
            ("Session", "Best Session", "Worst Session"),
            ("Entry Hour", "Best Entry Hour", "Worst Entry Hour"),
            ("Detector Combination", "Best Detector Combination", "Worst Detector Combination"),
            ("Score Band", "Best Score Band", "Worst Score Band"),
            ("Confidence Band", "Best Confidence Band", "Worst Confidence Band"),
        ):
            items = self.group_statistics(closed, dimension)
            rankings.extend(((best_label, sorted(items, key=lambda x: x["Net Profit"], reverse=True)[:1]), (worst_label, sorted(items, key=lambda x: x["Net Profit"])[:1])))
        loss_rows = [r for r in closed if r["Profit/Loss"] < 0]
        loss_groups = self.group_statistics(loss_rows, "Failure Category")
        rankings.extend((
            ("Most Common Loss Category", sorted(loss_groups, key=lambda x: x["Trade Count"], reverse=True)[:1]),
            ("Largest Monetary Loss Category", sorted(loss_groups, key=lambda x: x["Net Profit"])[:1]),
        ))
        for label, items in rankings:
            for item in items: ws.append([label, item["Group"], item["Trade Count"], item["Net Profit"], item["Reliability"]])

    def _executive_summary(self, closed):
        by_symbol = {x["Group"]: x for x in self.group_statistics(closed, "Symbol")}
        eur, xau = by_symbol.get("EURUSD", {}), by_symbol.get("XAUUSD", {})
        detectors = self.group_statistics(closed, "Detector Combination")
        best = max(detectors, key=lambda x: x["Net Profit"], default={})
        worst = min(detectors, key=lambda x: x["Net Profit"], default={})
        losses = Counter(r["Failure Category"] for r in closed if r["Profit/Loss"] < 0)
        return [
            ("1. Why does EURUSD lose?", f"{eur.get('Trade Count', 0)} trades; win rate {eur.get('Win Rate %', 0):.4f}%; PF {eur.get('Profit Factor', 0)}; expectancy {eur.get('Expectancy', 0):.4f}; net {eur.get('Net Profit', 0):.2f}."),
            ("2. Why does XAUUSD win?", f"{xau.get('Trade Count', 0)} trades; win rate {xau.get('Win Rate %', 0):.4f}%; PF {xau.get('Profit Factor', 0)}; expectancy {xau.get('Expectancy', 0):.4f}; net {xau.get('Net Profit', 0):.2f}."),
            ("3. What is the biggest weakness?", f"Most common loss category is {losses.most_common(1)[0][0] if losses else 'NONE'} ({losses.most_common(1)[0][1] if losses else 0} losses); detector context is sparse."),
            ("4. What is the biggest strength?", f"XAUUSD has positive expectancy {xau.get('Expectancy', 0):.4f} and PF {xau.get('Profit Factor', 0)} across {xau.get('Trade Count', 0)} trades."),
            ("5. Which detector contributes most?", f"Best measured combination: {best.get('Group', 'NONE')} (net {best.get('Net Profit', 0):.2f}, n={best.get('Trade Count', 0)}, {best.get('Reliability', '')})."),
            ("6. Which detector contributes least?", f"Worst measured combination: {worst.get('Group', 'NONE')} (net {worst.get('Net Profit', 0):.2f}, n={worst.get('Trade Count', 0)}, {worst.get('Reliability', '')})."),
            ("7. What should be investigated next?", "Capture currently NOT_CAPTURED context fields and investigate the reliable symbol/session/time/detector slices; do not infer causality from grouped associations."),
        ]

    @staticmethod
    def _write_rows(ws, rows):
        headers = list(rows[0]) if rows else []
        ws.append(headers)
        for row in rows: ws.append([row.get(h) for h in headers])

    def _write_analysis(self, ws, sections):
        ws.append(self.METRIC_HEADERS)
        for _, items in sections.items():
            for item in items: ws.append([item[h] for h in self.METRIC_HEADERS])

    def _write_loss_categories(self, ws, closed):
        losses = [r for r in closed if r["Profit/Loss"] < 0]
        ws.append(["Failure Category", "Trade Count", "Share Of Losses %", "Net Profit", "Monetary Loss", "Reliability"])
        for category in sorted(set(r["Failure Category"] for r in losses)):
            group = [r for r in losses if r["Failure Category"] == category]
            net = sum(r["Profit/Loss"] for r in group)
            ws.append([category, len(group), round(len(group)/len(losses)*100, 4) if losses else 0, round(net,2), round(abs(net),2), "RELIABLE" if len(group)>=20 else "LOW_SAMPLE"])

    def _write_cost_analysis(self, ws, closed):
        ws.append(["Symbol", "Trade Count", "Spread Amount Sum", "Commission Sum", "Slippage Amount x Volume", "Average Spread / Stop %", "Cost-Sensitive Losses", "Reliability"])
        for symbol in sorted(set(r["Symbol"] for r in closed)):
            group = [r for r in closed if r["Symbol"] == symbol]
            ratios = [r["Spread"] / r["Stop Distance"] * 100 for r in group if r["Stop Distance"]]
            ws.append([symbol, len(group), round(sum(r["Spread"] for r in group), 8), round(sum(r["Commission"] for r in group), 2), round(sum(r["Slippage"]*r["Volume"] for r in group), 8), round(sum(ratios)/len(ratios), 4) if ratios else 0, sum(r["Failure Category"] == "COST_SENSITIVE_LOSS" for r in group), "RELIABLE" if len(group)>=20 else "LOW_SAMPLE"])

    def _write_diagnostics(self, ws, rows, closed, source):
        not_captured = Counter()
        for row in rows:
            for field, value in row.items():
                if value == self.NOT_CAPTURED: not_captured[field] += 1
        diagnostics = [
            ("Classification Version", "SPRINT_78_V1"), ("Source Workbook", str(source) if source else "IN_MEMORY"),
            ("Source Rows", len(rows)), ("Rows Classified", len(rows)), ("Rows Removed", 0),
            ("Closed Rows", len(closed)), ("All Rows Have Category", all(r.get("Failure Category") for r in rows)),
            ("All Rows Have Evidence", all(r.get("Classification Evidence") for r in rows)),
            ("Statistical Population", "Status=CLOSED only"),
            ("Failure Rule Precedence", "UNCLASSIFIED(open) > CLEAN_WIN > WRONG_DIRECTION > COUNTER_TREND > RANGE_FALSE_BREAK > INVALID_CONTEXT > LOW_CONFLUENCE > CLEAN_LOSS"),
            ("COST_SENSITIVE_LOSS", "Not assigned without captured monetary counterfactual evidence"),
            ("AMBIGUOUS_CANDLE_LOSS", "Not assigned because same-candle hit state was not captured"),
            ("LATE_ENTRY", "Not assigned because expected tradable-candle sequence was not captured"),
        ]
        ws.append(["Diagnostic", "Value"])
        for item in diagnostics: ws.append(list(item))
        for field, count in sorted(not_captured.items()): ws.append([f"NOT_CAPTURED::{field}", count])

    @staticmethod
    def _format(ws):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True); cell.fill = fill
        for col in range(1, ws.max_column + 1):
            width = min(45, max(10, max(len(str(ws.cell(row, col).value or "")) for row in range(1, min(ws.max_row, 200) + 1)) + 2))
            ws.column_dimensions[get_column_letter(col)].width = width
