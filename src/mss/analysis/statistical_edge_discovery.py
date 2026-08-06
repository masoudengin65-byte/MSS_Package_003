"""Deterministic, diagnostic-only statistical edge discovery (Sprint 81)."""

from __future__ import annotations

import hashlib
import json
import math
import random
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class StatisticalEdgeDiscovery:
    VERSION = "SPRINT_81_V1"
    MIN_SAMPLE = 20
    BOOTSTRAP_ITERATIONS = 1000
    REQUIRED_SHEETS = [
        "Executive Summary", "Data Quality", "Feature Performance", "Numeric Bins",
        "Categorical Analysis", "Interaction Analysis", "EURUSD Analysis",
        "XAUUSD Analysis", "Temporal Stability", "Edge Ranking",
        "Insufficient Data", "Diagnostics",
    ]
    NUMERIC_FEATURES = (
        "trend_strength", "swing_count", "bos_strength", "liquidity_distance",
        "atr", "average_candle_size", "current_candle_size", "relative_volatility",
        "spread", "tick_volume", "sl_distance", "tp_distance", "rr",
        "structure_score", "bos_score", "choch_score", "liquidity_score",
        "ob_score", "fvg_score", "session_score", "htf_score", "risk_score",
        "unattributed_score", "final_score", "confidence", "position_size",
        "portfolio_exposure", "correlation_score", "entry_delay_minutes",
        "decision_hour", "entry_hour", "holding_duration_minutes",
    )
    CATEGORICAL_FEATURES = (
        "symbol", "direction", "structure", "bos", "bos_direction", "choch",
        "choch_direction", "liquidity_detected", "liquidity_side", "liquidity_sweep",
        "order_block_detected", "fvg_detected", "session", "kill_zone",
        "day_of_week", "risk_approved",
    )
    INTERACTIONS = (
        ("symbol", "session"), ("symbol", "direction"), ("symbol", "kill_zone"),
        ("symbol", "day_of_week"), ("direction", "session"),
        ("structure", "direction"), ("bos", "liquidity_detected"),
        ("bos", "fvg_detected"), ("bos", "order_block_detected"),
        ("session", "kill_zone"), ("relative_volatility_band", "direction"),
        ("relative_volatility_band", "symbol"),
        ("spread_band", "symbol"), ("final_score_band", "symbol"),
        ("confidence_band", "symbol"),
    )
    METRIC_FIELDS = (
        "trade_count", "wins", "losses", "win_rate", "profit_factor",
        "expectancy", "average_r", "median_r", "net_profit",
        "drawdown_contribution", "win_rate_ci_low", "win_rate_ci_high",
        "expectancy_ci_low", "expectancy_ci_high", "bootstrap_p_value",
        "adjusted_p_value", "effect_size", "sample_status", "temporal_stability",
        "edge_label",
    )

    def run(self, workbook_path, audit_path, output_xlsx, output_json):
        source_hash = self._sha256(workbook_path)
        audit_hash = self._sha256(audit_path)
        records, quality = self.load_and_validate(workbook_path, audit_path)
        result = self.analyze_records(records, quality)
        self.write_json(result, output_json)
        self.write_workbook(result, output_xlsx)
        if source_hash != self._sha256(workbook_path) or audit_hash != self._sha256(audit_path):
            raise RuntimeError("Source artifact mutation detected")
        return result

    def load_and_validate(self, workbook_path, audit_path):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        trades = self._sheet_records(workbook["Trades"])
        contexts = self._sheet_records(workbook["Context Snapshot"])
        context_map = {(str(r["Symbol"]), int(r["Trade ID"])): r for r in contexts}
        records = []
        for trade in trades:
            key = (str(trade["Symbol"]), int(trade["Trade ID"]))
            context = context_map.get(key)
            if context is None:
                raise ValueError(f"Missing context for {key}")
            record = {str(k).lower().replace(" ", "_").replace("/", "_"): v for k, v in context.items() if k not in ("Trade ID", "Symbol")}
            record.update({
                "trade_id": key[1], "symbol": key[0], "direction": trade["Direction"],
                "status": trade["Status"], "profit": float(trade["Profit/Loss"] or 0),
                "r_multiple": float(trade["R Multiple"] or 0),
                "decision_time_value": trade["Signal Time"], "entry_time_value": trade["Entry Time"],
                "exit_time_value": trade["Exit Time"],
            })
            record["decision_hour"] = trade["Signal Time"].hour
            record["entry_hour"] = trade["Entry Time"].hour
            record["holding_duration_minutes"] = ((trade["Exit Time"] - trade["Entry Time"]).total_seconds() / 60.0) if trade["Exit Time"] else None
            records.append(record)
        audit = json.loads(Path(audit_path).read_text(encoding="utf-8"))
        closed = [r for r in records if r["status"] == "CLOSED"]
        ids = [(r["symbol"], r["trade_id"]) for r in records]
        quality = {
            "opened_trades": len(records), "closed_trades": len(closed),
            "excluded_unresolved": len(records) - len(closed),
            "context_rows": len(contexts), "context_field_count": audit["context_field_count"],
            "duplicate_trade_ids": len(ids) - len(set(ids)),
            "dropped_closed_trades": audit["closed_trades"] - len(closed),
            "source_future_data_violations": audit["future_data_violations"],
            "source_mutation_violations": audit["context_mutation_violations"],
            "source_deterministic": audit["deterministic_repeated_output"],
            "not_available_by_field": audit.get("not_available_by_field", {}),
        }
        expected = (170, 169, 1, 87)
        actual = (quality["opened_trades"], quality["closed_trades"], quality["excluded_unresolved"], quality["context_field_count"])
        if actual != expected or quality["duplicate_trade_ids"] or quality["dropped_closed_trades"]:
            raise ValueError(f"Dataset validation failed: {quality}")
        return records, quality

    def analyze_records(self, records, quality=None):
        closed = sorted((dict(r) for r in records if r["status"] == "CLOSED"), key=lambda r: (r["entry_time_value"], r["symbol"], r["trade_id"]))
        for index, row in enumerate(closed): row["_index"] = index
        numeric, categorical, unavailable, bins = [], [], [], {}
        for field, count in sorted((quality or {}).get("not_available_by_field", {}).items()):
            unavailable.append({"feature": field, "reason": "SOURCE_NOT_AVAILABLE", "affected_snapshots": count})
        for feature in self.NUMERIC_FEATURES:
            usable = [r for r in closed if self._numeric(r.get(feature)) is not None]
            if not usable:
                unavailable.append({"feature": feature, "reason": "ALL_NOT_AVAILABLE_OR_NON_NUMERIC"}); continue
            assignments, boundaries = self.quantile_bins(usable, feature)
            bins[feature] = {"assignments": assignments, "boundaries": boundaries}
            for row in usable: row[f"{feature}_band"] = assignments[row["_index"]]
            groups = defaultdict(list)
            for row in usable: groups[assignments[row["_index"]]].append(row)
            for label, group in sorted(groups.items()):
                item = self._candidate("NUMERIC_BIN", feature, label, group, closed)
                item["boundaries"] = boundaries[label]
                numeric.append(item)
        for feature in self.CATEGORICAL_FEATURES:
            usable = [r for r in closed if not self._unavailable(r.get(feature))]
            if not usable:
                unavailable.append({"feature": feature, "reason": "ALL_NOT_AVAILABLE"}); continue
            groups = defaultdict(list)
            for row in usable: groups[str(row.get(feature))].append(row)
            for value, group in sorted(groups.items()): categorical.append(self._candidate("CATEGORICAL", feature, value, group, closed))
        interactions = []
        for left, right in self.INTERACTIONS:
            if not any(left in row and right in row and not self._unavailable(row.get(left)) and not self._unavailable(row.get(right)) for row in closed):
                unavailable.append({"feature": f"{left} × {right}", "reason": "INTERACTION_INPUT_UNAVAILABLE"}); continue
            groups = defaultdict(list)
            for row in closed:
                if self._unavailable(row.get(left)) or self._unavailable(row.get(right)): continue
                groups[f"{row.get(left)} × {row.get(right)}"].append(row)
            for value, group in sorted(groups.items()):
                if len(group) >= self.MIN_SAMPLE: interactions.append(self._candidate("INTERACTION", f"{left} × {right}", value, group, closed))
        all_candidates = numeric + categorical + interactions
        self._apply_bh(all_candidates)
        for item in all_candidates: self._label(item)
        stability = [self._stability(item, closed) for item in all_candidates if item["trade_count"] >= 20]
        stability_map = {(x["analysis_type"], x["feature"], x["value"]): x for x in stability}
        for item in all_candidates:
            stable = stability_map.get((item["analysis_type"], item["feature"], item["value"]))
            item["temporal_stability"] = stable["stability"] if stable else "INSUFFICIENT_DATA"
            self._label(item)
        ranking = sorted(all_candidates, key=self._rank_key)
        symbol_analysis = {symbol: self._symbol_findings(symbol, all_candidates) for symbol in ("EURUSD", "XAUUSD")}
        return {
            "schema_version": self.VERSION,
            "data_quality": quality or {"closed_trades": len(closed)},
            "analysis_config": {"minimum_sample": 20, "limited_max": 39, "reliable_min": 40, "quantile_bins": 4, "bootstrap_iterations": self.BOOTSTRAP_ITERATIONS, "multiple_comparison_method": "Benjamini-Hochberg", "edge_fdr_threshold": 0.10},
            "numeric_bins": self._clean(numeric), "categorical_analysis": self._clean(categorical),
            "interaction_analysis": self._clean(interactions), "temporal_stability": self._clean(stability),
            "edge_ranking": self._clean(ranking), "symbol_analysis": symbol_analysis,
            "unavailable_findings": unavailable,
            "diagnostics": {"candidate_count": len(all_candidates), "closed_trade_indices_preserved": [r["_index"] for r in closed], "deterministic_seed_method": "SHA256(feature|value)"},
        }

    def quantile_bins(self, rows, feature):
        pairs = sorted((float(r[feature]), r["_index"]) for r in rows if self._numeric(r.get(feature)) is not None)
        unique = sorted(set(v for v, _ in pairs))
        if len(unique) == 1:
            label = f"[{unique[0]:.10g}, {unique[0]:.10g}]"
            return {index: label for _, index in pairs}, {label: {"lower": unique[0], "upper": unique[0], "lower_inclusive": True, "upper_inclusive": True}}
        cuts = sorted(set(self._quantile([v for v, _ in pairs], q) for q in (.25, .5, .75)))
        groups = [[] for _ in range(len(cuts) + 1)]
        for value, index in pairs:
            bin_index = next((i for i, cut in enumerate(cuts) if value <= cut), len(cuts))
            groups[bin_index].append((value, index))
        groups = [g for g in groups if g]
        while len(groups) > 1 and any(len(g) < self.MIN_SAMPLE for g in groups):
            i = next(i for i, g in enumerate(groups) if len(g) < self.MIN_SAMPLE)
            neighbor = 1 if i == 0 else i - 1 if i == len(groups) - 1 else (i - 1 if len(groups[i-1]) <= len(groups[i+1]) else i + 1)
            lo, hi = sorted((i, neighbor)); groups[lo] = groups[lo] + groups[hi]; del groups[hi]
        assignments, boundaries = {}, {}
        for group in groups:
            lower, upper = min(v for v, _ in group), max(v for v, _ in group)
            label = f"[{lower:.10g}, {upper:.10g}]"
            boundaries[label] = {"lower": lower, "upper": upper, "lower_inclusive": True, "upper_inclusive": True}
            for _, index in group: assignments[index] = label
        return assignments, boundaries

    def _candidate(self, analysis_type, feature, value, group, population):
        metrics = self.metrics(group)
        complement = [r for r in population if r["_index"] not in {x["_index"] for x in group}]
        metrics.update({
            "analysis_type": analysis_type, "feature": feature, "value": str(value),
            "effect_size": self._effect_size(group, complement), "adjusted_p_value": None,
            "temporal_stability": "PENDING", "edge_label": "PENDING",
            "_indices": [r["_index"] for r in group],
        })
        return metrics

    def metrics(self, rows):
        rows = sorted(rows, key=lambda r: r["entry_time_value"])
        n = len(rows); profits = [r["profit"] for r in rows]; rs = [r["r_multiple"] for r in rows]
        wins = sum(x > 0 for x in profits); losses = sum(x < 0 for x in profits)
        gp = sum(max(0, x) for x in profits); gl = abs(sum(min(0, x) for x in profits))
        win_low, win_high = self._wilson(wins, n)
        exp_low, exp_high, p_value = self._bootstrap(profits, f"{rows[0]['_index'] if rows else 0}|{n}")
        return {
            "trade_count": n, "wins": wins, "losses": losses,
            "win_rate": wins / n * 100 if n else 0.0,
            "profit_factor": gp / gl if gl else (None if not gp else "INF"),
            "expectancy": sum(profits) / n if n else 0.0,
            "average_r": sum(rs) / n if n else 0.0,
            "median_r": statistics.median(rs) if rs else 0.0,
            "net_profit": sum(profits), "drawdown_contribution": self._max_drawdown(profits),
            "win_rate_ci_low": win_low * 100, "win_rate_ci_high": win_high * 100,
            "expectancy_ci_low": exp_low, "expectancy_ci_high": exp_high,
            "bootstrap_p_value": p_value,
            "sample_status": "LOW_SAMPLE" if n < 20 else "LIMITED" if n < 40 else "RELIABLE",
        }

    def _stability(self, item, closed):
        midpoint = len(closed) // 2
        selected = set(item["_indices"])
        first = [r for r in closed[:midpoint] if r["_index"] in selected]
        second = [r for r in closed[midpoint:] if r["_index"] in selected]
        fm, sm = self.metrics(first), self.metrics(second)
        same = self._sign(fm["expectancy"]) == self._sign(sm["expectancy"]) and self._sign(fm["expectancy"]) != 0
        stable = len(first) >= 10 and len(second) >= 10 and item["trade_count"] >= 20 and same
        return {"analysis_type": item["analysis_type"], "feature": item["feature"], "value": item["value"], "combined_trade_count": item["trade_count"], "first_half": fm, "second_half": sm, "direction_same": same, "stability": "STABLE" if stable else "UNSTABLE"}

    def _label(self, item):
        n=item["trade_count"]; exp=item["expectancy"]; stable=item.get("temporal_stability")=="STABLE"; adj=item.get("adjusted_p_value")
        if n < 20: item["edge_label"]="INSUFFICIENT_DATA"
        elif n >= 40 and stable and item["expectancy_ci_low"] > 0 and adj is not None and adj <= .10: item["edge_label"]="RELIABLE_EDGE"
        elif n >= 40 and stable and item["expectancy_ci_high"] < 0 and adj is not None and adj <= .10: item["edge_label"]="NEGATIVE_EDGE"
        elif 20 <= n < 40 and stable and (item["expectancy_ci_low"] > 0 or item["expectancy_ci_high"] < 0): item["edge_label"]="PROMISING_BUT_LIMITED"
        else: item["edge_label"]="NO_EDGE"

    def _symbol_findings(self, symbol, candidates):
        relevant = [x for x in candidates if (x["feature"] == "symbol" and x["value"] == symbol) or ("symbol" in x["feature"] and symbol in x["value"])]
        positive = sorted((x for x in relevant if x["expectancy"] > 0), key=lambda x: (x["sample_status"]=="RELIABLE", x["expectancy"], x["trade_count"]), reverse=True)
        negative = sorted((x for x in relevant if x["expectancy"] < 0), key=lambda x: (x["sample_status"]=="RELIABLE", -x["expectancy"], x["trade_count"]), reverse=True)
        comparison_features = {
            "direction": "symbol × direction", "session": "symbol × session",
            "kill_zone": "symbol × kill_zone", "day_of_week": "symbol × day_of_week",
            "volatility_band": "relative_volatility_band × symbol",
            "spread_band": "spread_band × symbol", "final_score_band": "final_score_band × symbol",
            "confidence_band": "confidence_band × symbol",
        }
        comparisons = {}
        for name, feature in comparison_features.items():
            items = [x for x in candidates if x["feature"] == feature and symbol in x["value"]]
            comparisons[name] = {
                "best": self._clean(max(items, key=lambda x: x["expectancy"])) if items else None,
                "worst": self._clean(min(items, key=lambda x: x["expectancy"])) if items else None,
                "discriminates": len({round(x["expectancy"], 10) for x in items}) > 1,
            }
        return {"strongest_positive": self._clean(positive[:10]), "strongest_negative": self._clean(negative[:10]), "comparisons": comparisons}

    def _apply_bh(self, items):
        ordered = sorted(enumerate(items), key=lambda x: x[1]["bootstrap_p_value"])
        m=len(ordered); adjusted=[0.0]*m; running=1.0
        for rank_index in range(m-1,-1,-1):
            original,item=ordered[rank_index]; rank=rank_index+1
            running=min(running,item["bootstrap_p_value"]*m/rank); adjusted[original]=min(1.0,running)
        for item,value in zip(items,adjusted): item["adjusted_p_value"]=value

    def write_json(self, result, path):
        path=Path(path); path.parent.mkdir(parents=True,exist_ok=True)
        path.write_text(json.dumps(result,indent=2,sort_keys=True,default=str),encoding="utf-8")

    def write_workbook(self, result, path):
        wb=Workbook(); ws=wb.active; ws.title="Executive Summary"
        ranking=result["edge_ranking"]
        ws.append(["Finding Class","Count"])
        for label in ("RELIABLE_EDGE","NEGATIVE_EDGE","PROMISING_BUT_LIMITED","NO_EDGE","INSUFFICIENT_DATA"):
            ws.append([label,sum(x["edge_label"]==label for x in ranking)])
        ws.append([]); ws.append(["Rank","Type","Feature","Value","N","Expectancy","PF","Average R","Stability","Label"])
        for i,x in enumerate(ranking[:25],1): ws.append([i,x["analysis_type"],x["feature"],x["value"],x["trade_count"],x["expectancy"],x["profit_factor"],x["average_r"],x["temporal_stability"],x["edge_label"]])
        self._key_value_sheet(wb.create_sheet("Data Quality"),result["data_quality"])
        self._table(wb.create_sheet("Feature Performance"),result["edge_ranking"])
        self._table(wb.create_sheet("Numeric Bins"),result["numeric_bins"])
        self._table(wb.create_sheet("Categorical Analysis"),result["categorical_analysis"])
        self._table(wb.create_sheet("Interaction Analysis"),result["interaction_analysis"])
        self._symbol_sheet(wb.create_sheet("EURUSD Analysis"),result["symbol_analysis"]["EURUSD"])
        self._symbol_sheet(wb.create_sheet("XAUUSD Analysis"),result["symbol_analysis"]["XAUUSD"])
        self._table(wb.create_sheet("Temporal Stability"),result["temporal_stability"])
        self._table(wb.create_sheet("Edge Ranking"),result["edge_ranking"])
        self._table(wb.create_sheet("Insufficient Data"),result["unavailable_findings"]+[x for x in result["edge_ranking"] if x["edge_label"]=="INSUFFICIENT_DATA"])
        self._key_value_sheet(wb.create_sheet("Diagnostics"),{**result["analysis_config"],**result["diagnostics"]})
        for sheet in wb.worksheets: self._format(sheet)
        path=Path(path);path.parent.mkdir(parents=True,exist_ok=True);wb.save(path)

    @staticmethod
    def _sheet_records(sheet):
        headers=[c.value for c in next(sheet.iter_rows())]
        return [dict(zip(headers,row)) for row in sheet.iter_rows(min_row=2,values_only=True)]
    @staticmethod
    def _numeric(value):
        if value in (None,"","NOT_AVAILABLE"): return None
        if isinstance(value,bool): return None
        try: return float(value)
        except (TypeError,ValueError): return None
    @staticmethod
    def _unavailable(value): return value in (None,"","NOT_AVAILABLE")
    @staticmethod
    def _quantile(values,q):
        pos=(len(values)-1)*q;lo=math.floor(pos);hi=math.ceil(pos)
        return values[lo] if lo==hi else values[lo]+(values[hi]-values[lo])*(pos-lo)
    @staticmethod
    def _wilson(wins,n,z=1.959963984540054):
        if not n:return 0.0,0.0
        p=wins/n;den=1+z*z/n;center=(p+z*z/(2*n))/den;margin=z*math.sqrt(p*(1-p)/n+z*z/(4*n*n))/den
        return max(0,center-margin),min(1,center+margin)
    def _bootstrap(self,values,key):
        if not values:return 0.0,0.0,1.0
        seed=int(hashlib.sha256(key.encode()).hexdigest()[:16],16);rng=random.Random(seed);n=len(values)
        means=sorted(sum(values[rng.randrange(n)] for _ in range(n))/n for _ in range(self.BOOTSTRAP_ITERATIONS))
        low=means[int(.025*(len(means)-1))];high=means[int(.975*(len(means)-1))]
        le=(sum(x<=0 for x in means)+1)/(len(means)+1);ge=(sum(x>=0 for x in means)+1)/(len(means)+1)
        return low,high,min(1.0,2*min(le,ge))
    @staticmethod
    def _max_drawdown(profits):
        equity=peak=drawdown=0.0
        for value in profits: equity+=value;peak=max(peak,equity);drawdown=max(drawdown,peak-equity)
        return drawdown
    @staticmethod
    def _effect_size(group,complement):
        a=[x["r_multiple"] for x in group];b=[x["r_multiple"] for x in complement]
        if len(a)<2 or len(b)<2:return 0.0
        pooled=math.sqrt(((len(a)-1)*statistics.variance(a)+(len(b)-1)*statistics.variance(b))/(len(a)+len(b)-2))
        return (statistics.mean(a)-statistics.mean(b))/pooled if pooled else 0.0
    @staticmethod
    def _sign(value): return 1 if value>0 else -1 if value<0 else 0
    @staticmethod
    def _rank_key(x):
        order={"RELIABLE_EDGE":0,"NEGATIVE_EDGE":1,"PROMISING_BUT_LIMITED":2,"NO_EDGE":3,"INSUFFICIENT_DATA":4}
        return (order[x["edge_label"]],-x["trade_count"],-abs(x["effect_size"]),-abs(x["expectancy"]),x["feature"],x["value"])
    @staticmethod
    def _clean(value):
        if isinstance(value,list):return [StatisticalEdgeDiscovery._clean(x) for x in value]
        if isinstance(value,dict):return {k:StatisticalEdgeDiscovery._clean(v) for k,v in value.items() if not k.startswith("_")}
        return value
    @staticmethod
    def _sha256(path): return hashlib.sha256(Path(path).read_bytes()).hexdigest()
    @staticmethod
    def _table(ws,rows):
        if not rows: ws.append(["No findings"]);return
        headers=[]
        for row in rows:
            for key in row:
                if key not in headers:headers.append(key)
        ws.append(headers)
        for row in rows:ws.append([json.dumps(row.get(h),sort_keys=True,default=str) if isinstance(row.get(h),(dict,list)) else row.get(h) for h in headers])
    @staticmethod
    def _key_value_sheet(ws,values):
        ws.append(["Metric","Value"])
        for key,value in values.items():ws.append([key,json.dumps(value,sort_keys=True,default=str) if isinstance(value,(dict,list)) else value])
    def _symbol_sheet(self,ws,values):
        rows=[]
        for section,items in values.items():
            if isinstance(items,list):
                for item in items:rows.append({"section":section,**item})
            elif isinstance(items,dict):
                for name,comparison in items.items():
                    for side in ("best","worst"):
                        item=comparison.get(side)
                        if item: rows.append({"section":f"{name}_{side}","discriminates":comparison["discriminates"],**item})
        self._table(ws,rows)
    @staticmethod
    def _format(ws):
        ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions;fill=PatternFill("solid",fgColor="1F4E78")
        for cell in ws[1]:cell.font=Font(color="FFFFFF",bold=True);cell.fill=fill
        for col in range(1,ws.max_column+1):
            width=min(45,max(10,max(len(str(ws.cell(row,col).value or "")) for row in range(1,min(ws.max_row,200)+1))+2));ws.column_dimensions[get_column_letter(col)].width=width
