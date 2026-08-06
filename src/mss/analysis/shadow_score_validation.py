"""Sprint 83 diagnostic validation of immutable Sprint 82 shadow scores."""

from __future__ import annotations

import copy
import hashlib
import json
import math
import random
import re
import statistics
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mss.analysis.score_engine import ScoreEngine


class ShadowScoreValidation:
    """Read-only, reproducible score/outcome diagnostics.

    The composite ``(symbol, trade_id)`` key is required because trade IDs are
    allocated per symbol in the canonical historical workbook.
    """

    VERSION = "SPRINT_83_V1"
    BOOTSTRAP_ITERATIONS = 2000
    PERMUTATION_ITERATIONS = 2000
    SEED = 830082
    MIN_GROUP = 20
    REQUIRED_SHEETS = (
        "Summary", "Data Validation", "Winner Loser Comparison",
        "Shadow Score Buckets", "Confidence Buckets", "Component Analysis",
        "Symbol Analysis", "Temporal Stability", "Statistical Tests",
        "Diagnostics", "Conclusions",
    )

    HISTORICAL_COLUMNS = (
        "Trade ID", "Symbol", "Direction", "Entry Time", "Status",
        "Exit Reason", "Profit/Loss", "R Multiple", "Score", "Confidence",
        "Frozen Context Snapshot",
    )
    SHADOW_COLUMNS = (
        "Trade ID", "Symbol", "Direction", "Entry Time", "Legacy Score",
        "Legacy Confidence", "Shadow Score", "Shadow Confidence",
    )

    def run(self, historical_path, shadow_path, excel_path, json_path):
        inputs = [Path(historical_path), Path(shadow_path)]
        outputs = [Path(excel_path), Path(json_path)]
        if set(map(Path.resolve, inputs)) & set(map(Path.resolve, outputs)):
            raise ValueError("Validation outputs must not overwrite immutable inputs")
        before = {str(path): self.file_sha256(path) for path in inputs}
        historical = self.load_historical(historical_path)
        shadow = self.load_shadow(shadow_path)
        result = self.analyze(historical, shadow)
        after = {str(path): self.file_sha256(path) for path in inputs}
        if before != after:
            raise RuntimeError("Immutable input artifact changed during validation")
        result["data_integrity"]["input_artifacts"] = [
            {"path": str(path), "sha256_before": before[str(path)],
             "sha256_after": after[str(path)], "unchanged": True}
            for path in inputs
        ]
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    @staticmethod
    def _key(row):
        return str(row["symbol"]), int(row["trade_id"])

    def load_historical(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Trades" not in workbook.sheetnames:
            raise ValueError("Historical artifact is missing required sheet: Trades")
        sheet = workbook["Trades"]
        headers = [c.value for c in next(sheet.iter_rows())]
        self.validate_columns(headers, self.HISTORICAL_COLUMNS, "historical Trades")
        records = []
        for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row = dict(zip(headers, values))
            records.append({
                "source_index": index, "trade_id": int(row["Trade ID"]),
                "symbol": str(row["Symbol"]), "direction": str(row["Direction"]).upper(),
                "entry_time": row["Entry Time"], "status": str(row["Status"]).upper(),
                "exit_reason": row["Exit Reason"], "profit": row["Profit/Loss"],
                "r_multiple": row["R Multiple"], "legacy_score": row["Score"],
                "legacy_confidence": row["Confidence"],
                "context_snapshot": row["Frozen Context Snapshot"],
            })
        return records

    def load_shadow(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Score Breakdown" not in workbook.sheetnames:
            raise ValueError("Shadow artifact is missing required sheet: Score Breakdown")
        sheet = workbook["Score Breakdown"]
        headers = [c.value for c in next(sheet.iter_rows())]
        self.validate_columns(headers, self.SHADOW_COLUMNS + tuple(ScoreEngine.COMPONENTS), "shadow Score Breakdown")
        records = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            records.append({
                "trade_id": int(row["Trade ID"]), "symbol": str(row["Symbol"]),
                "direction": str(row["Direction"]).upper(), "entry_time": row["Entry Time"],
                "legacy_score": row["Legacy Score"], "legacy_confidence": row["Legacy Confidence"],
                "shadow_score": row["Shadow Score"], "shadow_confidence": row["Shadow Confidence"],
                "components": {name: row.get(name, ScoreEngine.NOT_AVAILABLE) for name in ScoreEngine.COMPONENTS},
            })
        return records

    def join(self, historical, shadow):
        h_counts = Counter(self._key(x) for x in historical)
        s_counts = Counter(self._key(x) for x in shadow)
        duplicates = sorted({k for k, n in h_counts.items() if n > 1} | {k for k, n in s_counts.items() if n > 1})
        shadow_map = {self._key(x): x for x in shadow if s_counts[self._key(x)] == 1}
        joined, unmatched, conflicts = [], [], []
        for trade in historical:
            key = self._key(trade)
            score = shadow_map.get(key)
            if h_counts[key] != 1 or score is None:
                unmatched.append(key); continue
            conflict_fields = [name for name in ("direction", "entry_time", "legacy_score", "legacy_confidence") if trade[name] != score[name]]
            if conflict_fields:
                conflicts.append({"key": list(key), "fields": conflict_fields}); continue
            joined.append({**copy.deepcopy(trade), **copy.deepcopy(score)})
        extra = sorted(set(shadow_map) - set(h_counts))
        unmatched.extend(extra)
        joined.sort(key=lambda x: (self._time_value(x["entry_time"]), x["source_index"]))
        chronology = all(
            (self._time_value(left["entry_time"]), left["source_index"])
            <= (self._time_value(right["entry_time"]), right["source_index"])
            for left, right in zip(joined, joined[1:])
        )
        integrity = {
            "join_key": ["symbol", "trade_id"], "total_historical_trades": len(historical),
            "closed_trades": sum(x["status"] == "CLOSED" for x in historical),
            "unresolved_trades": sum(x["status"] != "CLOSED" for x in historical),
            "matched_trades": len(joined), "unmatched_trades": len(unmatched),
            "unmatched_identifiers": [list(x) for x in unmatched],
            "duplicate_identifiers": len(duplicates), "duplicate_keys": [list(x) for x in duplicates],
            "conflicting_records": len(conflicts), "conflicts": conflicts,
            "missing_scores": sum(x.get("shadow_score") is None or x.get("shadow_confidence") is None for x in joined),
            "missing_outcomes": sum(x["status"] == "CLOSED" and x.get("profit") is None for x in joined),
            "chronology_preserved": chronology,
        }
        return joined, integrity

    def analyze(self, historical, shadow):
        before_h, before_s = copy.deepcopy(historical), copy.deepcopy(shadow)
        rows, integrity = self.join(historical, shadow)
        closed = [x for x in rows if x["status"] == "CLOSED" and x["profit"] is not None]
        unresolved = [x for x in rows if x["status"] != "CLOSED"]
        score = self._predictor_analysis(closed, "shadow_score", 5.0)
        confidence = self._predictor_analysis(closed, "shadow_confidence", 5.0)
        legacy_score = self._predictor_analysis(closed, "legacy_score", 5.0)
        legacy_confidence = self._predictor_analysis(closed, "legacy_confidence", 5.0)
        symbols = {name: self._stratum([x for x in closed if x["symbol"] == name]) for name in ("EURUSD", "XAUUSD")}
        directions = {name: self._stratum([x for x in closed if x["direction"] == name]) for name in ("BUY", "SELL")}
        temporal = self._temporal(closed)
        components = self._components(closed, rows)
        redundancy = self._redundancy(closed)
        models = {
            "score_only": self._logistic(closed, "shadow_score"),
            "confidence_only": self._logistic(closed, "shadow_confidence"),
        }
        models["combined"] = self._logistic(closed, "shadow_score", "shadow_confidence") if abs(redundancy["pearson_correlation"] or 0) < .9 else {"status": "NOT_RUN_HIGH_COLLINEARITY"}
        outcome = self._outcomes(rows)
        score_verdict = self._verdict(score, symbols, temporal, "shadow_score")
        confidence_verdict = self._verdict(confidence, symbols, temporal, "shadow_confidence")
        conclusions = self._conclusions(score, confidence, components, symbols, temporal)
        result = {
            "schema_version": self.VERSION, "data_integrity": integrity,
            "overall_metrics": outcome, "unresolved_diagnostics": [{k: x[k] for k in ("symbol", "trade_id", "entry_time", "status")} for x in unresolved],
            "score_analysis": score, "confidence_analysis": confidence,
            "win_loss_comparison": {
                field: self._win_loss(closed, field) for field in (
                    "legacy_score", "shadow_score", "legacy_confidence", "shadow_confidence"
                )
            },
            "symbol_analysis": symbols, "direction_analysis": directions,
            "temporal_stability": temporal, "component_analysis": components,
            "score_confidence_redundancy": redundancy, "diagnostic_models": models,
            "predictor_comparison": {
                "legacy_score": legacy_score, "shadow_score": score,
                "legacy_confidence": legacy_confidence, "shadow_confidence": confidence,
            },
            "legacy_baseline": {
                "score_unique_values": len(set(x["legacy_score"] for x in rows)),
                "confidence_unique_values": len(set(x["legacy_confidence"] for x in rows)),
                "score_auc": legacy_score["auc"], "confidence_auc": legacy_confidence["auc"],
                "discrimination": "NULL_CONSTANT_BASELINE" if
                legacy_score["descriptive_statistics"]["unique_value_count"] <= 1 and
                legacy_confidence["descriptive_statistics"]["unique_value_count"] <= 1
                else "MEASURED_BASELINE",
            },
            "verdicts": {"shadow_score": score_verdict, "shadow_confidence": confidence_verdict},
            "conclusions": conclusions,
            "configuration": {"bootstrap_iterations": self.BOOTSTRAP_ITERATIONS, "permutation_iterations": self.PERMUTATION_ITERATIONS, "seed": self.SEED, "minimum_group_sample": self.MIN_GROUP, "closed_only_for_predictive_metrics": True, "unresolved_preserved": True, "mode": "DIAGNOSTIC_ONLY"},
            "diagnostics": {"historical_unchanged": historical == before_h, "shadow_scores_unchanged": shadow == before_s, "production_strategy_changed": False, "future_data_used_for_decisions": False},
        }
        return self._clean(result)

    def _predictor_analysis(self, rows, field, width):
        usable = [x for x in rows if self._numeric(x.get(field)) is not None]
        values = [float(x[field]) for x in usable]
        fixed = self.fixed_groups(rows, field, width)
        quantiles = self.quantile_groups(rows, field)
        labels = [1 if x["profit"] > 0 else 0 for x in usable]
        auc = self.auc(values, labels)
        auc_ci = self.bootstrap_auc(values, labels, field)
        permutation = self.permutation_auc(values, labels, field)
        profits = [float(x["profit"]) for x in usable]
        rho, rho_p = self.spearman(values, profits)
        comparison = self._win_loss(rows, field)
        return {
            "descriptive_statistics": self.descriptive(values, len(rows) - len(values)),
            "fixed_bins": fixed, "quantile_groups": quantiles,
            "auc": auc, "auc_ci_low": auc_ci[0], "auc_ci_high": auc_ci[1],
            "auc_permutation_p_value": permutation, "rank_biserial": 2 * auc - 1 if auc is not None else None,
            "mann_whitney_u": comparison["mann_whitney_u"], "mann_whitney_p_value": comparison["mann_whitney_p_value"],
            "spearman_profit": rho, "spearman_profit_p_value": rho_p,
            "fixed_bin_monotonicity": self.monotonicity(fixed), "quantile_monotonicity": self.monotonicity(quantiles),
        }

    @staticmethod
    def descriptive(values, missing=0):
        quartiles = statistics.quantiles(values, n=4, method="inclusive") if len(values) > 1 else ([values[0]] * 3 if values else [None] * 3)
        return {"count": len(values), "minimum": min(values) if values else None, "maximum": max(values) if values else None,
                "mean": statistics.mean(values) if values else None, "median": statistics.median(values) if values else None,
                "standard_deviation": statistics.stdev(values) if len(values) > 1 else 0.0,
                "first_quartile": quartiles[0], "third_quartile": quartiles[2],
                "unique_value_count": len(set(values)), "missing_value_count": missing}

    def fixed_groups(self, rows, field, width=5.0):
        values = [float(x[field]) for x in rows if x.get(field) is not None]
        if not values: return []
        lower = math.floor(min(values) / width) * width
        upper = math.ceil(max(values) / width) * width
        if upper == max(values): upper += width
        result = []
        edge = lower
        while edge < upper:
            group = [x for x in rows if x.get(field) is not None and edge <= float(x[field]) < edge + width]
            if group: result.append(self._group_metrics(group, field, edge, edge + width, False))
            edge += width
        return result

    def quantile_groups(self, rows, field):
        usable = [x for x in rows if x.get(field) is not None]
        unique = sorted(set(float(x[field]) for x in usable))
        if not unique: return []
        cuts = []
        ordered = sorted(float(x[field]) for x in usable)
        for q in (.2, .4, .6, .8):
            cut = ordered[min(len(ordered) - 1, math.ceil(q * len(ordered)) - 1)]
            if cut not in cuts and cut != unique[-1]: cuts.append(cut)
        groups = defaultdict(list)
        for row in usable:
            value = float(row[field]); index = next((i for i, cut in enumerate(cuts) if value <= cut), len(cuts))
            groups[index].append(row)
        result = []
        for index in sorted(groups):
            group = groups[index]
            result.append(self._group_metrics(group, field, min(float(x[field]) for x in group), max(float(x[field]) for x in group), True))
        return result

    def _group_metrics(self, rows, field, lower, upper, upper_inclusive):
        metrics = self.metrics(rows)
        metrics.update({"score_interval": f"[{lower:.10g}, {upper:.10g}{']' if upper_inclusive else ')'}", "lower": lower, "upper": upper, "upper_inclusive": upper_inclusive,
                        "symbol_composition": dict(sorted(Counter(x["symbol"] for x in rows).items())), "direction_composition": dict(sorted(Counter(x["direction"] for x in rows).items()))})
        return metrics

    def metrics(self, rows):
        profits = [float(x["profit"]) for x in rows]
        wins = sum(x > 0 for x in profits); losses = sum(x <= 0 for x in profits)
        gross_profit = sum(x for x in profits if x > 0); gross_loss = abs(sum(x for x in profits if x < 0))
        low, high = self.wilson(wins, len(profits))
        return {"trade_count": len(rows), "wins": wins, "losses": losses, "win_rate": wins / len(rows) if rows else None,
                "total_net_profit": sum(profits), "average_net_profit": statistics.mean(profits) if profits else None,
                "median_net_profit": statistics.median(profits) if profits else None,
                "profit_factor": gross_profit / gross_loss if gross_loss else None,
                "expectancy": statistics.mean(profits) if profits else None,
                "confidence_interval_low": low, "confidence_interval_high": high,
                "sample_status": "INSUFFICIENT" if len(rows) < self.MIN_GROUP else "SUFFICIENT"}

    def _outcomes(self, rows):
        closed = [x for x in rows if x["status"] == "CLOSED" and x["profit"] is not None]
        profits = [float(x["profit"]) for x in closed]; winners = [x for x in profits if x > 0]; losers = [x for x in profits if x <= 0]
        base = self.metrics(closed)
        base.update({"wins": len(winners), "losses": len(losers), "unresolved": len(rows) - len(closed), "average_winner": statistics.mean(winners) if winners else None, "average_loser": statistics.mean(losers) if losers else None})
        return base

    def _win_loss(self, rows, field):
        wins = [float(x[field]) for x in rows if x["profit"] > 0 and x.get(field) is not None]
        losses = [float(x[field]) for x in rows if x["profit"] <= 0 and x.get(field) is not None]
        u, p = self.mann_whitney(wins, losses)
        return {"winners": self.descriptive(wins), "losers": self.descriptive(losses),
                "winner_count": len(wins), "winner_mean": statistics.mean(wins) if wins else None, "winner_median": statistics.median(wins) if wins else None,
                "loser_count": len(losses), "loser_mean": statistics.mean(losses) if losses else None, "loser_median": statistics.median(losses) if losses else None,
                "mann_whitney_u": u, "mann_whitney_p_value": p}

    def _stratum(self, rows):
        if not rows: return {"trade_count": 0, "status": "INSUFFICIENT"}
        labels = [1 if x["profit"] > 0 else 0 for x in rows]
        result = self.metrics(rows)
        for field in ("shadow_score", "shadow_confidence"):
            values = [float(x[field]) for x in rows]
            result[field] = {"range": [min(values), max(values)], "auc": self.auc(values, labels), "spearman_profit": self.spearman(values, [x["profit"] for x in rows])[0], "direction": self._sign((self.auc(values, labels) or .5) - .5)}
        return result

    def _temporal(self, rows):
        ordered = sorted(rows, key=lambda x: (self._time_value(x["entry_time"]), x["source_index"]))
        midpoint = len(ordered) // 2
        periods = {"first_half": ordered[:midpoint], "second_half": ordered[midpoint:]}
        fold_size = math.ceil(len(ordered) / 4)
        for i in range(4): periods[f"fold_{i + 1}"] = ordered[i * fold_size:(i + 1) * fold_size]
        output = {}
        for name, group in periods.items():
            item = {"trade_count": len(group), "start": group[0]["entry_time"] if group else None, "end": group[-1]["entry_time"] if group else None}
            for field in ("shadow_score", "shadow_confidence"):
                values = [float(x[field]) for x in group]; labels = [1 if x["profit"] > 0 else 0 for x in group]
                item[field] = {"range": [min(values), max(values)] if values else None, "auc": self.auc(values, labels),
                               "spearman_profit": self.spearman(values, [x["profit"] for x in group])[0] if values else None,
                               "fixed_groups": self.fixed_groups(group, field, 5.0)}
            output[name] = item
        return output

    def _components(self, closed, all_rows):
        items = []
        for name in ScoreEngine.COMPONENTS:
            available = [x for x in closed if self._numeric(x["components"].get(name)) is not None]
            values = [float(x["components"][name]) for x in available]
            wins = [float(x["components"][name]) for x in available if x["profit"] > 0]
            losses = [float(x["components"][name]) for x in available if x["profit"] <= 0]
            u, p = self.mann_whitney(wins, losses)
            effect = (statistics.mean(wins) - statistics.mean(losses)) if wins and losses else None
            ci = self.bootstrap_difference(wins, losses, name)
            labels = [1 if x["profit"] > 0 else 0 for x in available]
            profits = [float(x["profit"]) for x in available]
            items.append({"component": name, "availability_count": len(available), "not_available_count": sum(self._numeric(x["components"].get(name)) is None for x in all_rows),
                          "minimum": min(values) if values else None, "maximum": max(values) if values else None, "mean": statistics.mean(values) if values else None,
                          "unique_values": len(set(values)), "contribution_frequency": sum(v != 0 for v in values) / len(values) if values else None,
                          "winner_statistics": self.descriptive(wins), "loser_statistics": self.descriptive(losses),
                          "winner_mean": statistics.mean(wins) if wins else None, "loser_mean": statistics.mean(losses) if losses else None,
                          "winner_expectancy": statistics.mean([x["profit"] for x in available if x["components"][name] != ScoreEngine.NOT_AVAILABLE and x["profit"] > 0]) if wins else None,
                          "loser_expectancy": statistics.mean([x["profit"] for x in available if x["components"][name] != ScoreEngine.NOT_AVAILABLE and x["profit"] <= 0]) if losses else None,
                          "auc": self.auc(values, labels),
                          "spearman_profit": self.spearman(values, profits)[0] if values else None,
                          "effect": effect, "effect_direction": self._sign(effect), "bootstrap_ci_low": ci[0], "bootstrap_ci_high": ci[1], "p_value": p, "mann_whitney_u": u})
        self.apply_bh(items)
        for item in items:
            if item["availability_count"] < self.MIN_GROUP or item["unique_values"] < 2:
                item["classification"] = "INSUFFICIENT_DATA" if item["availability_count"] < self.MIN_GROUP else "NOT_RELIABLE"
            elif item["adjusted_p_value"] is not None and item["adjusted_p_value"] < .05 and item["bootstrap_ci_low"] * item["bootstrap_ci_high"] > 0:
                item["classification"] = "PROMISING_BUT_LIMITED"
            else:
                item["classification"] = "NOT_RELIABLE"
        return items

    def _redundancy(self, rows):
        scores = [float(x["shadow_score"]) for x in rows]; confidence = [float(x["shadow_confidence"]) for x in rows]
        pearson = self.pearson(scores, confidence); spearman = self.spearman(scores, confidence)[0]
        score_auc = self.auc(scores, [x["profit"] > 0 for x in rows]); conf_auc = self.auc(confidence, [x["profit"] > 0 for x in rows])
        return {"pearson_correlation": pearson, "spearman_correlation": spearman, "unique_score_confidence_pairs": len(set(zip(scores, confidence))),
                "evidence_of_redundancy": "HIGH" if abs(pearson or 0) >= .9 else "MODERATE" if abs(pearson or 0) >= .7 else "LOW",
                "confidence_auc_minus_score_auc": conf_auc - score_auc if conf_auc is not None and score_auc is not None else None,
                "confidence_additional_discrimination": "NOT_DEMONSTRATED" if abs((conf_auc or .5) - (score_auc or .5)) < .03 else "POSSIBLE"}

    def _logistic(self, rows, *fields):
        # Diagnostic IRLS model: intercept + symbol + direction + requested predictors.
        if len(rows) < 40: return {"status": "INSUFFICIENT_SAMPLE"}
        columns = ["intercept", "symbol_XAUUSD", "direction_SELL", *fields]
        x = [[1.0, 1.0 if r["symbol"] == "XAUUSD" else 0.0, 1.0 if r["direction"] == "SELL" else 0.0, *[float(r[f]) for f in fields]] for r in rows]
        # Standardize continuous predictors for interpretable and stable diagnostics.
        for j in range(3, len(columns)):
            mean = statistics.mean(row[j] for row in x); sd = statistics.pstdev(row[j] for row in x) or 1.0
            for row in x: row[j] = (row[j] - mean) / sd
        y = [1.0 if r["profit"] > 0 else 0.0 for r in rows]; beta = [0.0] * len(columns)
        try:
            for _ in range(50):
                p = [1 / (1 + math.exp(-max(-30, min(30, sum(a*b for a,b in zip(row,beta)))))) for row in x]
                h = [[sum(x[i][j]*x[i][k]*max(1e-6,p[i]*(1-p[i])) for i in range(len(x))) for k in range(len(columns))] for j in range(len(columns))]
                g = [sum(x[i][j]*(y[i]-p[i]) for i in range(len(x))) for j in range(len(columns))]
                delta = self.solve(h, g); beta = [a+b for a,b in zip(beta,delta)]
                if max(abs(v) for v in delta) < 1e-8: break
            covariance = self.inverse(h); se = [math.sqrt(max(0, covariance[i][i])) for i in range(len(columns))]
            return {"status": "DIAGNOSTIC_ONLY", "trade_count": len(rows), "coefficients": {name: {"coefficient": beta[i], "standard_error": se[i], "ci_low": beta[i]-1.96*se[i], "ci_high": beta[i]+1.96*se[i], "direction": self._sign(beta[i])} for i,name in enumerate(columns)}, "predictors_standardized": list(fields)}
        except (ArithmeticError, ValueError): return {"status": "MODEL_SINGULAR"}

    def _verdict(self, analysis, symbols, temporal, field):
        auc = analysis["auc"]; low = analysis["auc_ci_low"]; high = analysis["auc_ci_high"]
        halves = [temporal[x][field]["auc"] for x in ("first_half", "second_half")]
        positive_halves = all(x is not None and x > .5 for x in halves)
        symbol_aucs = [symbols[x].get(field, {}).get("auc") for x in ("EURUSD", "XAUUSD")]
        positive_symbols = all(x is not None and x >= .5 for x in symbol_aucs)
        if auc is None: return "INSUFFICIENT_DATA"
        if low is not None and low > .5 and positive_halves and positive_symbols and auc >= .58:
            return "RELIABLE"
        if auc >= .55 and (analysis["spearman_profit"] or 0) > 0:
            return "PROMISING_BUT_LIMITED"
        return "NOT_RELIABLE"

    def _conclusions(self, score, confidence, components, symbols, temporal):
        def is_monotonic(analysis, metric):
            finding = analysis["fixed_bin_monotonicity"]
            return isinstance(finding, dict) and finding.get(metric) in ("strictly monotonic", "generally monotonic")

        score_class = self._verdict(score, symbols, temporal, "shadow_score")
        confidence_class = self._verdict(confidence, symbols, temporal, "shadow_confidence")
        reliable_components = [x["component"] for x in components if x["classification"] == "RELIABLE"]
        stable_score = all(
            temporal[name]["shadow_score"]["auc"] is not None and temporal[name]["shadow_score"]["auc"] > .5
            for name in ("first_half", "second_half")
        ) and all(symbols[name]["shadow_score"]["auc"] is not None and symbols[name]["shadow_score"]["auc"] > .5 for name in symbols)
        return {
            "higher_shadow_score_higher_win_rate": {"supported": is_monotonic(score, "win_rate"), "classification": score_class},
            "higher_shadow_score_better_expectancy": {"supported": is_monotonic(score, "expectancy"), "classification": score_class},
            "higher_shadow_confidence_better_outcomes": {"supported": confidence_class in ("RELIABLE", "PROMISING_BUT_LIMITED"), "classification": confidence_class},
            "individual_component_reliable_value": {"supported": bool(reliable_components), "components": reliable_components,
                                                       "classification": "RELIABLE" if reliable_components else "NOT_RELIABLE"},
            "stable_across_time_and_symbols": {"supported": stable_score, "classification": "RELIABLE" if stable_score and score_class == "RELIABLE" else "NOT_RELIABLE"},
            "shadow_score_validated_discriminative_power": {"supported": score_class == "RELIABLE", "classification": score_class},
            "production_use_justified": {"supported": False, "classification": "NOT_RELIABLE",
                                           "reason": "Sprint 83 is diagnostic-only and no stable out-of-sample edge was validated."},
        }

    @staticmethod
    def monotonicity(groups):
        if len(groups) < 3: return "insufficient evidence"
        verdict = {}
        for field in ("win_rate", "expectancy", "profit_factor", "average_net_profit"):
            values = [x[field] for x in groups if x[field] is not None]
            if len(values) < 3: verdict[field] = "insufficient evidence"
            elif all(b > a for a,b in zip(values,values[1:])): verdict[field] = "strictly monotonic"
            elif sum(b >= a for a,b in zip(values,values[1:])) >= max(1, len(values)-2): verdict[field] = "generally monotonic"
            else: verdict[field] = "non-monotonic"
        return verdict

    @staticmethod
    def ranks(values):
        order = sorted(range(len(values)), key=lambda i: values[i]); ranks = [0.0]*len(values); i=0
        while i < len(order):
            j=i+1
            while j < len(order) and values[order[j]] == values[order[i]]: j += 1
            rank=(i+j+1)/2
            for k in range(i,j): ranks[order[k]]=rank
            i=j
        return ranks

    @classmethod
    def auc(cls, values, labels):
        positives=sum(bool(x) for x in labels); negatives=len(labels)-positives
        if not positives or not negatives or len(values) != len(labels): return None
        rank=cls.ranks(values); return (sum(r for r,y in zip(rank,labels) if y)-positives*(positives+1)/2)/(positives*negatives)

    def bootstrap_auc(self, values, labels, key):
        rng=random.Random(self._seed("bootstrap_auc|"+key)); n=len(values); samples=[]
        for _ in range(self.BOOTSTRAP_ITERATIONS):
            indices=[rng.randrange(n) for _ in range(n)]; value=self.auc([values[i] for i in indices],[labels[i] for i in indices])
            if value is not None:samples.append(value)
        return self.percentile_ci(samples)

    def permutation_auc(self, values, labels, key):
        observed=self.auc(values,labels)
        if observed is None:return None
        rng=random.Random(self._seed("permutation|"+key)); perm=list(labels); extreme=0
        for _ in range(self.PERMUTATION_ITERATIONS):
            rng.shuffle(perm); value=self.auc(values,perm); extreme += abs(value-.5) >= abs(observed-.5)
        return (extreme+1)/(self.PERMUTATION_ITERATIONS+1)

    def bootstrap_difference(self, left, right, key):
        if not left or not right:return None,None
        rng=random.Random(self._seed("component|"+key)); values=[]
        for _ in range(self.BOOTSTRAP_ITERATIONS):
            values.append(statistics.mean(left[rng.randrange(len(left))] for _ in left)-statistics.mean(right[rng.randrange(len(right))] for _ in right))
        return self.percentile_ci(values)

    @staticmethod
    def percentile_ci(values):
        if not values:return None,None
        values=sorted(values); return values[int(.025*(len(values)-1))],values[int(.975*(len(values)-1))]

    @classmethod
    def mann_whitney(cls, left, right):
        if not left or not right:return None,None
        combined=left+right;ranks=cls.ranks(combined);u=sum(ranks[:len(left)])-len(left)*(len(left)+1)/2
        n1,n2=len(left),len(right);counts=Counter(combined);tie=sum(c**3-c for c in counts.values());n=n1+n2
        variance=n1*n2/12*((n+1)-tie/(n*(n-1))) if n>1 else 0
        z=(u-n1*n2/2)/math.sqrt(variance) if variance else 0;p=math.erfc(abs(z)/math.sqrt(2))
        return u,p

    @classmethod
    def spearman(cls, left, right):
        if len(left)<3 or len(left)!=len(right):return None,None
        rho=cls.pearson(cls.ranks(left),cls.ranks(right)); n=len(left)
        if rho is None or abs(rho)>=1:return rho,0.0 if rho else 1.0
        t=abs(rho)*math.sqrt((n-2)/(1-rho*rho)); p=math.erfc(t/math.sqrt(2))
        return rho,min(1.0,p)

    @staticmethod
    def pearson(left,right):
        if len(left)<2 or len(left)!=len(right):return None
        ml,mr=statistics.mean(left),statistics.mean(right);num=sum((a-ml)*(b-mr) for a,b in zip(left,right));den=math.sqrt(sum((a-ml)**2 for a in left)*sum((b-mr)**2 for b in right))
        return num/den if den else None

    @staticmethod
    def wilson(wins,n,z=1.959963984540054):
        if not n:return None,None
        p=wins/n;den=1+z*z/n;center=(p+z*z/(2*n))/den;margin=z*math.sqrt(p*(1-p)/n+z*z/(4*n*n))/den
        return max(0,center-margin),min(1,center+margin)

    @staticmethod
    def apply_bh(items):
        valid=sorted([(i,x["p_value"]) for i,x in enumerate(items) if x["p_value"] is not None],key=lambda x:x[1]);m=len(valid);running=1.0
        adjusted={}
        for pos in range(m-1,-1,-1):
            index,p=valid[pos];running=min(running,p*m/(pos+1));adjusted[index]=min(1.0,running)
        for i,item in enumerate(items):item["adjusted_p_value"]=adjusted.get(i)

    @staticmethod
    def solve(matrix, vector):
        n=len(vector);a=[list(matrix[i])+[vector[i]] for i in range(n)]
        for col in range(n):
            pivot=max(range(col,n),key=lambda r:abs(a[r][col]));a[col],a[pivot]=a[pivot],a[col]
            if abs(a[col][col])<1e-12:raise ArithmeticError("singular")
            scale=a[col][col];a[col]=[x/scale for x in a[col]]
            for row in range(n):
                if row!=col:
                    factor=a[row][col];a[row]=[x-factor*y for x,y in zip(a[row],a[col])]
        return [a[i][-1] for i in range(n)]

    @classmethod
    def inverse(cls,matrix):
        n=len(matrix);return [[cls.solve(matrix,[1.0 if i==j else 0.0 for i in range(n)])[k] for j in range(n)] for k in range(n)]

    def write_json(self,result,path):
        path=Path(path);path.parent.mkdir(parents=True,exist_ok=True);path.write_text(json.dumps(result,indent=2,sort_keys=True,default=str,allow_nan=False),encoding="utf-8")

    def write_workbook(self,result,path):
        wb=Workbook();wb.remove(wb.active)
        fixed_timestamp = datetime(2000, 1, 1)
        wb.properties.created = fixed_timestamp
        wb.properties.modified = fixed_timestamp
        summary=wb.create_sheet("Summary");summary.append(["Metric","Value"])
        for key,value in {"Shadow Score Classification":result["verdicts"]["shadow_score"],"Shadow Confidence Classification":result["verdicts"]["shadow_confidence"],"Trades":result["data_integrity"]["total_historical_trades"],"Closed":result["data_integrity"]["closed_trades"],"Unresolved":result["data_integrity"]["unresolved_trades"],"Winners":result["overall_metrics"]["wins"],"Losers":result["overall_metrics"]["losses"],"Shadow Score AUC":result["score_analysis"]["auc"],"Shadow Confidence AUC":result["confidence_analysis"]["auc"],"Mode":"DIAGNOSTIC_ONLY"}.items():summary.append([key,value])
        mappings={
            "Data Validation":{**result["data_integrity"],"unresolved_records":result["unresolved_diagnostics"]},
            "Winner Loser Comparison":[{"predictor":k,**v} for k,v in result["win_loss_comparison"].items()],
            "Shadow Score Buckets":[{"bucket_type":"fixed",**x} for x in result["score_analysis"]["fixed_bins"]]+[{"bucket_type":"quantile",**x} for x in result["score_analysis"]["quantile_groups"]],
            "Confidence Buckets":[{"bucket_type":"fixed",**x} for x in result["confidence_analysis"]["fixed_bins"]]+[{"bucket_type":"quantile",**x} for x in result["confidence_analysis"]["quantile_groups"]],
            "Symbol Analysis":[{"symbol":k,**v} for k,v in result["symbol_analysis"].items()],
            "Temporal Stability":[{"period":k,**v} for k,v in result["temporal_stability"].items()],"Component Analysis":result["component_analysis"],
            "Statistical Tests":[{"predictor":name,"auc":value["auc"],"auc_ci_low":value["auc_ci_low"],"auc_ci_high":value["auc_ci_high"],"auc_permutation_p":value["auc_permutation_p_value"],"mann_whitney_p":value["mann_whitney_p_value"],"spearman_profit":value["spearman_profit"],"spearman_profit_p":value["spearman_profit_p_value"]} for name,value in result["predictor_comparison"].items()],
            "Diagnostics":{**result["diagnostics"],"models":result["diagnostic_models"],"legacy_baseline":result["legacy_baseline"],"redundancy":result["score_confidence_redundancy"],"configuration":result["configuration"]},
            "Conclusions":result["conclusions"],
        }
        for name,value in mappings.items():
            ws=wb.create_sheet(name)
            if isinstance(value,list):self._table(ws,value)
            else:self._key_values(ws,value)
        for ws in wb.worksheets:self._format(ws)
        path=Path(path);path.parent.mkdir(parents=True,exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        wb.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @staticmethod
    def _table(ws,rows):
        if not rows:ws.append(["No data"]);return
        headers=[]
        for row in rows:
            for key in row:
                if key not in headers:headers.append(key)
        ws.append(headers)
        for row in rows:ws.append([json.dumps(row.get(k),sort_keys=True,default=str) if isinstance(row.get(k),(dict,list)) else row.get(k) for k in headers])

    @staticmethod
    def _key_values(ws,values):
        ws.append(["Metric","Value"])
        for key,value in values.items():ws.append([key,json.dumps(value,sort_keys=True,default=str) if isinstance(value,(dict,list)) else value])

    @staticmethod
    def _format(ws):
        ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions;fill=PatternFill("solid",fgColor="1F4E78")
        for c in ws[1]:c.font=Font(color="FFFFFF",bold=True);c.fill=fill
        for col in range(1,ws.max_column+1):ws.column_dimensions[get_column_letter(col)].width=min(60,max(12,max(len(str(ws.cell(r,col).value or "")) for r in range(1,min(ws.max_row,200)+1))+2))

    @staticmethod
    def validate_columns(headers, required, label):
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError(f"{label} is missing required columns: {', '.join(missing)}")

    @staticmethod
    def file_sha256(path):
        digest = hashlib.sha256()
        with Path(path).open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def normalize_xlsx(source, destination):
        """Rewrite the XLSX ZIP container with stable metadata and ordering."""
        with zipfile.ZipFile(source, "r") as input_zip, zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as output_zip:
            for name in sorted(input_zip.namelist()):
                original = input_zip.getinfo(name)
                info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = original.create_system
                info.external_attr = original.external_attr
                info.flag_bits = original.flag_bits
                content = input_zip.read(name)
                if name == "docProps/core.xml":
                    content = re.sub(
                        br"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                        content,
                    )
                output_zip.writestr(info, content, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)

    def _seed(self,key):return self.SEED+int(hashlib.sha256(key.encode()).hexdigest()[:12],16)
    @staticmethod
    def _numeric(value):
        if value in (None,"",ScoreEngine.NOT_AVAILABLE) or isinstance(value,bool):return None
        try:return float(value)
        except (TypeError,ValueError):return None
    @staticmethod
    def _time_value(value):
        if isinstance(value,datetime):return value
        return datetime.fromisoformat(str(value))
    @staticmethod
    def _sign(value):return 1 if value is not None and value>0 else -1 if value is not None and value<0 else 0
    @classmethod
    def _clean(cls,value):
        if isinstance(value,float) and (math.isnan(value) or math.isinf(value)):return None
        if isinstance(value,dict):return {str(k):cls._clean(v) for k,v in value.items()}
        if isinstance(value,(list,tuple)):return [cls._clean(v) for v in value]
        return value
