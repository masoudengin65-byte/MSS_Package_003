"""Sprint 87 diagnostic Smart Money lifecycle evidence collection."""

from __future__ import annotations

import copy
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine
from mss.analysis.displacement_detector import DisplacementDetector
from mss.analysis.liquidity_detector import LiquidityDetector
from mss.analysis.mtf_evidence_engine import HistoricalTimeframeLoader, MTFEvidenceEngine
from mss.analysis.real_swing_engine import RealSwingEngine
from mss.analysis.structure_engine import StructureEngine
from mss.domain.smart_money_evidence import SmartMoneyEvidence


class SmartMoneyEvidenceEngine(MTFEvidenceEngine):
    """Extract lifecycle evidence without entering any production decision path."""

    VERSION = "SPRINT_87_SMART_MONEY_EVIDENCE_V1"
    NOT_AVAILABLE = ContextCaptureEngine.NOT_AVAILABLE
    LOOKBACK_LIMIT = 500
    MINIMUM_CANDLES = 100
    METADATA_FIELDS = (
        "schema_version", "trade_id", "symbol", "direction", "decision_time",
        "session", "source_snapshot_version", "source_context_field_count",
        "source_expanded_field_count", "source_mtf_field_count",
    )
    INPUT_FIELDS = (
        "source", "completed_candle_count", "earliest_candle_open_time",
        "latest_candle_open_time", "latest_candle_close_time", "current_price",
    )
    ORDER_BLOCK_FIELDS = (
        "ob_detected", "ob_direction", "ob_formation_timestamp",
        "ob_detection_timestamp", "ob_source_candles", "ob_age_minutes",
        "ob_age_candles", "ob_distance_from_current_price", "ob_freshness",
        "ob_retest_count", "ob_mitigation_state", "ob_reaction_after_formation",
        "ob_departure_strength", "ob_zone_low", "ob_zone_high",
        "ob_invalidation", "ob_detector_valid_at_detection",
        "ob_detector_mitigated_at_detection",
    )
    FVG_FIELDS = (
        "fvg_detected", "fvg_direction", "fvg_formation_timestamp",
        "fvg_detection_timestamp", "fvg_width", "fvg_age_minutes",
        "fvg_age_candles", "fvg_distance_from_current_price",
        "fvg_fill_percentage", "fvg_partial_fill", "fvg_full_fill",
        "fvg_invalidation", "fvg_lifecycle_state", "fvg_gap_low",
        "fvg_gap_high", "fvg_source_candles", "fvg_detector_valid_at_detection",
        "fvg_detector_filled_at_detection",
    )
    LIQUIDITY_FIELDS = (
        "equal_highs", "equal_lows", "liquidity_pool_location",
        "liquidity_pool_side", "liquidity_formation_timestamp",
        "liquidity_distance_from_current_price", "liquidity_sweep_detected",
        "liquidity_sweep_direction", "liquidity_sweep_timestamp",
        "liquidity_rejection_after_sweep", "liquidity_time_since_sweep_minutes",
        "liquidity_lifecycle_state", "liquidity_source_swings",
        "liquidity_tolerance",
    )
    FIELDS = (
        *METADATA_FIELDS, *INPUT_FIELDS, *ORDER_BLOCK_FIELDS, *FVG_FIELDS,
        *LIQUIDITY_FIELDS,
    )
    REQUIRED_SHEETS = (
        "Summary", "Schema", "Availability", "Order Block Evidence",
        "FVG Evidence", "Liquidity Evidence", "Lifecycle Statistics",
        "Winner Loser Distribution", "Symbol Analysis", "Diagnostics",
        "Limitations",
    )
    OB_STATES = ("FRESH", "MITIGATED", "INVALIDATED", NOT_AVAILABLE)
    FVG_STATES = (
        "FRESH", "PARTIALLY_FILLED", "FILLED", "INVALIDATED", NOT_AVAILABLE,
    )
    LIQUIDITY_STATES = ("UNSWEPT", "SWEPT", "REJECTED", NOT_AVAILABLE)

    def __init__(self, *, minimum_candles=None, lookback_limit=None):
        self.minimum_candles = int(minimum_candles or self.MINIMUM_CANDLES)
        self.lookback_limit = int(lookback_limit or self.LOOKBACK_LIMIT)
        if self.minimum_candles < DisplacementDetector.LOOKBACK + 1:
            raise ValueError("minimum_candles must cover the displacement lookback")
        if self.lookback_limit < self.minimum_candles:
            raise ValueError("lookback_limit cannot be smaller than minimum_candles")
        self.loader = HistoricalTimeframeLoader()
        self.swing_engine = RealSwingEngine()
        self.structure_engine = StructureEngine()

    def load(self, path):
        rows, validation = super().load(path)
        validation["expanded_context_field_count"] = len(ContextExpansionEngine.FIELDS)
        validation["mtf_evidence_field_count"] = len(MTFEvidenceEngine.FIELDS)
        return rows, validation

    def extract(
        self, *, trade_id, symbol, direction, decision_time, frozen_context,
        candles, source_label="HISTORICAL_CANDLES",
    ):
        """Extract one snapshot using completed M15 candles only."""
        frozen_before = copy.deepcopy(frozen_context)
        series_before = self._series_signature({"M15": candles})
        decision = self._time_value(decision_time)
        latest_frozen = self._time_value(frozen_context["latest_visible_candle_time"])
        if latest_frozen > decision:
            raise ValueError("Frozen M15 context contains a future candle")
        selected = self.loader.completed(
            candles, "M15", decision, limit=self.lookback_limit,
        )
        values = {field: self.NOT_AVAILABLE for field in self.FIELDS}
        values.update({
            "schema_version": self.VERSION,
            "trade_id": int(trade_id), "symbol": str(symbol),
            "direction": str(direction).upper(), "decision_time": decision.isoformat(),
            "session": frozen_context.get("session", self.NOT_AVAILABLE),
            "source_snapshot_version": frozen_context.get("snapshot_version", self.NOT_AVAILABLE),
            "source_context_field_count": len(ContextCaptureEngine.FIELDS),
            "source_expanded_field_count": len(ContextExpansionEngine.FIELDS),
            "source_mtf_field_count": len(MTFEvidenceEngine.FIELDS),
            "source": str(source_label) if selected else self.NOT_AVAILABLE,
            "completed_candle_count": len(selected),
        })
        if selected:
            values.update({
                "earliest_candle_open_time": selected[0].time.isoformat(),
                "latest_candle_open_time": selected[-1].time.isoformat(),
                "latest_candle_close_time": self.loader.close_time(selected[-1], "M15").isoformat(),
                "current_price": float(selected[-1].close),
            })
        if len(selected) >= self.minimum_candles:
            ob_candidate, fvg_candidate = self._find_detector_candidates(str(symbol), selected)
            values.update(self._order_block_evidence(ob_candidate, selected, decision))
            values.update(self._fvg_evidence(fvg_candidate, selected, decision))
            values.update(self._liquidity_evidence(selected, decision))
        if frozen_context != frozen_before:
            raise RuntimeError("Immutable frozen context was mutated")
        if self._series_signature({"M15": candles}) != series_before:
            raise RuntimeError("Historical candle input was mutated")
        if set(values) != set(self.FIELDS):
            raise RuntimeError("Smart Money evidence does not match the stable schema")
        return SmartMoneyEvidence.create(values)

    def _find_detector_candidates(self, symbol, candles):
        """Replay existing detectors only on displacement-eligible prefixes."""
        latest_ob = None
        latest_fvg = None
        minimum = DisplacementDetector.LOOKBACK + 1
        for end in range(minimum, len(candles) + 1):
            prefix = candles[:end]
            previous = prefix[-DisplacementDetector.LOOKBACK - 1:-1]
            average = sum(abs(float(item.close) - float(item.open)) for item in previous) / len(previous)
            body = abs(float(prefix[-1].close) - float(prefix[-1].open))
            if average <= 0 or body / average < DisplacementDetector.RATIO:
                continue
            swings = self.swing_engine.detect(prefix)
            analysis = self.structure_engine.analyze(
                symbol=symbol, timeframe="M15", candles=prefix, swings=swings,
            )
            order_block = analysis.order_block
            if getattr(order_block, "candle_time", None) is not None:
                latest_ob = {
                    "object": copy.deepcopy(order_block),
                    "detection_candle": prefix[-1],
                    "departure_strength": float(analysis.displacement.ratio),
                }
            gap = analysis.fair_value_gap
            if getattr(gap, "candle_time", None) is not None and bool(gap.valid):
                latest_fvg = {
                    "object": copy.deepcopy(gap),
                    "detection_candle": prefix[-1],
                }
        return latest_ob, latest_fvg

    def _order_block_evidence(self, candidate, candles, decision):
        result = {field: self.NOT_AVAILABLE for field in self.ORDER_BLOCK_FIELDS}
        result["ob_detected"] = False
        if candidate is None:
            return result
        block = candidate["object"]
        formation_index = self._candle_index(candles, block.candle_time)
        detection_index = self._candle_index(candles, candidate["detection_candle"].time)
        if formation_index is None or detection_index is None:
            return result
        low, high = float(block.low), float(block.high)
        body_low = min(float(block.open), float(block.close))
        body_high = max(float(block.open), float(block.close))
        later = candles[formation_index + 1:]
        touches = [item for item in later if float(item.low) <= high and float(item.high) >= low]
        if str(block.direction).upper() == "BULLISH":
            mitigated = any(float(item.close) <= body_low for item in later)
            reaction = max([float(item.high) - high for item in later] or [0.0])
        else:
            mitigated = any(float(item.close) >= body_high for item in later)
            reaction = max([low - float(item.low) for item in later] or [0.0])
        state = "MITIGATED" if mitigated else "FRESH"
        formation_close = self.loader.close_time(candles[formation_index], "M15")
        result.update({
            "ob_detected": True, "ob_direction": str(block.direction).upper(),
            "ob_formation_timestamp": block.candle_time.isoformat(),
            "ob_detection_timestamp": candidate["detection_candle"].time.isoformat(),
            "ob_source_candles": [
                self._candle_payload(candles[formation_index]),
                self._candle_payload(candles[detection_index]),
            ],
            "ob_age_minutes": (decision - formation_close).total_seconds() / 60.0,
            "ob_age_candles": len(later),
            "ob_distance_from_current_price": self._distance_to_zone(float(candles[-1].close), low, high),
            "ob_freshness": state == "FRESH", "ob_retest_count": len(touches),
            "ob_mitigation_state": state,
            "ob_reaction_after_formation": max(0.0, reaction),
            "ob_departure_strength": candidate["departure_strength"],
            "ob_zone_low": low, "ob_zone_high": high,
            # The existing detector exposes mitigation, not an independent invalidation rule.
            "ob_invalidation": self.NOT_AVAILABLE,
            "ob_detector_valid_at_detection": bool(block.valid),
            "ob_detector_mitigated_at_detection": bool(block.mitigated),
        })
        return result

    def _fvg_evidence(self, candidate, candles, decision):
        result = {field: self.NOT_AVAILABLE for field in self.FVG_FIELDS}
        result["fvg_detected"] = False
        if candidate is None:
            return result
        gap = candidate["object"]
        formation_index = self._candle_index(candles, gap.candle_time)
        detection_index = self._candle_index(candles, candidate["detection_candle"].time)
        if formation_index is None or detection_index is None:
            return result
        low, high = float(gap.low), float(gap.high)
        width = high - low
        if width <= 0:
            return result
        later = candles[detection_index + 1:]
        direction = str(gap.direction).upper()
        if direction == "BULLISH":
            deepest = min([float(item.low) for item in later] or [high])
            fill = 100.0 * (high - deepest) / width
        else:
            deepest = max([float(item.high) for item in later] or [low])
            fill = 100.0 * (deepest - low) / width
        fill = max(0.0, min(100.0, fill))
        full = fill >= 100.0
        partial = 0.0 < fill < 100.0
        state = "FILLED" if full else "PARTIALLY_FILLED" if partial else "FRESH"
        formation_close = self.loader.close_time(candles[formation_index], "M15")
        first_source = max(0, detection_index - 2)
        result.update({
            "fvg_detected": True, "fvg_direction": direction,
            "fvg_formation_timestamp": gap.candle_time.isoformat(),
            "fvg_detection_timestamp": candidate["detection_candle"].time.isoformat(),
            "fvg_width": width,
            "fvg_age_minutes": (decision - formation_close).total_seconds() / 60.0,
            "fvg_age_candles": len(candles) - formation_index - 1,
            "fvg_distance_from_current_price": self._distance_to_zone(float(candles[-1].close), low, high),
            "fvg_fill_percentage": fill, "fvg_partial_fill": partial,
            "fvg_full_fill": full,
            # No independent invalidation rule exists in FairValueGapValidator.
            "fvg_invalidation": self.NOT_AVAILABLE,
            "fvg_lifecycle_state": state, "fvg_gap_low": low,
            "fvg_gap_high": high,
            "fvg_source_candles": [self._candle_payload(item) for item in candles[first_source:detection_index + 1]],
            "fvg_detector_valid_at_detection": bool(gap.valid),
            "fvg_detector_filled_at_detection": bool(gap.filled),
        })
        return result

    def _liquidity_evidence(self, candles, decision):
        result = {field: self.NOT_AVAILABLE for field in self.LIQUIDITY_FIELDS}
        result.update({
            "equal_highs": False, "equal_lows": False,
            "liquidity_sweep_detected": False,
            "liquidity_tolerance": float(LiquidityDetector.TOLERANCE),
        })
        swings = self.swing_engine.detect(candles)
        highs = [item for item in swings if getattr(item, "is_high", False)]
        lows = [item for item in swings if getattr(item, "is_low", False)]
        candidates = []
        for kind, points in (("HIGH", highs), ("LOW", lows)):
            for index in range(1, len(points)):
                first, second = points[index - 1], points[index]
                if abs(float(first.price) - float(second.price)) <= LiquidityDetector.TOLERANCE:
                    candidates.append((second.time, kind, first, second))
        if not candidates:
            return result
        _, kind, first, second = max(candidates, key=lambda item: (item[0], item[1]))
        level = float(second.price)
        after = [item for item in candles if item.time > second.time]
        if kind == "HIGH":
            sweep = next((item for item in after if float(item.high) > level and float(item.close) < level), None)
            side, direction = "BUY_SIDE", "BUY_SIDE"
            result["equal_highs"] = True
        else:
            sweep = next((item for item in after if float(item.low) < level and float(item.close) > level), None)
            side, direction = "SELL_SIDE", "SELL_SIDE"
            result["equal_lows"] = True
        state = "REJECTED" if sweep is not None else "UNSWEPT"
        result.update({
            "liquidity_pool_location": level, "liquidity_pool_side": side,
            "liquidity_formation_timestamp": second.time.isoformat(),
            "liquidity_distance_from_current_price": abs(float(candles[-1].close) - level),
            "liquidity_sweep_detected": sweep is not None,
            "liquidity_sweep_direction": direction if sweep is not None else self.NOT_AVAILABLE,
            "liquidity_sweep_timestamp": sweep.time.isoformat() if sweep is not None else self.NOT_AVAILABLE,
            "liquidity_rejection_after_sweep": True if sweep is not None else False,
            "liquidity_time_since_sweep_minutes": (
                (decision - self.loader.close_time(sweep, "M15")).total_seconds() / 60.0
                if sweep is not None else self.NOT_AVAILABLE
            ),
            "liquidity_lifecycle_state": state,
            "liquidity_source_swings": [self._swing_payload(first), self._swing_payload(second)],
        })
        return result

    def run(self, historical_path, candles_by_symbol, excel_path, json_path, source_metadata=None):
        source = Path(historical_path)
        outputs = {Path(excel_path).resolve(), Path(json_path).resolve()}
        if source.resolve() in outputs:
            raise ValueError("Outputs must not overwrite the immutable historical source")
        source_hash_before = self.file_sha256(source)
        rows, validation = self.load(source)
        series_before = self._nested_series_signature(candles_by_symbol)
        result = self.analyze(rows, validation, candles_by_symbol, source_metadata=source_metadata)
        source_hash_after = self.file_sha256(source)
        if source_hash_before != source_hash_after:
            raise RuntimeError("Immutable historical source changed")
        if self._nested_series_signature(candles_by_symbol) != series_before:
            raise RuntimeError("Historical candle inputs changed")
        result["diagnostics"]["input_artifact"] = {
            "path": str(source), "sha256_before": source_hash_before,
            "sha256_after": source_hash_after, "unchanged": True,
        }
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    def analyze(self, rows, validation, candles_by_symbol, source_metadata=None):
        rows_before = copy.deepcopy(rows)
        extracted = []
        for row in rows:
            symbol_data = candles_by_symbol.get(row["symbol"], {})
            candles = symbol_data.get("M15", ())
            snapshot = self.extract(
                trade_id=row["trade_id"], symbol=row["symbol"], direction=row["direction"],
                decision_time=row["decision_time"], frozen_context=row["frozen_context"],
                candles=candles,
                source_label=self._source_label(source_metadata, row["symbol"]),
            )
            extracted.append({
                "source_index": row["source_index"], "status": row["status"],
                "profit": row["profit"], "context": snapshot.to_dict(),
            })
        if rows != rows_before:
            raise RuntimeError("Historical trade rows were mutated")
        closed = [item for item in extracted if item["status"] == "CLOSED" and item["profit"] is not None]
        winners = [item for item in closed if item["profit"] > 0]
        losers = [item for item in closed if item["profit"] <= 0]
        result = {
            "schema_version": self.VERSION, "mode": "DIAGNOSTIC_ONLY",
            "schema": self._schema(),
            "data_validation": {
                **validation, "smart_money_evidence_field_count": len(self.FIELDS),
                "winner_count": len(winners), "loser_count": len(losers),
                "outcome_analysis_count": len(closed),
            },
            "availability": self._availability_smart_money(extracted, closed),
            "lifecycle_statistics": self._lifecycle_statistics(extracted, closed),
            "winner_loser_distribution": self._winner_loser_smart_money(winners, losers),
            "symbol_analysis": self._dimension_analysis(extracted, "symbol"),
            "session_analysis": self._dimension_analysis(extracted, "session"),
            "temporal_coverage": self._temporal_coverage_smart_money(extracted),
            "records": [item["context"] for item in extracted],
            "diagnostics": {
                "input_objects_unchanged": True, "future_candles_used": False,
                "candle_close_boundary_inclusive": True,
                "outcome_fields_in_evidence_schema": False,
                "production_decision_path_invoked": False,
                "production_decisions_changed": False,
                "existing_detectors_used_read_only": True,
                "existing_detector_logic_modified": False,
                "unresolved_excluded_from_outcome_analysis": len(closed) == validation["closed_trade_count"],
                "source_metadata": source_metadata or {},
                "series": self._series_metadata(candles_by_symbol),
            },
            "limitations": [
                "This evidence layer is diagnostic-only and is not imported by production decision paths.",
                "Only M15 candles whose derived close boundary is at or before the immutable decision time are used.",
                "Order Block and validated FVG candidates originate from the existing detectors replayed chronologically without rule changes.",
                "OB invalidation remains NOT_AVAILABLE because the existing detector exposes mitigation but no independent invalidation rule.",
                "FVG invalidation remains NOT_AVAILABLE because the existing validator exposes fill but no independent invalidation rule.",
                "Liquidity pools and sweeps use the existing absolute liquidity tolerance and sweep/rejection definition.",
                "Lifecycle evidence is limited to the 500 completed candles retained before each decision.",
                "These findings are descriptive and do not validate predictive power or justify production use.",
            ],
            "production_change_justified": False,
        }
        return self._clean(result)

    def _schema(self):
        output = []
        for field in self.FIELDS:
            if field in self.METADATA_FIELDS:
                category = "Metadata"
            elif field in self.INPUT_FIELDS:
                category = "Input"
            elif field in self.ORDER_BLOCK_FIELDS:
                category = "Order Block"
            elif field in self.FVG_FIELDS:
                category = "FVG"
            else:
                category = "Liquidity"
            output.append({
                "field": field, "category": category,
                "not_available_preserved": field not in self.METADATA_FIELDS,
                "diagnostic_only": True,
            })
        return output

    def _availability_smart_money(self, rows, closed):
        output = []
        categories = (
            ("Order Block", "ob_detected"),
            ("FVG", "fvg_detected"),
            ("Liquidity Pool", "liquidity_lifecycle_state"),
        )
        for category, field in categories:
            if field.endswith("detected"):
                available = [item for item in rows if item["context"][field] is True]
                closed_available = [item for item in closed if item["context"][field] is True]
            else:
                available = [item for item in rows if self._available(item["context"][field]) is not None]
                closed_available = [item for item in closed if self._available(item["context"][field]) is not None]
            output.append({
                "row_type": "category", "category": category, "field": field,
                "trade_count": len(rows), "available_count": len(available),
                "missing_count": len(rows) - len(available),
                "availability_percent": 100.0 * len(available) / len(rows) if rows else 0.0,
                "closed_available_count": len(closed_available),
                "winner_available_count": sum(item["profit"] > 0 for item in closed_available),
                "loser_available_count": sum(item["profit"] <= 0 for item in closed_available),
            })
        for field in (*self.ORDER_BLOCK_FIELDS, *self.FVG_FIELDS, *self.LIQUIDITY_FIELDS):
            available = sum(self._available(item["context"][field]) is not None for item in rows)
            output.append({
                "row_type": "field", "category": self._field_category(field),
                "field": field, "trade_count": len(rows), "available_count": available,
                "missing_count": len(rows) - available,
                "availability_percent": 100.0 * available / len(rows) if rows else 0.0,
            })
        return output

    def _lifecycle_statistics(self, rows, closed):
        output = []
        specifications = (
            ("Order Block", "ob_mitigation_state", self.OB_STATES),
            ("FVG", "fvg_lifecycle_state", self.FVG_STATES),
            ("Liquidity", "liquidity_lifecycle_state", self.LIQUIDITY_STATES),
        )
        for category, field, states in specifications:
            for state in states:
                group = [item for item in rows if item["context"][field] == state]
                closed_group = [item for item in closed if item["context"][field] == state]
                output.append({
                    "analysis_level": "overall", "category": category,
                    "segment": "ALL", "state": state, "trade_count": len(group),
                    "closed_count": len(closed_group),
                    "winner_count": sum(item["profit"] > 0 for item in closed_group),
                    "loser_count": sum(item["profit"] <= 0 for item in closed_group),
                })
        for item in self._dimension_analysis(rows, "session"):
            output.append({"analysis_level": "session", **item})
        return output

    def _winner_loser_smart_money(self, winners, losers):
        fields = (
            "ob_detected", "ob_direction", "ob_age_minutes", "ob_retest_count",
            "ob_mitigation_state", "ob_reaction_after_formation", "ob_departure_strength",
            "fvg_detected", "fvg_direction", "fvg_width", "fvg_age_minutes",
            "fvg_fill_percentage", "fvg_lifecycle_state", "equal_highs", "equal_lows",
            "liquidity_pool_side", "liquidity_sweep_detected",
            "liquidity_lifecycle_state", "liquidity_time_since_sweep_minutes",
        )
        return [{
            "field": field,
            "winner_distribution": self._distribution_field(winners, field),
            "loser_distribution": self._distribution_field(losers, field),
        } for field in fields]

    def _dimension_analysis(self, rows, dimension):
        values = sorted({str(item["context"].get(dimension, self.NOT_AVAILABLE)) for item in rows})
        output = []
        for value in values:
            group = [item for item in rows if str(item["context"].get(dimension, self.NOT_AVAILABLE)) == value]
            closed = [item for item in group if item["status"] == "CLOSED" and item["profit"] is not None]
            for category, detected_field, state_field in (
                ("Order Block", "ob_detected", "ob_mitigation_state"),
                ("FVG", "fvg_detected", "fvg_lifecycle_state"),
                ("Liquidity", None, "liquidity_lifecycle_state"),
            ):
                detected = (
                    sum(item["context"][detected_field] is True for item in group)
                    if detected_field else sum(self._available(item["context"][state_field]) is not None for item in group)
                )
                output.append({
                    dimension: value, "category": category, "trade_count": len(group),
                    "closed_count": len(closed),
                    "winner_count": sum(item["profit"] > 0 for item in closed),
                    "loser_count": sum(item["profit"] <= 0 for item in closed),
                    "available_count": detected,
                    "availability_percent": 100.0 * detected / len(group) if group else 0.0,
                    "state_distribution": dict(sorted(Counter(item["context"][state_field] for item in group).items())),
                })
        return output

    def _temporal_coverage_smart_money(self, rows):
        buckets = {}
        for item in rows:
            month = self._time_value(item["context"]["decision_time"]).strftime("%Y-%m")
            buckets.setdefault(month, []).append(item)
        output = []
        for month, group in sorted(buckets.items()):
            output.append({
                "period": month, "trade_count": len(group),
                "ob_available_count": sum(item["context"]["ob_detected"] is True for item in group),
                "fvg_available_count": sum(item["context"]["fvg_detected"] is True for item in group),
                "liquidity_available_count": sum(self._available(item["context"]["liquidity_lifecycle_state"]) is not None for item in group),
                "start": min(item["context"]["decision_time"] for item in group),
                "end": max(item["context"]["decision_time"] for item in group),
            })
        return output

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        summary = workbook.create_sheet("Summary")
        self._key_values(summary, {
            "Mode": "DIAGNOSTIC_ONLY", "Schema Version": self.VERSION,
            "Trades": result["data_validation"]["trade_count"],
            "Closed": result["data_validation"]["closed_trade_count"],
            "Unresolved": result["data_validation"]["unresolved_trade_count"],
            "Winners": result["data_validation"]["winner_count"],
            "Losers": result["data_validation"]["loser_count"],
            "Smart Money Evidence Fields": len(self.FIELDS),
            "Source Context Fields Preserved": len(ContextCaptureEngine.FIELDS),
            "Expanded Context Fields Preserved": len(ContextExpansionEngine.FIELDS),
            "MTF Context Fields Preserved": len(MTFEvidenceEngine.FIELDS),
            "Production Change Justified": False,
        })
        records = result["records"]
        mappings = {
            "Schema": result["schema"], "Availability": result["availability"],
            "Order Block Evidence": self._evidence_records(records, self.ORDER_BLOCK_FIELDS),
            "FVG Evidence": self._evidence_records(records, self.FVG_FIELDS),
            "Liquidity Evidence": self._evidence_records(records, self.LIQUIDITY_FIELDS),
            "Lifecycle Statistics": result["lifecycle_statistics"],
            "Winner Loser Distribution": result["winner_loser_distribution"],
            "Symbol Analysis": result["symbol_analysis"],
            "Diagnostics": {
                **result["data_validation"], **result["diagnostics"],
                "session_analysis": result["session_analysis"],
                "temporal_coverage": result["temporal_coverage"],
            },
            "Limitations": [{"limitation": value} for value in result["limitations"]],
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
        for sheet in workbook.worksheets:
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @classmethod
    def _evidence_records(cls, records, fields):
        keys = ("trade_id", "symbol", "direction", "decision_time", "session")
        return [{key: record[key] for key in (*keys, *fields)} for record in records]

    @classmethod
    def _distribution_field(cls, rows, field):
        values = [item["context"][field] for item in rows if cls._available(item["context"][field]) is not None]
        return cls._distribution(values)

    @classmethod
    def _field_category(cls, field):
        if field in cls.ORDER_BLOCK_FIELDS:
            return "Order Block"
        if field in cls.FVG_FIELDS:
            return "FVG"
        return "Liquidity"

    @staticmethod
    def _distance_to_zone(price, low, high):
        if low <= price <= high:
            return 0.0
        return low - price if price < low else price - high

    @staticmethod
    def _candle_index(candles, timestamp):
        for index, candle in enumerate(candles):
            if candle.time == timestamp:
                return index
        return None

    @staticmethod
    def _candle_payload(candle):
        return {
            "time": candle.time.isoformat(), "open": float(candle.open),
            "high": float(candle.high), "low": float(candle.low),
            "close": float(candle.close),
        }

    @staticmethod
    def _swing_payload(swing):
        return {
            "time": swing.time.isoformat(), "price": float(swing.price),
            "kind": "HIGH" if getattr(swing, "is_high", False) else "LOW",
            "index": int(swing.index),
        }

    @staticmethod
    def _source_label(metadata, symbol):
        if not metadata:
            return "HISTORICAL_CANDLES"
        resolved = metadata.get("symbols", {}).get(symbol, {}).get("resolved_symbol", symbol)
        return f"MT5:{resolved}:M15"
