"""Sprint 90 diagnostic-only historical multi-asset dataset preparation."""

from __future__ import annotations

import hashlib
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment

from mss.analysis.multi_asset_registry import MultiAssetRegistry
from mss.domain.multi_asset_dataset import MultiAssetDataset


class MultiAssetDatasetBuilder(MultiAssetRegistry):
    """Build immutable completed-candle datasets without production consumption."""

    VERSION = "SPRINT_90_MULTI_ASSET_DATASET_V1"
    MODE = "DIAGNOSTIC_ONLY"
    REQUIRED_SHEETS = (
        "Summary", "Dataset Metadata", "Coverage", "Data Quality",
        "M15 Candles", "H1 Candles", "H4 Candles", "D1 Candles",
        "Issues", "Configuration", "Diagnostics", "Audit",
    )
    RESULT_KEYS = (
        "schema_version", "mode", "generated_as_of", "supported_assets",
        "summary", "dataset_metadata", "coverage", "data_quality",
        "issues", "records", "configuration", "diagnostics", "audit",
        "production_change_justified",
    )
    RECORD_FIELDS = (
        "record_id", "canonical_symbol", "broker_symbol", "asset_class",
        "timeframe", "source_index", "open_time", "close_time", "open",
        "high", "low", "close", "tick_volume", "spread", "real_volume",
        "quality_flags",
    )

    def build(self, broker_symbols, history_results, as_of, runtime_metadata=None):
        as_of = self._timestamp(as_of)
        broker_symbols = tuple(broker_symbols or ())
        history_results = history_results or {}
        input_before = self.input_sha256(broker_symbols, history_results)
        assets, coverage_rows, records, issues = [], [], [], []

        for definition in self.universe:
            canonical = definition.canonical_symbol
            broker_symbol = self.resolve_symbol(canonical, broker_symbols)
            assets.append({
                "canonical_symbol": canonical,
                "broker_symbol": broker_symbol or self.NOT_AVAILABLE,
                "asset_class": definition.asset_class,
                "base_asset": definition.base_asset,
                "quote_asset": definition.quote_asset,
                "resolution_status": "RESOLVED" if broker_symbol else self.NOT_AVAILABLE,
            })
            if broker_symbol is None:
                issues.append(self._issue(
                    canonical, "ALL", "SYMBOL_NOT_RESOLVED", "CRITICAL",
                    "No deterministic broker symbol is available for dataset collection.",
                ))
            symbol_payload = self._value(history_results, canonical, {}) or {}
            for timeframe in self.TIMEFRAMES:
                payload = self._value(symbol_payload, timeframe, {}) or {}
                coverage = self.validate_history(
                    definition, timeframe, payload, as_of,
                    default_resolved=broker_symbol,
                )
                enriched = self._enrich_coverage(coverage, payload)
                coverage_rows.append(enriched)
                issues.extend(self._dataset_issues(enriched))
                records.extend(self._candle_records(
                    definition, broker_symbol, timeframe, payload,
                ))

        records = sorted(records, key=self._record_sort_key)
        coverage_rows = sorted(
            coverage_rows,
            key=lambda row: (
                self.supported_symbols.index(row["canonical_symbol"]),
                self.TIMEFRAMES.index(row["timeframe"]),
            ),
        )
        issues = sorted(
            issues,
            key=lambda row: (
                -self.SEVERITY_ORDER[row["severity"]],
                self.supported_symbols.index(row["canonical_symbol"]),
                row["timeframe"], row["issue_code"],
            ),
        )
        input_after = self.input_sha256(broker_symbols, history_results)
        if input_before != input_after:
            raise RuntimeError("Historical multi-asset inputs were mutated")

        records_sha256 = self._json_sha256(records)
        coverage_sha256 = self._json_sha256(coverage_rows)
        duplicate_keys = self._duplicate_record_keys(records)
        quality = self._data_quality(coverage_rows, records, issues, duplicate_keys)
        summary = self._summary(assets, coverage_rows, records, issues)
        payload = {
            "schema_version": self.VERSION,
            "mode": self.MODE,
            "generated_as_of": as_of.isoformat(),
            "supported_assets": assets,
            "summary": summary,
            "dataset_metadata": {
                "grain": "canonical_symbol + timeframe + candle_open_time",
                "canonical_symbol_count": len(assets),
                "asset_class_count": len({row["asset_class"] for row in assets}),
                "timeframe_count": len(self.TIMEFRAMES),
                "record_count": len(records),
                "source_snapshot_sha256": input_before,
                "records_sha256": records_sha256,
                "coverage_sha256": coverage_sha256,
                "immutable_snapshot": True,
                "completed_candles_only": True,
                "lookahead_permitted": False,
            },
            "coverage": coverage_rows,
            "data_quality": quality,
            "issues": issues,
            "records": records,
            "configuration": {
                "timeframes": list(self.TIMEFRAMES),
                "broker_aliases": {
                    key: list(value) for key, value in sorted(self.broker_aliases.items())
                },
                "future_candle_policy": "REJECT",
                "gap_policy": "REPORT_WITHOUT_IMPUTATION",
                "duplicate_policy": "REPORT_WITHOUT_DEDUPLICATION",
                "invalid_ohlc_policy": "REPORT_WITHOUT_RECONSTRUCTION",
                "production_consumption": False,
            },
            "diagnostics": {
                "input_snapshot_unchanged": input_before == input_after,
                "future_candle_count": sum(row["future_candle_count"] for row in coverage_rows),
                "lookahead_violation_count": 0,
                "duplicate_record_key_count": duplicate_keys,
                "record_schema_field_count": len(self.RECORD_FIELDS),
                "no_strategy_imports": True,
                "production_decision_consumption": False,
            },
            "audit": {
                "registry_schema_version": MultiAssetRegistry.VERSION,
                "runtime_metadata": self._clean(runtime_metadata or {}),
                "input_snapshot_sha256_before": input_before,
                "input_snapshot_sha256_after": input_after,
                "input_snapshot_unchanged": input_before == input_after,
                "records_sha256": records_sha256,
                "coverage_sha256": coverage_sha256,
                "source_values_imputed": False,
                "source_rows_removed": False,
            },
            "production_change_justified": False,
        }
        self.validate_dataset_schema(payload)
        return MultiAssetDataset.create(payload)

    def run(self, broker_symbols, history_results, as_of, excel_path, json_path, runtime_metadata=None):
        dataset = self.build(
            broker_symbols, history_results, as_of, runtime_metadata,
        )
        result = dataset.to_dict()
        self.write_json(result, json_path)
        self.write_workbook(result, excel_path)
        return result

    @classmethod
    def validate_dataset_schema(cls, result):
        if tuple(result) != cls.RESULT_KEYS:
            raise ValueError("Multi-asset dataset result schema is invalid")
        if result["mode"] != cls.MODE:
            raise ValueError("Dataset mode must remain diagnostic-only")
        if result["production_change_justified"] is not False:
            raise ValueError("Dataset production guardrail is invalid")
        if len(result["supported_assets"]) != len(cls.TARGET_UNIVERSE):
            raise ValueError("Dataset asset universe is incomplete")
        expected_coverage = len(cls.TARGET_UNIVERSE) * len(cls.TIMEFRAMES)
        if len(result["coverage"]) != expected_coverage:
            raise ValueError("Dataset timeframe coverage is incomplete")
        if any(tuple(record) != cls.RECORD_FIELDS for record in result["records"]):
            raise ValueError("Dataset candle record schema is invalid")
        if result["diagnostics"]["future_candle_count"]:
            raise ValueError("Dataset contains future candles")
        if result["diagnostics"]["lookahead_violation_count"]:
            raise ValueError("Dataset contains lookahead violations")

    def _enrich_coverage(self, coverage, payload):
        timeframe = coverage["timeframe"]
        duration = self.DURATIONS[timeframe]
        candles = tuple(self._value(payload, "candles", ()) or ())
        times = [self._timestamp(self._value(candle, "time")) for candle in candles]
        interval_seconds = duration.total_seconds()
        missing_intervals = 0
        irregular_gaps = 0
        for left, right in zip(times, times[1:]):
            delta = (right - left).total_seconds()
            if delta > interval_seconds:
                missing_intervals += max(0, int(delta // interval_seconds) - 1)
                if delta % interval_seconds:
                    irregular_gaps += 1
        return {
            **coverage,
            "expected_interval_minutes": int(interval_seconds // 60),
            "missing_candle_interval_count": missing_intervals,
            "irregular_gap_count": irregular_gaps,
            "record_count_matches_returned_count": (
                coverage["observed_candle_count"] == coverage["returned_count"]
            ),
        }

    def _dataset_issues(self, row):
        issues = [
            item for item in self._coverage_issues(row)
            if item["issue_code"] != "TIME_GAPS"
        ]
        if row["missing_candle_interval_count"]:
            severity = "HIGH" if row["asset_class"] == "CRYPTO" else "LOW"
            issues.append(self._issue(
                row["canonical_symbol"], row["timeframe"],
                "MISSING_CANDLE_INTERVALS", severity,
                f"Detected {row['missing_candle_interval_count']} absent expected intervals; no candles were imputed.",
            ))
        if row["irregular_gap_count"]:
            issues.append(self._issue(
                row["canonical_symbol"], row["timeframe"],
                "IRREGULAR_GAPS", "HIGH",
                f"Detected {row['irregular_gap_count']} gaps that are not exact timeframe multiples.",
            ))
        return issues

    def _candle_records(self, definition, broker_symbol, timeframe, payload):
        duration = self.DURATIONS[timeframe]
        candles = tuple(self._value(payload, "candles", ()) or ())
        times = [self._timestamp(self._value(candle, "time")) for candle in candles]
        counts = {}
        for value in times:
            counts[value] = counts.get(value, 0) + 1
        output = []
        prior = None
        resolved = (
            self._value(payload, "resolved_symbol", None)
            or broker_symbol
            or self.NOT_AVAILABLE
        )
        for index, (candle, open_time) in enumerate(zip(candles, times)):
            values = {
                field: self._finite_number(self._value(candle, field))
                for field in ("open", "high", "low", "close")
            }
            flags = []
            if any(value is None for value in values.values()):
                flags.append("NONFINITE_PRICE")
            elif (
                values["high"] < max(values["open"], values["close"])
                or values["low"] > min(values["open"], values["close"])
                or values["low"] > values["high"]
            ):
                flags.append("INVALID_OHLC")
            tick_volume = self._finite_number(self._value(candle, "tick_volume"))
            real_volume = self._finite_number(self._value(candle, "real_volume"))
            spread = self._finite_number(self._value(candle, "spread"))
            if (tick_volume is not None and tick_volume < 0) or (real_volume is not None and real_volume < 0):
                flags.append("NEGATIVE_VOLUME")
            if spread is not None and spread < 0:
                flags.append("NEGATIVE_SPREAD")
            if counts[open_time] > 1:
                flags.append("DUPLICATE_TIMESTAMP")
            if prior is not None and open_time <= prior:
                flags.append("NON_CHRONOLOGICAL")
            prior = open_time
            record_key = f"{definition.canonical_symbol}|{timeframe}|{open_time.isoformat()}|{index}"
            output.append({
                "record_id": hashlib.sha256(record_key.encode("utf-8")).hexdigest(),
                "canonical_symbol": definition.canonical_symbol,
                "broker_symbol": resolved,
                "asset_class": definition.asset_class,
                "timeframe": timeframe,
                "source_index": index,
                "open_time": open_time.isoformat(),
                "close_time": (open_time + duration).isoformat(),
                "open": values["open"],
                "high": values["high"],
                "low": values["low"],
                "close": values["close"],
                "tick_volume": self._number_value(tick_volume),
                "spread": self._number_value(spread),
                "real_volume": self._number_value(real_volume),
                "quality_flags": "|".join(flags) if flags else "PASS",
            })
        return output

    def _record_sort_key(self, row):
        return (
            self.supported_symbols.index(row["canonical_symbol"]),
            self.TIMEFRAMES.index(row["timeframe"]),
            row["open_time"], row["source_index"],
        )

    @staticmethod
    def _number_value(value):
        if value is None:
            return None
        return int(value) if float(value).is_integer() else value

    @staticmethod
    def _json_sha256(value):
        raw = json.dumps(
            value, sort_keys=True, separators=(",", ":"),
            default=str, allow_nan=False,
        ).encode("utf-8")
        return hashlib.sha256(raw).hexdigest()

    @staticmethod
    def _duplicate_record_keys(records):
        keys = [
            (row["canonical_symbol"], row["timeframe"], row["open_time"])
            for row in records
        ]
        return len(keys) - len(set(keys))

    def _data_quality(self, coverage, records, issues, duplicate_keys):
        return {
            "record_count": len(records),
            "valid_record_count": sum(row["quality_flags"] == "PASS" for row in records),
            "flagged_record_count": sum(row["quality_flags"] != "PASS" for row in records),
            "chronology_failure_count": sum(not row["chronological_order"] for row in coverage),
            "duplicate_timestamp_count": sum(row["duplicate_timestamp_count"] for row in coverage),
            "duplicate_record_key_count": duplicate_keys,
            "missing_candle_interval_count": sum(row["missing_candle_interval_count"] for row in coverage),
            "irregular_gap_count": sum(row["irregular_gap_count"] for row in coverage),
            "nonfinite_price_count": sum(row["nonfinite_price_count"] for row in coverage),
            "invalid_ohlc_count": sum(row["invalid_ohlc_count"] for row in coverage),
            "negative_volume_count": sum(row["negative_volume_count"] for row in coverage),
            "negative_spread_count": sum(row["negative_spread_count"] for row in coverage),
            "future_candle_count": sum(row["future_candle_count"] for row in coverage),
            "critical_issue_count": sum(row["severity"] == "CRITICAL" for row in issues),
            "high_issue_count": sum(row["severity"] == "HIGH" for row in issues),
            "medium_issue_count": sum(row["severity"] == "MEDIUM" for row in issues),
            "low_issue_count": sum(row["severity"] == "LOW" for row in issues),
        }

    @staticmethod
    def _summary(assets, coverage, records, issues):
        return {
            "supported_asset_count": len(assets),
            "resolved_asset_count": sum(row["resolution_status"] == "RESOLVED" for row in assets),
            "timeframe_slice_count": len(coverage),
            "available_timeframe_slice_count": sum(row["availability_status"] != "MISSING" for row in coverage),
            "missing_timeframe_slice_count": sum(row["availability_status"] == "MISSING" for row in coverage),
            "candle_record_count": len(records),
            "data_quality_issue_count": len(issues),
            "production_change_justified": False,
        }

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        mappings = {
            "Summary": result["summary"],
            "Dataset Metadata": result["dataset_metadata"],
            "Coverage": result["coverage"],
            "Data Quality": result["data_quality"],
            "M15 Candles": self._workbook_records(result["records"], "M15"),
            "H1 Candles": self._workbook_records(result["records"], "H1"),
            "H4 Candles": self._workbook_records(result["records"], "H4"),
            "D1 Candles": self._workbook_records(result["records"], "D1"),
            "Issues": result["issues"],
            "Configuration": result["configuration"],
            "Diagnostics": result["diagnostics"],
            "Audit": self._workbook_audit(result["audit"]),
        }
        for name, values in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, values) if isinstance(values, list) else self._key_values(sheet, values)
            if name.endswith("Candles"):
                self._format_candle_sheet(sheet)
            self._format(sheet)
            if not isinstance(values, list):
                self._format_key_value_sheet(sheet)
            sheet.sheet_view.showGridLines = False
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @classmethod
    def _workbook_records(cls, records, timeframe):
        return [
            {field: row.get(field) for field in cls.RECORD_FIELDS}
            for row in records if row["timeframe"] == timeframe
        ]

    @staticmethod
    def _workbook_audit(audit):
        output = {
            key: value for key, value in audit.items()
            if key != "runtime_metadata"
        }
        for key, value in sorted((audit.get("runtime_metadata") or {}).items()):
            output[f"runtime_{key}"] = value
        return output

    @staticmethod
    def _format_key_value_sheet(sheet):
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            length = len(str(cell.value or ""))
            if length > 55:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                sheet.row_dimensions[row].height = min(60, 15 * (1 + length // 55))

    @staticmethod
    def _format_candle_sheet(sheet):
        headers = {cell.value: cell.column for cell in sheet[1]}
        for field in ("open_time", "close_time"):
            column = headers.get(field)
            if column:
                for row in range(2, sheet.max_row + 1):
                    value = sheet.cell(row, column).value
                    if isinstance(value, str):
                        sheet.cell(row, column).value = datetime.fromisoformat(value)
                    sheet.cell(row, column).number_format = "yyyy-mm-dd hh:mm:ss"
        for field in ("open", "high", "low", "close"):
            column = headers.get(field)
            if column:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "0.00000"
        for field in ("tick_volume", "spread", "real_volume", "source_index"):
            column = headers.get(field)
            if column:
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, column).number_format = "#,##0"
