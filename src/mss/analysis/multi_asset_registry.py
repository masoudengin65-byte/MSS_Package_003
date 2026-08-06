"""Sprint 89 diagnostic-only multi-asset data expansion and validation."""

from __future__ import annotations

import hashlib
import json
import math
import re
import zipfile
from collections.abc import Iterable, Mapping
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mss.domain.asset_metadata import AssetDefinition, AssetMetadata


class MultiAssetRegistry:
    """Validate broker and candle coverage without entering a decision path."""

    VERSION = "SPRINT_89_1_MULTI_ASSET_DATA_V1"
    MODE = "DIAGNOSTIC_ONLY"
    NOT_AVAILABLE = "NOT_AVAILABLE"
    TIMEFRAMES = ("M15", "H1", "H4", "D1")
    DURATIONS = {
        "M15": timedelta(minutes=15),
        "H1": timedelta(hours=1),
        "H4": timedelta(hours=4),
        "D1": timedelta(days=1),
    }
    TARGET_UNIVERSE = (
        AssetDefinition("EURUSD", "FOREX", "EUR", "USD"),
        AssetDefinition("GBPUSD", "FOREX", "GBP", "USD"),
        AssetDefinition("USDJPY", "FOREX", "USD", "JPY"),
        AssetDefinition("AUDUSD", "FOREX", "AUD", "USD"),
        AssetDefinition("USDCAD", "FOREX", "USD", "CAD"),
        AssetDefinition("XAUUSD", "METAL", "XAU", "USD"),
        AssetDefinition("BTCUSD", "CRYPTO", "BTC", "USD"),
        AssetDefinition("ETHUSD", "CRYPTO", "ETH", "USD"),
    )
    BROKER_ALIASES = {
        "BTCUSD": ("BITCOIN",),
        "ETHUSD": ("ETHEREUM",),
    }
    TRADE_MODE_NAMES = {
        0: "DISABLED",
        1: "LONG_ONLY",
        2: "SHORT_ONLY",
        3: "CLOSE_ONLY",
        4: "FULL",
    }
    SYMBOL_INFO_FIELDS = (
        "name", "description", "path", "digits", "point", "spread",
        "trade_mode", "visible", "select", "volume_min", "volume_max",
        "volume_step", "volume_limit", "trade_contract_size",
        "trade_tick_size", "trade_tick_value", "trade_stops_level",
        "trade_freeze_level", "filling_mode", "order_mode", "swap_mode",
        "swap_long", "swap_short", "currency_base", "currency_profit",
        "currency_margin",
    )
    REQUIRED_SHEETS = (
        "Summary", "Symbol Registry", "History Availability",
        "Timeframe Coverage", "Candle Quality", "Trading Conditions",
        "Per-Symbol Quality", "Data Quality Issues", "Configuration",
        "Diagnostics", "Audit",
    )
    RESULT_KEYS = (
        "schema_version", "mode", "generated_as_of", "target_universe",
        "summary", "asset_class_summary", "symbol_registry",
        "history_availability", "timeframe_coverage", "candle_quality",
        "trading_conditions", "symbol_data_quality", "data_quality_issues",
        "configuration", "diagnostics", "audit",
        "production_change_justified",
    )
    SEVERITY_ORDER = {"PASS": 0, "LOW": 1, "MEDIUM": 2, "HIGH": 3, "CRITICAL": 4}

    def __init__(self, target_universe=None):
        self._universe = tuple(target_universe or self.TARGET_UNIVERSE)
        self.validate_registry(self._universe)
        supported = {item.canonical_symbol for item in self._universe}
        self._broker_aliases = {
            canonical: tuple(aliases)
            for canonical, aliases in self.BROKER_ALIASES.items()
            if canonical in supported
        }
        self.validate_aliases(self._universe, self._broker_aliases)

    @property
    def universe(self) -> tuple[AssetDefinition, ...]:
        return self._universe

    @property
    def supported_symbols(self) -> tuple[str, ...]:
        return tuple(item.canonical_symbol for item in self._universe)

    @property
    def broker_aliases(self) -> dict[str, tuple[str, ...]]:
        return dict(self._broker_aliases)

    @classmethod
    def validate_registry(cls, definitions) -> None:
        definitions = tuple(definitions)
        if not definitions:
            raise ValueError("The multi-asset registry cannot be empty")
        symbols = [item.canonical_symbol for item in definitions]
        if len(symbols) != len(set(symbols)):
            raise ValueError("Canonical symbols must be unique")
        for item in definitions:
            if item.asset_class not in {"FOREX", "METAL", "CRYPTO"}:
                raise ValueError(f"Unsupported asset class: {item.asset_class}")
            if not re.fullmatch(r"[A-Z0-9]+", item.canonical_symbol):
                raise ValueError(f"Invalid canonical symbol: {item.canonical_symbol}")
            if not item.base_asset or not item.quote_asset:
                raise ValueError(f"Missing asset identity for {item.canonical_symbol}")

    @classmethod
    def validate_aliases(cls, definitions, aliases) -> None:
        canonical_symbols = {item.canonical_symbol for item in definitions}
        seen = set()
        for canonical, values in aliases.items():
            if canonical not in canonical_symbols:
                raise ValueError(f"Alias mapping references an unregistered symbol: {canonical}")
            if not values:
                raise ValueError(f"Alias mapping cannot be empty: {canonical}")
            for alias in values:
                if not re.fullmatch(r"[A-Z0-9]+", alias):
                    raise ValueError(f"Invalid broker alias for {canonical}: {alias}")
                if alias in canonical_symbols and alias != canonical:
                    raise ValueError(f"Broker alias collides with canonical symbol: {alias}")
                if alias in seen:
                    raise ValueError(f"Broker aliases must be unique: {alias}")
                seen.add(alias)

    @staticmethod
    def _value(source, name, default=None):
        if isinstance(source, Mapping):
            return source.get(name, default)
        return getattr(source, name, default)

    @classmethod
    def _broker_name(cls, item) -> str:
        value = item if isinstance(item, str) else cls._value(item, "name", "")
        return str(value or "")

    def resolve_symbol(self, canonical_symbol: str, broker_symbols: Iterable) -> str | None:
        """Resolve canonical forms, then explicit aliases, using stable ranking."""
        canonical = str(canonical_symbol).upper()
        if canonical not in self.supported_symbols:
            raise KeyError(f"Symbol is not registered: {canonical_symbol}")
        candidates = []
        roots = ((canonical, 0),) + tuple(
            (alias, 10 + index * 10)
            for index, alias in enumerate(self._broker_aliases.get(canonical, ()))
        )
        for item in broker_symbols or ():
            name = self._broker_name(item)
            upper = name.upper()
            compact = re.sub(r"[^A-Z0-9]", "", upper)
            for root, base_rank in roots:
                if upper == root:
                    rank = base_rank
                elif compact == root:
                    rank = base_rank + 1
                elif upper.startswith(root) or compact.startswith(root):
                    rank = base_rank + 2
                elif upper.endswith(root) or compact.endswith(root):
                    rank = base_rank + 3
                elif root in upper or root in compact:
                    rank = base_rank + 4
                else:
                    continue
                candidates.append((rank, len(name), upper, name))
        return min(candidates)[3] if candidates else None

    def capture_metadata(self, definition: AssetDefinition, broker_info) -> AssetMetadata:
        resolved = self._broker_name(broker_info) if broker_info is not None else ""
        value = lambda name: self._value(broker_info, name, None) if broker_info is not None else None
        point, spread, trade_mode = value("point"), value("spread"), value("trade_mode")
        spread_price = None
        if self._finite_number(point) is not None and self._finite_number(spread) is not None:
            spread_price = self._finite_number(point) * self._finite_number(spread)
        trade_allowed = None if trade_mode is None else int(trade_mode) != 0
        selected = value("select")
        if selected is None:
            selected = value("selected")
        normalized = lambda item: self.NOT_AVAILABLE if item is None else item
        return AssetMetadata(
            canonical_symbol=definition.canonical_symbol,
            asset_class=definition.asset_class,
            base_asset=definition.base_asset,
            quote_asset=definition.quote_asset,
            broker_symbol=resolved or self.NOT_AVAILABLE,
            resolved_symbol=resolved or self.NOT_AVAILABLE,
            resolution_status="RESOLVED" if resolved else "NOT_AVAILABLE",
            description=normalized(value("description")),
            broker_path=normalized(value("path")),
            digits=normalized(value("digits")),
            point=normalized(point),
            spread_points=normalized(spread),
            spread_price=normalized(spread_price),
            trade_mode=normalized(trade_mode),
            trade_mode_name=normalized(self.TRADE_MODE_NAMES.get(trade_mode)),
            trade_allowed=normalized(trade_allowed),
            visible=normalized(value("visible")),
            selected=normalized(selected),
            volume_min=normalized(value("volume_min")),
            volume_max=normalized(value("volume_max")),
            volume_step=normalized(value("volume_step")),
            volume_limit=normalized(value("volume_limit")),
            trade_contract_size=normalized(value("trade_contract_size")),
            trade_tick_size=normalized(value("trade_tick_size")),
            trade_tick_value=normalized(value("trade_tick_value")),
            trade_stops_level=normalized(value("trade_stops_level")),
            trade_freeze_level=normalized(value("trade_freeze_level")),
            filling_mode=normalized(value("filling_mode")),
            order_mode=normalized(value("order_mode")),
            swap_mode=normalized(value("swap_mode")),
            swap_long=normalized(value("swap_long")),
            swap_short=normalized(value("swap_short")),
            currency_base=normalized(value("currency_base")),
            currency_profit=normalized(value("currency_profit")),
            currency_margin=normalized(value("currency_margin")),
        )

    def analyze(self, broker_symbols, history_results, as_of, runtime_metadata=None) -> dict:
        as_of = self._timestamp(as_of)
        broker_symbols = tuple(broker_symbols or ())
        history_results = history_results or {}
        input_before = self.input_sha256(broker_symbols, history_results)
        info_by_name = {self._broker_name(item): item for item in broker_symbols}
        registry_rows, trading_rows, coverage_rows, issues = [], [], [], []

        for definition in self._universe:
            resolved = self.resolve_symbol(definition.canonical_symbol, broker_symbols)
            broker_info = info_by_name.get(resolved) if resolved else None
            metadata = self.capture_metadata(definition, broker_info)
            metadata_row = metadata.to_dict()
            registry_rows.append({
                "canonical_symbol": definition.canonical_symbol,
                "asset_class": definition.asset_class,
                "base_asset": definition.base_asset,
                "quote_asset": definition.quote_asset,
                "broker_symbol": metadata.broker_symbol,
                "resolved_symbol": metadata.resolved_symbol,
                "resolution_status": metadata.resolution_status,
            })
            trading_rows.append(metadata_row)
            if resolved is None:
                issues.append(self._issue(
                    definition.canonical_symbol, "ALL", "SYMBOL_NOT_RESOLVED",
                    "CRITICAL", "No deterministic broker-symbol match was found.",
                ))
            symbol_payload = self._value(history_results, definition.canonical_symbol, {}) or {}
            for timeframe in self.TIMEFRAMES:
                payload = self._value(symbol_payload, timeframe, {}) or {}
                coverage_rows.append(self.validate_history(
                    definition=definition,
                    timeframe=timeframe,
                    payload=payload,
                    as_of=as_of,
                    default_resolved=resolved,
                ))

        for row in coverage_rows:
            issues.extend(self._coverage_issues(row))

        symbol_quality = self._symbol_quality(registry_rows, coverage_rows, issues)
        input_after = self.input_sha256(broker_symbols, history_results)
        if input_before != input_after:
            raise RuntimeError("Multi-asset validation mutated its input snapshot")

        result = {
            "schema_version": self.VERSION,
            "mode": self.MODE,
            "generated_as_of": as_of.isoformat(),
            "target_universe": [item.to_dict() for item in self._universe],
            "summary": self._summary(registry_rows, coverage_rows, symbol_quality, issues),
            "asset_class_summary": self._asset_class_summary(registry_rows, coverage_rows, issues),
            "symbol_registry": registry_rows,
            "history_availability": [self._availability_row(row) for row in coverage_rows],
            "timeframe_coverage": coverage_rows,
            "candle_quality": [self._quality_row(row) for row in coverage_rows],
            "trading_conditions": trading_rows,
            "symbol_data_quality": symbol_quality,
            "data_quality_issues": sorted(
                issues,
                key=lambda row: (
                    -self.SEVERITY_ORDER[row["severity"]], row["canonical_symbol"],
                    row["timeframe"], row["issue_code"],
                ),
            ),
            "configuration": {
                "timeframes": list(self.TIMEFRAMES),
                "completed_candles_only": True,
                "future_candle_policy": "REJECT",
                "symbol_resolution_policy": "CANONICAL_THEN_EXPLICIT_ALIAS",
                "broker_aliases": {
                    key: list(value) for key, value in sorted(self._broker_aliases.items())
                },
                "strategy_consumption": False,
            },
            "diagnostics": {
                "input_snapshot_sha256": input_before,
                "input_snapshot_unchanged": input_before == input_after,
                "future_candle_count": sum(row["future_candle_count"] for row in coverage_rows),
                "chronology_failure_count": sum(not row["chronological_order"] for row in coverage_rows),
                "duplicate_timestamp_count": sum(row["duplicate_timestamp_count"] for row in coverage_rows),
                "invalid_ohlc_count": sum(row["invalid_ohlc_count"] for row in coverage_rows),
                "no_strategy_imports": True,
                "production_decision_consumption": False,
            },
            "audit": {
                "runtime_metadata": self._clean(runtime_metadata or {}),
                "source_grain": "canonical_symbol + timeframe",
                "registry_definition_count": len(self._universe),
                "history_record_count": len(coverage_rows),
                "input_snapshot_sha256_before": input_before,
                "input_snapshot_sha256_after": input_after,
                "input_snapshot_unchanged": input_before == input_after,
            },
            "production_change_justified": False,
        }
        self.validate_result_schema(result)
        return result

    def validate_history(self, definition, timeframe, payload, as_of, default_resolved=None) -> dict:
        timeframe = str(timeframe).upper()
        if timeframe not in self.DURATIONS:
            raise ValueError(f"Unsupported validation timeframe: {timeframe}")
        as_of = self._timestamp(as_of)
        duration = self.DURATIONS[timeframe]
        candles = tuple(self._value(payload, "candles", ()) or ())
        times = [self._timestamp(self._value(candle, "time")) for candle in candles]
        close_times = [item + duration for item in times]
        future = [value for value in close_times if value > as_of]
        if future:
            raise ValueError(
                f"Future candle rejected for {definition.canonical_symbol} {timeframe}: "
                f"close={min(future).isoformat()}, as_of={as_of.isoformat()}"
            )

        chronological = all(left < right for left, right in zip(times, times[1:]))
        duplicates = len(times) - len(set(times))
        nonfinite, invalid_ohlc, negative_volume, negative_spread, zero_volume = 0, 0, 0, 0, 0
        for candle in candles:
            values = [self._finite_number(self._value(candle, field)) for field in ("open", "high", "low", "close")]
            if any(value is None for value in values):
                nonfinite += 1
            else:
                open_, high, low, close = values
                if high < max(open_, close) or low > min(open_, close) or low > high:
                    invalid_ohlc += 1
            tick_volume = self._finite_number(self._value(candle, "tick_volume"))
            real_volume = self._finite_number(self._value(candle, "real_volume"))
            spread = self._finite_number(self._value(candle, "spread"))
            if tick_volume is not None and tick_volume == 0:
                zero_volume += 1
            if (tick_volume is not None and tick_volume < 0) or (real_volume is not None and real_volume < 0):
                negative_volume += 1
            if spread is not None and spread < 0:
                negative_spread += 1

        gaps = [right - left for left, right in zip(times, times[1:]) if right - left > duration]
        requested = int(self._value(payload, "requested_count", 0) or 0)
        returned = int(self._value(payload, "returned_count", len(candles)) or 0)
        observed = len(candles)
        if observed == 0:
            availability = "MISSING"
        elif requested > 0 and returned < requested:
            availability = "PARTIAL"
        else:
            availability = "AVAILABLE"
        coverage_percent = None if requested <= 0 else round(min(100.0, returned * 100.0 / requested), 6)
        hard_failures = nonfinite + invalid_ohlc + negative_volume + negative_spread + duplicates + int(not chronological)
        if observed == 0 or hard_failures:
            quality_status = "FAIL"
        elif availability == "PARTIAL" or gaps or returned != observed:
            quality_status = "WARNING"
        else:
            quality_status = "PASS"
        resolved = self._value(payload, "resolved_symbol", None) or default_resolved or self.NOT_AVAILABLE
        return {
            "canonical_symbol": definition.canonical_symbol,
            "asset_class": definition.asset_class,
            "resolved_symbol": resolved,
            "broker_symbol": resolved,
            "timeframe": timeframe,
            "requested_count": requested,
            "returned_count": returned,
            "observed_candle_count": observed,
            "availability_status": availability,
            "coverage_percent": coverage_percent if coverage_percent is not None else self.NOT_AVAILABLE,
            "first_candle_open_time": times[0].isoformat() if times else self.NOT_AVAILABLE,
            "last_candle_open_time": times[-1].isoformat() if times else self.NOT_AVAILABLE,
            "last_candle_close_time": close_times[-1].isoformat() if close_times else self.NOT_AVAILABLE,
            "history_span_days": round((times[-1] - times[0]).total_seconds() / 86400.0, 6) if len(times) > 1 else 0.0,
            "chronological_order": chronological,
            "duplicate_timestamp_count": duplicates,
            "future_candle_count": 0,
            "nonfinite_price_count": nonfinite,
            "invalid_ohlc_count": invalid_ohlc,
            "negative_volume_count": negative_volume,
            "negative_spread_count": negative_spread,
            "zero_tick_volume_count": zero_volume,
            "gap_count": len(gaps),
            "maximum_gap_minutes": round(max((gap.total_seconds() / 60.0 for gap in gaps), default=0.0), 6),
            "attempt_count": int(self._value(payload, "attempts", 0) or 0),
            "error_code": int(self._value(payload, "error_code", 0) or 0),
            "error_message": str(self._value(payload, "error_message", "") or ""),
            "quality_status": quality_status,
        }

    def run(self, broker_symbols, history_results, as_of, excel_path, json_path, runtime_metadata=None) -> dict:
        result = self.analyze(broker_symbols, history_results, as_of, runtime_metadata)
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    @classmethod
    def validate_result_schema(cls, result) -> None:
        if tuple(result) != cls.RESULT_KEYS:
            raise ValueError("Multi-asset result schema keys or ordering are invalid")
        if len(result["target_universe"]) != len(cls.TARGET_UNIVERSE):
            raise ValueError("Multi-asset target universe is incomplete")
        expected_records = len(result["target_universe"]) * len(cls.TIMEFRAMES)
        if len(result["timeframe_coverage"]) != expected_records:
            raise ValueError("Timeframe coverage schema is incomplete")
        if result["mode"] != cls.MODE or result["production_change_justified"] is not False:
            raise ValueError("Diagnostic-only production guardrail is invalid")

    def _coverage_issues(self, row) -> list[dict]:
        symbol, timeframe = row["canonical_symbol"], row["timeframe"]
        issues = []
        if row["availability_status"] == "MISSING":
            message = row["error_message"] or "No completed candles were returned."
            issues.append(self._issue(symbol, timeframe, "HISTORY_MISSING", "HIGH", message))
        elif row["availability_status"] == "PARTIAL":
            issues.append(self._issue(
                symbol, timeframe, "HISTORY_PARTIAL", "MEDIUM",
                f"Returned {row['returned_count']} of {row['requested_count']} requested candles.",
            ))
        checks = (
            (not row["chronological_order"], "NON_CHRONOLOGICAL", "CRITICAL", "Candle timestamps are not strictly chronological."),
            (row["duplicate_timestamp_count"] > 0, "DUPLICATE_TIMESTAMPS", "CRITICAL", f"Duplicate timestamps: {row['duplicate_timestamp_count']}."),
            (row["nonfinite_price_count"] > 0, "NONFINITE_PRICES", "CRITICAL", f"Non-finite price candles: {row['nonfinite_price_count']}."),
            (row["invalid_ohlc_count"] > 0, "INVALID_OHLC", "CRITICAL", f"Invalid OHLC candles: {row['invalid_ohlc_count']}."),
            (row["negative_volume_count"] > 0, "NEGATIVE_VOLUME", "HIGH", f"Negative-volume candles: {row['negative_volume_count']}."),
            (row["negative_spread_count"] > 0, "NEGATIVE_SPREAD", "HIGH", f"Negative-spread candles: {row['negative_spread_count']}."),
            (row["returned_count"] != row["observed_candle_count"], "COUNT_MISMATCH", "HIGH", "Returned and observed candle counts differ."),
            (row["gap_count"] > 0, "TIME_GAPS", "LOW", f"Observed timestamp gaps: {row['gap_count']}; market closures may explain some gaps."),
        )
        for failed, code, severity, message in checks:
            if failed:
                issues.append(self._issue(symbol, timeframe, code, severity, message))
        return issues

    @staticmethod
    def _issue(symbol, timeframe, code, severity, message) -> dict:
        return {
            "canonical_symbol": symbol,
            "timeframe": timeframe,
            "issue_code": code,
            "severity": severity,
            "message": message,
        }

    def _symbol_quality(self, registry_rows, coverage_rows, issues) -> list[dict]:
        output = []
        for registry in registry_rows:
            symbol = registry["canonical_symbol"]
            records = [row for row in coverage_rows if row["canonical_symbol"] == symbol]
            symbol_issues = [row for row in issues if row["canonical_symbol"] == symbol]
            worst = max(
                (row["severity"] for row in symbol_issues),
                key=lambda item: self.SEVERITY_ORDER[item],
                default="PASS",
            )
            output.append({
                "canonical_symbol": symbol,
                "asset_class": registry["asset_class"],
                "resolved_symbol": registry["resolved_symbol"],
                "broker_symbol": registry["broker_symbol"],
                "resolution_status": registry["resolution_status"],
                "available_timeframe_count": sum(row["availability_status"] != "MISSING" for row in records),
                "missing_timeframe_count": sum(row["availability_status"] == "MISSING" for row in records),
                "passing_timeframe_count": sum(row["quality_status"] == "PASS" for row in records),
                "warning_timeframe_count": sum(row["quality_status"] == "WARNING" for row in records),
                "failing_timeframe_count": sum(row["quality_status"] == "FAIL" for row in records),
                "issue_count": len(symbol_issues),
                "highest_severity": worst,
                "overall_quality_status": "PASS" if worst in {"PASS", "LOW"} else "REVIEW_REQUIRED",
            })
        return output

    def _summary(self, registry, coverage, quality, issues) -> dict:
        return {
            "target_symbol_count": len(registry),
            "resolved_symbol_count": sum(row["resolution_status"] == "RESOLVED" for row in registry),
            "unresolved_symbol_count": sum(row["resolution_status"] != "RESOLVED" for row in registry),
            "timeframe_record_count": len(coverage),
            "available_timeframe_count": sum(row["availability_status"] != "MISSING" for row in coverage),
            "missing_timeframe_count": sum(row["availability_status"] == "MISSING" for row in coverage),
            "full_history_count": sum(row["availability_status"] == "AVAILABLE" for row in coverage),
            "partial_history_count": sum(row["availability_status"] == "PARTIAL" for row in coverage),
            "quality_pass_symbol_count": sum(row["overall_quality_status"] == "PASS" for row in quality),
            "review_required_symbol_count": sum(row["overall_quality_status"] != "PASS" for row in quality),
            "data_quality_issue_count": len(issues),
            "critical_issue_count": sum(row["severity"] == "CRITICAL" for row in issues),
            "high_issue_count": sum(row["severity"] == "HIGH" for row in issues),
            "production_change_justified": False,
        }

    def _asset_class_summary(self, registry, coverage, issues) -> list[dict]:
        rows = []
        for asset_class in ("FOREX", "METAL", "CRYPTO"):
            symbols = {row["canonical_symbol"] for row in registry if row["asset_class"] == asset_class}
            records = [row for row in coverage if row["canonical_symbol"] in symbols]
            class_issues = [row for row in issues if row["canonical_symbol"] in symbols]
            rows.append({
                "asset_class": asset_class,
                "target_symbol_count": len(symbols),
                "resolved_symbol_count": sum(row["resolution_status"] == "RESOLVED" for row in registry if row["canonical_symbol"] in symbols),
                "available_timeframe_count": sum(row["availability_status"] != "MISSING" for row in records),
                "missing_timeframe_count": sum(row["availability_status"] == "MISSING" for row in records),
                "issue_count": len(class_issues),
            })
        return rows

    @staticmethod
    def _availability_row(row) -> dict:
        keys = (
            "canonical_symbol", "asset_class", "broker_symbol", "resolved_symbol", "timeframe",
            "requested_count", "returned_count", "availability_status",
            "coverage_percent", "first_candle_open_time", "last_candle_close_time",
            "attempt_count", "error_code", "error_message",
        )
        return {key: row[key] for key in keys}

    @staticmethod
    def _quality_row(row) -> dict:
        keys = (
            "canonical_symbol", "asset_class", "timeframe", "observed_candle_count",
            "chronological_order", "duplicate_timestamp_count", "future_candle_count",
            "nonfinite_price_count", "invalid_ohlc_count", "negative_volume_count",
            "negative_spread_count", "zero_tick_volume_count", "gap_count",
            "maximum_gap_minutes", "quality_status",
        )
        return {key: row[key] for key in keys}

    def input_sha256(self, broker_symbols, history_results) -> str:
        brokers = []
        for item in broker_symbols or ():
            brokers.append({field: self._clean(self._value(item, field, None)) for field in self.SYMBOL_INFO_FIELDS})
        history = []
        for definition in self._universe:
            symbol_payload = self._value(history_results, definition.canonical_symbol, {}) or {}
            for timeframe in self.TIMEFRAMES:
                payload = self._value(symbol_payload, timeframe, {}) or {}
                candles = []
                for candle in tuple(self._value(payload, "candles", ()) or ()):
                    candles.append({
                        field: self._clean(self._value(candle, field, None))
                        for field in ("time", "open", "high", "low", "close", "tick_volume", "spread", "real_volume")
                    })
                history.append({
                    "canonical_symbol": definition.canonical_symbol,
                    "timeframe": timeframe,
                    "resolved_symbol": self._value(payload, "resolved_symbol", None),
                    "requested_count": self._value(payload, "requested_count", 0),
                    "returned_count": self._value(payload, "returned_count", 0),
                    "attempts": self._value(payload, "attempts", 0),
                    "error_code": self._value(payload, "error_code", 0),
                    "error_message": self._value(payload, "error_message", ""),
                    "candles": candles,
                })
        raw = json.dumps(
            {"brokers": sorted(brokers, key=lambda row: str(row.get("name") or "")), "history": history},
            sort_keys=True, separators=(",", ":"), default=str, allow_nan=False,
        ).encode("utf-8")
        return hashlib.sha256(raw).hexdigest()

    def write_json(self, result, path) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(result, indent=2, sort_keys=True, default=str, allow_nan=False) + "\n",
            encoding="utf-8",
        )

    def write_workbook(self, result, path) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        mappings = {
            "Summary": {**result["summary"], "asset_class_summary": result["asset_class_summary"]},
            "Symbol Registry": result["symbol_registry"],
            "History Availability": result["history_availability"],
            "Timeframe Coverage": result["timeframe_coverage"],
            "Candle Quality": result["candle_quality"],
            "Trading Conditions": result["trading_conditions"],
            "Per-Symbol Quality": result["symbol_data_quality"],
            "Data Quality Issues": result["data_quality_issues"],
            "Configuration": result["configuration"],
            "Diagnostics": result["diagnostics"],
            "Audit": result["audit"],
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @staticmethod
    def _table(sheet, rows) -> None:
        if not rows:
            sheet.append(["No data"])
            return
        headers = []
        for row in rows:
            for key in row:
                if key not in headers:
                    headers.append(key)
        sheet.append(headers)
        for row in rows:
            sheet.append([
                json.dumps(row.get(key), sort_keys=True, default=str)
                if isinstance(row.get(key), (dict, list, tuple)) else row.get(key)
                for key in headers
            ])

    @staticmethod
    def _key_values(sheet, values) -> None:
        sheet.append(["Metric", "Value"])
        for key, value in values.items():
            sheet.append([
                key,
                json.dumps(value, sort_keys=True, default=str)
                if isinstance(value, (dict, list, tuple)) else value,
            ])

    @staticmethod
    def _format(sheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
        for column in range(1, sheet.max_column + 1):
            width = max(
                len(str(sheet.cell(row, column).value or ""))
                for row in range(1, min(sheet.max_row, 200) + 1)
            ) + 2
            sheet.column_dimensions[get_column_letter(column)].width = min(60, max(12, width))

    @staticmethod
    def normalize_xlsx(source, destination) -> None:
        with zipfile.ZipFile(source, "r") as input_zip, zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9,
        ) as output_zip:
            for name in sorted(input_zip.namelist()):
                original = input_zip.getinfo(name)
                info = zipfile.ZipInfo(name, date_time=(2000, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = original.create_system
                info.external_attr = original.external_attr
                info.flag_bits = original.flag_bits
                content = input_zip.read(name)
                if name == "docProps/core.xml":
                    content = re.sub(
                        br"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                        content,
                    )
                output_zip.writestr(info, content, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)

    @staticmethod
    def _finite_number(value):
        if value is None or isinstance(value, bool):
            return None
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        return number if math.isfinite(number) else None

    @staticmethod
    def _timestamp(value) -> datetime:
        if not isinstance(value, datetime):
            value = datetime.fromisoformat(str(value))
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)
        return value

    @classmethod
    def _clean(cls, value):
        if isinstance(value, datetime):
            return cls._timestamp(value).isoformat()
        if isinstance(value, float) and not math.isfinite(value):
            return None
        if isinstance(value, Mapping):
            return {str(key): cls._clean(item) for key, item in value.items()}
        if isinstance(value, (list, tuple)):
            return [cls._clean(item) for item in value]
        return value
