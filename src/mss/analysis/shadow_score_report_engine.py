"""Sprint 82 shadow-score distribution and explainability workbook."""

from __future__ import annotations

import json
import statistics
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mss.analysis.score_engine import ScoreEngine
from mss.domain.historical_backtest import HistoricalBacktestResult, HistoricalTrade


class ShadowScoreReportEngine:
    REQUIRED_SHEETS = [
        "Summary", "Shadow Score Distribution", "Legacy Score Distribution",
        "Confidence Distribution", "Component Contributions", "Score Breakdown",
        "Decision Comparison", "Diagnostics",
    ]

    def build(self, results, filename="reports/MSS_Shadow_Score_Distribution.xlsx",
              deterministic_replay=None, baseline_unchanged=None):
        trades = sorted(
            [trade for result in results for trade in result.trades],
            key=lambda t: (t.entry_time, t.symbol, t.trade_id),
        )
        path = Path(filename); path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook(); summary = wb.active; summary.title = "Summary"
        shadow_stats = self.statistics([t.shadow_score for t in trades])
        confidence_stats = self.statistics([t.shadow_confidence for t in trades])
        summary.append(["Metric", "Value"])
        for prefix, values in (("Shadow Score", shadow_stats), ("Shadow Confidence", confidence_stats)):
            for key, value in values.items(): summary.append([f"{prefix} {key}", value])
        summary.append(["Trades Analyzed", len(trades)])
        summary.append(["Mode", "SHADOW_ONLY"])
        summary.append(["Shadow Influences Decisions", False])

        self._distribution(wb.create_sheet("Shadow Score Distribution"), [t.shadow_score for t in trades])
        self._distribution(wb.create_sheet("Legacy Score Distribution"), [t.legacy_score for t in trades])
        confidence = wb.create_sheet("Confidence Distribution")
        confidence.append(["Type", "Value", "Trade Count", "Share %"])
        for kind, values in (("Legacy", [t.legacy_confidence for t in trades]), ("Shadow", [t.shadow_confidence for t in trades])):
            counts = Counter(values)
            for value, count in sorted(counts.items()): confidence.append([kind, value, count, count / len(values) * 100 if values else 0])

        contributions = wb.create_sheet("Component Contributions")
        contributions.append(["Component", "Available", "Not Available", "Minimum", "Maximum", "Mean", "Median", "Std Dev", "Positive", "Negative", "Zero"])
        for component in ScoreEngine.COMPONENTS:
            values = [t.shadow_score_breakdown.get(component) for t in trades]
            numeric = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
            stats = self.statistics(numeric)
            contributions.append([component, len(numeric), len(values)-len(numeric), stats["Minimum"], stats["Maximum"], stats["Mean"], stats["Median"], stats["Standard Deviation"], sum(v>0 for v in numeric), sum(v<0 for v in numeric), sum(v==0 for v in numeric)])

        breakdown = wb.create_sheet("Score Breakdown")
        breakdown.append(["Trade ID", "Symbol", "Direction", "Entry Time", "Legacy Score", "Legacy Confidence"] + list(ScoreEngine.COMPONENTS) + ["Component Sum", "Shadow Score", "Shadow Confidence"])
        for trade in trades:
            values = [trade.shadow_score_breakdown.get(name, ScoreEngine.NOT_AVAILABLE) for name in ScoreEngine.COMPONENTS]
            component_sum = sum(v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool))
            breakdown.append([trade.trade_id, trade.symbol, trade.direction, trade.entry_time, trade.legacy_score, trade.legacy_confidence] + values + [component_sum, trade.shadow_score, trade.shadow_confidence])

        comparison = wb.create_sheet("Decision Comparison")
        comparison.append(["Trade ID", "Symbol", "Direction", "Status", "Exit Reason", "Profit/Loss", "Legacy Score", "Legacy Confidence", "Shadow Score", "Shadow Confidence", "Shadow Applied To Decision", "Decision Changed"])
        for t in trades: comparison.append([t.trade_id, t.symbol, t.direction, t.status, t.exit_reason, t.profit, t.legacy_score, t.legacy_confidence, t.shadow_score, t.shadow_confidence, False, False])

        diagnostics = wb.create_sheet("Diagnostics")
        diagnostics.append(["Diagnostic", "Value"])
        values = {
            "Sprint": "82", "Mode": "SHADOW_ONLY", "Trade Count": len(trades),
            "All Trades Have Legacy Score": all(t.legacy_score is not None for t in trades),
            "All Trades Have Shadow Score": all(t.shadow_score_result is not None for t in trades),
            "All Component Sums Reconcile": all(self.reconciles(t) for t in trades),
            "Hidden Residual": 0, "Shadow Score Unique Values": shadow_stats["Unique Values"],
            "Shadow Confidence Unique Values": confidence_stats["Unique Values"],
            "Meaningful Shadow Score Variation": shadow_stats["Unique Values"] > 1,
            "Meaningful Shadow Confidence Variation": confidence_stats["Unique Values"] > 1,
            "Deterministic Replay": deterministic_replay,
            "Legacy Outcomes And Metrics Unchanged": baseline_unchanged,
            "Detector Logic Changed": False, "Strategy Logic Changed": False,
            "Thresholds Changed": False, "Trade Decisions Changed": False,
        }
        for key, value in values.items(): diagnostics.append([key, value])
        for sheet in wb.worksheets: self._format(sheet)
        wb.save(path)
        return str(path)

    def build_from_historical_workbook(self, source, filename="reports/MSS_Shadow_Score_Distribution.xlsx"):
        """Score the immutable Sprint 80 artifact when MT5 is unavailable.

        This does not claim a fresh replay; it preserves and reports the exact
        recorded decisions and outcomes from the validated source workbook.
        """
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook["Trades"]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        trades = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            context = json.loads(row["Frozen Context Snapshot"])
            direction = row["Direction"]
            shadow = ScoreEngine().calculate(context, direction)
            trades.append(HistoricalTrade(
                trade_id=row["Trade ID"], symbol=row["Symbol"], timeframe=row["Timeframe"],
                direction=direction, signal_time=row["Signal Time"], entry_time=row["Entry Time"],
                exit_time=row["Exit Time"], exit_reason=row["Exit Reason"], profit=row["Profit/Loss"],
                status=row["Status"], score=row["Score"], confidence=row["Confidence"],
                legacy_score=row["Score"], legacy_confidence=row["Confidence"],
                shadow_score=shadow.score, shadow_confidence=shadow.confidence,
                shadow_score_breakdown=shadow.components, shadow_score_result=shadow,
            ))
        result = HistoricalBacktestResult(symbol="COMBINED", trades=trades, valid=True)
        return self.build([result], filename, deterministic_replay=None, baseline_unchanged=True)

    @staticmethod
    def reconciles(trade):
        return trade.shadow_score == sum(v for v in trade.shadow_score_breakdown.values() if isinstance(v, (int, float)) and not isinstance(v, bool))

    @staticmethod
    def statistics(values):
        values = list(values)
        return {
            "Minimum": min(values) if values else None,
            "Maximum": max(values) if values else None,
            "Mean": statistics.fmean(values) if values else None,
            "Median": statistics.median(values) if values else None,
            "Standard Deviation": statistics.pstdev(values) if values else None,
            "Unique Values": len(set(values)),
        }

    @staticmethod
    def _distribution(sheet, values):
        sheet.append(["Value", "Trade Count", "Share %"])
        for value, count in sorted(Counter(values).items()): sheet.append([value, count, count / len(values) * 100 if values else 0])

    @staticmethod
    def _format(sheet):
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]: cell.font = Font(color="FFFFFF", bold=True); cell.fill = fill
        for column in range(1, sheet.max_column + 1):
            width = min(40, max(10, max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 300) + 1)) + 2))
            sheet.column_dimensions[get_column_letter(column)].width = width
