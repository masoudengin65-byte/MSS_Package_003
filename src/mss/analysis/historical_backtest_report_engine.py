"""Six-sheet Excel report for historical backtest baselines."""

import json
from pathlib import Path

from openpyxl import Workbook

from mss.analysis.historical_backtest_engine import HistoricalBacktestEngine
from mss.domain.trade_statistics import TradeStatistics
from mss.analysis.context_capture_engine import ContextCaptureEngine


class HistoricalBacktestReportEngine:

    SUMMARY_FIELDS = [
        ("Total Trades", "total_trades"),
        ("Winning Trades", "winning_trades"),
        ("Losing Trades", "losing_trades"),
        ("Win Rate %", "win_rate"),
        ("Gross Profit", "gross_profit"),
        ("Gross Loss", "gross_loss"),
        ("Net Profit", "net_profit"),
        ("Profit Factor", "profit_factor"),
        ("Expectancy", "expectancy"),
        ("Average R", "average_r"),
        ("Maximum Drawdown", "maximum_drawdown"),
        ("Maximum Drawdown %", "maximum_drawdown_percent"),
        ("Maximum Consecutive Wins", "maximum_consecutive_wins"),
        ("Maximum Consecutive Losses", "maximum_consecutive_losses"),
        ("Average Holding Minutes", "average_holding_minutes"),
        ("Ending Balance", "ending_balance"),
        ("Return %", "return_percent"),
    ]

    def build(
        self,
        results,
        filename="reports/MSS_Historical_Backtest.xlsx",
    ) -> str:
        path = Path(filename)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"

        combined_trades = sorted(
            [
                trade
                for result in results
                for trade in result.trades
                if trade.status == "CLOSED"
            ],
            key=lambda trade: trade.exit_time,
        )
        combined_start = sum(result.config.starting_balance for result in results)
        combined_metrics = HistoricalBacktestEngine._calculate_metrics(
            combined_trades,
            combined_start,
            TradeStatistics(),
        )

        summary.append(["Metric", "Combined"] + [r.symbol for r in results])
        for label, attribute in self.SUMMARY_FIELDS:
            summary.append(
                [label, getattr(combined_metrics, attribute)]
                + [getattr(result.metrics, attribute) for result in results]
            )

        trades_sheet = workbook.create_sheet("Trades")
        trades_sheet.append([
            "Trade ID", "Symbol", "Timeframe", "Direction", "Signal Time",
            "Entry Time", "Entry Price", "Stop Loss", "Take Profit",
            "Exit Time", "Exit Price", "Exit Reason", "Spread",
            "Commission", "Slippage", "Volume", "Profit/Loss", "R Multiple",
            "Score", "Confidence", "Status", "Detector and Context States",
            "Frozen Context Snapshot", "Legacy Score", "Legacy Confidence",
            "Shadow Score", "Shadow Confidence", "Shadow Score Breakdown",
        ])
        for result in results:
            for trade in result.trades:
                trades_sheet.append([
                    trade.trade_id,
                    trade.symbol,
                    trade.timeframe,
                    trade.direction,
                    trade.signal_time,
                    trade.entry_time,
                    trade.entry_price,
                    trade.stop_loss,
                    trade.take_profit,
                    trade.exit_time,
                    trade.exit_price,
                    trade.exit_reason,
                    trade.spread,
                    trade.commission,
                    trade.slippage,
                    trade.volume,
                    trade.profit,
                    trade.r_multiple,
                    trade.score,
                    trade.confidence,
                    trade.status,
                    json.dumps(trade.detector_states, sort_keys=True),
                    trade.context_snapshot.payload_json if trade.context_snapshot else None,
                    trade.legacy_score,
                    trade.legacy_confidence,
                    trade.shadow_score,
                    trade.shadow_confidence,
                    json.dumps(trade.shadow_score_breakdown, sort_keys=True),
                ])

        equity_sheet = workbook.create_sheet("Equity Curve")
        equity_sheet.append(["Scope", "Trade Number", "Balance"])
        for trade_number, balance in combined_metrics.equity_curve:
            equity_sheet.append(["Combined", trade_number, balance])
        for result in results:
            for trade_number, balance in result.metrics.equity_curve:
                equity_sheet.append([result.symbol, trade_number, balance])

        detector_sheet = workbook.create_sheet("Detector Performance")
        detector_sheet.append([
            "Symbol", "Detector/Context", "State", "Trades", "Wins",
            "Losses", "Net Profit",
        ])
        for result in results:
            grouped = {}
            for trade in result.trades:
                if trade.status != "CLOSED":
                    continue
                for detector, state in trade.detector_states.items():
                    key = (detector, str(state))
                    values = grouped.setdefault(key, [0, 0, 0, 0.0])
                    values[0] += 1
                    values[1] += int(trade.profit > 0)
                    values[2] += int(trade.profit < 0)
                    values[3] += trade.profit
            for (detector, state), values in sorted(grouped.items()):
                detector_sheet.append([
                    result.symbol,
                    detector,
                    state,
                    values[0],
                    values[1],
                    values[2],
                    round(values[3], 2),
                ])

        config_sheet = workbook.create_sheet("Configuration")
        config_sheet.append(["Symbol", "Setting", "Value"])
        for result in results:
            config_values = {
                **vars(result.config),
                "point": result.metadata.point,
                "contract_size": result.metadata.contract_size,
                "volume_min": result.metadata.volume_min,
                "volume_max": result.metadata.volume_max,
                "volume_step": result.metadata.volume_step,
                "broker_spread_points": result.metadata.spread_points,
                "spread_source": (
                    "historical candle spread"
                    if result.config.spread_points is None
                    else "configured fixed spread"
                ),
                "entry_timing": "Decision candle close; next candle open",
                "completed_candles_only": True,
                "paper_trading_only": True,
                "real_orders_sent": False,
            }
            for setting, value in config_values.items():
                config_sheet.append([result.symbol, setting, value])

        diagnostics_sheet = workbook.create_sheet("Diagnostics")
        diagnostics_sheet.append(["Symbol", "Diagnostic", "Value"])
        for result in results:
            for name, value in vars(result.diagnostics).items():
                if isinstance(value, dict):
                    value = json.dumps(value, sort_keys=True)
                diagnostics_sheet.append([result.symbol, name, value])

        all_trades = [trade for result in results for trade in result.trades]
        context_fields = list(ContextCaptureEngine.FIELDS)
        self._context_sheet(workbook, "Context Snapshot", all_trades, context_fields)
        detector_fields = [f for f in context_fields if f.startswith(("structure", "bos", "choch", "liquidity", "order_block", "fvg")) or f in {"trend_strength", "swing_count", "equal_high", "equal_low", "final_score", "confidence", "unattributed_score"}]
        risk_fields = ["risk_approved", "position_size", "sl_distance", "tp_distance", "rr", "portfolio_exposure", "correlation_score", "risk_score"]
        session_fields = ["session", "kill_zone", "session_bias", "time_of_day", "day_of_week", "news_allowed", "minutes_to_next_news", "minutes_since_last_news", "news_severity", "decision_time", "entry_time", "entry_delay_minutes", "decision_candle", "entry_candle"]
        htf_fields = [f for f in context_fields if f.startswith(("h1_", "h4_", "daily_"))]
        self._context_sheet(workbook, "Detector Context", all_trades, detector_fields)
        self._context_sheet(workbook, "Risk Context", all_trades, risk_fields)
        self._context_sheet(workbook, "Session Context", all_trades, session_fields)
        self._context_sheet(workbook, "HTF Context", all_trades, htf_fields)

        workbook.save(path)
        return str(path)

    @staticmethod
    def _context_sheet(workbook, name, trades, fields):
        sheet = workbook.create_sheet(name)
        sheet.append(["Trade ID", "Symbol"] + fields)
        for trade in trades:
            values = trade.context_snapshot.to_dict() if trade.context_snapshot else {}
            sheet.append([trade.trade_id, trade.symbol] + [
                json.dumps(values.get(field), sort_keys=True) if isinstance(values.get(field), (dict, list)) else values.get(field)
                for field in fields
            ])
