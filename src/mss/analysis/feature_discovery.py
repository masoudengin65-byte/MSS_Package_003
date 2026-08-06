"""Sprint 84 diagnostic-only discovery from immutable decision-time evidence."""

from __future__ import annotations

import copy
import hashlib
import json
import math
import random
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.shadow_score_validation import ShadowScoreValidation


class FeatureDiscovery(ShadowScoreValidation):
    """Extract and test pre-outcome features without changing production code."""

    VERSION = "SPRINT_84_V1"
    SEED = 840084
    REQUIRED_SHEETS = (
        "Summary", "Feature Availability", "Feature Statistics",
        "Winner Loser Comparison", "Feature Buckets", "Symbol Analysis",
        "Temporal Stability", "Statistical Tests", "Missing Context Analysis",
        "Recommendations", "Diagnostics",
    )
    REQUIRED_TRADE_COLUMNS = (
        "Trade ID", "Symbol", "Direction", "Entry Time", "Profit/Loss",
        "Status", "Frozen Context Snapshot",
    )
    MIN_FEATURE_SAMPLE = 40
    MIN_GROUP_SAMPLE = 10

    FEATURE_SPECS = (
        ("m15_structure_alignment", "binary", "AVAILABLE_NOW"),
        ("m15_bos_alignment", "binary", "AVAILABLE_NOW"),
        ("liquidity_detected", "binary", "AVAILABLE_NOW"),
        ("liquidity_distance_atr", "numeric", "POSSIBLE_WITH_EXISTING_DATA"),
        ("market_zone", "categorical", "AVAILABLE_NOW"),
        ("market_zone_alignment", "binary", "POSSIBLE_WITH_EXISTING_DATA"),
        ("range_position", "numeric", "POSSIBLE_WITH_EXISTING_DATA"),
        ("relative_volatility", "numeric", "AVAILABLE_NOW"),
        ("volatility_regime", "categorical", "POSSIBLE_WITH_EXISTING_DATA"),
        ("decision_body_fraction", "numeric", "POSSIBLE_WITH_EXISTING_DATA"),
        ("decision_wick_imbalance", "numeric", "POSSIBLE_WITH_EXISTING_DATA"),
        ("decision_candle_alignment", "binary", "POSSIBLE_WITH_EXISTING_DATA"),
        ("tick_volume", "numeric", "AVAILABLE_NOW"),
        ("session", "categorical", "AVAILABLE_NOW"),
        ("kill_zone", "categorical", "AVAILABLE_NOW"),
        ("order_block_detected", "binary", "AVAILABLE_NOW"),
        ("fvg_detected", "binary", "AVAILABLE_NOW"),
        ("liquidity_sweep", "binary", "AVAILABLE_NOW"),
    )

    AVAILABILITY = (
        ("H1 trend", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "H1 candle sequence is not stored"),
        ("H4 trend", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "H4 candle sequence is not stored"),
        ("Daily trend", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "daily candle sequence is not stored"),
        ("Higher timeframe BOS/CHoCH", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "HTF candles and states are NOT_AVAILABLE"),
        ("Multi-timeframe alignment", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "HTF states are NOT_AVAILABLE"),
        ("M15 structure/BOS alignment", "AVAILABLE_NOW", "frozen decision-time structure and BOS direction"),
        ("Order Block detected", "AVAILABLE_NOW", "frozen boolean; only one positive record"),
        ("Order Block direction/quality/age/mitigation/distance", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "all detailed OB fields are NOT_AVAILABLE"),
        ("FVG detected", "AVAILABLE_NOW", "frozen boolean; constant false"),
        ("FVG direction/width/fill/quality", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "all detailed FVG fields are NOT_AVAILABLE"),
        ("Liquidity pool detected", "AVAILABLE_NOW", "frozen liquidity detection"),
        ("Equal highs/equal lows", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "states are NOT_AVAILABLE and candle history is absent"),
        ("Liquidity sweep confirmation", "AVAILABLE_NOW", "frozen boolean; constant false"),
        ("Distance to liquidity", "AVAILABLE_NOW", "84 frozen distances"),
        ("ATR-normalized liquidity distance", "POSSIBLE_WITH_EXISTING_DATA", "derived only from frozen distance and ATR"),
        ("Premium/Discount location", "AVAILABLE_NOW", "frozen current_zone and range bounds"),
        ("Range position", "POSSIBLE_WITH_EXISTING_DATA", "derived from frozen range and decision close"),
        ("Displacement quality", "POSSIBLE_WITH_EXISTING_DATA", "decision-candle geometry and frozen relative volatility only"),
        ("Volatility regime", "POSSIBLE_WITH_EXISTING_DATA", "derived from frozen relative volatility"),
        ("Session", "AVAILABLE_NOW", "frozen session label"),
        ("Kill-zone interaction", "AVAILABLE_NOW", "frozen kill-zone label"),
        ("Session bias", "NOT_FEASIBLE_WITH_CURRENT_HISTORY", "field is NOT_AVAILABLE"),
        ("London/New York overlap", "REQUIRES_NEW_DATA_SOURCE", "decision timestamps have no declared broker timezone/DST mapping"),
        ("Economic news proximity/impact/distance", "REQUIRES_NEW_DATA_SOURCE", "all news fields are NOT_AVAILABLE; calendar source required"),
    )

    def run(self, historical_path, excel_path, json_path):
        source = Path(historical_path)
        outputs = [Path(excel_path), Path(json_path)]
        if source.resolve() in {path.resolve() for path in outputs}:
            raise ValueError("Outputs must not overwrite the immutable input")
        before = self.file_sha256(source)
        rows, validation = self.load(source)
        result = self.analyze(rows, validation)
        after = self.file_sha256(source)
        if before != after:
            raise RuntimeError("Immutable historical input changed")
        result["data_validation"]["input_artifact"] = {
            "path": str(source), "sha256_before": before,
            "sha256_after": after, "unchanged": True,
        }
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    def load(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        required_sheets = {"Trades", "Context Snapshot"}
        missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
        if missing_sheets:
            raise ValueError(f"Missing required sheets: {', '.join(missing_sheets)}")
        context_sheet = workbook["Context Snapshot"]
        context_headers = [cell.value for cell in next(context_sheet.iter_rows())]
        expected_context = ("Trade ID", "Symbol", *ContextCaptureEngine.FIELDS)
        self.validate_columns(context_headers, expected_context, "Context Snapshot")
        if len(context_headers) != 89:
            raise ValueError(f"Expected 89 context columns (2 keys + 87 fields), found {len(context_headers)}")
        context_keys = [
            (str(values[1]), int(values[0]))
            for values in context_sheet.iter_rows(min_row=2, values_only=True)
        ]

        sheet = workbook["Trades"]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        self.validate_columns(headers, self.REQUIRED_TRADE_COLUMNS, "Trades")
        rows = []
        for source_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            raw = dict(zip(headers, values))
            try:
                context = json.loads(raw["Frozen Context Snapshot"])
            except (TypeError, json.JSONDecodeError) as exc:
                raise ValueError(f"Invalid frozen context JSON at trade row {source_index + 2}") from exc
            if set(context) != set(ContextCaptureEngine.FIELDS):
                missing = sorted(set(ContextCaptureEngine.FIELDS) - set(context))
                extra = sorted(set(context) - set(ContextCaptureEngine.FIELDS))
                raise ValueError(f"Frozen context schema mismatch: missing={missing}, extra={extra}")
            rows.append({
                "source_index": source_index, "trade_id": int(raw["Trade ID"]),
                "symbol": str(raw["Symbol"]), "direction": str(raw["Direction"]).upper(),
                "entry_time": raw["Entry Time"], "status": str(raw["Status"]).upper(),
                "profit": raw["Profit/Loss"], "context": context,
            })

        keys = [(row["symbol"], row["trade_id"]) for row in rows]
        duplicates = [list(key) for key, count in Counter(keys).items() if count > 1]
        context_duplicates = [list(key) for key, count in Counter(context_keys).items() if count > 1]
        missing_context_keys = sorted(set(keys) - set(context_keys))
        extra_context_keys = sorted(set(context_keys) - set(keys))
        future_context = []
        for row in rows:
            decision_time = self._time_value(row["context"]["decision_time"])
            latest_visible = self._time_value(row["context"]["latest_visible_candle_time"])
            entry_time = self._time_value(row["context"]["entry_time"])
            if latest_visible > decision_time or decision_time >= entry_time:
                future_context.append([row["symbol"], row["trade_id"]])
        if duplicates or context_duplicates or missing_context_keys or extra_context_keys:
            raise ValueError("Trade and Context Snapshot keys must be unique and identical")
        if future_context:
            raise ValueError(f"Future-data chronology violation for trades: {future_context}")
        validation = {
            "trade_rows": len(rows), "context_field_count": len(ContextCaptureEngine.FIELDS),
            "context_sheet_rows": len(context_keys), "context_sheet_columns": len(context_headers), "duplicate_trade_keys": duplicates,
            "duplicate_trade_key_count": len(duplicates),
            "duplicate_context_key_count": len(context_duplicates),
            "missing_context_key_count": len(missing_context_keys),
            "extra_context_key_count": len(extra_context_keys),
            "closed_trades": sum(row["status"] == "CLOSED" and row["profit"] is not None for row in rows),
            "unresolved_trades": sum(row["status"] != "CLOSED" for row in rows),
            "missing_closed_outcomes": sum(row["status"] == "CLOSED" and row["profit"] is None for row in rows),
            "future_context_record_count": len(future_context), "chronology_valid": not future_context,
        }
        return rows, validation

    def analyze(self, rows, validation=None):
        before = copy.deepcopy(rows)
        closed = [row for row in rows if row["status"] == "CLOSED" and row["profit"] is not None]
        unresolved = [row for row in rows if row["status"] != "CLOSED"]
        extracted = [self.extract(row) for row in rows]
        closed_features = [item for item in extracted if item["status"] == "CLOSED" and item["profit"] is not None]
        analyses = []
        for name, kind, availability in self.FEATURE_SPECS:
            analyses.append(self._analyze_feature(closed_features, name, kind, availability))
        self._apply_multiple_comparison(analyses)
        temporal = {item["feature"]: self._temporal_feature(closed_features, item["feature"], item["kind"]) for item in analyses}
        symbols = {
            symbol: {item["feature"]: self._feature_slice([r for r in closed_features if r["symbol"] == symbol], item["feature"], item["kind"])
                     for item in analyses}
            for symbol in ("EURUSD", "XAUUSD")
        }
        for item in analyses:
            item["classification"] = self._classify(item, temporal[item["feature"]], symbols)
        availability = [
            {"feature_category": name, "availability": status, "evidence": evidence}
            for name, status, evidence in self.AVAILABILITY
        ]
        measurable = [item["feature"] for item in analyses if item["classification"] in ("RELIABLE", "PROMISING_BUT_LIMITED")]
        result = {
            "schema_version": self.VERSION,
            "data_validation": validation or {"trade_rows": len(rows), "closed_trades": len(closed), "unresolved_trades": len(unresolved)},
            "feature_availability": availability,
            "feature_statistics": analyses,
            "symbol_analysis": symbols,
            "temporal_stability": temporal,
            "unresolved_trades": [{k: row[k] for k in ("symbol", "trade_id", "entry_time", "status")} for row in unresolved],
            "recommendations": {
                "measurable_features": measurable,
                "production_consideration_justified": any(item["classification"] == "RELIABLE" for item in analyses),
                "decision": "DIAGNOSTIC_ONLY_NO_PRODUCTION_CHANGE",
                "next_evidence": "Persist full pre-decision M15/H1/H4/D1 windows and timezone-qualified news/calendar context in a future evidence sprint.",
            },
            "diagnostics": {
                "input_objects_unchanged": rows == before, "future_candles_used": False,
                "decision_candle_latest_allowed": True, "entry_candle_used_for_features": False,
                "unresolved_excluded_from_statistics": len(closed_features) == len(closed),
                "production_decisions_consumed_features": False,
                "strategy_or_detector_changed": False,
                "feature_count": len(analyses), "bootstrap_seed": self.SEED,
                "bootstrap_iterations": self.BOOTSTRAP_ITERATIONS,
                "permutation_iterations": self.PERMUTATION_ITERATIONS,
            },
        }
        return self._clean(result)

    def extract(self, row):
        context = row["context"]
        decision = context.get("decision_candle") or {}
        high, low = self._number(decision.get("high")), self._number(decision.get("low"))
        open_, close = self._number(decision.get("open")), self._number(decision.get("close"))
        candle_range = high - low if high is not None and low is not None and high > low else None
        body_fraction = abs(close - open_) / candle_range if candle_range and open_ is not None and close is not None else None
        upper_wick = high - max(open_, close) if candle_range and open_ is not None and close is not None else None
        lower_wick = min(open_, close) - low if candle_range and open_ is not None and close is not None else None
        wick_imbalance = (upper_wick - lower_wick) / candle_range if candle_range else None
        direction = row["direction"]
        structure = str(context.get("structure", "")).upper()
        bos_direction = str(context.get("bos_direction", "")).upper()
        zone = self._available(context.get("current_zone"))
        atr = self._number(context.get("atr"))
        liquidity_distance = self._number(context.get("liquidity_distance"))
        relative_volatility = self._number(context.get("relative_volatility"))
        premium, discount = context.get("premium"), context.get("discount")
        range_position = None
        if isinstance(premium, list) and isinstance(discount, list) and len(premium) == 2 and len(discount) == 2 and close is not None:
            range_low, range_high = self._number(discount[0]), self._number(premium[1])
            if range_low is not None and range_high is not None and range_high > range_low:
                range_position = (close - range_low) / (range_high - range_low)
        features = {
            "m15_structure_alignment": int((direction == "BUY" and structure == "UPTREND") or (direction == "SELL" and structure == "DOWNTREND")),
            "m15_bos_alignment": int((direction == "BUY" and bos_direction == "BULLISH") or (direction == "SELL" and bos_direction == "BEARISH")),
            "liquidity_detected": int(bool(context.get("liquidity_detected"))),
            "liquidity_distance_atr": liquidity_distance / atr if liquidity_distance is not None and atr else None,
            "market_zone": zone,
            "market_zone_alignment": int((direction == "BUY" and zone == "DISCOUNT") or (direction == "SELL" and zone == "PREMIUM")) if zone else None,
            "range_position": range_position,
            "relative_volatility": relative_volatility,
            "volatility_regime": self._volatility_regime(relative_volatility),
            "decision_body_fraction": body_fraction,
            "decision_wick_imbalance": wick_imbalance,
            "decision_candle_alignment": int((direction == "BUY" and close > open_) or (direction == "SELL" and close < open_)) if close is not None and open_ is not None else None,
            "tick_volume": self._number(context.get("tick_volume")),
            "session": self._available(context.get("session")),
            "kill_zone": self._available(context.get("kill_zone")),
            "order_block_detected": int(bool(context.get("order_block_detected"))),
            "fvg_detected": int(bool(context.get("fvg_detected"))),
            "liquidity_sweep": int(bool(context.get("liquidity_sweep"))),
        }
        return {k: row[k] for k in ("source_index", "trade_id", "symbol", "direction", "entry_time", "status", "profit")} | {"decision_time": context["decision_time"], "features": features}

    def _analyze_feature(self, rows, name, kind, availability):
        usable = [row for row in rows if row["features"].get(name) is not None]
        values = [row["features"][name] for row in usable]
        numeric = kind in ("numeric", "binary")
        numbers = [float(value) for value in values] if numeric else []
        profits = [float(row["profit"]) for row in usable]
        labels = [1 if row["profit"] > 0 else 0 for row in usable]
        wins = [float(row["features"][name]) for row in usable if row["profit"] > 0] if numeric else []
        losses = [float(row["features"][name]) for row in usable if row["profit"] <= 0] if numeric else []
        groups = self._groups(usable, name, kind)
        auc = self.auc(numbers, labels) if numeric else None
        auc_ci = self.bootstrap_auc(numbers, labels, name) if numeric and auc is not None else (None, None)
        u, p = self.mann_whitney(wins, losses) if numeric else (None, self._permutation_group_p(usable, name))
        rho, rho_p = self.spearman(numbers, profits) if numeric else (None, None)
        return {
            "feature": name, "kind": kind, "availability_category": availability,
            "available_count": len(usable), "missing_count": len(rows) - len(usable),
            "unique_value_count": len(set(values)),
            "overall_statistics": self.descriptive(numbers, len(rows) - len(usable)) if numeric else {"count": len(values), "unique_value_count": len(set(values)), "missing_value_count": len(rows) - len(usable)},
            "population_metrics": self.metrics(usable),
            "winner_statistics": self.descriptive(wins) if numeric else self._category_distribution(usable, name, True),
            "loser_statistics": self.descriptive(losses) if numeric else self._category_distribution(usable, name, False),
            "groups": groups, "auc": auc, "auc_ci_low": auc_ci[0], "auc_ci_high": auc_ci[1],
            "spearman_profit": rho, "spearman_profit_p_value": rho_p,
            "test_statistic": u, "raw_p_value": p,
            "sample_status": self._sample_status(groups, len(usable), len(set(values))),
        }

    def _groups(self, rows, name, kind):
        if not rows:
            return []
        if kind == "numeric":
            return self.quantile_groups(self._feature_rows(rows, name), name)
        grouped = defaultdict(list)
        for row in rows:
            grouped[str(row["features"][name])].append(row)
        return [{"group": key, **self.metrics(group)} for key, group in sorted(grouped.items())]

    @staticmethod
    def _feature_rows(rows, name):
        return [{**row, name: row["features"][name]} for row in rows]

    def _feature_slice(self, rows, name, kind):
        usable = [row for row in rows if row["features"].get(name) is not None]
        result = {"trade_count": len(usable), "groups": self._groups(usable, name, kind)}
        if kind in ("numeric", "binary") and usable:
            values = [float(row["features"][name]) for row in usable]
            labels = [1 if row["profit"] > 0 else 0 for row in usable]
            result.update({"auc": self.auc(values, labels), "spearman_profit": self.spearman(values, [row["profit"] for row in usable])[0]})
        else:
            result.update({"auc": None, "spearman_profit": None})
        return result

    def _temporal_feature(self, rows, name, kind):
        ordered = sorted(rows, key=lambda row: (self._time_value(row["decision_time"]), row["source_index"]))
        midpoint = len(ordered) // 2
        return {
            "first_half": self._feature_slice(ordered[:midpoint], name, kind),
            "second_half": self._feature_slice(ordered[midpoint:], name, kind),
        }

    def _permutation_group_p(self, rows, name):
        if not rows or len(set(row["features"][name] for row in rows)) < 2:
            return None
        labels = [1 if row["profit"] > 0 else 0 for row in rows]
        categories = [str(row["features"][name]) for row in rows]
        observed = self._group_signal(categories, labels)
        rng = random.Random(self._seed("categorical|" + name))
        permuted = labels[:]
        extreme = 0
        for _ in range(self.PERMUTATION_ITERATIONS):
            rng.shuffle(permuted)
            extreme += self._group_signal(categories, permuted) >= observed
        return (extreme + 1) / (self.PERMUTATION_ITERATIONS + 1)

    @staticmethod
    def _group_signal(categories, labels):
        overall = statistics.mean(labels)
        groups = defaultdict(list)
        for category, label in zip(categories, labels):
            groups[category].append(label)
        return sum(len(values) * (statistics.mean(values) - overall) ** 2 for values in groups.values())

    @staticmethod
    def _category_distribution(rows, name, winner):
        values = [str(row["features"][name]) for row in rows if (row["profit"] > 0) == winner]
        return {"count": len(values), "distribution": dict(sorted(Counter(values).items()))}

    def _apply_multiple_comparison(self, analyses):
        items = [{"p_value": item["raw_p_value"]} for item in analyses]
        self.apply_bh(items)
        for analysis, adjusted in zip(analyses, items):
            analysis["adjusted_p_value"] = adjusted["adjusted_p_value"]

    def _classify(self, item, temporal, symbols):
        if item["sample_status"] == "INSUFFICIENT":
            return "INSUFFICIENT_DATA"
        p = item["adjusted_p_value"]
        if p is None or p >= .05:
            return "NOT_RELIABLE"
        if item["kind"] == "categorical":
            return "PROMISING_BUT_LIMITED"
        auc = item["auc"]
        if auc is None or item["auc_ci_low"] is None or item["auc_ci_high"] is None:
            return "INSUFFICIENT_DATA"
        excludes_null = item["auc_ci_low"] > .5 or item["auc_ci_high"] < .5
        direction = self._sign(auc - .5)
        temporal_stable = all(
            part["auc"] is not None and self._sign(part["auc"] - .5) == direction
            for part in temporal.values()
        )
        symbol_stable = all(
            symbols[symbol][item["feature"]]["auc"] is not None and
            self._sign(symbols[symbol][item["feature"]]["auc"] - .5) == direction
            for symbol in symbols
        )
        if excludes_null and temporal_stable and symbol_stable:
            return "RELIABLE"
        return "PROMISING_BUT_LIMITED" if excludes_null else "NOT_RELIABLE"

    def _sample_status(self, groups, available, unique):
        if available < self.MIN_FEATURE_SAMPLE or unique < 2:
            return "INSUFFICIENT"
        if any(group["trade_count"] < self.MIN_GROUP_SAMPLE for group in groups):
            return "INSUFFICIENT"
        return "SUFFICIENT"

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        stats = result["feature_statistics"]
        summary = workbook.create_sheet("Summary")
        self._key_values(summary, {
            "Mode": "DIAGNOSTIC_ONLY", "Trades": result["data_validation"]["trade_rows"],
            "Closed": result["data_validation"]["closed_trades"], "Unresolved": result["data_validation"]["unresolved_trades"],
            "Features Tested": len(stats), "Reliable Features": sum(x["classification"] == "RELIABLE" for x in stats),
            "Promising Features": sum(x["classification"] == "PROMISING_BUT_LIMITED" for x in stats),
            "Production Consideration Justified": result["recommendations"]["production_consideration_justified"],
        })
        mappings = {
            "Feature Availability": result["feature_availability"],
            "Feature Statistics": [{k: v for k, v in item.items() if k not in ("groups", "winner_statistics", "loser_statistics")} for item in stats],
            "Winner Loser Comparison": [{"feature": item["feature"], "winners": item["winner_statistics"], "losers": item["loser_statistics"]} for item in stats],
            "Feature Buckets": [{"feature": item["feature"], **group} for item in stats for group in item["groups"]],
            "Symbol Analysis": [{"symbol": symbol, "feature": feature, **values} for symbol, features in result["symbol_analysis"].items() for feature, values in features.items()],
            "Temporal Stability": [{"feature": feature, "period": period, **values} for feature, periods in result["temporal_stability"].items() for period, values in periods.items()],
            "Statistical Tests": [{"feature": item["feature"], "auc": item["auc"], "auc_ci_low": item["auc_ci_low"], "auc_ci_high": item["auc_ci_high"], "spearman_profit": item["spearman_profit"], "raw_p_value": item["raw_p_value"], "adjusted_p_value": item["adjusted_p_value"], "classification": item["classification"]} for item in stats],
            "Missing Context Analysis": [item for item in result["feature_availability"] if item["availability"] not in ("AVAILABLE_NOW", "POSSIBLE_WITH_EXISTING_DATA")],
            "Recommendations": result["recommendations"],
            "Diagnostics": {**result["data_validation"], **result["diagnostics"]},
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
        for sheet in workbook.worksheets:
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @staticmethod
    def _volatility_regime(value):
        if value is None:
            return None
        if value < .75:
            return "LOW"
        if value <= 1.5:
            return "NORMAL"
        return "HIGH"

    @staticmethod
    def _available(value):
        return None if value in (None, "", ContextCaptureEngine.NOT_AVAILABLE) else value

    @classmethod
    def _number(cls, value):
        value = cls._available(value)
        if value is None or isinstance(value, bool):
            return None
        try:
            number = float(value)
            return number if math.isfinite(number) else None
        except (TypeError, ValueError):
            return None
