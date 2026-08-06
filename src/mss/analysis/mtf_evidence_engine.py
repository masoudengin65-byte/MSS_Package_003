"""Sprint 86 diagnostic historical multi-timeframe evidence collection."""

from __future__ import annotations

import copy
import hashlib
import json
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine
from mss.analysis.real_swing_engine import RealSwingEngine
from mss.analysis.structure_engine import StructureEngine
from mss.domain.mtf_evidence import MTFEvidence


class HistoricalTimeframeLoader:
    """Select only bars whose explicit close boundary is at/before a decision."""

    DURATIONS = {
        "M15": timedelta(minutes=15),
        "H1": timedelta(hours=1),
        "H4": timedelta(hours=4),
        "D1": timedelta(days=1),
    }

    @classmethod
    def completed(cls, candles, timeframe, decision_time, limit=None):
        name = str(timeframe).upper()
        if name not in cls.DURATIONS:
            raise ValueError(f"Unsupported historical timeframe: {timeframe}")
        if not isinstance(decision_time, datetime):
            decision_time = datetime.fromisoformat(str(decision_time))
        ordered = tuple(candles or ())
        previous = None
        for candle in ordered:
            if not isinstance(candle.time, datetime):
                raise ValueError(f"{name} candle time must be a datetime")
            if previous is not None and candle.time <= previous:
                raise ValueError(f"{name} candles must be strictly chronological")
            previous = candle.time
        duration = cls.DURATIONS[name]
        selected = tuple(candle for candle in ordered if candle.time + duration <= decision_time)
        if limit is not None:
            selected = selected[-max(0, int(limit)):]
        if selected and selected[-1].time + duration > decision_time:
            raise RuntimeError(f"{name} future candle escaped the boundary filter")
        return selected

    @classmethod
    def close_time(cls, candle, timeframe):
        return candle.time + cls.DURATIONS[str(timeframe).upper()]


class MTFEvidenceEngine(ContextExpansionEngine):
    """Build evidence snapshots without importing or changing decision paths."""

    VERSION = "SPRINT_86_MTF_CONTEXT_V1"
    NOT_AVAILABLE = ContextCaptureEngine.NOT_AVAILABLE
    TIMEFRAMES = ("M15", "H1", "H4", "D1")
    HIGHER_TIMEFRAMES = ("H1", "H4", "D1")
    LOOKBACK_LIMIT = 500
    MINIMUM_CANDLES = 100
    ALIGNMENTS = ("ALIGNED", "CONFLICTING", "NEUTRAL", NOT_AVAILABLE)
    METADATA_FIELDS = (
        "schema_version", "trade_id", "symbol", "direction", "decision_time",
        "source_snapshot_version", "source_context_field_count",
        "source_expanded_schema_version", "source_expanded_field_count",
    )
    TIMEFRAME_ATTRIBUTES = (
        "availability", "source", "completed_candle_count",
        "earliest_candle_open_time", "latest_candle_open_time",
        "latest_candle_close_time", "market_structure", "trend_direction",
        "bos_state", "bos_direction", "choch_state", "choch_direction",
        "swing_count", "trend_strength", "latest_swing_high",
        "latest_swing_low", "alignment_with_m15",
    )
    TIMEFRAME_FIELDS = tuple(
        f"{timeframe}_{attribute}"
        for timeframe in ("m15", "h1", "h4", "d1")
        for attribute in (
            "availability", "source", "completed_candle_count",
            "earliest_candle_open_time", "latest_candle_open_time",
            "latest_candle_close_time", "market_structure", "trend_direction",
            "bos_state", "bos_direction", "choch_state", "choch_direction",
            "swing_count", "trend_strength", "latest_swing_high",
            "latest_swing_low", "alignment_with_m15",
        )
    )
    FIELDS = (*METADATA_FIELDS, *TIMEFRAME_FIELDS, "overall_timeframe_agreement")
    REQUIRED_SHEETS = (
        "Summary", "Schema", "Timeframe Availability", "M15 Context",
        "H1 Context", "H4 Context", "D1 Context", "Alignment Analysis",
        "Winner Loser Distribution", "Symbol Analysis", "Diagnostics",
        "Limitations",
    )

    def __init__(self, *, minimum_candles=None, lookback_limit=None):
        self.minimum_candles = int(minimum_candles or self.MINIMUM_CANDLES)
        self.lookback_limit = int(lookback_limit or self.LOOKBACK_LIMIT)
        if self.minimum_candles < 5:
            raise ValueError("minimum_candles must be at least 5")
        if self.lookback_limit < self.minimum_candles:
            raise ValueError("lookback_limit cannot be smaller than minimum_candles")
        self.loader = HistoricalTimeframeLoader()
        self.swing_engine = RealSwingEngine()
        self.structure_engine = StructureEngine()

    def load(self, path):
        """Load the immutable source while retaining the Sprint 85 schema count."""
        rows, validation = super().load(path)
        validation["expanded_context_field_count"] = len(ContextExpansionEngine.FIELDS)
        return rows, validation

    def extract(
        self, *, trade_id, symbol, direction, decision_time, frozen_context,
        candles_by_timeframe, source_labels=None,
    ):
        """Extract one immutable snapshot from bars completed by decision time."""
        frozen_before = copy.deepcopy(frozen_context)
        series_before = self._series_signature(candles_by_timeframe)
        decision = self._time_value(decision_time)
        latest_frozen = self._time_value(frozen_context["latest_visible_candle_time"])
        if latest_frozen > decision:
            raise ValueError("Frozen M15 context contains a future candle")
        values = {field: self.NOT_AVAILABLE for field in self.FIELDS}
        values.update({
            "schema_version": self.VERSION,
            "trade_id": int(trade_id),
            "symbol": str(symbol),
            "direction": str(direction).upper(),
            "decision_time": decision.isoformat(),
            "source_snapshot_version": frozen_context.get("snapshot_version", self.NOT_AVAILABLE),
            "source_context_field_count": len(ContextCaptureEngine.FIELDS),
            "source_expanded_schema_version": ContextExpansionEngine.VERSION,
            "source_expanded_field_count": len(ContextExpansionEngine.FIELDS),
        })
        evidence = {}
        normalized = {str(key).upper(): value for key, value in (candles_by_timeframe or {}).items()}
        labels = {str(key).upper(): value for key, value in (source_labels or {}).items()}
        for timeframe in self.TIMEFRAMES:
            selected = self.loader.completed(
                normalized.get(timeframe, ()), timeframe, decision,
                limit=self.lookback_limit,
            )
            evidence[timeframe] = self._extract_timeframe(
                str(symbol), timeframe, selected,
                labels.get(timeframe, "HISTORICAL_CANDLES"),
            )
            for attribute, value in evidence[timeframe].items():
                values[f"{timeframe.lower()}_{attribute}"] = value
        m15_trend = self._available(evidence["M15"]["trend_direction"])
        values["m15_alignment_with_m15"] = "ALIGNED" if m15_trend is not None else self.NOT_AVAILABLE
        comparisons = []
        for timeframe in self.HIGHER_TIMEFRAMES:
            alignment = self._alignment(m15_trend, self._available(evidence[timeframe]["trend_direction"]))
            values[f"{timeframe.lower()}_alignment_with_m15"] = alignment
            comparisons.append(alignment)
        values["overall_timeframe_agreement"] = self._overall_alignment(comparisons)
        if frozen_context != frozen_before:
            raise RuntimeError("Immutable frozen context was mutated")
        if self._series_signature(candles_by_timeframe) != series_before:
            raise RuntimeError("Historical candle input was mutated")
        if set(values) != set(self.FIELDS):
            raise RuntimeError("MTF evidence does not match the stable schema")
        return MTFEvidence.create(values)

    def _extract_timeframe(self, symbol, timeframe, candles, source_label):
        result = {attribute: self.NOT_AVAILABLE for attribute in self.TIMEFRAME_ATTRIBUTES}
        count = len(candles)
        result["completed_candle_count"] = count
        result["source"] = str(source_label) if count else self.NOT_AVAILABLE
        if count:
            result.update({
                "earliest_candle_open_time": candles[0].time.isoformat(),
                "latest_candle_open_time": candles[-1].time.isoformat(),
                "latest_candle_close_time": self.loader.close_time(candles[-1], timeframe).isoformat(),
            })
        if count < self.minimum_candles:
            return result
        swings = self.swing_engine.detect(candles)
        analysis = self.structure_engine.analyze(
            symbol=symbol, timeframe=timeframe, candles=candles, swings=swings,
        )
        structure = analysis.structure.state.value if analysis.structure is not None else self.NOT_AVAILABLE
        highs = [swing for swing in swings if getattr(swing, "is_high", False)]
        lows = [swing for swing in swings if getattr(swing, "is_low", False)]
        result.update({
            "availability": "AVAILABLE",
            "market_structure": structure,
            "trend_direction": structure,
            "bos_state": analysis.bos is not None,
            "bos_direction": analysis.bos.direction if analysis.bos is not None else self.NOT_AVAILABLE,
            "choch_state": analysis.choch is not None,
            "choch_direction": analysis.choch.direction if analysis.choch is not None else self.NOT_AVAILABLE,
            "swing_count": len(swings),
            # No native trend-strength metric exists in StructureEngine.
            "trend_strength": self.NOT_AVAILABLE,
            "latest_swing_high": highs[-1].price if highs else self.NOT_AVAILABLE,
            "latest_swing_low": lows[-1].price if lows else self.NOT_AVAILABLE,
        })
        return result

    @classmethod
    def _alignment(cls, m15_trend, other_trend):
        if m15_trend is None or other_trend is None:
            return cls.NOT_AVAILABLE
        left, right = str(m15_trend).upper(), str(other_trend).upper()
        directional = {"UPTREND", "DOWNTREND"}
        if left not in directional or right not in directional:
            return "NEUTRAL"
        return "ALIGNED" if left == right else "CONFLICTING"

    @classmethod
    def _overall_alignment(cls, values):
        if len(values) != len(cls.HIGHER_TIMEFRAMES) or any(value == cls.NOT_AVAILABLE for value in values):
            return cls.NOT_AVAILABLE
        if all(value == "ALIGNED" for value in values):
            return "ALIGNED"
        if all(value == "CONFLICTING" for value in values):
            return "CONFLICTING"
        return "NEUTRAL"

    def run(self, historical_path, candles_by_symbol, excel_path, json_path, source_metadata=None):
        source = Path(historical_path)
        outputs = {Path(excel_path).resolve(), Path(json_path).resolve()}
        if source.resolve() in outputs:
            raise ValueError("Outputs must not overwrite the immutable historical source")
        source_hash_before = self.file_sha256(source)
        rows, validation = self.load(source)
        series_before = self._nested_series_signature(candles_by_symbol)
        result = self.analyze(rows, validation, candles_by_symbol, source_metadata=source_metadata)
        source_hash_after = self.file_sha256(source)
        if source_hash_before != source_hash_after:
            raise RuntimeError("Immutable historical source changed")
        if self._nested_series_signature(candles_by_symbol) != series_before:
            raise RuntimeError("Historical candle inputs changed")
        result["diagnostics"]["input_artifact"] = {
            "path": str(source), "sha256_before": source_hash_before,
            "sha256_after": source_hash_after, "unchanged": True,
        }
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    def analyze(self, rows, validation, candles_by_symbol, source_metadata=None):
        rows_before = copy.deepcopy(rows)
        extracted = []
        for row in rows:
            symbol_series = candles_by_symbol.get(row["symbol"], {})
            snapshot = self.extract(
                trade_id=row["trade_id"], symbol=row["symbol"], direction=row["direction"],
                decision_time=row["decision_time"], frozen_context=row["frozen_context"],
                candles_by_timeframe=symbol_series,
                source_labels={timeframe: self._source_label(source_metadata, row["symbol"], timeframe) for timeframe in self.TIMEFRAMES},
            )
            extracted.append({
                "source_index": row["source_index"], "status": row["status"],
                "profit": row["profit"], "context": snapshot.to_dict(),
            })
        if rows != rows_before:
            raise RuntimeError("Historical trade rows were mutated")
        closed = [row for row in extracted if row["status"] == "CLOSED" and row["profit"] is not None]
        winners = [row for row in closed if row["profit"] > 0]
        losers = [row for row in closed if row["profit"] <= 0]
        diagnostics = {
            "input_objects_unchanged": True,
            "future_candles_used": False,
            "candle_close_boundary_inclusive": True,
            "outcome_fields_in_evidence_schema": False,
            "production_decision_path_invoked": False,
            "production_decisions_changed": False,
            "existing_structure_engine_used_read_only": True,
            "existing_detector_logic_modified": False,
            "unresolved_excluded_from_outcome_analysis": len(closed) == validation["closed_trade_count"],
            "source_metadata": source_metadata or {},
            "series": self._series_metadata(candles_by_symbol),
        }
        result = {
            "schema_version": self.VERSION,
            "mode": "DIAGNOSTIC_ONLY",
            "schema": self._schema(),
            "data_validation": {
                **validation,
                "mtf_evidence_field_count": len(self.FIELDS),
                "winner_count": len(winners), "loser_count": len(losers),
                "outcome_analysis_count": len(closed),
            },
            "timeframe_availability": self._timeframe_availability(extracted, closed),
            "alignment_analysis": self._alignment_analysis(extracted, closed),
            "winner_loser_distribution": self._winner_loser_distribution(winners, losers),
            "symbol_analysis": self._symbol_analysis(extracted),
            "temporal_coverage": self._temporal_coverage_mtf(extracted),
            "missing_evidence": self._missing_evidence(extracted),
            "records": [row["context"] for row in extracted],
            "diagnostics": diagnostics,
            "limitations": [
                "This layer is diagnostic-only and is not imported by production decision paths.",
                "MT5 bar timestamps are treated as bar-open timestamps; close boundaries are derived from the declared timeframe.",
                "Only bars whose derived close time is at or before the immutable decision timestamp are used.",
                "Trend strength remains NOT_AVAILABLE because StructureEngine exposes no native trend-strength measure.",
                "BOS and CHoCH fields report the existing detector outputs without changing detector rules.",
                "Historical terminal coverage is limited to the candles returned by the verified broker history source.",
                "This artifact is descriptive; it does not statistically validate predictive power or justify production use.",
            ],
            "production_change_justified": False,
        }
        return self._clean(result)

    def _schema(self):
        rows = []
        for field in self.FIELDS:
            if field in self.METADATA_FIELDS:
                category = "Metadata"
            elif field == "overall_timeframe_agreement":
                category = "Alignment"
            else:
                category = field.split("_", 1)[0].upper()
            rows.append({
                "field": field, "category": category,
                "not_available_preserved": field not in self.METADATA_FIELDS,
                "diagnostic_only": True,
            })
        return rows

    def _timeframe_availability(self, rows, closed):
        output = []
        for timeframe in self.TIMEFRAMES:
            field = f"{timeframe.lower()}_availability"
            available = [row for row in rows if row["context"][field] == "AVAILABLE"]
            closed_available = [row for row in closed if row["context"][field] == "AVAILABLE"]
            output.append({
                "timeframe": timeframe, "trade_count": len(rows),
                "available_count": len(available), "missing_count": len(rows) - len(available),
                "availability_percent": 100.0 * len(available) / len(rows) if rows else 0.0,
                "closed_available_count": len(closed_available),
                "winner_available_count": sum(row["profit"] > 0 for row in closed_available),
                "loser_available_count": sum(row["profit"] <= 0 for row in closed_available),
            })
        return output

    def _alignment_analysis(self, rows, closed):
        output = []
        for label, field in [
            ("M15_vs_H1", "h1_alignment_with_m15"),
            ("M15_vs_H4", "h4_alignment_with_m15"),
            ("M15_vs_D1", "d1_alignment_with_m15"),
            ("OVERALL", "overall_timeframe_agreement"),
        ]:
            for value in self.ALIGNMENTS:
                all_group = [row for row in rows if row["context"][field] == value]
                closed_group = [row for row in closed if row["context"][field] == value]
                output.append({
                    "comparison": label, "value": value,
                    "trade_count": len(all_group), "closed_count": len(closed_group),
                    "winner_count": sum(row["profit"] > 0 for row in closed_group),
                    "loser_count": sum(row["profit"] <= 0 for row in closed_group),
                    "availability_percent": 100.0 * len(all_group) / len(rows) if rows else 0.0,
                })
        return output

    def _winner_loser_distribution(self, winners, losers):
        fields = []
        for timeframe in self.TIMEFRAMES:
            prefix = timeframe.lower()
            fields.extend(f"{prefix}_{name}" for name in (
                "market_structure", "trend_direction", "bos_state", "choch_state",
                "swing_count", "latest_swing_high", "latest_swing_low", "alignment_with_m15",
            ))
        fields.append("overall_timeframe_agreement")
        return [{
            "field": field,
            "winner_distribution": self._distribution_available(winners, field),
            "loser_distribution": self._distribution_available(losers, field),
        } for field in fields]

    def _symbol_analysis(self, rows):
        output = []
        for symbol in sorted({row["context"]["symbol"] for row in rows}):
            symbol_rows = [row for row in rows if row["context"]["symbol"] == symbol]
            closed = [row for row in symbol_rows if row["status"] == "CLOSED" and row["profit"] is not None]
            for timeframe in self.TIMEFRAMES:
                availability = f"{timeframe.lower()}_availability"
                alignment = f"{timeframe.lower()}_alignment_with_m15"
                available = sum(row["context"][availability] == "AVAILABLE" for row in symbol_rows)
                counts = Counter(row["context"][alignment] for row in symbol_rows)
                output.append({
                    "symbol": symbol, "timeframe": timeframe,
                    "trade_count": len(symbol_rows), "closed_count": len(closed),
                    "winner_count": sum(row["profit"] > 0 for row in closed),
                    "loser_count": sum(row["profit"] <= 0 for row in closed),
                    "available_count": available,
                    "availability_percent": 100.0 * available / len(symbol_rows) if symbol_rows else 0.0,
                    "alignment_distribution": dict(sorted(counts.items())),
                })
        return output

    def _temporal_coverage_mtf(self, rows):
        buckets = {}
        for row in rows:
            month = self._time_value(row["context"]["decision_time"]).strftime("%Y-%m")
            buckets.setdefault(month, []).append(row)
        output = []
        for month, month_rows in sorted(buckets.items()):
            for timeframe in self.TIMEFRAMES:
                field = f"{timeframe.lower()}_availability"
                available = sum(row["context"][field] == "AVAILABLE" for row in month_rows)
                output.append({
                    "period": month, "timeframe": timeframe, "trade_count": len(month_rows),
                    "available_count": available,
                    "availability_percent": 100.0 * available / len(month_rows) if month_rows else 0.0,
                    "start": min(row["context"]["decision_time"] for row in month_rows),
                    "end": max(row["context"]["decision_time"] for row in month_rows),
                })
        return output

    def _missing_evidence(self, rows):
        output = []
        for field in (*self.TIMEFRAME_FIELDS, "overall_timeframe_agreement"):
            missing = sum(self._available(row["context"][field]) is None for row in rows)
            output.append({
                "field": field, "missing_count": missing,
                "missing_percent": 100.0 * missing / len(rows) if rows else 0.0,
            })
        return output

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        summary = workbook.create_sheet("Summary")
        self._key_values(summary, {
            "Mode": "DIAGNOSTIC_ONLY", "Schema Version": self.VERSION,
            "Trades": result["data_validation"]["trade_count"],
            "Closed": result["data_validation"]["closed_trade_count"],
            "Unresolved": result["data_validation"]["unresolved_trade_count"],
            "Winners": result["data_validation"]["winner_count"],
            "Losers": result["data_validation"]["loser_count"],
            "MTF Evidence Fields": len(self.FIELDS),
            "Source Context Fields Preserved": len(ContextCaptureEngine.FIELDS),
            "Expanded Context Fields Preserved": len(ContextExpansionEngine.FIELDS),
            "Production Change Justified": False,
        })
        mappings = {
            "Schema": result["schema"],
            "Timeframe Availability": result["timeframe_availability"],
            "M15 Context": self._timeframe_records(result["records"], "M15"),
            "H1 Context": self._timeframe_records(result["records"], "H1"),
            "H4 Context": self._timeframe_records(result["records"], "H4"),
            "D1 Context": self._timeframe_records(result["records"], "D1"),
            "Alignment Analysis": result["alignment_analysis"],
            "Winner Loser Distribution": result["winner_loser_distribution"],
            "Symbol Analysis": result["symbol_analysis"],
            "Diagnostics": {
                **result["data_validation"], **result["diagnostics"],
                "temporal_coverage": result["temporal_coverage"],
                "missing_evidence": result["missing_evidence"],
            },
            "Limitations": [{"limitation": value} for value in result["limitations"]],
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
        for sheet in workbook.worksheets:
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @classmethod
    def _timeframe_records(cls, records, timeframe):
        prefix = timeframe.lower()
        fields = [field for field in cls.TIMEFRAME_FIELDS if field.startswith(prefix + "_")]
        keys = ("trade_id", "symbol", "direction", "decision_time")
        return [{key: record[key] for key in (*keys, *fields)} for record in records]

    @classmethod
    def _distribution_available(cls, rows, field):
        values = [row["context"][field] for row in rows if cls._available(row["context"][field]) is not None]
        return cls._distribution(values)

    @classmethod
    def _series_signature(cls, mapping):
        output = []
        for timeframe, candles in sorted((mapping or {}).items(), key=lambda item: str(item[0])):
            output.append((str(timeframe), tuple(
                (candle.time, candle.open, candle.high, candle.low, candle.close,
                 candle.tick_volume, candle.spread, candle.real_volume)
                for candle in candles
            )))
        return tuple(output)

    @classmethod
    def _nested_series_signature(cls, mapping):
        return tuple(
            (symbol, cls._series_signature(series))
            for symbol, series in sorted((mapping or {}).items())
        )

    @classmethod
    def _series_metadata(cls, mapping):
        output = []
        for symbol, series in sorted((mapping or {}).items()):
            for timeframe, candles in sorted(series.items()):
                digest = hashlib.sha256()
                for candle in candles:
                    digest.update(repr((
                        candle.time.isoformat(), float(candle.open), float(candle.high),
                        float(candle.low), float(candle.close), int(candle.tick_volume),
                        int(candle.spread), int(candle.real_volume),
                    )).encode("utf-8"))
                output.append({
                    "symbol": symbol, "timeframe": str(timeframe).upper(),
                    "candle_count": len(candles),
                    "first_open_time": candles[0].time.isoformat() if candles else None,
                    "last_open_time": candles[-1].time.isoformat() if candles else None,
                    "sha256": digest.hexdigest(),
                })
        return output

    @staticmethod
    def _source_label(metadata, symbol, timeframe):
        if not metadata:
            return "HISTORICAL_CANDLES"
        symbols = metadata.get("symbols", {})
        resolved = symbols.get(symbol, {}).get("resolved_symbol", symbol)
        return f"MT5:{resolved}:{timeframe}"
