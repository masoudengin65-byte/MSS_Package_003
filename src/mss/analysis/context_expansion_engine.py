"""Sprint 85 diagnostic expansion of immutable decision-time context."""

from __future__ import annotations

import copy
import json
import math
import statistics
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.shadow_score_validation import ShadowScoreValidation
from mss.domain.expanded_context import ExpandedContext


class ContextExpansionEngine(ShadowScoreValidation):
    """Create evidence fields without importing or changing decision paths."""

    VERSION = "SPRINT_85_CONTEXT_V1"
    NOT_AVAILABLE = ContextCaptureEngine.NOT_AVAILABLE
    TIMEFRAMES = ("m15", "h1", "h4", "d1")
    MTF_ATTRIBUTES = (
        "market_structure", "trend_direction", "bos_state", "choch_state",
        "swing_count", "trend_strength", "alignment_with_m15",
    )
    METADATA_FIELDS = (
        "schema_version", "trade_id", "symbol", "direction", "decision_time",
        "latest_visible_m15_time", "source_snapshot_version",
    )
    MTF_FIELDS = tuple(
        f"{timeframe}_{attribute}"
        for timeframe in ("m15", "h1", "h4", "d1")
        for attribute in (
            "market_structure", "trend_direction", "bos_state", "choch_state",
            "swing_count", "trend_strength", "alignment_with_m15",
        )
    )
    ORDER_BLOCK_FIELDS = (
        "ob_detected", "ob_direction", "ob_distance_from_entry", "ob_freshness",
        "ob_mitigation_status", "ob_age", "ob_quality",
    )
    FVG_FIELDS = (
        "fvg_detected", "fvg_direction", "fvg_width", "fvg_distance_from_entry",
        "fvg_fill_percentage", "fvg_freshness", "fvg_quality",
    )
    LIQUIDITY_FIELDS = (
        "equal_highs", "equal_lows", "liquidity_pool_detected", "liquidity_pool_location",
        "liquidity_sweep_occurrence", "liquidity_sweep_direction",
        "distance_to_liquidity", "rejection_after_sweep",
    )
    SESSION_FIELDS = (
        "session_timezone", "session_dst_handling", "london_session",
        "new_york_session", "session_overlap_period", "session_bias",
        "session_label", "kill_zone",
    )
    ECONOMIC_FIELDS = (
        "nearby_economic_event", "economic_currency_affected", "economic_impact_level",
        "minutes_before_event", "minutes_after_event",
    )
    FIELDS = (
        *METADATA_FIELDS, *MTF_FIELDS, *ORDER_BLOCK_FIELDS, *FVG_FIELDS,
        *LIQUIDITY_FIELDS, *SESSION_FIELDS, *ECONOMIC_FIELDS,
    )
    REQUIRED_SHEETS = (
        "Summary", "Schema", "Availability", "MTF Context",
        "Order Block Context", "FVG Context", "Liquidity Context",
        "Session Context", "Economic Context", "Diagnostics", "Limitations",
    )
    REQUIRED_TRADE_COLUMNS = (
        "Trade ID", "Symbol", "Direction", "Entry Time", "Profit/Loss",
        "Status", "Frozen Context Snapshot",
    )

    FIELD_CATEGORIES = {
        **{name: "Metadata" for name in METADATA_FIELDS},
        **{name: "MTF" for name in MTF_FIELDS},
        **{name: "Order Block" for name in ORDER_BLOCK_FIELDS},
        **{name: "FVG" for name in FVG_FIELDS},
        **{name: "Liquidity" for name in LIQUIDITY_FIELDS},
        **{name: "Session" for name in SESSION_FIELDS},
        **{name: "Economic" for name in ECONOMIC_FIELDS},
    }

    def expand(
        self, *, trade_id, symbol, direction, decision_time, frozen_context,
        timeframe_evidence=None, order_block_evidence=None, fvg_evidence=None,
        liquidity_evidence=None, session_evidence=None, economic_evidence=None,
    ):
        """Expand one frozen snapshot from evidence visible by decision time."""
        before = copy.deepcopy(frozen_context)
        decision_time = self._time_value(decision_time)
        latest_m15 = self._time_value(frozen_context["latest_visible_candle_time"])
        if latest_m15 > decision_time:
            raise ValueError("M15 evidence contains a future candle")
        values = {field: self.NOT_AVAILABLE for field in self.FIELDS}
        values.update({
            "schema_version": self.VERSION, "trade_id": int(trade_id),
            "symbol": str(symbol), "direction": str(direction).upper(),
            "decision_time": decision_time.isoformat(),
            "latest_visible_m15_time": latest_m15.isoformat(),
            "source_snapshot_version": frozen_context.get("snapshot_version", self.NOT_AVAILABLE),
        })
        self._populate_m15(values, frozen_context)
        self._populate_timeframes(values, timeframe_evidence or {}, decision_time)
        self._populate_existing_context(values, frozen_context)
        self._populate_optional(values, order_block_evidence, self.ORDER_BLOCK_FIELDS, decision_time)
        self._populate_optional(values, fvg_evidence, self.FVG_FIELDS, decision_time)
        self._populate_optional(values, liquidity_evidence, self.LIQUIDITY_FIELDS, decision_time)
        self._populate_optional(values, session_evidence, self.SESSION_FIELDS, decision_time)
        self._populate_optional(values, economic_evidence, self.ECONOMIC_FIELDS, decision_time)
        if frozen_context != before:
            raise RuntimeError("Frozen context was mutated")
        return ExpandedContext.create(values)

    def _populate_m15(self, values, context):
        structure = self._available(context.get("structure"))
        values.update({
            "m15_market_structure": structure,
            "m15_trend_direction": structure,
            "m15_bos_state": self._available(context.get("bos")),
            "m15_choch_state": self._available(context.get("choch")),
            "m15_swing_count": self._available(context.get("swing_count")),
            "m15_trend_strength": self._available(context.get("trend_strength")),
            "m15_alignment_with_m15": "BASELINE" if structure is not None else self.NOT_AVAILABLE,
        })

    def _populate_timeframes(self, values, evidence, decision_time):
        m15_direction = self._available(values["m15_trend_direction"])
        for timeframe in ("h1", "h4", "d1"):
            item = evidence.get(timeframe)
            if not item:
                continue
            self._validate_evidence_time(item, decision_time, f"{timeframe.upper()} evidence")
            for attribute in self.MTF_ATTRIBUTES[:-1]:
                value = self._available(item.get(attribute))
                if value is not None:
                    values[f"{timeframe}_{attribute}"] = value
            trend = self._available(values[f"{timeframe}_trend_direction"])
            if trend is not None and m15_direction is not None:
                values[f"{timeframe}_alignment_with_m15"] = "ALIGNED" if str(trend).upper() == str(m15_direction).upper() else "CONFLICTING"

    def _populate_existing_context(self, values, context):
        mapping = {
            "ob_detected": "order_block_detected", "ob_direction": "order_block_type",
            "ob_mitigation_status": "order_block_mitigation_state", "ob_age": "order_block_age",
            "ob_quality": "order_block_quality_score", "fvg_detected": "fvg_detected",
            "fvg_direction": "fvg_direction", "fvg_width": "fvg_width",
            "fvg_fill_percentage": "fvg_fill_percentage", "fvg_quality": "fvg_quality_score",
            "equal_highs": "equal_high", "equal_lows": "equal_low",
            "liquidity_pool_detected": "liquidity_detected",
            "liquidity_sweep_occurrence": "liquidity_sweep", "distance_to_liquidity": "liquidity_distance",
            "session_bias": "session_bias", "session_label": "session", "kill_zone": "kill_zone",
        }
        for target, source in mapping.items():
            value = self._available(context.get(source))
            if value is not None:
                values[target] = value

    def _populate_optional(self, values, evidence, allowed_fields, decision_time):
        if not evidence:
            return
        self._validate_evidence_time(evidence, decision_time, "optional evidence")
        for field in allowed_fields:
            value = self._available(evidence.get(field))
            if value is not None:
                values[field] = value

    def _validate_evidence_time(self, evidence, decision_time, label):
        captured_at = evidence.get("captured_at")
        if captured_at is None:
            raise ValueError(f"{label} must include captured_at")
        if self._time_value(captured_at) > decision_time:
            raise ValueError(f"{label} was captured after decision time")
        latest = evidence.get("latest_visible_candle_time")
        if latest is not None and self._time_value(latest) > decision_time:
            raise ValueError(f"{label} contains a future candle")

    def run(self, historical_path, excel_path, json_path):
        source = Path(historical_path)
        if source.resolve() in {Path(excel_path).resolve(), Path(json_path).resolve()}:
            raise ValueError("Outputs must not overwrite the immutable source")
        before = self.file_sha256(source)
        rows, validation = self.load(source)
        result = self.analyze(rows, validation)
        after = self.file_sha256(source)
        if before != after:
            raise RuntimeError("Immutable historical source changed")
        result["diagnostics"]["input_artifact"] = {
            "path": str(source), "sha256_before": before,
            "sha256_after": after, "unchanged": True,
        }
        self.write_workbook(result, excel_path)
        self.write_json(result, json_path)
        return result

    def load(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Trades" not in workbook.sheetnames or "Context Snapshot" not in workbook.sheetnames:
            raise ValueError("Historical source must contain Trades and Context Snapshot sheets")
        context_sheet = workbook["Context Snapshot"]
        context_headers = [cell.value for cell in next(context_sheet.iter_rows())]
        self.validate_columns(context_headers, ("Trade ID", "Symbol", *ContextCaptureEngine.FIELDS), "Context Snapshot")
        if len(context_headers) != 89:
            raise ValueError("Expected 2 keys plus the immutable 87-field context schema")
        context_keys = [(str(row[1]), int(row[0])) for row in context_sheet.iter_rows(min_row=2, values_only=True)]
        sheet = workbook["Trades"]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        self.validate_columns(headers, self.REQUIRED_TRADE_COLUMNS, "Trades")
        rows = []
        for source_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            raw = dict(zip(headers, values))
            try:
                context = json.loads(raw["Frozen Context Snapshot"])
            except (TypeError, json.JSONDecodeError) as exc:
                raise ValueError(f"Invalid frozen context at row {source_index + 2}") from exc
            if set(context) != set(ContextCaptureEngine.FIELDS):
                raise ValueError("Frozen context does not match the immutable 87-field schema")
            decision_time = self._time_value(context["decision_time"])
            latest = self._time_value(context["latest_visible_candle_time"])
            entry_time = self._time_value(context["entry_time"])
            if latest > decision_time or decision_time >= entry_time:
                raise ValueError(f"Future-data chronology violation at row {source_index + 2}")
            rows.append({
                "source_index": source_index, "trade_id": int(raw["Trade ID"]),
                "symbol": str(raw["Symbol"]), "direction": str(raw["Direction"]).upper(),
                "status": str(raw["Status"]).upper(), "profit": raw["Profit/Loss"],
                "decision_time": decision_time, "frozen_context": context,
            })
        keys = [(row["symbol"], row["trade_id"]) for row in rows]
        if len(keys) != len(set(keys)) or len(context_keys) != len(set(context_keys)) or set(keys) != set(context_keys):
            raise ValueError("Trade and context keys must be unique and identical")
        validation = {
            "trade_count": len(rows), "context_row_count": len(context_keys),
            "source_context_field_count": len(ContextCaptureEngine.FIELDS),
            "expanded_context_field_count": len(self.FIELDS),
            "closed_trade_count": sum(row["status"] == "CLOSED" and row["profit"] is not None for row in rows),
            "unresolved_trade_count": sum(row["status"] != "CLOSED" for row in rows),
            "duplicate_key_count": len(keys) - len(set(keys)),
            "future_context_count": 0, "chronology_valid": True,
        }
        return rows, validation

    def analyze(self, rows, validation=None):
        before = copy.deepcopy(rows)
        expanded = []
        for row in rows:
            snapshot = self.expand(
                trade_id=row["trade_id"], symbol=row["symbol"], direction=row["direction"],
                decision_time=row["decision_time"], frozen_context=row["frozen_context"],
            )
            expanded.append({
                "source_index": row["source_index"], "status": row["status"],
                "profit": row["profit"], "context": snapshot.to_dict(),
            })
        closed = [row for row in expanded if row["status"] == "CLOSED" and row["profit"] is not None]
        availability = self._availability(expanded, closed)
        symbols = self._symbol_split(expanded)
        temporal = self._temporal_coverage(expanded)
        result = {
            "schema_version": self.VERSION,
            "schema": [{"field": field, "category": self.FIELD_CATEGORIES[field]} for field in self.FIELDS],
            "data_validation": validation or {
                "trade_count": len(rows), "closed_trade_count": len(closed),
                "unresolved_trade_count": len(rows) - len(closed),
            },
            "availability": availability,
            "winner_loser_distribution": self._winner_loser(closed),
            "symbol_split": symbols, "temporal_coverage": temporal,
            "records": [row["context"] for row in expanded],
            "diagnostics": {
                "input_objects_unchanged": rows == before, "future_candles_used": False,
                "entry_candle_used": False, "production_decisions_changed": False,
                "detectors_invoked": False, "scores_changed": False,
                "unresolved_excluded_from_outcome_analysis": len(closed) == sum(row["status"] == "CLOSED" and row["profit"] is not None for row in expanded),
            },
            "limitations": [
                "The validated artifact stores no immutable H1/H4/D1 candle windows.",
                "Detailed OB and FVG lifecycle evidence was not captured historically.",
                "Equal-high/equal-low and sweep-rejection evidence cannot be reconstructed.",
                "Session timestamps lack a declared broker timezone and DST mapping.",
                "No reliable immutable historical economic-calendar source is present.",
                "NOT_AVAILABLE values are preserved; no unavailable feature is inferred.",
            ],
        }
        return self._clean(result)

    def _availability(self, expanded, closed):
        output = []
        for field in self.FIELDS:
            if field in self.METADATA_FIELDS:
                continue
            values = [row["context"][field] for row in expanded]
            available = [value for value in values if self._available(value) is not None]
            closed_available = [row for row in closed if self._available(row["context"][field]) is not None]
            output.append({
                "field": field, "category": self.FIELD_CATEGORIES[field],
                "available_count": len(available), "missing_count": len(values) - len(available),
                "availability_percent": 100 * len(available) / len(values) if values else 0.0,
                "missing_percent": 100 * (len(values) - len(available)) / len(values) if values else 0.0,
                "closed_available_count": len(closed_available),
                "winner_available_count": sum(row["profit"] > 0 for row in closed_available),
                "loser_available_count": sum(row["profit"] <= 0 for row in closed_available),
                "unique_value_count": len({json.dumps(value, sort_keys=True, default=str) for value in available}),
            })
        return output

    def _winner_loser(self, closed):
        output = []
        for field in self.FIELDS:
            if field in self.METADATA_FIELDS:
                continue
            winners = [row["context"][field] for row in closed if row["profit"] > 0 and self._available(row["context"][field]) is not None]
            losers = [row["context"][field] for row in closed if row["profit"] <= 0 and self._available(row["context"][field]) is not None]
            output.append({
                "field": field, "category": self.FIELD_CATEGORIES[field],
                "winner_distribution": self._distribution(winners),
                "loser_distribution": self._distribution(losers),
            })
        return output

    def _symbol_split(self, expanded):
        output = []
        for symbol in ("EURUSD", "XAUUSD"):
            rows = [row for row in expanded if row["context"]["symbol"] == symbol]
            for field in self.FIELDS:
                if field in self.METADATA_FIELDS:
                    continue
                count = sum(self._available(row["context"][field]) is not None for row in rows)
                output.append({"symbol": symbol, "field": field, "trade_count": len(rows), "available_count": count, "availability_percent": 100 * count / len(rows) if rows else 0.0})
        return output

    def _temporal_coverage(self, expanded):
        ordered = sorted(expanded, key=lambda row: (self._time_value(row["context"]["decision_time"]), row["source_index"]))
        midpoint = len(ordered) // 2
        output = []
        for period, rows in (("first_half", ordered[:midpoint]), ("second_half", ordered[midpoint:])):
            for field in self.FIELDS:
                if field in self.METADATA_FIELDS:
                    continue
                count = sum(self._available(row["context"][field]) is not None for row in rows)
                output.append({
                    "period": period, "field": field, "trade_count": len(rows),
                    "available_count": count, "availability_percent": 100 * count / len(rows) if rows else 0.0,
                    "start": rows[0]["context"]["decision_time"] if rows else None,
                    "end": rows[-1]["context"]["decision_time"] if rows else None,
                })
        return output

    @classmethod
    def _distribution(cls, values):
        numbers = [cls._number(value) for value in values]
        if values and all(value is not None for value in numbers):
            return {
                "count": len(numbers), "minimum": min(numbers), "maximum": max(numbers),
                "mean": statistics.mean(numbers), "median": statistics.median(numbers),
                "standard_deviation": statistics.stdev(numbers) if len(numbers) > 1 else 0.0,
            }
        return {"count": len(values), "values": dict(sorted(Counter(str(value) for value in values).items()))}

    def write_workbook(self, result, path):
        workbook = Workbook()
        workbook.remove(workbook.active)
        fixed = datetime(2000, 1, 1)
        workbook.properties.created = fixed
        workbook.properties.modified = fixed
        summary = workbook.create_sheet("Summary")
        self._key_values(summary, {
            "Mode": "DIAGNOSTIC_ONLY", "Schema Version": self.VERSION,
            "Trades": result["data_validation"]["trade_count"],
            "Closed": result["data_validation"]["closed_trade_count"],
            "Unresolved": result["data_validation"]["unresolved_trade_count"],
            "Expanded Fields": len(self.FIELDS),
            "Production Change Justified": False,
        })
        records = result["records"]
        mappings = {
            "Schema": result["schema"], "Availability": result["availability"],
            "MTF Context": self._category_records(records, self.MTF_FIELDS),
            "Order Block Context": self._category_records(records, self.ORDER_BLOCK_FIELDS),
            "FVG Context": self._category_records(records, self.FVG_FIELDS),
            "Liquidity Context": self._category_records(records, self.LIQUIDITY_FIELDS),
            "Session Context": self._category_records(records, self.SESSION_FIELDS),
            "Economic Context": self._category_records(records, self.ECONOMIC_FIELDS),
            "Diagnostics": {**result["data_validation"], **result["diagnostics"], "winner_loser_distribution": result["winner_loser_distribution"], "symbol_split": result["symbol_split"], "temporal_coverage": result["temporal_coverage"]},
            "Limitations": [{"limitation": value} for value in result["limitations"]],
        }
        for name, value in mappings.items():
            sheet = workbook.create_sheet(name)
            self._table(sheet, value) if isinstance(value, list) else self._key_values(sheet, value)
        for sheet in workbook.worksheets:
            self._format(sheet)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".tmp")
        workbook.save(temporary)
        self.normalize_xlsx(temporary, path)
        temporary.unlink()

    @staticmethod
    def _category_records(records, fields):
        keys = ("trade_id", "symbol", "direction", "decision_time")
        return [{key: record[key] for key in (*keys, *fields)} for record in records]

    @classmethod
    def _available(cls, value):
        return None if value in (None, "", cls.NOT_AVAILABLE) else value

    @classmethod
    def _number(cls, value):
        value = cls._available(value)
        if value is None or isinstance(value, bool):
            return None
        try:
            number = float(value)
            return number if math.isfinite(number) else None
        except (TypeError, ValueError):
            return None
