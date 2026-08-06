from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
import json
from pathlib import Path
import shutil

from openpyxl import load_workbook
import pytest

from mss.analysis.context_combination_analysis import ContextCombinationAnalysis
from mss.domain.context_combination_result import ContextCombinationResult


ROOT = Path(__file__).resolve().parents[1]


def source_paths(**updates):
    values = {
        "historical": ROOT / "reports/MSS_Historical_Backtest_Context_v1.xlsx",
        "expanded": ROOT / "reports/MSS_Context_Expansion_v1.xlsx",
        "mtf": ROOT / "reports/MSS_MTF_Context_v1.xlsx",
        "smart_money": ROOT / "reports/MSS_SmartMoney_Evidence_v1.xlsx",
        "feature_discovery": ROOT / "reports/MSS_Feature_Discovery.json",
        "shadow_validation": ROOT / "reports/MSS_Shadow_Score_Validation.json",
    }
    values.update(updates)
    return values


def synthetic_records(count=60):
    start = datetime(2026, 1, 1)
    records = []
    for index in range(count):
        first = index < count // 2
        winner = (index % 3 != 0) if first else (index % 4 == 0)
        records.append({
            "source_index": index, "trade_id": index + 1,
            "symbol": "EURUSD" if index % 2 else "XAUUSD",
            "direction": "BUY" if index % 2 else "SELL",
            "status": "CLOSED", "profit": 10.0 if winner else -7.0,
            "decision_time": start + timedelta(hours=index),
            "features": {
                "m15_trend": "UPTREND" if first else "DOWNTREND",
                "h1_trend": "UPTREND" if first else "DOWNTREND",
                "h4_trend": "UPTREND" if index % 2 else "RANGE",
                "d1_trend": "UPTREND" if first else "DOWNTREND",
                "h1_alignment": "ALIGNED" if first else "CONFLICTING",
                "h4_alignment": "ALIGNED" if index % 2 else "NEUTRAL",
                "d1_alignment": "ALIGNED" if first else "CONFLICTING",
                "overall_alignment": "ALIGNED" if first else "CONFLICTING",
                "ob_state": "FRESH" if first else "MITIGATED",
                "ob_present": "PRESENT",
                "liquidity_state": "REJECTED" if first else "UNSWEPT",
                "structure": "UPTREND" if first else "DOWNTREND",
                "bos_direction": "BULLISH" if first else "BEARISH",
                "market_zone": "DISCOUNT" if first else "PREMIUM",
                "session": "LONDON" if index % 2 else "NEWYORK",
                "kill_zone": "LONDON_OPEN" if index % 2 else "NONE",
                "volatility_regime": "NORMAL" if first else "HIGH",
            },
        })
    records.append({
        **deepcopy(records[-1]), "source_index": count, "trade_id": count + 1,
        "status": "OPEN", "profit": None,
        "decision_time": start + timedelta(hours=count),
    })
    return records


def fast_engine():
    engine = ContextCombinationAnalysis()
    engine.BOOTSTRAP_ITERATIONS = 50
    engine.PERMUTATION_ITERATIONS = 50
    return engine


def test_result_value_is_immutable_and_serialization_is_deterministic():
    first = ContextCombinationResult.create({"b": 2, "a": 1})
    second = ContextCombinationResult.create({"a": 1, "b": 2})
    assert first.payload_json == second.payload_json
    with pytest.raises(TypeError):
        first["a"] = 2


def test_validated_sources_join_exactly_and_remain_unchanged():
    paths = source_paths()
    before = {key: hashlib.sha256(Path(path).read_bytes()).hexdigest() for key, path in paths.items()}
    records, validation, audit, prior = fast_engine().load_sources(paths)
    after = {key: hashlib.sha256(Path(path).read_bytes()).hexdigest() for key, path in paths.items()}
    assert before == after
    assert len(records) == validation["joined_unique_key_count"] == 170
    assert validation["closed_trade_count"] == 169
    assert validation["unresolved_trade_count"] == 1
    assert validation["join_key_mismatch_count"] == 0
    assert validation["future_data_violation_count"] == 0
    assert prior["shadow_score_classification"] == "NOT_RELIABLE"
    assert prior["feature_discovery_measurable_features"] == []


def test_future_auxiliary_evidence_is_rejected(tmp_path):
    contaminated = tmp_path / "smart.xlsx"
    shutil.copy2(source_paths()["smart_money"], contaminated)
    workbook = load_workbook(contaminated)
    sheet = workbook["Order Block Evidence"]
    headers = [cell.value for cell in sheet[1]]
    field = headers.index("ob_detection_timestamp") + 1
    changed = False
    for row_index in range(2, sheet.max_row + 1):
        if sheet.cell(row_index, field).value not in (None, "NOT_AVAILABLE"):
            sheet.cell(row_index, field).value = "2099-01-01T00:00:00"
            changed = True
            break
    assert changed
    workbook.save(contaminated)
    with pytest.raises(ValueError, match="Future-data contamination"):
        fast_engine().load_sources(source_paths(smart_money=contaminated))


def test_join_key_mismatch_is_rejected(tmp_path):
    mismatched = tmp_path / "mtf.xlsx"
    shutil.copy2(source_paths()["mtf"], mismatched)
    workbook = load_workbook(mismatched)
    workbook["D1 Context"].delete_rows(2)
    workbook.save(mismatched)
    with pytest.raises(ValueError, match="identical keys"):
        fast_engine().load_sources(source_paths(mtf=mismatched))


def test_closed_and_unresolved_are_separated_and_inputs_unchanged():
    records = synthetic_records()
    before = deepcopy(records)
    result = fast_engine().analyze(records)
    assert records == before
    assert result["summary"]["closed_trade_count"] == 60
    assert result["summary"]["unresolved_trade_count"] == 1
    assert result["diagnostics"]["unresolved_excluded_from_statistics"] is True
    assert all("profit" not in definition["fields"] for definition in result["combination_definitions"])
    assert all("status" not in definition["fields"] for definition in result["combination_definitions"])


def test_statistics_bootstrap_and_multiple_correction_are_reproducible():
    records = synthetic_records()
    engine = fast_engine()
    first = engine.analyze(records)
    second = engine.analyze(records)
    assert first == second
    tests = [
        {"p_value": 0.01, "permutation_win_rate_p_value": 0.02},
        {"p_value": 0.04, "permutation_win_rate_p_value": 0.03},
        {"p_value": 0.50, "permutation_win_rate_p_value": 0.80},
    ]
    engine._apply_corrections(tests)
    assert all(0 <= row["adjusted_p_value"] <= 1 for row in tests)
    assert all(0 <= row["permutation_adjusted_p_value"] <= 1 for row in tests)
    assert tests[0]["adjusted_p_value"] <= tests[1]["adjusted_p_value"] <= tests[2]["adjusted_p_value"]


def test_sparse_patterns_are_warned_and_never_reliable():
    engine = fast_engine()
    row = {
        "sample_warning": "GROUP_BELOW_15",
        "win_rate_difference_ci_low": 1.0, "win_rate_difference_ci_high": 10.0,
        "expectancy_difference_ci_low": 1.0, "expectancy_difference_ci_high": 10.0,
        "mann_whitney_adjusted_p_value": 0.001,
        "permutation_adjusted_p_value": 0.001,
        "symbol_stable": True, "temporal_stable": True,
    }
    assert engine._classification(row) == "INSUFFICIENT_DATA"


def test_real_artifacts_are_byte_reproducible_and_have_exact_sheets(tmp_path):
    engine = fast_engine()
    first_xlsx, second_xlsx = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    first = engine.run(source_paths(), first_xlsx, first_json)
    second = engine.run(source_paths(), second_xlsx, second_json)
    assert first == second
    assert first_xlsx.read_bytes() == second_xlsx.read_bytes()
    assert first_json.read_bytes() == second_json.read_bytes()
    assert tuple(load_workbook(first_xlsx, read_only=True).sheetnames) == engine.REQUIRED_SHEETS
    assert json.loads(first_json.read_text(encoding="utf-8"))["summary"]["trade_count"] == 170


def test_production_modules_do_not_import_combination_analysis():
    production_files = (
        ROOT / "src/mss/analysis/smart_money_pipeline.py",
        ROOT / "src/mss/analysis/structure_engine.py",
        ROOT / "src/mss/engine/signal_engine.py",
        ROOT / "src/mss/analysis/risk_engine.py",
        ROOT / "src/mss/analysis/historical_backtest_engine.py",
    )
    for path in production_files:
        assert "context_combination_analysis" not in path.read_text(encoding="utf-8").lower()
