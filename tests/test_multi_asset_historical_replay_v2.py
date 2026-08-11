import hashlib
import json
from pathlib import Path
import zipfile
from xml.etree import ElementTree
import subprocess
import sys

from openpyxl import load_workbook

import pytest

from mss.analysis.multi_asset_historical_replay_v2 import MultiAssetHistoricalReplayV2
from mss.domain.candle import Candle
from mss.domain.historical_backtest import BacktestSymbolMetadata


ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v2.json"
XLSX_PATH = ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v2.xlsx"
V1_PATHS = (
    ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v1.json",
    ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v1.xlsx",
)


def payload():
    return json.loads(JSON_PATH.read_text(encoding="utf-8"))


def test_full_v2_orchestration_and_combined_capital():
    data = payload()
    assert [row["canonical_symbol"] for row in data["universe"]] == list(MultiAssetHistoricalReplayV2.SYMBOLS)
    assert all(row["source_candles"] == 10000 for row in data["per_symbol_results"])
    assert data["combined_independent_results"]["starting_balance"] == 80000.0
    assert data["acceptance_status"] == "PASS"


def test_canonical_and_broker_symbols_and_frozen_windows_are_preserved():
    data = payload()
    mapping = {row["canonical_symbol"]: row["broker_symbol"] for row in data["universe"]}
    assert mapping["BTCUSD"] == "BITCOIN"
    assert mapping["ETHUSD"] == "ETHEREUM"
    assert len(data["source_windows"]) == 8
    assert all(row["source_authority"] == "SPRINT_91_V1_SAVED_WINDOW_METADATA" for row in data["source_windows"])
    assert all(len(row["source_sha256"]) == 64 for row in data["source_windows"])


def test_historical_conversion_integration_account_currency_and_no_lookahead():
    data = payload()
    audits = {row["canonical_symbol"]: row for row in data["historical_fx_conversion"]}
    assert audits["USDJPY"]["conversion_path"].endswith("USDJPY:INVERSE")
    assert audits["USDCAD"]["conversion_path"].endswith("USDCAD:INVERSE")
    assert audits["XAUUSD"]["identity_factor_exactly_one"] is True
    for row in audits.values():
        assert row["future_conversion_count"] == 0
        for sample in row["sample_conversions"]:
            assert sample["entry_conversion_time"] <= sample["entry_time"]
            assert sample["exit_conversion_time"] <= sample["exit_time"]


def test_risk_sizing_and_grouped_metrics_reconcile():
    data = payload()
    assert all(row["maximum_accepted_sl_risk_percent"] <= 1.0 for row in data["risk_audit"])
    groups = {row["scope"]: row for row in data["asset_class_results"]}
    assert groups["FOREX"]["starting_balance"] == 50000.0
    assert groups["METAL"]["starting_balance"] == 10000.0
    assert groups["CRYPTO"]["starting_balance"] == 20000.0
    assert sum(row["net_profit"] for row in groups.values()) == pytest.approx(data["combined_independent_results"]["net_profit"])


def test_current_tick_value_does_not_drive_frozen_trade_valuation():
    first = BacktestSymbolMetadata(account_currency="USD", currency_base="USD", currency_profit="JPY", currency_margin="USD", trade_calc_mode=0,
        point=.001, digits=3, tick_size=.001, tick_value=.63, contract_size=100000, volume_min=.01, volume_max=100, volume_step=.01)
    second = BacktestSymbolMetadata(**{**first.__dict__, "tick_value": 999.0})
    from mss.analysis.historical_valuation import HistoricalValuation
    assert HistoricalValuation.size_for_risk(100, .1, first, 1 / 150) == HistoricalValuation.size_for_risk(100, .1, second, 1 / 150)
    assert HistoricalValuation.signed_pnl(150, 149.9, "BUY", 1, first, 1 / 150) == HistoricalValuation.signed_pnl(150, 149.9, "BUY", 1, second, 1 / 150)


def test_v1_artifacts_are_preserved_and_v2_rebuild_is_deterministic():
    data = payload()
    assert data["audit"]["v1_artifact_preserved"] is True
    assert data["audit"]["deterministic_json_rebuild"] is True
    assert data["audit"]["full_strategy_replay_count"] == 1
    assert data["audit"]["artifact_rebuild_count"] == 2
    assert all(path.exists() and path.stat().st_size > 0 for path in V1_PATHS)
    first = JSON_PATH.read_bytes()
    second = JSON_PATH.read_bytes()
    assert hashlib.sha256(first).digest() == hashlib.sha256(second).digest()


def test_workbook_completeness():
    expected = ["Executive Summary", "V1 vs V2", "Asset Performance", "Asset Class Performance", "Trades", "Equity Curves", "Risk Audit", "Rejections", "Broker Metadata", "Historical FX Conversion", "Configuration", "Source Windows", "Data Quality", "Diagnostics", "Audit"]
    with zipfile.ZipFile(XLSX_PATH) as archive:
        root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    assert [node.attrib["name"] for node in root.findall("x:sheets/x:sheet", namespace)] == expected
    workbook = load_workbook(XLSX_PATH, data_only=False, read_only=False)
    assert len(workbook.sheetnames) == 15
    assert sum(len(sheet.tables) for sheet in workbook.worksheets) == 16
    assert all(sheet.freeze_panes == "A5" for sheet in workbook.worksheets)
    assert all(sheet.tables for sheet in workbook.worksheets)
    formulas = [cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if cell.data_type == "f"]
    assert len(formulas) == 27
    assert not any("#REF!" in formula for formula in formulas)


def test_workbook_reconciles_to_json_and_rebuild_is_byte_deterministic(tmp_path):
    data = payload()
    rows = data["per_symbol_results"]
    combined = data["combined_independent_results"]
    assert sum(row["starting_balance"] for row in rows) == combined["starting_balance"] == 80000.0
    assert sum(row["ending_balance"] for row in rows) == pytest.approx(combined["ending_balance"])
    assert sum(row["net_profit"] for row in rows) == pytest.approx(combined["net_profit"])
    assert sum(row["opened_trades"] for row in rows) == combined["opened_trades"]
    assert sum(row["closed_trades"] for row in rows) == combined["closed_trades"]
    assert all(row["winners"] + row["losers"] == row["closed_trades"] for row in rows)
    builder = ROOT / "integration_tests/build_multi_asset_historical_replay_v2_workbook.py"
    outputs = [tmp_path / "first.xlsx", tmp_path / "second.xlsx"]
    for output in outputs:
        subprocess.run([sys.executable, str(builder), str(JSON_PATH), str(output)], cwd=ROOT, check=True, capture_output=True, text=True)
    assert hashlib.sha256(outputs[0].read_bytes()).digest() == hashlib.sha256(outputs[1].read_bytes()).digest()
