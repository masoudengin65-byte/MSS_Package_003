from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
import json
from math import sin
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine
from mss.analysis.mtf_evidence_engine import HistoricalTimeframeLoader, MTFEvidenceEngine
from mss.domain.candle import Candle


def candle(at, price=100.0):
    return Candle(
        time=at, open=price, high=price + 1.0, low=price - 1.0,
        close=price + 0.25, tick_volume=100, spread=2, real_volume=0,
    )


def series(decision, timeframe, count=120, slope=0.05):
    duration = HistoricalTimeframeLoader.DURATIONS[timeframe]
    start = decision - duration * count
    values = []
    for index in range(count):
        price = 100.0 + slope * index + sin(index / 3.0) * 2.0
        values.append(candle(start + duration * index, price))
    return values


def frozen_context(decision):
    values = {field: ContextCaptureEngine.NOT_AVAILABLE for field in ContextCaptureEngine.FIELDS}
    values.update({
        "snapshot_version": "SPRINT_79_V1",
        "captured_at": decision.isoformat(),
        "available_candle_count": 100,
        "latest_visible_candle_time": decision.isoformat(),
        "structure": "UPTREND",
        "trend_strength": 50.0,
        "swing_count": 10,
        "bos": True,
        "bos_direction": "BULLISH",
        "choch": False,
        "decision_time": decision.isoformat(),
        "entry_time": (decision + timedelta(minutes=15)).isoformat(),
        "entry_delay_minutes": 15.0,
        "decision_candle": {"time": decision.isoformat(), "open": 1, "high": 2, "low": 0, "close": 1.5},
        "entry_candle": {"time": (decision + timedelta(minutes=15)).isoformat(), "open": 1, "high": 2, "low": 0, "close": 1.5},
    })
    return values


def row(trade_id, decision, *, status="CLOSED", profit=10.0, symbol="EURUSD"):
    return {
        "source_index": trade_id - 1, "trade_id": trade_id, "symbol": symbol,
        "direction": "BUY", "status": status, "profit": profit,
        "decision_time": decision, "frozen_context": frozen_context(decision),
    }


def all_series(decision, slope=0.05):
    return {timeframe: series(decision, timeframe, slope=slope) for timeframe in MTFEvidenceEngine.TIMEFRAMES}


def extract(engine, decision, candle_sets):
    return engine.extract(
        trade_id=1, symbol="EURUSD", direction="BUY", decision_time=decision,
        frozen_context=frozen_context(decision), candles_by_timeframe=candle_sets,
    )


def write_source(path, rows):
    workbook = Workbook()
    trades = workbook.active
    trades.title = "Trades"
    trades.append(ContextExpansionEngine.REQUIRED_TRADE_COLUMNS)
    for item in rows:
        context = item["frozen_context"]
        trades.append([
            item["trade_id"], item["symbol"], item["direction"], context["entry_time"],
            item["profit"], item["status"], json.dumps(context),
        ])
    snapshots = workbook.create_sheet("Context Snapshot")
    snapshots.append(["Trade ID", "Symbol", *ContextCaptureEngine.FIELDS])
    for item in rows:
        context = item["frozen_context"]
        snapshots.append([
            item["trade_id"], item["symbol"],
            *[json.dumps(context[field]) if isinstance(context[field], (dict, list)) else context[field] for field in ContextCaptureEngine.FIELDS],
        ])
    workbook.save(path)


def test_completed_candle_boundary_is_inclusive_and_future_bar_is_rejected():
    start = datetime(2026, 1, 1)
    candles = [candle(start + timedelta(minutes=15 * index)) for index in range(3)]
    selected = HistoricalTimeframeLoader.completed(candles, "M15", start + timedelta(minutes=30))
    assert [item.time for item in selected] == [start, start + timedelta(minutes=15)]
    assert HistoricalTimeframeLoader.close_time(selected[-1], "M15") == start + timedelta(minutes=30)


def test_unordered_or_duplicate_candles_are_rejected():
    at = datetime(2026, 1, 1)
    with pytest.raises(ValueError, match="strictly chronological"):
        HistoricalTimeframeLoader.completed([candle(at), candle(at)], "H1", at + timedelta(hours=2))


def test_future_candles_cannot_change_extracted_evidence():
    decision = datetime(2026, 6, 1, 12)
    engine = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120)
    inputs = all_series(decision)
    first = extract(engine, decision, inputs).to_dict()
    with_future = {key: list(value) for key, value in inputs.items()}
    for timeframe in engine.TIMEFRAMES:
        with_future[timeframe].append(candle(decision, 999.0))
    second = extract(engine, decision, with_future).to_dict()
    assert first == second
    for timeframe in engine.TIMEFRAMES:
        assert datetime.fromisoformat(first[f"{timeframe.lower()}_latest_candle_close_time"]) <= decision


def test_snapshot_and_inputs_are_immutable():
    decision = datetime(2026, 6, 1, 12)
    engine = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120)
    context = frozen_context(decision)
    inputs = all_series(decision)
    context_before = deepcopy(context)
    inputs_before = deepcopy(inputs)
    snapshot = engine.extract(
        trade_id=1, symbol="EURUSD", direction="BUY", decision_time=decision,
        frozen_context=context, candles_by_timeframe=inputs,
    )
    assert context == context_before
    assert inputs == inputs_before
    with pytest.raises(TypeError):
        snapshot["h1_trend_direction"] = "DOWNTREND"


def test_missing_timeframes_preserve_not_available():
    decision = datetime(2026, 6, 1, 12)
    engine = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120)
    values = extract(engine, decision, {"M15": series(decision, "M15")}).to_dict()
    for timeframe in ("h1", "h4", "d1"):
        assert values[f"{timeframe}_availability"] == engine.NOT_AVAILABLE
        assert values[f"{timeframe}_trend_direction"] == engine.NOT_AVAILABLE
        assert values[f"{timeframe}_alignment_with_m15"] == engine.NOT_AVAILABLE
    assert values["overall_timeframe_agreement"] == engine.NOT_AVAILABLE
    assert values["h1_trend_strength"] == engine.NOT_AVAILABLE


def test_extraction_is_deterministic_and_schema_is_exact():
    decision = datetime(2026, 6, 1, 12)
    engine = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120)
    inputs = all_series(decision)
    first = extract(engine, decision, inputs)
    second = extract(engine, decision, inputs)
    assert first.payload_json == second.payload_json
    assert set(first.to_dict()) == set(engine.FIELDS)
    assert len(engine.FIELDS) == len(set(engine.FIELDS))
    assert len(ContextCaptureEngine.FIELDS) == 87
    assert len(ContextExpansionEngine.FIELDS) == 70
    assert "profit" not in first.to_dict()
    assert "status" not in first.to_dict()


def test_closed_and_unresolved_outcomes_are_separated():
    base = datetime(2026, 6, 1, 12)
    rows = [
        row(1, base, profit=10.0),
        row(2, base + timedelta(days=1), profit=-5.0),
        row(3, base + timedelta(days=2), status="OPEN", profit=None),
    ]
    candle_data = {
        "EURUSD": {timeframe: series(base + timedelta(days=3), timeframe, count=300) for timeframe in MTFEvidenceEngine.TIMEFRAMES}
    }
    validation = {
        "trade_count": 3, "context_row_count": 3, "source_context_field_count": 87,
        "expanded_context_field_count": 70, "closed_trade_count": 2,
        "unresolved_trade_count": 1, "duplicate_key_count": 0,
        "future_context_count": 0, "chronology_valid": True,
    }
    result = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120).analyze(rows, validation, candle_data)
    assert result["data_validation"]["outcome_analysis_count"] == 2
    assert result["data_validation"]["winner_count"] == 1
    assert result["data_validation"]["loser_count"] == 1
    assert result["diagnostics"]["unresolved_excluded_from_outcome_analysis"] is True


def test_artifacts_are_reproducible_and_source_is_unchanged(tmp_path):
    base = datetime(2026, 6, 1, 12)
    rows = [row(1, base, profit=10.0), row(2, base + timedelta(days=1), profit=-5.0)]
    source = tmp_path / "source.xlsx"
    write_source(source, rows)
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    candle_data = {
        "EURUSD": {timeframe: series(base + timedelta(days=2), timeframe, count=300) for timeframe in MTFEvidenceEngine.TIMEFRAMES}
    }
    first_xlsx, second_xlsx = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    engine = MTFEvidenceEngine(minimum_candles=20, lookback_limit=120)
    first = engine.run(source, candle_data, first_xlsx, first_json)
    second = engine.run(source, candle_data, second_xlsx, second_json)
    assert first == second
    assert first["data_validation"]["source_context_field_count"] == 87
    assert first["data_validation"]["expanded_context_field_count"] == 70
    assert first["data_validation"]["mtf_evidence_field_count"] == 78
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash
    assert first_xlsx.read_bytes() == second_xlsx.read_bytes()
    assert first_json.read_bytes() == second_json.read_bytes()
    assert tuple(load_workbook(first_xlsx, read_only=True).sheetnames) == engine.REQUIRED_SHEETS


def test_production_decision_modules_do_not_import_mtf_evidence():
    root = Path(__file__).resolve().parents[1]
    production_files = (
        root / "src/mss/analysis/smart_money_pipeline.py",
        root / "src/mss/analysis/structure_engine.py",
        root / "src/mss/analysis/historical_backtest_engine.py",
        root / "src/mss/analysis/context_capture_engine.py",
    )
    for path in production_files:
        assert "mtf_evidence" not in path.read_text(encoding="utf-8").lower()
