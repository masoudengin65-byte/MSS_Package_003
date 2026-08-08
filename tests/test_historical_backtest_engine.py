from datetime import datetime, timedelta

from mss.analysis.historical_backtest_engine import HistoricalBacktestEngine
from mss.analysis.historical_backtest_report_engine import (
    HistoricalBacktestReportEngine,
)
from mss.domain.candle import Candle
from mss.domain.historical_backtest import (
    BacktestSymbolMetadata,
    HistoricalBacktestConfig,
    HistoricalTrade,
)
from mss.domain.pipeline_result import PipelineResult
from mss.domain.trade_statistics import TradeStatistics


START = datetime(2026, 1, 1, 0, 0)


def candle(index, open_price=100.0, high=101.0, low=99.5, close=100.5):
    return Candle(
        time=START + timedelta(minutes=15 * index),
        open=open_price,
        high=high,
        low=low,
        close=close,
        tick_volume=100,
        spread=0,
        real_volume=100,
    )


class PipelineStub:
    def __init__(self, signal_calls=None, stop=99.0):
        self.signal_calls = set(signal_calls or [])
        self.stop = stop
        self.calls = []

    def run(self, symbol, timeframe, candles):
        self.calls.append(tuple(item.time for item in candles))
        call_number = len(self.calls)
        if call_number in self.signal_calls:
            return PipelineResult(
                symbol=symbol,
                timeframe=timeframe,
                valid=True,
                bos_detected=True,
                bos_direction="BULLISH",
                last_low=self.stop,
                structure_state="UPTREND",
                score=80,
                confidence=75.0,
                recommendation="TRADE",
            )
        return PipelineResult(
            symbol=symbol,
            timeframe=timeframe,
            valid=True,
            recommendation="WAIT",
        )


def config(**overrides):
    values = dict(
        warmup_candles=2,
        analysis_lookback=10,
        starting_balance=10000.0,
        risk_percent=1.0,
        reward_risk_ratio=2.0,
        spread_points=0,
        commission_per_lot=0.0,
        slippage_points=0.0,
        ambiguous_policy="STOP_LOSS_FIRST",
    )
    values.update(overrides)
    return HistoricalBacktestConfig(**values)


def metadata():
    return BacktestSymbolMetadata(
        point=0.01,
        digits=2,
        tick_size=0.01,
        tick_value=1.0,
        contract_size=100.0,
        volume_min=0.01,
        volume_max=100.0,
        volume_step=0.01,
        spread_points=0,
    )


def test_chronological_replay_and_no_lookahead_access():
    pipeline = PipelineStub()
    candles = [candle(3), candle(0), candle(2), candle(1)]

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), metadata()
    )

    assert result.valid
    assert result.diagnostics.reordered_candles > 0
    assert [call[-1] for call in pipeline.calls] == [
        START + timedelta(minutes=15),
        START + timedelta(minutes=30),
        START + timedelta(minutes=45),
    ]
    assert all(list(call) == sorted(call) for call in pipeline.calls)
    assert [len(call) for call in pipeline.calls] == [2, 3, 4]


def test_decision_enters_on_next_candle_open():
    pipeline = PipelineStub(signal_calls={1})
    candles = [
        candle(0),
        candle(1),
        candle(2, open_price=100, high=101, low=99.5, close=100.5),
        candle(3, open_price=100.5, high=103, low=100, close=102),
    ]

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), metadata()
    )

    trade = result.trades[0]
    assert trade.signal_time == candles[1].time
    assert trade.entry_time == candles[2].time
    assert trade.entry_price == candles[2].open
    assert trade.entry_time > trade.signal_time
    from mss.analysis.context_capture_engine import ContextCaptureEngine
    assert set(trade.context_snapshot.to_dict()) == set(ContextCaptureEngine.FIELDS)
    assert trade.context_snapshot["decision_time"] == trade.signal_time.isoformat()
    assert trade.context_snapshot["entry_time"] == trade.entry_time.isoformat()
    assert trade.context_snapshot["latest_visible_candle_time"] <= trade.context_snapshot["decision_time"]


def test_stop_loss_hit():
    pipeline = PipelineStub(signal_calls={1})
    candles = [candle(0), candle(1), candle(2, low=98.5)]

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), metadata()
    )

    assert result.trades[0].exit_reason == "STOP_LOSS"
    assert result.trades[0].profit < 0


def test_take_profit_hit():
    pipeline = PipelineStub(signal_calls={1})
    candles = [candle(0), candle(1), candle(2, high=103, low=99.5)]

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), metadata()
    )

    assert result.trades[0].exit_reason == "TAKE_PROFIT"
    assert result.trades[0].profit > 0


def test_same_candle_stop_loss_first_policy():
    pipeline = PipelineStub(signal_calls={1})
    candles = [candle(0), candle(1), candle(2, high=103, low=98.5)]

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), metadata()
    )

    assert result.trades[0].exit_reason == "STOP_LOSS"
    assert result.config.ambiguous_policy == "STOP_LOSS_FIRST"


def test_spread_slippage_and_commission_reduce_result():
    candles = [candle(0), candle(1), candle(2, high=103, low=98.0)]
    no_cost = HistoricalBacktestEngine(PipelineStub({1})).run(
        "TEST", "M15", candles, config(), metadata()
    )
    with_cost = HistoricalBacktestEngine(PipelineStub({1})).run(
        "TEST",
        "M15",
        candles,
        config(spread_points=2, slippage_points=1, commission_per_lot=1.0),
        metadata(),
    )

    assert with_cost.trades[0].spread == 0.02
    assert with_cost.trades[0].slippage == 0.01
    assert with_cost.trades[0].commission > 0
    assert with_cost.trades[0].profit < no_cost.trades[0].profit


def test_minimum_volume_that_exceeds_risk_is_rejected():
    pipeline = PipelineStub(signal_calls={1}, stop=80.0)
    candles = [candle(0), candle(1), candle(2, high=103, low=79.0)]
    expensive_tick = BacktestSymbolMetadata(
        point=1.0, digits=0, tick_size=1.0, tick_value=1000.0,
        contract_size=1.0, volume_min=0.01, volume_max=100.0,
        volume_step=0.01, spread_points=0,
    )

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), expensive_tick,
    )

    assert result.trades == []
    assert result.diagnostics.rejection_reasons == {
        "MIN_VOLUME_EXCEEDS_RISK": 1,
    }


def test_missing_tick_metadata_rejects_safely():
    pipeline = PipelineStub(signal_calls={1})
    candles = [candle(0), candle(1), candle(2, high=103, low=98.5)]
    incomplete = metadata()
    incomplete.tick_value = None

    result = HistoricalBacktestEngine(pipeline).run(
        "TEST", "M15", candles, config(), incomplete,
    )

    assert result.trades == []
    assert result.diagnostics.rejection_reasons == {
        "VALUATION_METADATA_UNAVAILABLE:MISSING_TICK_VALUE": 1,
    }


def test_equity_drawdown_and_extended_metrics():
    trades = []
    for index, profit in enumerate((100.0, -50.0, -25.0, 200.0), start=1):
        trades.append(HistoricalTrade(
            trade_id=index,
            entry_time=START + timedelta(minutes=index * 15),
            exit_time=START + timedelta(minutes=(index + 1) * 15),
            profit=profit,
            r_multiple=profit / 100.0,
            status="CLOSED",
        ))

    metrics = HistoricalBacktestEngine._calculate_metrics(
        trades, 10000.0, TradeStatistics()
    )

    assert metrics.net_profit == 225.0
    assert metrics.ending_balance == 10225.0
    assert metrics.maximum_drawdown == 75.0
    assert metrics.maximum_drawdown_percent == 0.7426
    assert metrics.maximum_consecutive_losses == 2
    assert metrics.average_holding_minutes == 15.0


def test_repeated_runs_are_deterministic():
    candles = [candle(0), candle(1), candle(2, high=103, low=99.5)]

    first = HistoricalBacktestEngine(PipelineStub({1})).run(
        "TEST", "M15", candles, config(), metadata()
    )
    second = HistoricalBacktestEngine(PipelineStub({1})).run(
        "TEST", "M15", candles, config(), metadata()
    )

    assert first.trades == second.trades
    assert first.metrics == second.metrics
    assert first.diagnostics.rejection_reasons == second.diagnostics.rejection_reasons


def test_historical_report_contains_required_sheets(tmp_path):
    candles = [candle(0), candle(1), candle(2, high=103, low=99.5)]
    result = HistoricalBacktestEngine(PipelineStub({1})).run(
        "TEST", "M15", candles, config(), metadata()
    )

    filename = HistoricalBacktestReportEngine().build(
        [result],
        tmp_path / "MSS_Historical_Backtest.xlsx",
    )

    from openpyxl import load_workbook
    workbook = load_workbook(filename, read_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Trades",
        "Equity Curve",
        "Detector Performance",
        "Configuration",
        "Diagnostics",
        "Context Snapshot",
        "Detector Context",
        "Risk Context",
        "Session Context",
        "HTF Context",
    ]
    context_headers = [cell.value for cell in next(workbook["Context Snapshot"].iter_rows())]
    from mss.analysis.context_capture_engine import ContextCaptureEngine
    assert context_headers == ["Trade ID", "Symbol"] + list(ContextCaptureEngine.FIELDS)
