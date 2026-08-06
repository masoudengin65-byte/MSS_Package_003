from copy import deepcopy
from dataclasses import FrozenInstanceError
from datetime import datetime, timedelta
import hashlib
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from mss.analysis.multi_asset_registry import MultiAssetRegistry
from mss.domain.asset_metadata import AssetDefinition
from mss.domain.candle import Candle


def broker(symbol, **overrides):
    values = {
        "name": symbol,
        "description": f"{symbol} test instrument",
        "path": "Test\\Research",
        "digits": 5,
        "point": 0.00001,
        "spread": 12,
        "trade_mode": 4,
        "visible": True,
        "select": True,
        "volume_min": 0.01,
        "volume_max": 100.0,
        "volume_step": 0.01,
        "volume_limit": 0.0,
        "trade_contract_size": 100000.0,
        "trade_tick_size": 0.00001,
        "trade_tick_value": 1.0,
        "trade_stops_level": 0,
        "trade_freeze_level": 0,
        "filling_mode": 1,
        "order_mode": 127,
        "swap_mode": 1,
        "swap_long": -1.0,
        "swap_short": 0.5,
        "currency_base": symbol[:3],
        "currency_profit": "USD",
        "currency_margin": "USD",
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def candle(at, price=100.0, **overrides):
    values = {
        "time": at,
        "open": price,
        "high": price + 1.0,
        "low": price - 1.0,
        "close": price + 0.25,
        "tick_volume": 100,
        "spread": 2,
        "real_volume": 0,
    }
    values.update(overrides)
    return Candle(**values)


def payload(as_of, timeframe, count=4, requested=None):
    duration = MultiAssetRegistry.DURATIONS[timeframe]
    candles = [candle(as_of - duration * (count - index)) for index in range(count)]
    return {
        "resolved_symbol": "EURUSD",
        "requested_count": count if requested is None else requested,
        "returned_count": count,
        "attempts": 1,
        "error_code": 1,
        "error_message": "Success",
        "candles": candles,
    }


def complete_inputs(as_of):
    registry = MultiAssetRegistry()
    brokers = tuple(broker(symbol) for symbol in registry.supported_symbols)
    history = {
        symbol: {
            timeframe: {
                **payload(as_of, timeframe),
                "resolved_symbol": symbol,
            }
            for timeframe in registry.TIMEFRAMES
        }
        for symbol in registry.supported_symbols
    }
    return brokers, history


def test_registry_has_exact_multi_asset_universe_and_classification():
    registry = MultiAssetRegistry()
    assert registry.supported_symbols == (
        "EURUSD", "GBPUSD", "USDJPY", "AUDUSD", "USDCAD",
        "XAUUSD", "BTCUSD", "ETHUSD",
    )
    classes = {item.canonical_symbol: item.asset_class for item in registry.universe}
    assert [classes[symbol] for symbol in registry.supported_symbols[:5]] == ["FOREX"] * 5
    assert classes["XAUUSD"] == "METAL"
    assert classes["BTCUSD"] == classes["ETHUSD"] == "CRYPTO"


def test_registry_schema_rejects_duplicate_or_unknown_asset_class():
    duplicate = (
        AssetDefinition("EURUSD", "FOREX", "EUR", "USD"),
        AssetDefinition("EURUSD", "FOREX", "EUR", "USD"),
    )
    with pytest.raises(ValueError, match="unique"):
        MultiAssetRegistry(duplicate)
    with pytest.raises(ValueError, match="Unsupported asset class"):
        MultiAssetRegistry((AssetDefinition("SPXUSD", "INDEX", "SPX", "USD"),))


def test_broker_resolution_is_deterministic_and_prefers_exact_match():
    registry = MultiAssetRegistry()
    symbols = [broker("EURUSD.pro"), broker("xEURUSD"), broker("EURUSD")]
    assert registry.resolve_symbol("EURUSD", symbols) == "EURUSD"
    assert registry.resolve_symbol("GBPUSD", [broker("GBPUSD.z"), broker("GBPUSD.a")]) == "GBPUSD.a"
    assert registry.resolve_symbol("ETHUSD", symbols) is None
    with pytest.raises(KeyError, match="not registered"):
        registry.resolve_symbol("SPXUSD", symbols)


def test_explicit_crypto_aliases_resolve_without_changing_canonical_identity():
    registry = MultiAssetRegistry()
    symbols = [broker("ETHEREUM"), broker("BITCOIN")]
    assert registry.resolve_symbol("BTCUSD", symbols) == "BITCOIN"
    assert registry.resolve_symbol("ETHUSD", symbols) == "ETHEREUM"
    assert registry.broker_aliases == {
        "BTCUSD": ("BITCOIN",),
        "ETHUSD": ("ETHEREUM",),
    }
    metadata = registry.capture_metadata(registry.universe[-2], broker("BITCOIN"))
    assert metadata.canonical_symbol == "BTCUSD"
    assert metadata.broker_symbol == "BITCOIN"


def test_unknown_crypto_broker_names_are_not_guessed():
    registry = MultiAssetRegistry()
    unknown = [broker("LITECOIN"), broker("BTCEUR"), broker("ETHER")]
    assert registry.resolve_symbol("BTCUSD", unknown) is None
    assert registry.resolve_symbol("ETHUSD", unknown) is None


def test_asset_metadata_is_immutable_and_captures_spread_and_conditions():
    registry = MultiAssetRegistry()
    metadata = registry.capture_metadata(registry.universe[0], broker("EURUSD"))
    assert metadata.spread_points == 12
    assert metadata.spread_price == pytest.approx(0.00012)
    assert metadata.trade_mode_name == "FULL"
    assert metadata.trade_allowed is True
    with pytest.raises(FrozenInstanceError):
        metadata.spread_points = 99


def test_future_candles_are_rejected_at_completed_bar_boundary():
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    valid = payload(as_of, "M15", count=2)
    row = registry.validate_history(registry.universe[0], "M15", valid, as_of)
    assert row["future_candle_count"] == 0
    assert row["last_candle_close_time"] == as_of.isoformat()
    contaminated = deepcopy(valid)
    contaminated["candles"].append(candle(as_of))
    contaminated["returned_count"] += 1
    with pytest.raises(ValueError, match="Future candle rejected"):
        registry.validate_history(registry.universe[0], "M15", contaminated, as_of)


def test_candle_quality_checks_invalid_ohlc_duplicates_volume_and_spread():
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    at = as_of - timedelta(minutes=30)
    bad = {
        "requested_count": 2,
        "returned_count": 2,
        "candles": [
            candle(at, open=100.0, high=99.0, low=101.0, close=100.0, tick_volume=-1, spread=-2),
            candle(at),
        ],
    }
    row = registry.validate_history(registry.universe[0], "M15", bad, as_of)
    assert not row["chronological_order"]
    assert row["duplicate_timestamp_count"] == 1
    assert row["invalid_ohlc_count"] == 1
    assert row["negative_volume_count"] == 1
    assert row["negative_spread_count"] == 1
    assert row["quality_status"] == "FAIL"


def test_missing_and_partial_history_are_preserved_not_invented():
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    missing = registry.validate_history(
        registry.universe[0], "H1",
        {"requested_count": 100, "returned_count": 0, "error_code": -4, "error_message": "No history"},
        as_of,
    )
    partial = registry.validate_history(
        registry.universe[0], "H1", payload(as_of, "H1", count=4, requested=100), as_of,
    )
    assert missing["availability_status"] == "MISSING"
    assert missing["first_candle_open_time"] == registry.NOT_AVAILABLE
    assert partial["availability_status"] == "PARTIAL"
    assert partial["coverage_percent"] == 4.0


def test_analysis_is_deterministic_and_does_not_mutate_inputs():
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    before_brokers, before_history = deepcopy(brokers), deepcopy(history)
    first = registry.analyze(brokers, history, as_of, {"terminal_build": 6090})
    second = registry.analyze(brokers, history, as_of, {"terminal_build": 6090})
    assert first == second
    assert brokers == before_brokers
    assert history == before_history
    assert tuple(first) == registry.RESULT_KEYS
    assert first["summary"]["target_symbol_count"] == 8
    assert first["summary"]["available_timeframe_count"] == 32
    assert first["diagnostics"]["future_candle_count"] == 0
    assert first["production_change_justified"] is False


def test_alias_resolution_is_preserved_in_deterministic_report_rows():
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    brokers = tuple(
        broker("BITCOIN") if item.name == "BTCUSD"
        else broker("ETHEREUM") if item.name == "ETHUSD"
        else item
        for item in brokers
    )
    for timeframe in registry.TIMEFRAMES:
        history["BTCUSD"][timeframe]["resolved_symbol"] = "BITCOIN"
        history["ETHUSD"][timeframe]["resolved_symbol"] = "ETHEREUM"
    first = registry.analyze(brokers, history, as_of)
    second = registry.analyze(brokers, history, as_of)
    assert first == second
    by_canonical = {row["canonical_symbol"]: row for row in first["symbol_registry"]}
    assert by_canonical["BTCUSD"]["broker_symbol"] == "BITCOIN"
    assert by_canonical["ETHUSD"]["broker_symbol"] == "ETHEREUM"
    assert {row["canonical_symbol"] for row in first["symbol_registry"]} == set(registry.supported_symbols)


def test_json_and_xlsx_outputs_are_reproducible(tmp_path):
    registry = MultiAssetRegistry()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    first_xlsx, second_xlsx = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    first = registry.run(brokers, history, as_of, first_xlsx, first_json)
    second = registry.run(brokers, history, as_of, second_xlsx, second_json)
    assert first == second
    assert hashlib.sha256(first_json.read_bytes()).digest() == hashlib.sha256(second_json.read_bytes()).digest()
    assert hashlib.sha256(first_xlsx.read_bytes()).digest() == hashlib.sha256(second_xlsx.read_bytes()).digest()
    assert tuple(load_workbook(first_xlsx, read_only=True).sheetnames) == registry.REQUIRED_SHEETS


def test_production_modules_do_not_import_multi_asset_layer():
    root = Path(__file__).resolve().parents[1]
    production_files = (
        root / "src/mss/analysis/smart_money_pipeline.py",
        root / "src/mss/analysis/structure_engine.py",
        root / "src/mss/engine/signal_engine.py",
        root / "src/mss/analysis/risk_engine.py",
        root / "src/mss/analysis/execution_pipeline.py",
        root / "src/mss/execution/mt5_executor.py",
    )
    for path in production_files:
        content = path.read_text(encoding="utf-8").lower()
        assert "multi_asset_registry" not in content
        assert "asset_metadata" not in content
    layer = (root / "src/mss/analysis/multi_asset_registry.py").read_text(encoding="utf-8").lower()
    for prohibited in ("smart_money_pipeline", "structure_engine", "signal_engine", "risk_engine", "mt5_executor"):
        assert prohibited not in layer
