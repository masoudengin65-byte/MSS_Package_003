from copy import deepcopy
import json

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from mss.analysis.score_engine import ScoreEngine
from mss.analysis.shadow_score_validation import ShadowScoreValidation


def row(i, *, symbol="EURUSD", status="CLOSED", profit=1.0, score=None, confidence=None, component=1):
    score = i if score is None else score
    confidence = float(score) if confidence is None else confidence
    return {
        "source_index": i - 1, "trade_id": i, "symbol": symbol,
        "direction": "BUY" if i % 2 else "SELL", "entry_time": f"2026-01-{i:02d}T00:00:00",
        "status": status, "exit_reason": "TAKE_PROFIT" if profit > 0 else "STOP_LOSS",
        "profit": profit, "r_multiple": profit, "legacy_score": 25,
        "legacy_confidence": 22.73, "context_snapshot": "{}",
        "shadow_score": score, "shadow_confidence": confidence,
        "components": {name: (component if name == "structure" else ScoreEngine.NOT_AVAILABLE) for name in ScoreEngine.COMPONENTS},
    }


def shadow_of(item):
    return {k: deepcopy(item[k]) for k in ("trade_id", "symbol", "direction", "entry_time", "legacy_score", "legacy_confidence", "shadow_score", "shadow_confidence", "components")}


def test_deterministic_trade_to_score_join_and_order():
    engine = ShadowScoreValidation(); trades = [row(2), row(1)]; scores = [shadow_of(x) for x in reversed(trades)]
    joined, quality = engine.join(trades, scores)
    assert [x["trade_id"] for x in joined] == [1, 2]
    assert quality["matched_trades"] == 2 and quality["chronology_preserved"]


def test_composite_key_allows_same_trade_id_across_symbols():
    engine = ShadowScoreValidation(); trades = [row(1), row(1, symbol="XAUUSD")]
    joined, quality = engine.join(trades, [shadow_of(x) for x in trades])
    assert len(joined) == 2 and quality["duplicate_identifiers"] == 0


def test_duplicate_detection_does_not_silently_discard():
    engine = ShadowScoreValidation(); trade = row(1)
    joined, quality = engine.join([trade, deepcopy(trade)], [shadow_of(trade)])
    assert joined == [] and quality["duplicate_identifiers"] == 1 and quality["unmatched_trades"] == 2


def test_unresolved_preserved_and_closed_population_selected():
    engine = ShadowScoreValidation(); trades = [row(1), row(2, status="OPEN", profit=0)]
    result = engine.analyze(trades, [shadow_of(x) for x in trades])
    assert result["data_integrity"]["unresolved_trades"] == 1
    assert len(result["unresolved_diagnostics"]) == 1
    assert result["score_analysis"]["descriptive_statistics"]["count"] == 1


def test_fixed_bin_boundaries_cover_maximum():
    groups = ShadowScoreValidation().fixed_groups([row(1, score=29), row(2, score=30), row(3, score=59)], "shadow_score", 5)
    assert groups[0]["lower"] == 25 and groups[-1]["upper"] == 60
    assert sum(x["trade_count"] for x in groups) == 3


def test_quantiles_are_deterministic_and_do_not_split_ties():
    rows = [row(i, score=10 if i < 6 else 20) for i in range(1, 11)]
    engine = ShadowScoreValidation(); first = engine.quantile_groups(rows, "shadow_score"); second = engine.quantile_groups(rows, "shadow_score")
    assert first == second and len(first) == 2 and [x["trade_count"] for x in first] == [5, 5]


def test_bootstrap_is_deterministic():
    engine = ShadowScoreValidation(); values = list(range(20)); labels = [i % 2 for i in values]
    assert engine.bootstrap_auc(values, labels, "x") == engine.bootstrap_auc(values, labels, "x")


def test_permutation_is_deterministic():
    engine = ShadowScoreValidation(); values = list(range(20)); labels = [i % 2 for i in values]
    assert engine.permutation_auc(values, labels, "x") == engine.permutation_auc(values, labels, "x")


def test_score_outcome_metrics():
    metrics = ShadowScoreValidation().metrics([row(1, profit=10), row(2, profit=-5)])
    assert metrics["wins"] == 1 and metrics["losses"] == 1
    assert metrics["total_net_profit"] == 5 and metrics["expectancy"] == 2.5


def test_symbol_stratification():
    rows = [row(i, symbol="EURUSD" if i < 5 else "XAUUSD", profit=1 if i % 2 else -1) for i in range(1, 9)]
    result = ShadowScoreValidation().analyze(rows, [shadow_of(x) for x in rows])
    assert result["symbol_analysis"]["EURUSD"]["trade_count"] == 4
    assert result["symbol_analysis"]["XAUUSD"]["trade_count"] == 4


def test_chronological_split_integrity():
    rows = [row(i, profit=1 if i % 2 else -1) for i in range(1, 9)]
    periods = ShadowScoreValidation()._temporal(rows)
    assert periods["first_half"]["trade_count"] == 4 and periods["second_half"]["trade_count"] == 4
    assert periods["first_half"]["end"] < periods["second_half"]["start"]


def test_not_available_remains_distinct_from_zero():
    available = row(1, component=0); missing = row(2); missing["components"]["structure"] = ScoreEngine.NOT_AVAILABLE
    item = ShadowScoreValidation()._components([available, missing], [available, missing])[0]
    assert item["availability_count"] == 1 and item["not_available_count"] == 1
    assert item["minimum"] == 0 and item["contribution_frequency"] == 0


def test_multiple_comparison_correction():
    items = [{"p_value": .01}, {"p_value": .04}, {"p_value": None}]
    ShadowScoreValidation.apply_bh(items)
    assert items[0]["adjusted_p_value"] <= items[1]["adjusted_p_value"]
    assert items[2]["adjusted_p_value"] is None


def test_report_generation_and_json_serialization(tmp_path):
    rows = [row(i, profit=1 if i % 2 else -1) for i in range(1, 9)]
    engine = ShadowScoreValidation(); result = engine.analyze(rows, [shadow_of(x) for x in rows])
    xlsx = tmp_path / "report.xlsx"; js = tmp_path / "report.json"
    engine.write_workbook(result, xlsx); engine.write_json(result, js)
    assert set(engine.REQUIRED_SHEETS) == set(load_workbook(xlsx, read_only=True).sheetnames)
    assert json.loads(js.read_text())["schema_version"] == engine.VERSION


def test_repeated_run_equality_and_no_input_mutation():
    rows = [row(i, profit=1 if i % 2 else -1) for i in range(1, 9)]; scores = [shadow_of(x) for x in rows]
    before_rows, before_scores = deepcopy(rows), deepcopy(scores); engine = ShadowScoreValidation()
    assert engine.analyze(rows, scores) == engine.analyze(rows, scores)
    assert rows == before_rows and scores == before_scores


def test_no_production_strategy_changes_are_reported():
    trade = row(1); result = ShadowScoreValidation().analyze([trade], [shadow_of(trade)])
    assert result["diagnostics"]["production_strategy_changed"] is False
    assert result["configuration"]["mode"] == "DIAGNOSTIC_ONLY"


def test_winner_loser_statistics_include_required_distribution_fields():
    rows = [row(i, profit=1 if i <= 4 else -1) for i in range(1, 9)]
    comparison = ShadowScoreValidation()._win_loss(rows, "shadow_score")
    required = {"count", "minimum", "maximum", "mean", "median", "standard_deviation",
                "first_quartile", "third_quartile", "unique_value_count"}
    assert required <= comparison["winners"].keys()
    assert required <= comparison["losers"].keys()


def test_missing_predictor_does_not_misalign_outcome_labels():
    rows = [row(1, profit=-1), row(2, profit=1), row(3, profit=1)]
    rows[0]["shadow_score"] = None
    analysis = ShadowScoreValidation()._predictor_analysis(rows, "shadow_score", 5)
    assert analysis["descriptive_statistics"]["count"] == 2
    assert analysis["auc"] is None  # Remaining usable records contain only winners.


def test_conclusions_use_only_required_reliability_classes():
    rows = [row(i, profit=1 if i % 2 else -1) for i in range(1, 9)]
    result = ShadowScoreValidation().analyze(rows, [shadow_of(x) for x in rows])
    allowed = {"RELIABLE", "PROMISING_BUT_LIMITED", "NOT_RELIABLE", "INSUFFICIENT_DATA"}
    assert all(item["classification"] in allowed for item in result["conclusions"].values())
    assert result["conclusions"]["production_use_justified"]["supported"] is False


def test_schema_validation_rejects_missing_columns(tmp_path):
    path = tmp_path / "bad.xlsx"
    workbook = Workbook(); workbook.active.title = "Trades"; workbook.active.append(["Trade ID"]); workbook.save(path)
    with pytest.raises(ValueError, match="missing required columns"):
        ShadowScoreValidation().load_historical(path)


def test_workbook_output_is_byte_deterministic(tmp_path):
    rows = [row(i, profit=1 if i % 2 else -1) for i in range(1, 9)]
    engine = ShadowScoreValidation(); result = engine.analyze(rows, [shadow_of(x) for x in rows])
    first = tmp_path / "first.xlsx"; second = tmp_path / "second.xlsx"
    engine.write_workbook(result, first); engine.write_workbook(result, second)
    assert first.read_bytes() == second.read_bytes()
