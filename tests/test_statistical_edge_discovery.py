import json
from copy import deepcopy
from datetime import datetime, timedelta

from openpyxl import load_workbook

from mss.analysis.statistical_edge_discovery import StatisticalEdgeDiscovery


def records(count=41, unresolved=1):
    start=datetime(2026,1,1)
    rows=[]
    for i in range(count):
        rows.append({
            "trade_id":i+1,"symbol":"EURUSD" if i%2 else "XAUUSD","direction":"BUY" if i%3 else "SELL",
            "status":"CLOSED","profit":100.0 if i%2 else -50.0,"r_multiple":1.0 if i%2 else -.5,
            "entry_time_value":start+timedelta(hours=i),"decision_time_value":start+timedelta(hours=i)-timedelta(minutes=15),
            "exit_time_value":start+timedelta(hours=i+1),"trend_strength":float(i),"swing_count":i%5,
            "bos_strength":float(i),"atr":float(i+1),"average_candle_size":1.0,"current_candle_size":1.0,
            "relative_volatility":1.0,"spread":2.0,"tick_volume":100+i,"sl_distance":1.0,"tp_distance":2.0,
            "rr":2.0,"structure_score":0,"bos_score":25,"choch_score":0,"liquidity_score":0,"ob_score":0,
            "fvg_score":0,"session_score":0,"htf_score":0,"risk_score":0,"unattributed_score":0,
            "final_score":25,"confidence":22.73,"position_size":1.0,"portfolio_exposure":0.0,"correlation_score":0.0,
            "entry_delay_minutes":15.0,"decision_hour":i%24,"entry_hour":i%24,"holding_duration_minutes":60.0,
            "structure":"UPTREND","bos":True,"bos_direction":"BULLISH","choch":False,"choch_direction":"NOT_AVAILABLE",
            "liquidity_detected":False,"liquidity_side":"NOT_AVAILABLE","liquidity_sweep":False,
            "order_block_detected":False,"fvg_detected":False,"session":"LONDON","kill_zone":"NONE",
            "day_of_week":"Monday","risk_approved":True,"liquidity_distance":"NOT_AVAILABLE",
        })
    for i in range(unresolved):
        row=deepcopy(rows[-1]);row["trade_id"]=count+i+1;row["status"]="OPEN";rows.append(row)
    return rows


def test_analysis_is_deterministic_and_does_not_mutate_records():
    engine=StatisticalEdgeDiscovery();source=records();before=deepcopy(source)
    first=engine.analyze_records(source);second=engine.analyze_records(source)
    assert first==second
    assert source==before


def test_closed_preservation_and_unresolved_exclusion():
    result=StatisticalEdgeDiscovery().analyze_records(records())
    assert result["diagnostics"]["closed_trade_indices_preserved"]==list(range(41))
    symbol_total=sum(x["trade_count"] for x in result["categorical_analysis"] if x["feature"]=="symbol")
    assert symbol_total==41


def test_quantile_bins_reproducible_and_minimum_sample_merged():
    engine=StatisticalEdgeDiscovery();rows=records(80,0)
    for i,row in enumerate(rows):row["_index"]=i
    first=engine.quantile_bins(rows,"trend_strength");second=engine.quantile_bins(rows,"trend_strength")
    assert first==second
    assert all(list(first[0].values()).count(label)>=20 for label in set(first[0].values()))


def test_minimum_sample_labels():
    engine=StatisticalEdgeDiscovery();rows=records(19,0)
    for i,row in enumerate(rows):row["_index"]=i
    assert engine.metrics(rows)["sample_status"]=="LOW_SAMPLE"
    rows=records(20,0)
    for i,row in enumerate(rows):row["_index"]=i
    assert engine.metrics(rows)["sample_status"]=="LIMITED"
    rows=records(40,0)
    for i,row in enumerate(rows):row["_index"]=i
    assert engine.metrics(rows)["sample_status"]=="RELIABLE"


def test_interactions_filter_groups_below_twenty():
    result=StatisticalEdgeDiscovery().analyze_records(records(30,0))
    assert all(x["trade_count"]>=20 for x in result["interaction_analysis"])


def test_temporal_split_integrity_and_ranking_determinism():
    engine=StatisticalEdgeDiscovery();result=engine.analyze_records(records(40,0))
    for item in result["temporal_stability"]:
        assert item["first_half"]["trade_count"]+item["second_half"]["trade_count"]==item["combined_trade_count"]
    assert result["edge_ranking"]==engine.analyze_records(records(40,0))["edge_ranking"]


def test_workbook_and_json_schema_completeness(tmp_path):
    engine=StatisticalEdgeDiscovery();result=engine.analyze_records(records(40,0),{"closed_trades":40,"excluded_unresolved":0})
    xlsx=tmp_path/"report.xlsx";js=tmp_path/"report.json"
    engine.write_workbook(result,xlsx);engine.write_json(result,js)
    assert load_workbook(xlsx,read_only=True).sheetnames==engine.REQUIRED_SHEETS
    payload=json.loads(js.read_text())
    assert set(("schema_version","data_quality","analysis_config","numeric_bins","categorical_analysis","interaction_analysis","temporal_stability","edge_ranking","symbol_analysis","unavailable_findings","diagnostics"))<=set(payload)
