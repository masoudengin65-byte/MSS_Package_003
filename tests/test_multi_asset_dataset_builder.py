from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from mss.analysis.multi_asset_dataset_builder import MultiAssetDatasetBuilder
from mss.analysis.multi_asset_registry import MultiAssetRegistry
from mss.domain.candle import Candle


def broker(name):
    return SimpleNamespace(name=name)


def candle(at, price=100.0, **overrides):
    values = {
        "time": at,
        "open": price,
        "high": price + 1.0,
        "low": price - 1.0,
        "close": price + 0.25,
        "tick_volume": 100,
        "spread": 2,
        "real_volume": 0,
    }
    values.update(overrides)
    return Candle(**values)


def payload(as_of, timeframe, broker_symbol, count=4):
    duration = MultiAssetRegistry.DURATIONS[timeframe]
    candles = [
        candle(as_of - duration * (count - index), 100.0 + index)
        for index in range(count)
    ]
    return {
        "resolved_symbol": broker_symbol,
        "requested_count": count,
        "returned_count": count,
        "attempts": 1,
        "error_code": 1,
        "error_message": "Success",
        "candles": candles,
    }


def complete_inputs(as_of):
    builder = MultiAssetDatasetBuilder()
    broker_map = {
        symbol: "BITCOIN" if symbol == "BTCUSD"
        else "ETHEREUM" if symbol == "ETHUSD"
        else symbol
        for symbol in builder.supported_symbols
    }
    brokers = tuple(broker(name) for name in broker_map.values())
    history = {
        canonical: {
            timeframe: payload(as_of, timeframe, broker_symbol)
            for timeframe in builder.TIMEFRAMES
        }
        for canonical, broker_symbol in broker_map.items()
    }
    return brokers, history


def test_builder_loads_registry_assets_and_preserves_broker_identity():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    result = builder.build(brokers, history, as_of).to_dict()
    assets = {row["canonical_symbol"]: row for row in result["supported_assets"]}
    assert tuple(row["canonical_symbol"] for row in result["supported_assets"]) == builder.supported_symbols
    assert assets["BTCUSD"]["broker_symbol"] == "BITCOIN"
    assert assets["ETHUSD"]["broker_symbol"] == "ETHEREUM"
    assert {row["asset_class"] for row in result["records"]} == {"FOREX", "METAL", "CRYPTO"}
    assert result["summary"]["candle_record_count"] == 8 * 4 * 4


def test_dataset_snapshot_is_immutable_and_inputs_are_unchanged():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    before_brokers, before_history = deepcopy(brokers), deepcopy(history)
    first = builder.build(brokers, history, as_of, {"terminal_build": 6090})
    second = builder.build(brokers, history, as_of, {"terminal_build": 6090})
    assert first.payload_json == second.payload_json
    assert first.sha256 == second.sha256
    assert brokers == before_brokers
    assert history == before_history
    with pytest.raises(TypeError):
        first["records"] = []
    extracted = first.to_dict()
    extracted["records"].clear()
    assert first.to_dict()["records"]


def test_future_candle_is_rejected_and_cannot_create_lookahead():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    history["EURUSD"]["M15"]["candles"].append(candle(as_of))
    history["EURUSD"]["M15"]["returned_count"] += 1
    with pytest.raises(ValueError, match="Future candle rejected"):
        builder.build(brokers, history, as_of)


def test_duplicate_and_non_chronological_timestamps_are_detected_not_removed():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    source = history["EURUSD"]["M15"]
    duplicate = source["candles"][1]
    source["candles"] = [source["candles"][0], duplicate, duplicate, source["candles"][2]]
    source["returned_count"] = len(source["candles"])
    source["requested_count"] = len(source["candles"])
    result = builder.build(brokers, history, as_of).to_dict()
    assert result["data_quality"]["chronology_failure_count"] == 1
    assert result["data_quality"]["duplicate_timestamp_count"] == 1
    assert result["data_quality"]["duplicate_record_key_count"] == 1
    affected = [
        row for row in result["records"]
        if row["canonical_symbol"] == "EURUSD" and row["timeframe"] == "M15"
        and "DUPLICATE_TIMESTAMP" in row["quality_flags"]
    ]
    assert len(affected) == 2
    assert len([row for row in result["records"] if row["canonical_symbol"] == "EURUSD" and row["timeframe"] == "M15"]) == 4


def test_missing_candle_intervals_are_counted_without_imputation():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    source = history["BTCUSD"]["M15"]
    source["candles"].pop(1)
    source["returned_count"] -= 1
    result = builder.build(brokers, history, as_of).to_dict()
    coverage = next(
        row for row in result["coverage"]
        if row["canonical_symbol"] == "BTCUSD" and row["timeframe"] == "M15"
    )
    assert coverage["missing_candle_interval_count"] == 1
    assert len([row for row in result["records"] if row["canonical_symbol"] == "BTCUSD" and row["timeframe"] == "M15"]) == 3
    issue = next(row for row in result["issues"] if row["issue_code"] == "MISSING_CANDLE_INTERVALS")
    assert issue["severity"] == "HIGH"


def test_invalid_ohlc_is_flagged_without_reconstruction():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    original = history["XAUUSD"]["H1"]["candles"][0]
    history["XAUUSD"]["H1"]["candles"][0] = candle(
        original.time, open=100.0, high=99.0, low=101.0, close=100.0,
    )
    result = builder.build(brokers, history, as_of).to_dict()
    assert result["data_quality"]["invalid_ohlc_count"] == 1
    record = next(
        row for row in result["records"]
        if row["canonical_symbol"] == "XAUUSD" and row["timeframe"] == "H1"
        and row["source_index"] == 0
    )
    assert record["quality_flags"] == "INVALID_OHLC"
    assert record["open"] == 100.0
    assert record["high"] == 99.0


def test_missing_timeframe_is_preserved_as_unavailable():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    history["ETHUSD"]["D1"] = {
        "resolved_symbol": "ETHEREUM", "requested_count": 4,
        "returned_count": 0, "attempts": 3, "error_code": -4,
        "error_message": "No history", "candles": [],
    }
    result = builder.build(brokers, history, as_of).to_dict()
    assert result["summary"]["missing_timeframe_slice_count"] == 1
    assert not [row for row in result["records"] if row["canonical_symbol"] == "ETHUSD" and row["timeframe"] == "D1"]
    coverage = next(
        row for row in result["coverage"]
        if row["canonical_symbol"] == "ETHUSD" and row["timeframe"] == "D1"
    )
    assert coverage["availability_status"] == "MISSING"
    assert coverage["first_candle_open_time"] == builder.NOT_AVAILABLE


def test_dataset_schema_and_metadata_hashes_are_exact():
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    result = builder.build(brokers, history, as_of).to_dict()
    assert set(result) == set(builder.RESULT_KEYS)
    assert len(result["coverage"]) == 32
    assert all(set(row) == set(builder.RECORD_FIELDS) for row in result["records"])
    assert result["dataset_metadata"]["immutable_snapshot"] is True
    assert result["dataset_metadata"]["lookahead_permitted"] is False
    assert len(result["dataset_metadata"]["records_sha256"]) == 64
    assert result["diagnostics"]["future_candle_count"] == 0
    assert result["diagnostics"]["lookahead_violation_count"] == 0
    assert result["production_change_justified"] is False


def test_json_and_xlsx_artifacts_are_reproducible(tmp_path):
    builder = MultiAssetDatasetBuilder()
    as_of = datetime(2026, 1, 2, 12)
    brokers, history = complete_inputs(as_of)
    first_xlsx, second_xlsx = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    first = builder.run(brokers, history, as_of, first_xlsx, first_json)
    second = builder.run(brokers, history, as_of, second_xlsx, second_json)
    assert first == second
    assert hashlib.sha256(first_json.read_bytes()).digest() == hashlib.sha256(second_json.read_bytes()).digest()
    assert hashlib.sha256(first_xlsx.read_bytes()).digest() == hashlib.sha256(second_xlsx.read_bytes()).digest()
    workbook = load_workbook(first_xlsx, read_only=True, data_only=True)
    assert tuple(workbook.sheetnames) == builder.REQUIRED_SHEETS
    assert workbook["M15 Candles"].max_row == 8 * 4 + 1
    workbook.close()


def test_production_modules_do_not_import_dataset_builder():
    root = Path(__file__).resolve().parents[1]
    production_files = (
        root / "src/mss/analysis/smart_money_pipeline.py",
        root / "src/mss/analysis/structure_engine.py",
        root / "src/mss/engine/signal_engine.py",
        root / "src/mss/analysis/risk_engine.py",
        root / "src/mss/analysis/execution_pipeline.py",
        root / "src/mss/execution/mt5_executor.py",
    )
    for path in production_files:
        content = path.read_text(encoding="utf-8").lower()
        assert "multi_asset_dataset_builder" not in content
        assert "multi_asset_dataset" not in content
    layer = (root / "src/mss/analysis/multi_asset_dataset_builder.py").read_text(encoding="utf-8").lower()
    for prohibited in (
        "smart_money_pipeline", "structure_engine", "signal_engine",
        "risk_engine", "mt5_executor",
    ):
        assert prohibited not in layer
