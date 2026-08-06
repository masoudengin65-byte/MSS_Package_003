from pathlib import Path

from mss.analysis.excel_report_engine import ExcelReportEngine
from mss.domain.report import Report
from mss.domain.trade_statistics import TradeStatistics
from mss.domain.premium_discount import PremiumDiscount
from mss.domain.kill_zone_status import KillZoneStatus
from mss.domain.session_bias import SessionBias
from mss.domain.news_risk_status import NewsRiskStatus
from mss.domain.portfolio_exposure import PortfolioExposure
from openpyxl import load_workbook
from datetime import timedelta


def build_report():

    stats = TradeStatistics()

    stats.total_trades = 10
    stats.winning_trades = 7
    stats.losing_trades = 3
    stats.breakeven_trades = 0

    stats.gross_profit = 850
    stats.gross_loss = 250
    stats.net_profit = 600

    stats.win_rate = 70.0
    stats.profit_factor = 3.40
    stats.expectancy = 60.0
    stats.max_drawdown = 120.0

    stats.equity_curve = [

        0,
        100,
        180,
        250,
        220,
        350,
        470,
        600,

    ]

    report = Report()

    report.processed_candles = 500

    report.generated_signals = 18

    report.executed_trades = 10

    report.execution_time = 0.42

    report.statistics = stats

    report.premium_discount = PremiumDiscount(
        swing_high=110.0,
        swing_low=90.0,
        premium_zone=(100.0, 110.0),
        discount_zone=(90.0, 100.0),
        equilibrium=100.0,
        current_zone="PREMIUM",
        distance_to_equilibrium=5.0,
        valid=True,
    )

    report.kill_zone_status = KillZoneStatus(
        current_session="LONDON",
        active_kill_zone="LONDON_OPEN",
        remaining_time=timedelta(hours=1, minutes=15),
        active=True,
        valid=True,
    )

    report.session_bias = SessionBias(
        current_session="LONDON",
        bias="Bullish",
        strength=82.5,
        confidence=76.25,
        valid=True,
    )

    report.news_risk_status = NewsRiskStatus(
        next_event="US Nonfarm Payrolls",
        event_impact="HIGH",
        minutes_remaining=18,
        trading_status="BLOCKED",
        valid=True,
    )

    report.portfolio_exposure = PortfolioExposure(
        portfolio_exposure=2.5,
        currency_exposure={"EUR": 1.0, "USD": -2.5},
        asset_exposure={"EURUSD": 1.0, "XAUUSD": 1.5},
        correlation_level="HIGH",
        correlation_percent=100.0,
        portfolio_risk_score=62.5,
        risk_level="HIGH",
        open_positions=2,
        valid=True,
    )

    return report


def test_excel_report_created(tmp_path):

    report = build_report()

    filename = ExcelReportEngine().build(

        report,

        output_folder=tmp_path,

    )

    assert Path(filename).exists()


def test_excel_report_extension(tmp_path):

    report = build_report()

    filename = ExcelReportEngine().build(

        report,

        output_folder=tmp_path,

    )

    assert filename.endswith(".xlsx")


def test_excel_report_not_empty(tmp_path):

    report = build_report()

    filename = ExcelReportEngine().build(

        report,

        output_folder=tmp_path,

    )

    assert Path(filename).stat().st_size > 0


def test_excel_report_stores_premium_discount(tmp_path):

    filename = ExcelReportEngine().build(
        build_report(),
        output_folder=tmp_path,
    )

    sheet = load_workbook(filename).active

    assert sheet["E3"].value == "100.0 - 110.0"
    assert sheet["E4"].value == "90.0 - 100.0"
    assert sheet["E5"].value == 100.0
    assert sheet["E6"].value == "PREMIUM"
    assert sheet["E7"].value == 5.0


def test_excel_report_stores_kill_zone_status(tmp_path):

    filename = ExcelReportEngine().build(
        build_report(),
        output_folder=tmp_path,
    )

    sheet = load_workbook(filename).active

    assert sheet["E9"].value == "LONDON"
    assert sheet["E10"].value == "LONDON_OPEN"
    assert sheet["E11"].value == "01:15:00"
    assert sheet["E12"].value is True


def test_excel_report_stores_session_bias(tmp_path):

    filename = ExcelReportEngine().build(
        build_report(),
        output_folder=tmp_path,
    )

    sheet = load_workbook(filename).active

    assert sheet["E14"].value == "LONDON"
    assert sheet["E15"].value == "Bullish"
    assert sheet["E16"].value == 82.5
    assert sheet["E17"].value == 76.25


def test_excel_report_stores_news_risk(tmp_path):

    filename = ExcelReportEngine().build(
        build_report(),
        output_folder=tmp_path,
    )

    sheet = load_workbook(filename).active

    assert sheet["E19"].value == "US Nonfarm Payrolls"
    assert sheet["E20"].value == "HIGH"
    assert sheet["E21"].value == 18
    assert sheet["E22"].value == "BLOCKED"


def test_excel_report_stores_portfolio_exposure(tmp_path):

    filename = ExcelReportEngine().build(
        build_report(),
        output_folder=tmp_path,
    )

    sheet = load_workbook(filename).active

    assert sheet["E24"].value == 2.5
    assert sheet["E25"].value == "EUR: 1.0, USD: -2.5"
    assert sheet["E26"].value == "EURUSD: 1.0, XAUUSD: 1.5"
    assert sheet["E27"].value == "HIGH"
    assert sheet["E28"].value == 62.5
    assert sheet["E29"].value == "HIGH"
