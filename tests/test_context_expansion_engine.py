from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
import json

from openpyxl import Workbook, load_workbook
import pytest

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine


def frozen_context(i=1, **updates):
    values = {field: ContextCaptureEngine.NOT_AVAILABLE for field in ContextCaptureEngine.FIELDS}
    decision = datetime(2026, 1, 1) + timedelta(minutes=15 * i)
    values.update({
        "snapshot_version": "SPRINT_79_V1",
        "captured_at": decision.isoformat(),
        "available_candle_count": 100,
        "latest_visible_candle_time": decision.isoformat(),
        "structure": "UPTREND" if i % 2 else "DOWNTREND",
        "trend_strength": 80.0,
        "swing_count": 20,
        "bos": True,
        "bos_direction": "BULLISH" if i % 2 else "BEARISH",
        "choch": False,
        "liquidity_detected": i % 2 == 0,
        "liquidity_sweep": False,
        "liquidity_distance": 1.5 if i % 2 == 0 else ContextCaptureEngine.NOT_AVAILABLE,
        "order_block_detected": False,
        "fvg_detected": False,
        "session": "LONDON",
        "kill_zone": "LONDON_OPEN",
        "decision_time": decision.isoformat(),
        "entry_time": (decision + timedelta(minutes=15)).isoformat(),
        "entry_delay_minutes": 15.0,
        "decision_candle": {"time": decision.isoformat(), "open": 1, "high": 2, "low": 0, "close": 1.5},
        "entry_candle": {"time": (decision + timedelta(minutes=15)).isoformat(), "open": 9, "high": 10, "low": 8, "close": 9.5},
    })
    values.update(updates)
    return values


def source_row(i=1, *, status="CLOSED", profit=None, **updates):
    context = frozen_context(i, **updates)
    return {
        "source_index": i - 1, "trade_id": i,
        "symbol": "EURUSD" if i % 2 else "XAUUSD",
        "direction": "BUY" if i % 2 else "SELL",
        "status": status,
        "profit": (10.0 if i % 3 else -5.0) if profit is None else profit,
        "decision_time": datetime.fromisoformat(context["decision_time"]),
        "frozen_context": context,
    }


def expand(engine, row, **kwargs):
    return engine.expand(
        trade_id=row["trade_id"], symbol=row["symbol"], direction=row["direction"],
        decision_time=row["decision_time"], frozen_context=row["frozen_context"], **kwargs,
    )


def write_source(path, rows):
    workbook = Workbook()
    trades = workbook.active
    trades.title = "Trades"
    trades.append(ContextExpansionEngine.REQUIRED_TRADE_COLUMNS)
    for row in rows:
        context = row["frozen_context"]
        trades.append([
            row["trade_id"], row["symbol"], row["direction"], context["entry_time"],
            row["profit"], row["status"], json.dumps(context),
        ])
    snapshot = workbook.create_sheet("Context Snapshot")
    snapshot.append(["Trade ID", "Symbol", *ContextCaptureEngine.FIELDS])
    for row in rows:
        context = row["frozen_context"]
        snapshot.append([
            row["trade_id"], row["symbol"],
            *[json.dumps(context[field]) if isinstance(context[field], (dict, list)) else context[field] for field in ContextCaptureEngine.FIELDS],
        ])
    workbook.save(path)


def test_expanded_snapshot_is_immutable_and_source_unchanged():
    engine = ContextExpansionEngine()
    row = source_row(1)
    before = deepcopy(row["frozen_context"])
    snapshot = expand(engine, row)
    assert row["frozen_context"] == before
    with pytest.raises(TypeError):
        snapshot["h1_trend_direction"] = "UPTREND"


def test_future_timeframe_candle_is_rejected():
    engine = ContextExpansionEngine()
    row = source_row(1)
    future = row["decision_time"] + timedelta(minutes=1)
    evidence = {"h1": {"captured_at": row["decision_time"], "latest_visible_candle_time": future, "trend_direction": "UPTREND"}}
    with pytest.raises(ValueError, match="future candle"):
        expand(engine, row, timeframe_evidence=evidence)


def test_timeframe_boundary_is_inclusive_and_alignment_is_deterministic():
    engine = ContextExpansionEngine()
    row = source_row(1)
    at_decision = row["decision_time"]
    evidence = {"h1": {
        "captured_at": at_decision, "latest_visible_candle_time": at_decision,
        "market_structure": "UPTREND", "trend_direction": "UPTREND",
        "bos_state": True, "choch_state": False, "swing_count": 8,
        "trend_strength": 70.0,
    }}
    first = expand(engine, row, timeframe_evidence=evidence).to_dict()
    second = expand(engine, row, timeframe_evidence=evidence).to_dict()
    assert first == second
    assert first["h1_trend_direction"] == "UPTREND"
    assert first["h1_alignment_with_m15"] == "ALIGNED"


def test_unavailable_context_is_not_inferred():
    values = expand(ContextExpansionEngine(), source_row(1)).to_dict()
    for field in (
        "h1_trend_direction", "h4_trend_direction", "d1_trend_direction",
        "ob_distance_from_entry", "fvg_fill_percentage", "equal_highs",
        "session_timezone", "session_dst_handling", "nearby_economic_event",
    ):
        assert values[field] == ContextExpansionEngine.NOT_AVAILABLE


def test_schema_has_unique_stable_fields():
    engine = ContextExpansionEngine()
    assert len(engine.FIELDS) == len(set(engine.FIELDS))
    assert set(engine.FIELDS) == set(expand(engine, source_row(1)).to_dict())


def test_loader_rejects_future_m15_context(tmp_path):
    path = tmp_path / "future.xlsx"
    row = source_row(1)
    row["frozen_context"]["latest_visible_candle_time"] = "2099-01-01T00:00:00"
    write_source(path, [row])
    with pytest.raises(ValueError, match="Future-data chronology violation"):
        ContextExpansionEngine().load(path)


def test_analysis_excludes_unresolved_from_outcome_distribution():
    rows = [source_row(i) for i in range(1, 6)] + [source_row(6, status="OPEN")]
    result = ContextExpansionEngine().analyze(rows)
    assert result["data_validation"]["closed_trade_count"] == 5
    assert result["data_validation"]["unresolved_trade_count"] == 1
    assert result["diagnostics"]["unresolved_excluded_from_outcome_analysis"] is True


def test_run_is_reproducible_and_preserves_input(tmp_path):
    source = tmp_path / "source.xlsx"
    first_excel, second_excel = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    write_source(source, [source_row(i) for i in range(1, 9)])
    input_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    engine = ContextExpansionEngine()
    first = engine.run(source, first_excel, first_json)
    second = engine.run(source, second_excel, second_json)
    assert first == second
    assert hashlib.sha256(source.read_bytes()).hexdigest() == input_hash
    assert first_excel.read_bytes() == second_excel.read_bytes()
    assert first_json.read_bytes() == second_json.read_bytes()
    assert set(load_workbook(first_excel, read_only=True).sheetnames) == set(engine.REQUIRED_SHEETS)
