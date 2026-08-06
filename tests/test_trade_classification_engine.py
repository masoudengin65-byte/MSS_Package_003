from datetime import datetime, timedelta

from openpyxl import load_workbook

from mss.analysis.trade_classification_engine import TradeClassificationEngine


def trade(profit=-100.0, status="CLOSED", liquidity=False):
    start = datetime(2026, 1, 5, 13, 15)
    return {
        "Trade ID": 1, "Symbol": "EURUSD", "Timeframe": "M15", "Direction": "BUY",
        "Signal Time": start - timedelta(minutes=15), "Entry Time": start,
        "Entry Price": 1.1000, "Stop Loss": 1.0980, "Take Profit": 1.1040,
        "Exit Time": start + timedelta(hours=1), "Exit Price": 1.0980,
        "Exit Reason": "STOP_LOSS", "Spread": 0.00002, "Commission": 0,
        "Slippage": 0.00001, "Volume": .5, "Profit/Loss": profit,
        "R Multiple": profit / 100, "Score": 25, "Confidence": 22.73,
        "Status": status,
        "Detector and Context States": '{"structure":"UPTREND","bos":true,"bos_direction":"BULLISH","choch":false,"liquidity":%s,"order_block":false,"fair_value_gap":false}' % str(liquidity).lower(),
    }


def test_classification_is_deterministic_and_evidenced():
    engine = TradeClassificationEngine()
    assert engine.classify_trade(trade(), 1) == engine.classify_trade(trade(), 1)
    row = engine.classify_trade(trade(), 1)
    assert row["Failure Category"] == "LOW_CONFLUENCE"
    assert row["Classification Evidence"]


def test_all_source_trades_preserved_and_classified(tmp_path):
    engine = TradeClassificationEngine()
    rows = [engine.classify_trade(trade(i), i) for i in (-100, 200)]
    rows.append(engine.classify_trade(trade(0, "OPEN"), 3))
    output = tmp_path / "classification.xlsx"
    engine.build(rows, output)
    wb = load_workbook(output, read_only=True)
    assert wb["Classified Trades"].max_row - 1 == 3
    assert all(row["Failure Category"] for row in rows)
    assert rows[-1]["Failure Category"] == "UNCLASSIFIED"


def test_workbook_completeness(tmp_path):
    engine = TradeClassificationEngine()
    output = tmp_path / "classification.xlsx"
    engine.build([engine.classify_trade(trade(), 1)], output)
    assert load_workbook(output, read_only=True).sheetnames == engine.REQUIRED_SHEETS


def test_statistical_calculations_and_low_sample():
    engine = TradeClassificationEngine()
    rows = [engine.classify_trade(trade(200), i) for i in range(5)] + [engine.classify_trade(trade(-100), i) for i in range(5, 10)]
    stats = engine.group_statistics(rows, "Symbol")[0]
    assert stats["Trade Count"] == 10
    assert stats["Win Rate %"] == 50.0
    assert stats["Profit Factor"] == 2.0
    assert stats["Expectancy"] == 50.0
    assert stats["Average R"] == 0.5
    assert stats["Net Profit"] == 500.0
    assert stats["Drawdown Contribution"] == 500.0
    assert stats["Reliability"] == "LOW_SAMPLE"


def test_twenty_trades_are_reliable():
    engine = TradeClassificationEngine()
    rows = [engine.classify_trade(trade(10), i) for i in range(20)]
    assert engine.group_statistics(rows, "Symbol")[0]["Reliability"] == "RELIABLE"
