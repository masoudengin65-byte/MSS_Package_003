from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
import json

from openpyxl import Workbook, load_workbook
import pytest

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.feature_discovery import FeatureDiscovery


def context(i=1, **updates):
    values = {field: ContextCaptureEngine.NOT_AVAILABLE for field in ContextCaptureEngine.FIELDS}
    decision_time = datetime(2026, 1, 1) + timedelta(minutes=15 * i)
    values.update({
        "snapshot_version": "SPRINT_79_V1",
        "captured_at": decision_time.isoformat(),
        "available_candle_count": 100 + i,
        "latest_visible_candle_time": decision_time.isoformat(),
        "structure": "UPTREND" if i % 2 else "DOWNTREND",
        "bos": True,
        "bos_direction": "BULLISH" if i % 2 else "BEARISH",
        "liquidity_detected": i % 3 == 0,
        "liquidity_distance": 0.5 if i % 3 == 0 else ContextCaptureEngine.NOT_AVAILABLE,
        "liquidity_sweep": False,
        "order_block_detected": i == 1,
        "fvg_detected": False,
        "equilibrium_price": 100.0,
        "premium": [100.0, 110.0],
        "discount": [90.0, 100.0],
        "current_zone": "PREMIUM" if i % 2 else "DISCOUNT",
        "session": ("ASIA", "LONDON", "NEWYORK")[i % 3],
        "kill_zone": "LONDON_OPEN" if i % 3 == 1 else "NONE",
        "atr": 1.0,
        "average_candle_size": 2.0,
        "current_candle_size": 1.0 + (i % 4),
        "relative_volatility": (1.0 + (i % 4)) / 2.0,
        "tick_volume": 100 + i,
        "decision_time": decision_time.isoformat(),
        "entry_time": (decision_time + timedelta(minutes=15)).isoformat(),
        "entry_delay_minutes": 15.0,
        "decision_candle": {
            "time": decision_time.isoformat(), "open": 99.0,
            "high": 102.0, "low": 98.0, "close": 101.0,
            "spread": 2, "tick_volume": 100 + i,
        },
        "entry_candle": {
            "time": (decision_time + timedelta(minutes=15)).isoformat(),
            "open": 1000.0, "high": 2000.0, "low": 0.0, "close": 1500.0,
            "spread": 999, "tick_volume": 999999,
        },
    })
    values.update(updates)
    return values


def trade(i=1, *, status="CLOSED", profit=None, **context_updates):
    ctx = context(i, **context_updates)
    return {
        "source_index": i - 1,
        "trade_id": i,
        "symbol": "EURUSD" if i % 2 else "XAUUSD",
        "direction": "BUY" if i % 2 else "SELL",
        "entry_time": datetime.fromisoformat(ctx["entry_time"]),
        "status": status,
        "profit": (10.0 if i % 3 else -5.0) if profit is None else profit,
        "context": ctx,
    }


def write_fixture(path, rows):
    workbook = Workbook()
    trades = workbook.active
    trades.title = "Trades"
    trade_headers = [
        "Trade ID", "Symbol", "Direction", "Entry Time", "Profit/Loss",
        "Status", "Frozen Context Snapshot",
    ]
    trades.append(trade_headers)
    for item in rows:
        trades.append([
            item["trade_id"], item["symbol"], item["direction"], item["entry_time"],
            item["profit"], item["status"], json.dumps(item["context"]),
        ])
    snapshot = workbook.create_sheet("Context Snapshot")
    snapshot.append(["Trade ID", "Symbol", *ContextCaptureEngine.FIELDS])
    for item in rows:
        snapshot.append([item["trade_id"], item["symbol"], *[json.dumps(item["context"][field]) if isinstance(item["context"][field], (dict, list)) else item["context"][field] for field in ContextCaptureEngine.FIELDS]])
    workbook.save(path)


def test_feature_extraction_is_deterministic_and_input_is_immutable():
    engine = FeatureDiscovery()
    item = trade(1)
    before = deepcopy(item)
    assert engine.extract(item) == engine.extract(item)
    assert item == before


def test_no_lookahead_entry_candle_or_outcome_use():
    engine = FeatureDiscovery()
    first = trade(1, profit=10)
    second = deepcopy(first)
    second["profit"] = -99999
    second["context"]["entry_candle"] = {
        "time": "2099-01-01T00:00:00", "open": -1, "high": 999999,
        "low": -999999, "close": 42, "spread": 999, "tick_volume": 999999,
    }
    assert engine.extract(first)["features"] == engine.extract(second)["features"]


def test_missing_values_are_preserved_not_guessed():
    item = trade(
        1, liquidity_distance=ContextCaptureEngine.NOT_AVAILABLE,
        atr=ContextCaptureEngine.NOT_AVAILABLE,
        premium=ContextCaptureEngine.NOT_AVAILABLE,
        discount=ContextCaptureEngine.NOT_AVAILABLE,
    )
    features = FeatureDiscovery().extract(item)["features"]
    assert features["liquidity_distance_atr"] is None
    assert features["range_position"] is None


def test_statistics_and_randomized_calculations_are_reproducible():
    engine = FeatureDiscovery()
    rows = [trade(i) for i in range(1, 61)]
    first = engine.analyze(rows)
    second = engine.analyze(rows)
    assert first == second
    values = list(range(20))
    labels = [i % 2 for i in values]
    assert engine.bootstrap_auc(values, labels, "repeat") == engine.bootstrap_auc(values, labels, "repeat")
    assert engine.permutation_auc(values, labels, "repeat") == engine.permutation_auc(values, labels, "repeat")


def test_closed_and_unresolved_trade_separation():
    rows = [trade(i) for i in range(1, 6)] + [trade(6, status="OPEN", profit=None)]
    result = FeatureDiscovery().analyze(rows)
    assert result["data_validation"]["closed_trades"] == 5
    assert result["data_validation"]["unresolved_trades"] == 1
    assert len(result["unresolved_trades"]) == 1
    assert result["diagnostics"]["unresolved_excluded_from_statistics"] is True


def test_schema_validation_rejects_missing_context_columns(tmp_path):
    path = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    workbook.active.title = "Trades"
    workbook.active.append(FeatureDiscovery.REQUIRED_TRADE_COLUMNS)
    context_sheet = workbook.create_sheet("Context Snapshot")
    context_sheet.append(["Trade ID", "Symbol", "snapshot_version"])
    workbook.save(path)
    with pytest.raises(ValueError, match="missing required columns"):
        FeatureDiscovery().load(path)


def test_schema_validation_rejects_future_visible_candle(tmp_path):
    path = tmp_path / "future.xlsx"
    item = trade(1)
    item["context"]["latest_visible_candle_time"] = "2099-01-01T00:00:00"
    write_fixture(path, [item])
    with pytest.raises(ValueError, match="Future-data chronology violation"):
        FeatureDiscovery().load(path)


def test_run_preserves_immutable_workbook_and_writes_required_sheets(tmp_path):
    source = tmp_path / "history.xlsx"
    excel = tmp_path / "feature.xlsx"
    json_path = tmp_path / "feature.json"
    write_fixture(source, [trade(i) for i in range(1, 9)])
    before = hashlib.sha256(source.read_bytes()).hexdigest()
    engine = FeatureDiscovery()
    result = engine.run(source, excel, json_path)
    after = hashlib.sha256(source.read_bytes()).hexdigest()
    assert before == after
    assert result["data_validation"]["input_artifact"]["unchanged"] is True
    assert set(load_workbook(excel, read_only=True).sheetnames) == set(engine.REQUIRED_SHEETS)
    assert json.loads(json_path.read_text(encoding="utf-8"))["schema_version"] == engine.VERSION


def test_run_rejects_overwriting_input(tmp_path):
    source = tmp_path / "same.xlsx"
    source.touch()
    with pytest.raises(ValueError, match="must not overwrite"):
        FeatureDiscovery().run(source, source, tmp_path / "result.json")
