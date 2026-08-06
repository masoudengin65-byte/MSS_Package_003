from copy import deepcopy
from datetime import datetime, timedelta
import hashlib
import json
from math import sin
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from mss.analysis.context_capture_engine import ContextCaptureEngine
from mss.analysis.context_expansion_engine import ContextExpansionEngine
from mss.analysis.mtf_evidence_engine import HistoricalTimeframeLoader, MTFEvidenceEngine
from mss.analysis.smart_money_evidence_engine import SmartMoneyEvidenceEngine
from mss.domain.candle import Candle
from mss.domain.fair_value_gap import FairValueGap
from mss.domain.order_block import OrderBlock
from mss.domain.swing_point import SwingPoint


def candle(at, price=100.0, *, open_=None, high=None, low=None, close=None):
    open_value = price if open_ is None else open_
    return Candle(
        time=at, open=open_value,
        high=price + 1.0 if high is None else high,
        low=price - 1.0 if low is None else low,
        close=price + 0.2 if close is None else close,
        tick_volume=100, spread=2, real_volume=0,
    )


def series(decision, count=120):
    duration = HistoricalTimeframeLoader.DURATIONS["M15"]
    start = decision - duration * count
    return [
        candle(start + duration * index, 100.0 + index * 0.03 + sin(index / 3.0))
        for index in range(count)
    ]


def frozen_context(decision, session="LONDON"):
    values = {field: ContextCaptureEngine.NOT_AVAILABLE for field in ContextCaptureEngine.FIELDS}
    values.update({
        "snapshot_version": "SPRINT_79_V1", "captured_at": decision.isoformat(),
        "available_candle_count": 100,
        "latest_visible_candle_time": decision.isoformat(),
        "structure": "UPTREND", "trend_strength": 50.0, "swing_count": 10,
        "bos": True, "bos_direction": "BULLISH", "choch": False,
        "session": session, "decision_time": decision.isoformat(),
        "entry_time": (decision + timedelta(minutes=15)).isoformat(),
        "entry_delay_minutes": 15.0,
        "decision_candle": {"time": decision.isoformat(), "open": 1, "high": 2, "low": 0, "close": 1.5},
        "entry_candle": {"time": (decision + timedelta(minutes=15)).isoformat(), "open": 1, "high": 2, "low": 0, "close": 1.5},
    })
    return values


def row(trade_id, decision, *, status="CLOSED", profit=10.0, symbol="EURUSD", session="LONDON"):
    return {
        "source_index": trade_id - 1, "trade_id": trade_id, "symbol": symbol,
        "direction": "BUY", "status": status, "profit": profit,
        "decision_time": decision, "frozen_context": frozen_context(decision, session),
    }


def extract(engine, decision, candles):
    return engine.extract(
        trade_id=1, symbol="EURUSD", direction="BUY", decision_time=decision,
        frozen_context=frozen_context(decision), candles=candles,
    )


def write_source(path, rows):
    workbook = Workbook()
    trades = workbook.active
    trades.title = "Trades"
    trades.append(ContextExpansionEngine.REQUIRED_TRADE_COLUMNS)
    for item in rows:
        context = item["frozen_context"]
        trades.append([
            item["trade_id"], item["symbol"], item["direction"], context["entry_time"],
            item["profit"], item["status"], json.dumps(context),
        ])
    snapshots = workbook.create_sheet("Context Snapshot")
    snapshots.append(["Trade ID", "Symbol", *ContextCaptureEngine.FIELDS])
    for item in rows:
        context = item["frozen_context"]
        snapshots.append([
            item["trade_id"], item["symbol"],
            *[json.dumps(context[field]) if isinstance(context[field], (dict, list)) else context[field] for field in ContextCaptureEngine.FIELDS],
        ])
    workbook.save(path)


def test_future_candle_cannot_change_evidence_and_close_boundary_is_exact():
    decision = datetime(2026, 6, 1, 12)
    engine = SmartMoneyEvidenceEngine(minimum_candles=21, lookback_limit=120)
    candles = series(decision)
    first = extract(engine, decision, candles).to_dict()
    second = extract(engine, decision, candles + [candle(decision, 999.0)]).to_dict()
    assert first == second
    assert datetime.fromisoformat(first["latest_candle_close_time"]) == decision


def test_snapshot_and_source_inputs_are_immutable():
    decision = datetime(2026, 6, 1, 12)
    engine = SmartMoneyEvidenceEngine(minimum_candles=21, lookback_limit=120)
    context = frozen_context(decision)
    candles = series(decision)
    before_context, before_candles = deepcopy(context), deepcopy(candles)
    snapshot = engine.extract(
        trade_id=1, symbol="EURUSD", direction="BUY", decision_time=decision,
        frozen_context=context, candles=candles,
    )
    assert context == before_context
    assert candles == before_candles
    with pytest.raises(TypeError):
        snapshot["ob_detected"] = True


def test_missing_history_preserves_not_available():
    decision = datetime(2026, 6, 1, 12)
    engine = SmartMoneyEvidenceEngine(minimum_candles=100, lookback_limit=120)
    values = extract(engine, decision, series(decision, count=30)).to_dict()
    assert values["ob_detected"] == engine.NOT_AVAILABLE
    assert values["fvg_detected"] == engine.NOT_AVAILABLE
    assert values["liquidity_lifecycle_state"] == engine.NOT_AVAILABLE
    assert values["ob_invalidation"] == engine.NOT_AVAILABLE
    assert values["fvg_invalidation"] == engine.NOT_AVAILABLE


def test_order_block_lifecycle_uses_only_available_candles():
    start = datetime(2026, 1, 1)
    candles = [candle(start + timedelta(minutes=15 * index), 105.0) for index in range(30)]
    candles[10] = candle(
        candles[10].time, 100.0, open_=100.5, high=101.0, low=99.0, close=99.5,
    )
    candles[20] = candle(
        candles[20].time, 99.0, open_=100.0, high=100.5, low=98.0, close=98.5,
    )
    block = OrderBlock(
        direction="BULLISH", high=101.0, low=99.0, open=100.5,
        close=99.5, candle_time=candles[10].time, valid=True,
    )
    candidate = {"object": block, "detection_candle": candles[12], "departure_strength": 2.5}
    decision = HistoricalTimeframeLoader.close_time(candles[-1], "M15")
    values = SmartMoneyEvidenceEngine(minimum_candles=21)._order_block_evidence(candidate, candles, decision)
    assert values["ob_mitigation_state"] == "MITIGATED"
    assert values["ob_freshness"] is False
    assert values["ob_retest_count"] >= 1
    assert values["ob_invalidation"] == SmartMoneyEvidenceEngine.NOT_AVAILABLE


def test_fvg_fill_boundaries_are_deterministic():
    start = datetime(2026, 1, 1)
    candles = [candle(start + timedelta(minutes=15 * index), 105.0) for index in range(20)]
    gap = FairValueGap(
        direction="BULLISH", low=100.0, high=102.0,
        candle_time=candles[10].time, valid=True,
    )
    candidate = {"object": gap, "detection_candle": candles[11]}
    candles[12] = candle(candles[12].time, 103.0, high=104.0, low=101.0, close=103.0)
    partial_decision = HistoricalTimeframeLoader.close_time(candles[12], "M15")
    partial = SmartMoneyEvidenceEngine(minimum_candles=21)._fvg_evidence(candidate, candles[:13], partial_decision)
    assert partial["fvg_fill_percentage"] == 50.0
    assert partial["fvg_lifecycle_state"] == "PARTIALLY_FILLED"
    candles[13] = candle(candles[13].time, 101.0, high=103.0, low=100.0, close=101.0)
    full_decision = HistoricalTimeframeLoader.close_time(candles[13], "M15")
    full = SmartMoneyEvidenceEngine(minimum_candles=21)._fvg_evidence(candidate, candles[:14], full_decision)
    assert full["fvg_fill_percentage"] == 100.0
    assert full["fvg_lifecycle_state"] == "FILLED"
    assert full["fvg_invalidation"] == SmartMoneyEvidenceEngine.NOT_AVAILABLE


def test_liquidity_sweep_rejection_uses_existing_tolerance_and_completed_data():
    start = datetime(2026, 1, 1)
    candles = [candle(start + timedelta(minutes=15 * index), 104.0) for index in range(30)]
    candles[20] = candle(candles[20].time, 104.0, high=106.0, low=103.0, close=104.0)
    first = SwingPoint(index=5, price=105.0, time=candles[5].time, is_high=True, valid=True)
    second = SwingPoint(index=10, price=105.05, time=candles[10].time, is_high=True, valid=True)

    class FixedSwings:
        @staticmethod
        def detect(_candles):
            return [first, second]

    engine = SmartMoneyEvidenceEngine(minimum_candles=21)
    engine.swing_engine = FixedSwings()
    decision = HistoricalTimeframeLoader.close_time(candles[-1], "M15")
    values = engine._liquidity_evidence(candles, decision)
    assert values["equal_highs"] is True
    assert values["liquidity_sweep_detected"] is True
    assert values["liquidity_rejection_after_sweep"] is True
    assert values["liquidity_lifecycle_state"] == "REJECTED"


def test_schema_and_repeated_extraction_are_deterministic():
    decision = datetime(2026, 6, 1, 12)
    engine = SmartMoneyEvidenceEngine(minimum_candles=21, lookback_limit=120)
    candles = series(decision)
    first = extract(engine, decision, candles)
    second = extract(engine, decision, candles)
    assert first.payload_json == second.payload_json
    assert set(first.to_dict()) == set(engine.FIELDS)
    assert len(engine.FIELDS) == len(set(engine.FIELDS)) == 66
    assert len(ContextCaptureEngine.FIELDS) == 87
    assert len(ContextExpansionEngine.FIELDS) == 70
    assert len(MTFEvidenceEngine.FIELDS) == 78
    assert "profit" not in first.to_dict()
    assert "status" not in first.to_dict()


def test_artifacts_are_reproducible_and_unresolved_is_excluded(tmp_path):
    base = datetime(2026, 6, 1, 12)
    rows = [
        row(1, base, profit=10.0),
        row(2, base + timedelta(days=1), profit=-5.0),
        row(3, base + timedelta(days=2), status="OPEN", profit=None),
    ]
    source = tmp_path / "source.xlsx"
    write_source(source, rows)
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    candles = {"EURUSD": {"M15": series(base + timedelta(days=3), count=500)}}
    first_xlsx, second_xlsx = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    first_json, second_json = tmp_path / "first.json", tmp_path / "second.json"
    engine = SmartMoneyEvidenceEngine(minimum_candles=21, lookback_limit=120)
    first = engine.run(source, candles, first_xlsx, first_json)
    second = engine.run(source, candles, second_xlsx, second_json)
    assert first == second
    assert first["data_validation"]["outcome_analysis_count"] == 2
    assert first["data_validation"]["unresolved_trade_count"] == 1
    assert first["diagnostics"]["unresolved_excluded_from_outcome_analysis"] is True
    assert first["data_validation"]["source_context_field_count"] == 87
    assert first["data_validation"]["expanded_context_field_count"] == 70
    assert first["data_validation"]["mtf_evidence_field_count"] == 78
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash
    assert first_xlsx.read_bytes() == second_xlsx.read_bytes()
    assert first_json.read_bytes() == second_json.read_bytes()
    assert tuple(load_workbook(first_xlsx, read_only=True).sheetnames) == engine.REQUIRED_SHEETS


def test_production_modules_do_not_import_smart_money_evidence():
    root = Path(__file__).resolve().parents[1]
    production_files = (
        root / "src/mss/analysis/smart_money_pipeline.py",
        root / "src/mss/analysis/structure_engine.py",
        root / "src/mss/analysis/historical_backtest_engine.py",
        root / "src/mss/analysis/context_capture_engine.py",
    )
    for path in production_files:
        assert "smart_money_evidence" not in path.read_text(encoding="utf-8").lower()
