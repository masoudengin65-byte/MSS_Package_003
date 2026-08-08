"""Sprint 80: build and audit the context-enriched paper replay."""

import json
from pathlib import Path

import MetaTrader5 as mt5
from openpyxl import load_workbook

from mss.adapters.mt5.history import HistoryLoadError, HistoryService
from mss.analysis.historical_backtest_engine import HistoricalBacktestEngine
from mss.analysis.historical_backtest_report_engine import HistoricalBacktestReportEngine
from mss.analysis.historical_context_audit import HistoricalContextAudit
from mss.domain.historical_backtest import BacktestSymbolMetadata, HistoricalBacktestConfig
from mss.domain.trade_statistics import TradeStatistics


TERMINAL_PATH = r"C:\Program Files\Alpari MT5\terminal64.exe"
SYMBOLS = ("EURUSD", "XAUUSD")
CANDLE_COUNT = 10_000
OUTPUT = Path("reports/MSS_Historical_Backtest_Context_v1.xlsx")
AUDIT_OUTPUT = Path("reports/MSS_Historical_Backtest_Context_v1_audit.json")
BASELINE = Path("reports/MSS_Historical_Backtest.xlsx")


def config():
    return HistoricalBacktestConfig(
        warmup_candles=200, analysis_lookback=500, starting_balance=10_000.0,
        risk_percent=1.0, reward_risk_ratio=2.0, spread_points=None,
        commission_per_lot=0.0, slippage_points=1.0,
        ambiguous_policy="STOP_LOSS_FIRST",
    )


def _metadata(symbol):
    info = mt5.symbol_info(symbol)
    if info is None:
        raise RuntimeError(f"symbol_info failed for {symbol}: {mt5.last_error()}")
    return BacktestSymbolMetadata(point=info.point, digits=info.digits,
        tick_size=info.trade_tick_size, tick_value=info.trade_tick_value,
        contract_size=info.trade_contract_size,
        volume_min=info.volume_min, volume_max=info.volume_max,
        volume_step=info.volume_step, spread_points=info.spread)


def _baseline_metrics(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Summary"]
    headers = [cell.value for cell in next(sheet.iter_rows())]
    return {row[0].value: dict(zip(headers[1:], [cell.value for cell in row[1:]])) for row in sheet.iter_rows(min_row=2)}


def _metrics(result):
    return {name: getattr(result.metrics, name) for name in (
        "total_trades", "win_rate", "net_profit", "profit_factor", "expectancy",
        "average_r", "maximum_drawdown", "ending_balance",
    )}


def main():
    if not mt5.initialize(path=TERMINAL_PATH, timeout=60_000):
        raise RuntimeError(f"MT5 initialize failed: {mt5.last_error()}")
    try:
        print("MT5_VERSION", mt5.version())
        print("TERMINAL_INFO", mt5.terminal_info())
        print("ACCOUNT_INFO", mt5.account_info())
        histories = {}
        metadata = {}
        loader = HistoryService(max_attempts=3, retry_delay=0.25)
        for symbol in SYMBOLS:
            history = loader.load(symbol, mt5.TIMEFRAME_M15, CANDLE_COUNT, start_position=1)
            if not history.success:
                raise HistoryLoadError(history)
            times = [c.time for c in history.candles]
            if len(times) != CANDLE_COUNT or times != sorted(times) or len(times) != len(set(times)):
                raise RuntimeError(f"Invalid chronological history for {symbol}")
            histories[symbol] = history.candles
            metadata[symbol] = _metadata(history.resolved_symbol)
            print("HISTORY", symbol, history.diagnostic)

        results = [HistoricalBacktestEngine().run(s, "M15", histories[s], config(), metadata[s]) for s in SYMBOLS]
        repeated = [HistoricalBacktestEngine().run(s, "M15", histories[s], config(), metadata[s]) for s in SYMBOLS]
        deterministic = all(a.trades == b.trades and a.metrics == b.metrics for a, b in zip(results, repeated))
        if not deterministic:
            raise RuntimeError("Repeated replay output is not deterministic")

        HistoricalBacktestReportEngine().build(results, OUTPUT)
        audit = HistoricalContextAudit().calculate(results)
        audit["deterministic_repeated_output"] = deterministic
        audit["real_orders_sent"] = False
        audit["configuration"] = vars(config())
        audit["metrics"] = {result.symbol: _metrics(result) for result in results}
        combined_trades = sorted(
            [trade for result in results for trade in result.trades if trade.status == "CLOSED"],
            key=lambda trade: trade.exit_time,
        )
        combined = HistoricalBacktestEngine._calculate_metrics(
            combined_trades, 20_000.0, TradeStatistics()
        )
        audit["metrics"]["Combined"] = {name: getattr(combined, name) for name in (
            "total_trades", "win_rate", "net_profit", "profit_factor", "expectancy",
            "average_r", "maximum_drawdown", "ending_balance",
        )}
        audit["sprint_77_baseline"] = _baseline_metrics(BASELINE)
        AUDIT_OUTPUT.write_text(json.dumps(audit, indent=2, sort_keys=True), encoding="utf-8")
        print("AUDIT", json.dumps(audit, sort_keys=True, default=str))
        print("OUTPUT", OUTPUT)
        print("REAL_ORDERS_SENT", False)
        return results, audit
    finally:
        mt5.shutdown()


if __name__ == "__main__":
    main()
