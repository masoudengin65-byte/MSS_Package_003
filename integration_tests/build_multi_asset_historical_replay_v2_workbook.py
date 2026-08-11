"""Build the deterministic Sprint 92A.3 v2 workbook from frozen JSON only."""

from __future__ import annotations

from copy import copy
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from normalize_xlsx_package import normalize


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v2.json"
DEFAULT_OUTPUT = ROOT / "reports/MSS_Multi_Asset_Historical_Replay_v2.xlsx"
SHEETS = (
    "Executive Summary", "V1 vs V2", "Asset Performance",
    "Asset Class Performance", "Trades", "Equity Curves", "Risk Audit",
    "Rejections", "Broker Metadata", "Historical FX Conversion",
    "Configuration", "Source Windows", "Data Quality", "Diagnostics", "Audit",
)
NAVY, BLUE, PALE_BLUE = "17365D", "1F4E78", "D9EAF7"
PALE_GREEN, PALE_RED, BORDER = "E2F0D9", "FCE4D6", "D9E2F3"
THIN = Side(style="thin", color=BORDER)
FIXED_TIME = datetime(2026, 8, 7, 17, 15, tzinfo=timezone.utc)


def scalar(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, sort_keys=True, separators=(",", ":"), allow_nan=False)


def flatten(row, exclude=()):
    output = {}
    for key, value in row.items():
        if key in exclude:
            continue
        if isinstance(value, dict):
            for child, child_value in value.items():
                output[f"{key}.{child}"] = scalar(child_value)
        else:
            output[key] = scalar(value)
    return output


def safe_table_name(name):
    return re.sub(r"[^A-Za-z0-9]", "", name) + "Table"


def style_title(sheet, width, title, subtitle):
    last = get_column_letter(max(2, width))
    sheet.merge_cells(f"A1:{last}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(f"A2:{last}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Aptos", italic=True, color="404040")
    sheet["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 30


def number_format(header):
    lower = header.lower()
    if "percent" in lower or lower.endswith("_rate"):
        return '0.0000"%";[Red](0.0000)"%";-'
    if any(word in lower for word in (
        "balance", "profit", "loss", "expectancy", "drawdown", "risk",
        "price", "commission", "slippage", "spread", "volume", "equity",
    )):
        return '#,##0.00####;[Red](#,##0.00####);-'
    if "factor" in lower or lower.endswith("_r") or lower in {"r_multiple", "profit_factor"}:
        return '0.000000;[Red](0.000000);-'
    if any(word in lower for word in ("count", "trades", "winners", "losers", "decisions", "signals", "results")):
        return '#,##0'
    return "General"


def add_table_sheet(workbook, name, title, subtitle, raw_rows, headers=None):
    sheet = workbook.create_sheet(name)
    rows = [flatten(row) for row in raw_rows]
    headers = list(headers or (rows[0].keys() if rows else ["status"]))
    style_title(sheet, len(headers), title, subtitle)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 32
    values = rows or [{headers[0]: "NO_DATA"}]
    for row_index, row in enumerate(values, start=5):
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_index, column, scalar(row.get(header)))
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.number_format = number_format(header)
    last_row = 4 + len(values)
    last_column = get_column_letter(len(headers))
    table = Table(displayName=safe_table_name(name), ref=f"A4:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A4:{last_column}{last_row}"
    for column, header in enumerate(headers, start=1):
        observed = [len(str(header))]
        for row_index in range(5, min(last_row, 104) + 1):
            observed.append(len(str(sheet.cell(row_index, column).value or "")))
        sheet.column_dimensions[get_column_letter(column)].width = min(42, max(11, max(observed) + 2))
    return sheet, headers, last_row


def key_value_rows(values):
    return [{"metric": key, "value": scalar(value)} for key, value in values.items()]


def build(input_path=DEFAULT_INPUT, output_path=DEFAULT_OUTPUT):
    input_path, output_path = Path(input_path), Path(output_path)
    data = json.loads(input_path.read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "MSS Sprint 92A.3b"
    workbook.properties.lastModifiedBy = "MSS Sprint 92A.3b"
    workbook.properties.created = FIXED_TIME
    workbook.properties.modified = FIXED_TIME
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    combined = data["combined_independent_results"]
    conclusions = data["research_conclusions"]
    add_table_sheet(workbook, "Executive Summary", "Sprint 92A.3 Corrected Multi-Asset Replay v2",
        "Authoritative frozen replay; eight independent $10,000 accounts. All values source from the v2 JSON.", key_value_rows({
            "schema_version": data["schema_version"], "acceptance_status": data["acceptance_status"],
            "generated_as_of": data["generated_as_of"], "combined_starting_balance": combined["starting_balance"],
            "combined_ending_balance": combined["ending_balance"], "combined_net_profit": combined["net_profit"],
            "combined_return_percent": combined["return_percent"], "combined_profit_factor": combined["profit_factor"],
            "closed_trades": combined["closed_trades"], "profitable_symbols": conclusions["profitable_symbols"],
            "losing_symbols": conclusions["losing_symbols"], "xauusd_only_profitable": conclusions["xauusd_only_profitable"],
            "real_orders_sent": data["diagnostics"]["real_orders_sent"], "strategy_parameters_changed": data["diagnostics"]["strategy_parameters_changed"],
        }))
    add_table_sheet(workbook, "V1 vs V2", "Sprint 91 v1 vs Corrected v2",
        "Trade selection, monetary valuation, and risk-sizing/rejection effects are distinguished explicitly.", data["v1_vs_v2"])
    asset_headers = [
        "canonical_symbol", "broker_symbol", "asset_class", "account_currency", "currency_profit",
        "source_candles", "decisions", "buy_signals", "sell_signals", "wait_results",
        "opened_trades", "closed_trades", "unresolved_trades", "rejected_trades",
        "minimum_volume_rejections", "starting_balance", "ending_balance", "return_percent",
        "winners", "losers", "win_rate_percent", "gross_profit", "gross_loss", "net_profit",
        "profit_factor", "expectancy", "average_r", "median_r", "maximum_drawdown",
        "maximum_drawdown_percent", "maximum_consecutive_wins", "maximum_consecutive_losses",
        "average_holding_minutes",
    ]
    asset_sheet, asset_headers, asset_last = add_table_sheet(workbook, "Asset Performance", "Corrected Per-Asset Performance",
        "Canonical and broker identities remain separate; current tick_value is reference metadata only.", data["per_symbol_results"], asset_headers)
    net_col = get_column_letter(asset_headers.index("net_profit") + 1)
    asset_sheet.conditional_formatting.add(f"{net_col}5:{net_col}{asset_last}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    asset_sheet.conditional_formatting.add(f"{net_col}5:{net_col}{asset_last}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=PALE_RED)))
    add_table_sheet(workbook, "Asset Class Performance", "Independent-Account Asset-Class Aggregates",
        "FOREX, METAL, and CRYPTO are sums of independent accounts, not shared-capital allocations.", data["asset_class_results"])
    trade_headers = [
        "canonical_symbol", "broker_symbol", "asset_class", "trade_id", "timeframe", "direction",
        "signal_time", "entry_time", "entry_price", "stop_loss", "take_profit", "exit_time",
        "exit_price", "exit_reason", "spread", "commission", "slippage", "volume", "profit",
        "r_multiple", "status", "entry_conversion_factor", "entry_conversion_time",
        "entry_conversion_path", "exit_conversion_factor", "exit_conversion_time",
        "exit_conversion_path", "account_currency_stop_risk",
    ]
    add_table_sheet(workbook, "Trades", "Frozen Corrected Trade Dataset",
        "Entry and exit historical conversion evidence is retained for each accepted trade.", data["trades"], trade_headers)
    curves = []
    for row in data["per_symbol_results"]:
        curves.extend({"canonical_symbol": row["canonical_symbol"], "trade_index": point[0], "equity": point[1]} for point in row["equity_curve"])
    add_table_sheet(workbook, "Equity Curves", "Independent Symbol Equity Curves",
        "Each symbol starts at $10,000; these are not shared-capital portfolio curves.", curves)
    risk_rows = []
    result_by_symbol = {row["canonical_symbol"]: row for row in data["per_symbol_results"]}
    for row in data["risk_audit"]:
        risk_rows.append({**row, "minimum_volume_rejection_count": result_by_symbol[row["canonical_symbol"]]["minimum_volume_rejections"]})
    add_table_sheet(workbook, "Risk Audit", "Losing-Trade Risk Consistency",
        "Threshold counts and accepted stop risks use pre-trade account equity.", risk_rows)
    add_table_sheet(workbook, "Rejections", "Replay Rejection Reasons",
        "Includes minimum-volume and any deterministic valuation rejections.", data["rejections"])
    add_table_sheet(workbook, "Broker Metadata", "Broker Metadata Snapshot",
        "Current tick_value is retained solely as reference metadata.", data["broker_metadata"])
    fx_rows = []
    for row in data["historical_fx_conversion"]:
        base = {key: value for key, value in row.items() if key != "sample_conversions"}
        samples = row["sample_conversions"] or [{}]
        fx_rows.extend({**base, **sample} for sample in samples)
    add_table_sheet(workbook, "Historical FX Conversion", "Historical Account-Currency Conversion Audit",
        "Latest completed conversion candle at or before valuation time; no current-rate fallback.", fx_rows)
    add_table_sheet(workbook, "Configuration", "Frozen Sprint 91 Configuration",
        "No parameter optimization or production behavior change.", key_value_rows(data["configuration"]))
    add_table_sheet(workbook, "Source Windows", "Exact Sprint 91 Frozen Source Windows",
        "Saved Sprint 91 timestamps are authoritative; history was not shifted, expanded, or imputed.", data["source_windows"])
    add_table_sheet(workbook, "Data Quality", "Frozen Source Data Quality",
        "Known market gaps are preserved and no intervals were filled.", data["data_quality"])
    add_table_sheet(workbook, "Diagnostics", "Replay Diagnostics",
        "One full replay only; workbook values originate exclusively from the frozen v2 JSON.", key_value_rows(data["diagnostics"]))

    audit_rows = key_value_rows({**data["audit"], **{f"acceptance.{key}": value for key, value in data["acceptance"].items()}})
    audit_sheet, _, audit_last = add_table_sheet(workbook, "Audit", "Lineage, Determinism, and Reconciliation Audit",
        "Formula checks reconcile the workbook tables to frozen JSON control totals.", audit_rows)
    start = audit_last + 3
    checks = [
        ("Combined starting capital", "=SUM('Asset Performance'!P5:P12)", combined["starting_balance"], "Sum of eight independent starts"),
        ("Combined ending capital", "=SUM('Asset Performance'!Q5:Q12)", combined["ending_balance"], "Sum of per-symbol ending balances"),
        ("Combined net PnL", "=SUM('Asset Performance'!X5:X12)", combined["net_profit"], "Sum of per-symbol net profits"),
        ("Opened trade count", "=SUM('Asset Performance'!K5:K12)", combined["opened_trades"], "All accepted entries"),
        ("Closed trade count", "=SUM('Asset Performance'!L5:L12)", combined["closed_trades"], "All closed trades"),
        ("Unresolved trade count", "=SUM('Asset Performance'!M5:M12)", combined["unresolved_trades"], "All unresolved trades"),
        ("Winners plus losers", "=SUM('Asset Performance'!S5:S12)+SUM('Asset Performance'!T5:T12)", combined["closed_trades"], "Outcomes reconcile to closed trades"),
        ("Per-symbol ending balance", "=SUMPRODUCT(ABS('Asset Performance'!Q5:Q12-'Asset Performance'!P5:P12-'Asset Performance'!X5:X12))", 0.0, "Each ending balance equals start plus net PnL"),
        ("Combined balance equation", "='Executive Summary'!B9-'Executive Summary'!B8-'Executive Summary'!B10", 0.0, "Ending minus starting minus net PnL"),
    ]
    audit_sheet.cell(start, 1, "reconciliation_check")
    audit_sheet.cell(start, 2, "actual_formula")
    audit_sheet.cell(start, 3, "expected_from_json")
    audit_sheet.cell(start, 4, "difference")
    audit_sheet.cell(start, 5, "status")
    audit_sheet.cell(start, 6, "notes")
    for cell in audit_sheet[start]:
        cell.fill = PatternFill("solid", fgColor=BLUE); cell.font = Font(bold=True, color="FFFFFF")
    for offset, (label, formula, expected, notes) in enumerate(checks, start=1):
        row = start + offset
        audit_sheet.cell(row, 1, label); audit_sheet.cell(row, 2, formula); audit_sheet.cell(row, 3, expected)
        audit_sheet.cell(row, 4, f"=B{row}-C{row}"); audit_sheet.cell(row, 5, f'=IF(ABS(D{row})<0.005,"PASS","FAIL")'); audit_sheet.cell(row, 6, notes)
        audit_sheet.cell(row, 2).number_format = '#,##0.00####;[Red](#,##0.00####);-'
        audit_sheet.cell(row, 3).number_format = '#,##0.00####;[Red](#,##0.00####);-'
        audit_sheet.cell(row, 4).number_format = '#,##0.00####;[Red](#,##0.00####);-'
    recon_table = Table(displayName="ReconciliationChecksTable", ref=f"A{start}:F{start + len(checks)}")
    recon_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    audit_sheet.add_table(recon_table)
    audit_sheet.conditional_formatting.add(f"E{start + 1}:E{start + len(checks)}", CellIsRule(operator="equal", formula=['"PASS"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    for column, width in enumerate((31, 58, 23, 18, 12, 46), start=1):
        audit_sheet.column_dimensions[get_column_letter(column)].width = width

    assert tuple(workbook.sheetnames) == SHEETS
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    normalize(output_path)
    return output_path


if __name__ == "__main__":
    input_file = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_file = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    print(build(input_file, output_file))
