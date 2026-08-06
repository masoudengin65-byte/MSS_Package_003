"""Sprint 82: unchanged replay plus shadow-score distribution report."""

import json
from pathlib import Path

from openpyxl import load_workbook

from run_mt5_context_backtest import main as run_existing_replay
from mss.analysis.shadow_score_report_engine import ShadowScoreReportEngine


SOURCE = Path("reports/MSS_Historical_Backtest_Context_v1.xlsx")
OUTPUT = Path("reports/MSS_Shadow_Score_Distribution.xlsx")


def _legacy_rows(path):
    if not path.exists(): return None
    sheet = load_workbook(path, read_only=True, data_only=True)["Trades"]
    headers = [cell.value for cell in next(sheet.iter_rows())]
    keep = ["Trade ID", "Symbol", "Direction", "Signal Time", "Entry Time", "Entry Price", "Stop Loss", "Take Profit", "Exit Time", "Exit Price", "Exit Reason", "Profit/Loss", "R Multiple", "Score", "Confidence", "Status", "Detector and Context States"]
    indices = [headers.index(name) for name in keep]
    return [tuple(row[index].value for index in indices) for row in sheet.iter_rows(min_row=2)]


def _result_rows(results):
    return [(
        t.trade_id, t.symbol, t.direction, t.signal_time, t.entry_time, t.entry_price,
        t.stop_loss, t.take_profit, t.exit_time, t.exit_price, t.exit_reason, t.profit,
        t.r_multiple, t.legacy_score, t.legacy_confidence, t.status,
        json.dumps(t.detector_states, sort_keys=True),
    ) for result in results for t in result.trades]


def main():
    baseline = _legacy_rows(SOURCE)
    results, audit = run_existing_replay()
    unchanged = baseline is None or baseline == _result_rows(results)
    deterministic = bool(audit["deterministic_repeated_output"])
    if not unchanged: raise RuntimeError("Legacy outcomes, detectors, or decisions changed")
    report = ShadowScoreReportEngine().build(results, OUTPUT, deterministic, unchanged)
    trades = [t for result in results for t in result.trades]
    if not all(ShadowScoreReportEngine.reconciles(t) for t in trades): raise RuntimeError("Shadow component reconciliation failed")
    if len({t.shadow_score for t in trades}) <= 1: raise RuntimeError("Shadow score has no variation")
    if len({t.shadow_confidence for t in trades}) <= 1: raise RuntimeError("Shadow confidence has no variation")
    print("OUTPUT", report)
    print("UNIQUE_SHADOW_SCORES", len({t.shadow_score for t in trades}))
    print("UNIQUE_SHADOW_CONFIDENCE", len({t.shadow_confidence for t in trades}))
    print("LEGACY_UNCHANGED", unchanged)
    return results


if __name__ == "__main__": main()
