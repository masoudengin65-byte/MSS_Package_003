from datetime import datetime
from types import SimpleNamespace

from openpyxl import load_workbook

from mss.analysis.bos_detector import BOS
from mss.analysis.execution_pipeline import ExecutionPipeline
from mss.analysis.excel_report_engine import ExcelReportEngine
from mss.analysis.news_risk_filter import NewsRiskFilter
from mss.analysis.performance_analyzer import PerformanceAnalyzer
from mss.analysis.portfolio_exposure_engine import PortfolioExposureEngine
from mss.analysis.position_manager import PositionManager
from mss.analysis.risk_engine import RiskEngine
from mss.analysis.session_bias_engine import SessionBiasEngine
from mss.analysis.strategy_optimizer import StrategyOptimizer
from mss.analysis.structure_engine import StructureEngine
from mss.domain.analysis_result import AnalysisResult
from mss.domain.economic_event import EconomicEvent
from mss.domain.kill_zone_status import KillZoneStatus
from mss.domain.optimization_result import OptimizationCase
from mss.domain.pipeline_result import PipelineResult
from mss.domain.portfolio_exposure import PortfolioExposure
from mss.domain.premium_discount import PremiumDiscount
from mss.domain.report import Report
from mss.domain.trade_setup import TradeSetup
from mss.domain.candle import Candle
from mss.domain.order_block import OrderBlock
from mss.domain.fair_value_gap import FairValueGap
from mss.engine.signal_engine import SignalEngine


def test_personal_edition_complete_paper_flow(tmp_path):
    now = datetime(2026, 8, 5, 14, 0)
    news = NewsRiskFilter().evaluate(
        now,
        [
            EconomicEvent(
                name="US Nonfarm Payrolls",
                scheduled_at=datetime(2026, 8, 5, 16, 0),
                impact="HIGH",
                currency="USD",
            )
        ],
    )
    assert news.trading_status == "ALLOWED"

    premium_discount = PremiumDiscount(
        current_zone="DISCOUNT",
        valid=True,
    )
    kill_zone = KillZoneStatus(
        broker_time=now,
        current_session="NEWYORK",
        active_kill_zone="NEWYORK_OPEN",
        active=True,
        valid=True,
    )
    session_bias = SessionBiasEngine().calculate(
        [
            PipelineResult(valid=True, structure_state="UPTREND"),
            PipelineResult(
                valid=True,
                structure_state="UPTREND",
                bos_detected=True,
                bos_direction="BULLISH",
            ),
        ],
        premium_discount,
        kill_zone,
    )
    assert session_bias.bias == "Bullish"

    analysis = AnalysisResult(
        bos=BOS(
            direction="BULLISH",
            broken_level=1.1000,
            break_price=1.1010,
            break_time=now,
            reference_index=10,
        ),
        premium_discount=premium_discount,
    )
    decision = SignalEngine().generate(
        SimpleNamespace(
            analysis=analysis,
            kill_zone_status=kill_zone,
            session_bias=session_bias,
        )
    )
    assert decision.signal == "BULLISH"

    portfolio = PortfolioExposureEngine().calculate([])
    assert portfolio.risk_level == "LOW"

    setup = TradeSetup(
        direction="BUY",
        entry=1.1000,
        stop_loss=1.0950,
        take_profit_1=1.1100,
        take_profit_2=1.1150,
        risk=0.005,
        reward=0.015,
        rr=3.0,
        valid=True,
    )
    risk = RiskEngine().calculate(
        balance=10000,
        risk_percent=1,
        stop_distance=0.005,
        news_risk_status=news,
        portfolio_exposure=portfolio,
    )
    assert risk.valid

    order, paper_position = ExecutionPipeline().execute(
        symbol="EURUSD",
        trade_setup=setup,
        account_balance=10000,
        risk_percent=1,
        ticket=75,
        news_risk_status=news,
        portfolio_exposure=portfolio,
    )
    assert order.valid
    assert paper_position.valid

    closed_position = PositionManager().close_position(
        paper_position,
        close_price=1.1100,
        profit=200.0,
    )
    performance = PerformanceAnalyzer().calculate([closed_position])
    assert performance.valid
    assert performance.net_profit == 200.0

    optimization = StrategyOptimizer().optimize([
        OptimizationCase(
            parameters={"risk_percent": 1.0, "rr": 3.0},
            profit=performance.net_profit,
            drawdown=performance.max_drawdown,
            win_rate=performance.win_rate,
            profit_factor=performance.profit_factor,
        )
    ])
    assert optimization.valid

    report = Report(
        processed_candles=100,
        generated_signals=1,
        executed_trades=1,
        premium_discount=premium_discount,
        kill_zone_status=kill_zone,
        session_bias=session_bias,
        news_risk_status=news,
        portfolio_exposure=portfolio,
        risk_profile=risk,
        optimization_result=optimization,
        paper_positions=[closed_position],
        final_decision="BUY",
        decision_reason="PAPER_TRADE_OPENED",
        statistics=performance,
        title="MSS PERSONAL EDITION v1.0",
        valid=True,
    )
    filename = ExcelReportEngine().build(report, output_folder=tmp_path)
    workbook = load_workbook(filename)

    assert workbook["Summary"]["E41"].value == "BUY"
    assert workbook["Summary"]["E34"].value == "ALLOWED"
    assert workbook["Paper Trades"].max_row == 2


def test_structure_engine_connects_completed_ob_and_fvg_detectors():
    engine = StructureEngine()
    calls = []

    def detect_order_block(context, analysis):
        calls.append("ORDER_BLOCK")
        return OrderBlock(direction="BULLISH", valid=True)

    def detect_fvg(context, analysis):
        calls.append("FVG")
        assert analysis.order_block.valid
        return FairValueGap(direction="BULLISH", valid=True)

    engine.order_block_detector.detect = detect_order_block
    engine.fvg_detector.detect = detect_fvg
    candle = Candle(
        time=datetime(2026, 8, 5, 14, 0),
        open=1.1,
        high=1.2,
        low=1.0,
        close=1.15,
        tick_volume=1,
        spread=0,
        real_volume=1,
    )

    analysis = engine.analyze("EURUSD", "M1", [candle])

    assert calls == ["ORDER_BLOCK", "FVG"]
    assert analysis.order_block.valid
    assert analysis.fair_value_gap.valid
